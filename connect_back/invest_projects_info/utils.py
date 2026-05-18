from django.db.models import Sum, F, Q, Case, When, Value, DecimalField
from django.http import Http404
from django.core.exceptions import ObjectDoesNotExist

import openpyxl
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side

from common.humanize import get_humanized_month, get_humanized_month_year

from common.openpyxl_utils import get_height_for_row, factor_of_font_size_to_width
from common.accounting_catalogs.models import KATOCodesModel

from math import ceil
from datetime import datetime
from bpms.workgroups.serializers import WorkgroupCreateSerializer
from bpms.workgroups.models import (WorkgroupMembershipRole,
                                    WorkgroupMembershipStatus,
                                    WorkgroupMembersModel)

from . import models, humanize


def get_file_from_invest_projects(invest_projects):
    workbook = openpyxl.load_workbook('invest_animals.xlsx')

    common_font = Font(
        name='Arial',
        size=14,
        bold=False,
        italic=False,
        vertAlign=None,
        underline='none',
        strike=False,
        color='00000000'
    )
    footer_font = Font(
        name='Arial',
        size=14,
        bold=True,
        italic=False,
        vertAlign=None,
        underline='none',
        strike=False,
        color='00000000'
    )
    footer_fill = PatternFill(
        start_color='FFFF00',
        fill_type='solid',
    )

    center_aligment = Alignment(
        horizontal='center',
        vertical='center',
        wrap_text=True,
        shrink_to_fit=True,
        indent=0
    )
    thin_border = Border(
        left=Side(border_style='thin', color='00000000'),
        right=Side(border_style='thin', color='00000000'),
        top=Side(border_style='thin', color='00000000'),
        bottom=Side(border_style='thin', color='00000000'),
    )
    common_style = NamedStyle(name='common_style')
    common_style.font = common_font
    common_style.alignment = center_aligment
    common_style.border = thin_border

    footer_style = NamedStyle(name='footer_style')
    footer_style.font = footer_font
    footer_style.alignment = center_aligment
    footer_style.border = thin_border
    footer_style.fill = footer_fill

    sheet = workbook.active
    row = 18
    row_count = 1
    aggr = invest_projects.aggregate(
        funds_sum=Sum('funds'),
        own_funds_sum=Sum('own_funds'),
        borrowed_funds_sum=Sum('borrowed_funds'),
        bank_funds_sum=Sum('bank_funds'),
        project_capacity_sum=Sum('project_capacity')
    )
    for each in invest_projects:
        sheet.cell(row=row, column=1, value=row_count).style = common_style
        sheet.cell(row=row, column=2, value=each.district.name).style = common_style
        sheet.cell(row=row, column=3, value=each.company_name).style = common_style
        sheet.cell(row=row, column=4, value=each.project_name).style = common_style
        sheet.cell(row=row, column=5, value=each.direction.name).style = common_style
        sheet.cell(row=row, column=6, value=each.types_of_products).style = common_style
        sheet.cell(row=row, column=7, value=each.company_director_name).style = common_style
        sheet.cell(row=row, column=8, value=each.company_phone).style = common_style
        sheet.cell(row=row, column=9, value='новое' if each.is_new else 'расширение').style = common_style
        sheet.cell(
            row=row, column=10, value=each.foreign_investor_info if each.foreign_investor_info else 'нет'
        ).style = common_style
        sheet.cell(row=row, column=11, value=each.funds).style = common_style
        sheet.cell(row=row, column=12, value=each.own_funds).style = common_style
        sheet.cell(row=row, column=13, value=each.borrowed_funds).style = common_style
        sheet.cell(row=row, column=14, value=each.bank_funds).style = common_style
        sheet.cell(
            row=row, column=15, value='разработана' if each.has_documentation else 'не разработана'
        ).style = common_style
        sheet.cell(row=row, column=16, value=each.installation_stage).style = common_style
        sheet.cell(row=row, column=17, value=each.infrastructure_info).style = common_style
        sheet.cell(row=row, column=18, value=each.cadaster).style = common_style
        sheet.cell(row=row, column=19, value=each.jobs_permanent).style = common_style
        sheet.cell(row=row, column=20, value=each.jobs_temporary).style = common_style
        sheet.cell(row=row, column=21, value=each.project_capacity).style = common_style
        purchases = list(each.equipment_purchases.all().order_by('created_at'))
        for purchase in purchases:
            sheet.cell(row=row, column=22, value=purchase.country.name).style = common_style
            sheet.cell(row=row, column=23, value=purchase.supplier_name).style = common_style
            sheet.cell(row=row, column=24, value=get_humanized_month_year(purchase.delivery_date)).style = common_style
            row += 1
        if purchases:
            current_row = row - len(purchases)
        else:
            current_row = row
        sheet.cell(row=current_row, column=25, value=each.plowed_field_quantity).style = common_style
        sheet.cell(row=current_row, column=26, value=each.pasture_quantity).style = common_style
        sheet.cell(row=current_row, column=27, value=each.questions if each.questions else 'нет').style = common_style
        sheet.cell(row=current_row, column=28, value=get_humanized_month_year(each.dead_line)).style = common_style
        for i in range(current_row, row):
            sheet.row_dimensions[i].height = get_height_for_row(sheet, i-1)
        if purchases:
            end_row = row - 1
            sheet.merge_cells(start_row=current_row, end_row=end_row, start_column=1, end_column=1)
            sheet.merge_cells(start_row=current_row, end_row=end_row, start_column=2, end_column=2)
            sheet.merge_cells(start_row=current_row, end_row=end_row, start_column=3, end_column=3)
            sheet.merge_cells(start_row=current_row, end_row=end_row, start_column=4, end_column=4)
            sheet.merge_cells(start_row=current_row, end_row=end_row, start_column=5, end_column=5)
            sheet.merge_cells(start_row=current_row, end_row=end_row, start_column=6, end_column=6)
            sheet.merge_cells(start_row=current_row, end_row=end_row, start_column=7, end_column=7)
            sheet.merge_cells(start_row=current_row, end_row=end_row, start_column=8, end_column=8)
            sheet.merge_cells(start_row=current_row, end_row=end_row, start_column=9, end_column=9)
            sheet.merge_cells(start_row=current_row, end_row=end_row, start_column=10, end_column=10)
            sheet.merge_cells(start_row=current_row, end_row=end_row, start_column=11, end_column=11)
            sheet.merge_cells(start_row=current_row, end_row=end_row, start_column=12, end_column=12)
            sheet.merge_cells(start_row=current_row, end_row=end_row, start_column=13, end_column=13)
            sheet.merge_cells(start_row=current_row, end_row=end_row, start_column=14, end_column=14)
            sheet.merge_cells(start_row=current_row, end_row=end_row, start_column=15, end_column=15)
            sheet.merge_cells(start_row=current_row, end_row=end_row, start_column=16, end_column=16)
            sheet.merge_cells(start_row=current_row, end_row=end_row, start_column=17, end_column=17)
            sheet.merge_cells(start_row=current_row, end_row=end_row, start_column=18, end_column=18)
            sheet.merge_cells(start_row=current_row, end_row=end_row, start_column=19, end_column=19)
            sheet.merge_cells(start_row=current_row, end_row=end_row, start_column=20, end_column=20)
            sheet.merge_cells(start_row=current_row, end_row=end_row, start_column=21, end_column=21)
            sheet.merge_cells(start_row=current_row, end_row=end_row, start_column=25, end_column=25)
            sheet.merge_cells(start_row=current_row, end_row=end_row, start_column=26, end_column=26)
            sheet.merge_cells(start_row=current_row, end_row=end_row, start_column=27, end_column=27)
            sheet.merge_cells(start_row=current_row, end_row=end_row, start_column=28, end_column=28)
        else:
            row += 1
        row_count += 1
    sheet.cell(row=row, column=1).style = footer_style
    sheet.cell(row=row, column=2).style = footer_style
    sheet.cell(row=row, column=3).style = footer_style
    sheet.cell(row=row, column=4).style = footer_style
    sheet.cell(row=row, column=5).style = footer_style
    sheet.cell(row=row, column=6).style = footer_style
    sheet.cell(row=row, column=7).style = footer_style
    sheet.cell(row=row, column=8).style = footer_style
    sheet.cell(row=row, column=9).style = footer_style
    sheet.cell(row=row, column=10).style = footer_style
    sheet.cell(row=row, column=11, value=aggr.get('funds_sum')).style = footer_style
    sheet.cell(row=row, column=12, value=aggr.get('own_funds_sum')).style = footer_style
    sheet.cell(row=row, column=13, value=aggr.get('borrowed_funds_sum')).style = footer_style
    sheet.cell(row=row, column=14, value=aggr.get('bank_funds_sum')).style = footer_style
    sheet.cell(row=row, column=15).style = footer_style
    sheet.cell(row=row, column=16).style = footer_style
    sheet.cell(row=row, column=17).style = footer_style
    sheet.cell(row=row, column=18).style = footer_style
    sheet.cell(row=row, column=19).style = footer_style
    sheet.cell(row=row, column=20).style = footer_style
    sheet.cell(row=row, column=21, value=aggr.get('project_capacity_sum')).style = footer_style
    sheet.cell(row=row, column=22).style = footer_style
    sheet.cell(row=row, column=23).style = footer_style
    sheet.cell(row=row, column=24).style = footer_style
    sheet.cell(row=row, column=25).style = footer_style
    sheet.cell(row=row, column=26).style = footer_style
    sheet.cell(row=row, column=27).style = footer_style
    sheet.cell(row=row, column=28).style = footer_style
    sheet.row_dimensions[row].height = get_height_for_row(sheet, row-1)
    sheet.merge_cells(start_row=row, end_row=row, start_column=2, end_column=9)
    return workbook


measure_unit_map = {'tonne': 10, 'heads': 11, 'hectares': 12}


def get_file_roadmap_from_invest_projects_2(invest_projects):
    invest_projects = invest_projects.annotate(
        common_funds=F('own_funds') + F('borrowed_funds') + F('bank_funds')
    )
    workbook = openpyxl.load_workbook('invest_roadmap.xlsx')
    sheet = workbook.active
    row = 12
    industries = models.InvestProjectIndustryModel.objects.filter(
        is_active=True,
        invest_projects_info__in=invest_projects
    ).distinct()

    title_font = Font(
        name='Arial',
        size=16,
        bold=True,
        italic=False,
        vertAlign=None,
        underline='none',
        strike=False,
        color='00000000'
    )
    common_font = Font(
        name='Arial',
        size=13,
        bold=False,
        italic=False,
        vertAlign=None,
        underline='none',
        strike=False,
        color='00000000',
    )
    title_fill = PatternFill(
        start_color='FFFF00',
        fill_type='solid',
    )
    center_alignment = Alignment(
        horizontal='center',
        vertical='center',
        wrap_text=True,
        shrink_to_fit=True,
        indent=0
    )
    thin_border = Border(
        left=Side(border_style='thin', color='00000000'),
        right=Side(border_style='thin', color='00000000'),
        top=Side(border_style='thin', color='00000000'),
        bottom=Side(border_style='thin', color='00000000'),
    )
    title_style = NamedStyle(name='title_style')
    title_style.font = title_font
    title_style.alignment = center_alignment
    title_style.border = thin_border
    title_style.fill = title_fill

    common_style = NamedStyle(name='common_style')
    common_style.font = common_font
    common_style.alignment = center_alignment
    common_style.border = thin_border

    for industry in industries:
        industry_projects = invest_projects.filter(industry=industry)
        industry_projects_aggregate = industry_projects.aggregate(
            common_funds_sum=Sum('common_funds'),
            own_funds_sum=Sum('own_funds'),
            funds_sum=Sum('funds'),
        )
        # Заголовочная строка
        projects_count = invest_projects.filter(industry=industry).count()
        sheet.cell(row=row, column=1,).style = title_style
        sheet.cell(row=row, column=2,).style = title_style
        sheet.cell(
            row=row, column=3, value=f"{projects_count} {humanize.get_humanized_enumeration_project(projects_count)} "
                                     f"по {industry.dative_case}".upper()
        ).style = title_style
        # sheet.cell(row=row, column=4, ).style = title_style
        # sheet.cell(row=row, column=5, ).style = title_style
        # sheet.cell(row=row, column=6, ).style = title_style
        sheet.cell(row=row, column=7, value=industry_projects_aggregate['common_funds_sum']).style = title_style
        sheet.cell(row=row, column=8, value=industry_projects_aggregate['own_funds_sum']).style = title_style
        sheet.cell(row=row, column=9, value=industry_projects_aggregate['funds_sum']).style = title_style
        sheet.cell(row=row, column=10, ).style = title_style
        sheet.cell(row=row, column=11, ).style = title_style
        sheet.cell(row=row, column=12, ).style = title_style
        sheet.cell(row=row, column=13, ).style = title_style
        sheet.cell(row=row, column=14, ).style = title_style
        sheet.cell(row=row, column=15, ).style = title_style
        sheet.cell(row=row, column=16, ).style = title_style
        sheet.cell(row=row, column=17, ).style = title_style
        sheet.cell(row=row, column=18, ).style = title_style
        sheet.cell(row=row, column=19, ).style = title_style
        sheet.cell(row=row, column=20, ).style = title_style

        # Объединяем ячейки в заголовочной строке с переносом строк:
        merged_cells_width = sheet.column_dimensions[sheet.cell(row=row, column=3).column_letter].width + \
                             sheet.column_dimensions[sheet.cell(row=row, column=4).column_letter].width + \
                             sheet.column_dimensions[sheet.cell(row=row, column=5).column_letter].width + \
                             sheet.column_dimensions[sheet.cell(row=row, column=6).column_letter].width
        font_params = factor_of_font_size_to_width[16]
        words_count_at_one_row = merged_cells_width / font_params.get('factor')
        lines = ceil(len(str(sheet.cell(row=row, column=3).value)) / words_count_at_one_row)
        sheet.row_dimensions[row].height = max(20, lines * font_params.get('height'))
        sheet.merge_cells(start_row=row, end_row=row, start_column=3, end_column=6)

        # Список проектов по отрасли:
        row += 1
        row_count = 1
        for industry_project in industry_projects:
            sheet.cell(row=row, column=1, value=row_count).style = common_style
            sheet.cell(row=row, column=2, value=industry_project.district.region.name).style = common_style
            sheet.cell(row=row, column=3, value=f'Наименование: {industry_project.project_name}\n'
                                                f'БИН: {industry_project.company_bin}\n'
                                                f'Учредитель: {industry_project.company_director_name}\n'
                                                f'Контакты: {industry_project.company_phone}').style = common_style
            sheet.cell(row=row, column=4, value=industry_project.direction.name).style = common_style
            sheet.cell(row=row, column=5, value=industry_project.types_of_products).style = common_style
            sheet.cell(row=row, column=6, value=industry_project.fin_institute).style = common_style
            sheet.cell(
                row=row,
                column=7,
                value=industry_project.own_funds + industry_project.borrowed_funds + industry_project.bank_funds
            ).style = common_style  # TODO Сделал как понял
            sheet.cell(row=row, column=8, value=industry_project.own_funds).style = common_style
            sheet.cell(row=row, column=9, value=industry_project.funds).style = common_style
            measure_unit_column = measure_unit_map.get(industry_project.measure_unit_id, 10)
            sheet.cell(
                row=row,
                column=10,
                value=industry_project.project_capacity if measure_unit_column == 10 else ''
            ).style = common_style
            sheet.cell(
                row=row,
                column=11,
                value=industry_project.project_capacity if measure_unit_column == 11 else ''
            ).style = common_style
            sheet.cell(
                row=row,
                column=12,
                value=industry_project.project_capacity if measure_unit_column == 12 else ''
            ).style = common_style
            sheet.cell(row=row, column=13, value=industry_project.jobs_temporary).style = common_style
            sheet.cell(row=row, column=14, value=industry_project.jobs_permanent).style = common_style
            sheet.cell(
                row=row, column=15, value=get_humanized_month(industry_project.dead_line.month)
            ).style = common_style
            sheet.cell(row=row, column=16, value=industry_project.dead_line.year).style = common_style
            sheet.cell(
                row=row, column=17, value='есть' if industry_project.has_documentation else 'нет'
            ).style = common_style

            sheet.cell(
                row=row, column=18, value=industry_project.land_plot
            ).style = common_style

            sheet.cell(row=row, column=19, value=industry_project.work_experience).style = common_style

            sheet.cell(row=row, column=20, value='').style = common_style  # TODO Текущий статус?
            sheet.row_dimensions[row].height = get_height_for_row(sheet, row - 1, font_size=13)
            row += 1
            row_count += 1

    # Подвал:
    sheet.cell(row=row, column=1).style = title_style
    total_count = invest_projects.count()
    sheet.cell(
        row=row,
        column=2,
        value=f'итого {total_count} {humanize.get_humanized_enumeration_project(total_count)}'.upper()
    ).style = title_style
    sheet.cell(row=row, column=3).style = title_style
    sheet.cell(row=row, column=4).style = title_style
    sheet.cell(row=row, column=5).style = title_style
    sheet.cell(row=row, column=6).style = title_style
    invest_projects_aggregate = invest_projects.aggregate(
        common_funds_sum=Sum('common_funds'),
        own_funds_sum=Sum('own_funds'),
        funds_sum=Sum('funds'),
    )  # TODO сделал как понял
    sheet.cell(row=row, column=7, value=invest_projects_aggregate['common_funds_sum']).style = title_style
    sheet.cell(row=row, column=8, value=invest_projects_aggregate['own_funds_sum']).style = title_style
    sheet.cell(row=row, column=9, value=invest_projects_aggregate['funds_sum']).style = title_style
    sheet.cell(row=row, column=10).style = title_style
    sheet.cell(row=row, column=11).style = title_style
    sheet.cell(row=row, column=12).style = title_style
    sheet.cell(row=row, column=13).style = title_style
    sheet.cell(row=row, column=14).style = title_style
    sheet.cell(row=row, column=15).style = title_style
    sheet.cell(row=row, column=16).style = title_style
    sheet.cell(row=row, column=17).style = title_style
    sheet.cell(row=row, column=18).style = title_style
    sheet.cell(row=row, column=19).style = title_style
    sheet.cell(row=row, column=20).style = title_style
    sheet.row_dimensions[row].height = 24
    sheet.merge_cells(start_row=row, end_row=row, start_column=2, end_column=4)
    return workbook


def add_project(instance, request):
    # user=request.user.profile
    description = instance.category.name
    if instance.subcategory:
        description += f', {instance.subcategory.name}'
    project_serializer = WorkgroupCreateSerializer(
        context = {'request': request},
        data={
            'is_project': True,
            'name': instance.project_name,
            'description': description,
            'organization': instance.organization,
            'date_start_plan': datetime.combine(instance.date_start, datetime.min.time()),
            'dead_line': datetime.combine(instance.dead_line, datetime.min.time())
        }
    )
    project_serializer.is_valid(raise_exception=True)
    project = project_serializer.save()
    # member = WorkgroupMembersModel.objects.create(
    #     member=user,
    #     work_group=project,
    #     membership_role=WorkgroupMembershipRole.objects.get(code="FOUNDER"),
    #     membership_request_status=WorkgroupMembershipStatus.objects.get(code="APPROVED")
    # )
    # member.save()
    instance.project = project
    instance.save(update_fields=('project',))


def get_invest_project_creator_organizations(request) -> set:
    """Возвращает множество организаций, где пользователь может создавать инвестпроекты.
    Включает только те организации, где явно назначена эта роль"""
    from contractor_permissions.utils import contractors_where_user_has_permission
    user = request.user.profile
    create_permission_id = models.InvestProjectPermissionTypeModel.objects.get(code='create')
    creator_organizations = contractors_where_user_has_permission(
        user.pk,
        'create_invest_projects_info',
        create_permission_id
        )
    return creator_organizations


def get_invest_project_approver_organizations(request) -> set:
    """Возвращает множество организаций, где пользователь может одобрять инвестпроекты.
    Это организации, где у него есть роль одобрения инвестпроектов, а также НИЖЕСТОЯЩИЕ."""
    from contractor_permissions.utils import contractors_where_user_has_permission
    from users.utils import get_descendants_departments_related_organizations
    user = request.user.profile
    approve_permission_id = models.InvestProjectPermissionTypeModel.objects.get(code='approve')
    approver_organizations = contractors_where_user_has_permission(
        user.pk,
        'create_invest_projects_info',
        approve_permission_id
        )
    approver_organizations = get_descendants_departments_related_organizations(approver_organizations)
    return approver_organizations


def get_funding_statistics(queryset=None, funding_codes=[]) -> dict:
    if queryset is None:
        return {}
    return queryset.aggregate(
        grand_total=Sum('funding_sources__amount'),
        nf=Sum(
            Case(
                When(
                    funding_sources__funding_source__code='nf',
                    then='funding_sources__amount'
                ),
                default=Value(0),
                output_field=DecimalField()
            )
        ),
        rb=Sum(
            Case(
                When(
                    funding_sources__funding_source__code='rb',
                    then='funding_sources__amount'
                ),
                default=Value(0),
                output_field=DecimalField()
            )
        ),
        mb=Sum(
            Case(
                When(
                    funding_sources__funding_source__code='mb',
                    then='funding_sources__amount'
                ),
                default=Value(0),
                output_field=DecimalField()
            )
        ),
        other=Sum(
            Case(
                When(
                    ~Q(funding_sources__funding_source__code__in=funding_codes),
                    then='funding_sources__amount'
                ),
                default=Value(0),
                output_field=DecimalField()
            )
        ),
    )
