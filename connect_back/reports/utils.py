# 🔹 Стандартная библиотека
import calendar
import copy
import io
import ast
import json
import uuid
import re
from datetime import datetime, date, timedelta

# 🔹 Django
from django.conf import settings
from django.contrib.postgres.aggregates import StringAgg
from django.db import models
from django.db.models import (
    Sum, Avg, Count, Min, Max, Q, F,
)
from django.utils import timezone
from django.utils.dateparse import parse_date, parse_datetime
from django.utils.translation import gettext_lazy as _
from django.http import HttpResponse
from django.utils.encoding import force_str

# 🔹 Django REST framework
from rest_framework.exceptions import ValidationError
from rest_framework.response import Response

# 🔹 Третьи библиотеки
from mptt.models import TreeForeignKey
import pandas as pd
import plotly.graph_objects as go
import plotly.io as pio
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import NamedStyle, Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.dimensions import SheetFormatProperties
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_to_tuple


# 🔹 Локальные импорты
from .utils_meta import get_field_meta, get_report_field_type


def build_related_computed_annotations(model_class, requested_names, request):
    """
    Создаёт аннотации для связанных вычисляемых полей, запрошенных в отчёте.
    Пример: 'task__task_link', 'customer_card__work_duration'.
    Когда отчёт строится по основной модели (например по специалистам), подзапросы
    в аннотациях связанной модели (карточки) видят «внешней» строку основной модели,
    поэтому OuterRef('pk') дал бы pk специалиста. Передаём outer_ref_column (например
    customer_card_id), чтобы подзапрос привязался к FK текущей строки.
    """
    from django.db.models import F, OuterRef  # noqa: WPS433 (local import by design)
    from django.db.models.expressions import Expression, Func  # noqa: WPS433

    def prefix_expr(expr: Expression, prefix: str) -> Expression:
        # Рекурсивно префиксируем F('field') → F(f'{prefix}{field}')
        if isinstance(expr, OuterRef):
            # OuterRef указывает на колонку outer-query и уже сформирован целевой моделью.
            # Префиксация здесь ломает пути вида task_id -> task__task_id.
            return expr
        if isinstance(expr, F):
            field_name = expr.name
            if not field_name.startswith(prefix):
                field_name = f"{prefix}{field_name}"
            return F(field_name)
        if isinstance(expr, Func):
            # Функции / агрегаты: префиксим их аргументы
            expr.set_source_expressions([prefix_expr(e, prefix) for e in expr.get_source_expressions()])
            return expr
        if isinstance(expr, Expression):
            # Общее правило для Expression: префиксим дочерние выражения, если есть
            children = expr.get_source_expressions()
            if children:
                expr.set_source_expressions([prefix_expr(e, prefix) for e in children])
            return expr
        # Value, литералы и т.п. префиксации не требуют
        return expr
    
    annotations = {}
    for name in requested_names:
        # интересуют только связанные пути 'a__b__computed' (а не локальные)
        if '__' not in name:
            continue
        prefix_path, computed = name.rsplit('__', 1)

        # Определяем целевую связанную модель по пути до computed
        try:
            related_field = model_class._meta.get_field(prefix_path)
            _, related_model = get_field_meta(prefix_path, model_class)
            outer_ref_column = getattr(related_field, 'attname', None)  # e.g. customer_card_id
        except Exception:
            continue

        get_ann = getattr(related_model, 'get_report_annotations', None)
        if not callable(get_ann):
            continue

        # Вызываем с outer_ref_column, чтобы подзапросы использовали FK основной модели, а не pk
        try:
            rel_annotations = get_ann(request, [computed], outer_ref_column=outer_ref_column)
        except TypeError:
            rel_annotations = get_ann(request, [computed])
        expr = rel_annotations.get(computed) if rel_annotations else None
        if expr is None:
            continue  # целевая модель не объявила такое вычисляемое поле

        prefixed_expr = prefix_expr(expr, f"{prefix_path}__")
        # АЛИАС должeн совпасть с запрошенным именем, чтобы values('a__b__computed') сработал
        annotations[name] = prefixed_expr
        
        # Если вычисляемое поле имеет order_by_field, добавляем его в аннотации для сортировки
        order_by_field = get_computed_order_field(related_model, computed)
        if order_by_field and order_by_field != computed:
            # Добавляем order_by_field как аннотацию с префиксом
            order_by_annotation_name = f"{prefix_path}__{order_by_field}"
            if order_by_annotation_name not in annotations:
                annotations[order_by_annotation_name] = F(f"{prefix_path}__{order_by_field}")

    return annotations


def _get_duration_computed_field_names(model_class, prefix_path=None):
    """Имена вычисляемых полей типа DurationField у модели (с опциональным префиксом для связанной модели)."""
    if not model_class or not hasattr(model_class, "get_report_computed_fields_meta"):
        return set()
    result = set()
    for cf in model_class.get_report_computed_fields_meta():
        if cf.get("type") == "DurationField" and cf.get("name"):
            name = cf["name"]
            if prefix_path:
                name = f"{prefix_path}__{name}"
            result.add(name)
    return result


def _format_duration_seconds(seconds):
    """Форматирует секунды в формат '35:12:13' (часы могут быть больше 24)"""
    if seconds is None:
        return ""
    try:
        total_seconds = int(float(seconds))
    except (ValueError, TypeError):
        return str(seconds) if seconds else ""
    
    sign = '-' if total_seconds < 0 else ''
    total_seconds = abs(total_seconds)
    
    # Вычисляем общее количество часов (может быть больше 24)
    hours = total_seconds // 3600
    rem = total_seconds % 3600
    minutes, secs = divmod(rem, 60)
    
    result = f"{hours}:{minutes:02d}:{secs:02d}"
    
    return f"{sign}{result}" if sign else result


def _is_duration_field(field_name, duration_fields, aggregate_fields_raw):
    """Проверяет, является ли поле duration полем (включая агрегационные)."""
    if not duration_fields:
        return False
    # Прямая проверка
    if field_name in duration_fields:
        return True
    # Проверка через агрегационные поля
    for agg_def in aggregate_fields_raw:
        if agg_def.get("name") == field_name:
            source_field = agg_def.get("sum") or agg_def.get("avg") or agg_def.get("min") or agg_def.get("max")
            if source_field and source_field in duration_fields:
                return True
    return False


def _format_duration_value(value):
    """Форматирует значение, если это число (секунды), иначе возвращает как есть."""
    if value == "" or value is None:
        return value
    # Пропускаем уже отформатированные значения
    if isinstance(value, str) and ("д." in value or ":" in value):
        return value
    try:
        if isinstance(value, (int, float)):
            return _format_duration_seconds(value)
        elif isinstance(value, str):
            val_clean = value.replace('.', '').replace('-', '').replace(' ', '')
            if val_clean.isdigit():
                return _format_duration_seconds(float(value))
    except (ValueError, TypeError):
        pass
    return value


def _build_default_export_styles():
    """
    Стандартные стили для экспорта (используются:
    - как дефолтные при отсутствии шаблона
    - как fallback, если в шаблоне нет какого-то именованного стиля
    """
    thin_side = Side(style="thin")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    report_title_style = NamedStyle(name="report_title")
    report_title_style.font = Font(size=14, bold=True)
    report_title_style.alignment = Alignment(horizontal="left", vertical="center")

    report_footer_style = NamedStyle(name="report_footer")
    report_footer_style.font = Font(size=10, italic=True)
    report_footer_style.alignment = Alignment(horizontal="left", vertical="center")

    filter_header_style = NamedStyle(name="filter_header")
    filter_header_style.font = Font(bold=True)
    filter_header_style.border = thin_border
    filter_header_style.alignment = Alignment(horizontal="center", vertical="center")

    filter_row_style = NamedStyle(name="filter_row")
    filter_row_style.font = Font()
    filter_row_style.border = thin_border
    filter_row_style.alignment = Alignment(horizontal="left", vertical="center")

    table_header_style = NamedStyle(name="table_header")
    table_header_style.font = Font(bold=True)
    table_header_style.border = thin_border
    table_header_style.alignment = Alignment(horizontal="center", vertical="center")

    group_header_level_1_style = NamedStyle(name="group_header_level_1")
    group_header_level_1_style.font = Font(bold=True, color="000080")
    group_header_level_1_style.fill = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")
    group_header_level_1_style.border = thin_border
    group_header_level_1_style.alignment = Alignment(horizontal="left", vertical="center")

    group_header_level_2_style = NamedStyle(name="group_header_level_2")
    group_header_level_2_style.font = Font(bold=True, color="006600")
    group_header_level_2_style.fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")
    group_header_level_2_style.border = thin_border
    group_header_level_2_style.alignment = Alignment(horizontal="left", vertical="center")

    group_header_level_3_style = NamedStyle(name="group_header_level_3")
    group_header_level_3_style.font = Font(bold=True, color="660066")
    group_header_level_3_style.fill = PatternFill(start_color="FFE6FF", end_color="FFE6FF", fill_type="solid")
    group_header_level_3_style.border = thin_border
    group_header_level_3_style.alignment = Alignment(horizontal="left", vertical="center")

    group_footer_level_1_style = NamedStyle(name="group_footer_level_1")
    group_footer_level_1_style.font = Font(bold=True, color="000080")
    group_footer_level_1_style.border = thin_border
    group_footer_level_1_style.alignment = Alignment(horizontal="left", vertical="center")

    group_footer_level_2_style = NamedStyle(name="group_footer_level_2")
    group_footer_level_2_style.font = Font(bold=True, color="006600")
    group_footer_level_2_style.border = thin_border
    group_footer_level_2_style.alignment = Alignment(horizontal="left", vertical="center")

    group_footer_level_3_style = NamedStyle(name="group_footer_level_3")
    group_footer_level_3_style.font = Font(bold=True, color="660066")
    group_footer_level_3_style.border = thin_border
    group_footer_level_3_style.alignment = Alignment(horizontal="left", vertical="center")

    detail_rows_style = NamedStyle(name="detail_rows")
    detail_rows_style.font = Font()
    detail_rows_style.border = thin_border
    detail_rows_style.alignment = Alignment(horizontal="left", vertical="center")

    grand_totals_style = NamedStyle(name="grand_totals")
    grand_totals_style.font = Font(bold=True)
    grand_totals_style.border = thin_border
    grand_totals_style.alignment = Alignment(horizontal="right", vertical="center")

    named_styles_dict = {
        "report_title": report_title_style,
        "report_footer": report_footer_style,
        "filter_header": filter_header_style,
        "filter_row": filter_row_style,
        "table_header": table_header_style,
        "group_header_level_1": group_header_level_1_style,
        "group_header_level_2": group_header_level_2_style,
        "group_header_level_3": group_header_level_3_style,
        "group_footer_level_1": group_footer_level_1_style,
        "group_footer_level_2": group_footer_level_2_style,
        "group_footer_level_3": group_footer_level_3_style,
        "detail_rows": detail_rows_style,
        "grand_totals": grand_totals_style,
    }

    return named_styles_dict, filter_header_style, filter_row_style

def _normalize_cell_value(val):
    """
    Унифицированная нормализация значений для HTML и Excel:
    - np.nan / pandas.NA / None -> "" (пустая строка)
    - строковые "nan"/"None" -> "" (пустая строка)
    Остальные значения возвращаются как есть.
    """
    if val is None:
        return ""
    try:
        # Обрабатываем np.nan / pandas.NA
        if pd.isna(val):
            return ""
    except TypeError:
        # Для некоторых типов (dict и т.п.) pd.isna может выбросить ошибку — игнорируем
        pass
    if isinstance(val, str) and val in ("nan", "None"):
        return ""
    # Булевы значения → Да/Нет
    if isinstance(val, bool):
        return "Да" if val else "Нет"
    # Даты и дата-время как объекты:
    # - date → DD.MM.YYYY
    # - datetime → DD.MM.YYYY HH:MM
    if isinstance(val, datetime):
        return val.strftime("%d.%m.%Y %H:%M")
    if isinstance(val, date):
        return val.strftime("%d.%m.%Y")
    # Строковые представления дат / даты-времени в формате YYYY-MM-DD[ HH:MM[:SS]]
    if isinstance(val, str):
        stripped = val.strip()
        try:
            if re.fullmatch(r"\d{4}-\d{2}-\d{2}$", stripped):
                parsed = datetime.strptime(stripped, "%Y-%m-%d").date()
                return parsed.strftime("%d.%m.%Y")
            if re.fullmatch(r"\d{4}-\d{2}-\d{2} \d{2}:\d{2}$", stripped):
                parsed = datetime.strptime(stripped, "%Y-%m-%d %H:%M")
                return parsed.strftime("%d.%m.%Y %H:%M")
            if re.fullmatch(r"\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}$", stripped):
                parsed = datetime.strptime(stripped, "%Y-%m-%d %H:%M:%S")
                return parsed.strftime("%d.%m.%Y %H:%M")
        except Exception:
            # Если не смогли распарсить — оставляем как есть
            pass
    # Сокращение до 2 знаков после запятой для числовых значений
    if isinstance(val, (float, np.floating)):
        try:
            return round(float(val), 2)
        except Exception:
            return val
    return val


def _prepare_excel_rows_for_repr_only_columns(output_rows, model_class):
    """
    Централизованно подготавливает output_rows для Excel:
    для столбцов LinkListField заменяет список ссылок на строку repr через запятую.
    """
    if not model_class or not hasattr(model_class, 'get_report_computed_fields_meta'):
        return output_rows

    excel_repr_only_columns = {
        field_meta.get("name")
        for field_meta in model_class.get_report_computed_fields_meta()
        if field_meta.get("type") == "LinkListField" and field_meta.get("name")
    }
    if not excel_repr_only_columns:
        return output_rows

    def repr_text_from_link_list(val):
        items = _extract_link_list(val)
        if items is None:
            return None
        return ", ".join(
            force_str(item.get("repr") or item.get("url") or "")
            for item in items
        )

    prepared_rows = []
    for row in output_rows:
        prepared_row = dict(row)
        for column_name in excel_repr_only_columns:
            if column_name not in prepared_row:
                continue
            repr_text = repr_text_from_link_list(prepared_row.get(column_name))
            if repr_text is not None:
                prepared_row[column_name] = repr_text
        prepared_rows.append(prepared_row)
    return prepared_rows


def _extract_link_dict(val):
    """Возвращает dict ссылки {'repr','url'} из dict или строкового repr dict."""
    if isinstance(val, dict) and val.get("url"):
        return val
    if isinstance(val, str) and val.strip().startswith("{") and "'url'" in val:
        try:
            parsed = ast.literal_eval(val)
            if isinstance(parsed, dict) and parsed.get("url"):
                return parsed
        except Exception:
            return None
    return None


def _extract_link_list(val):
    """Возвращает список dict ссылок из list или строкового repr list."""
    if isinstance(val, tuple):
        val = list(val)
    if isinstance(val, np.ndarray):
        try:
            val = val.tolist()
        except Exception:
            return None
    if isinstance(val, list) and all(isinstance(item, dict) for item in val):
        return val
    if isinstance(val, str) and val.strip().startswith("["):
        try:
            parsed = ast.literal_eval(val)
            if isinstance(parsed, list) and all(isinstance(item, dict) for item in parsed):
                return parsed
        except Exception:
            return None
    return None


def _excel_numeric_columns(all_columns, model_class, numeric_aggregate_fields):
    """
    Множество имён столбцов, которые по метаданным являются числовыми
    (модель: IntegerField, FloatField, DecimalField; плюс агрегатные числовые поля).
    """
    numeric = set(numeric_aggregate_fields)
    if not model_class:
        return numeric
    numeric_field_types = (
        models.IntegerField, models.BigIntegerField, models.SmallIntegerField,
        models.PositiveIntegerField, models.PositiveSmallIntegerField, models.PositiveBigIntegerField,
        models.FloatField, models.DecimalField,
    )
    for col in all_columns:
        if col in numeric:
            continue
        model_field, _ = get_field_meta(col, model_class)
        if model_field and isinstance(model_field, numeric_field_types):
            numeric.add(col)
    return numeric


def _apply_style(cell, style_key, named_styles_dict):
    """
    Унифицированное применение стиля:
    - если в named_styles_dict есть NamedStyle → копируем его атрибуты (font/fill/border/alignment/number_format)
    - иначе используем стандартный стиль 'Normal'
    """
    style_obj = named_styles_dict.get(style_key)
    if isinstance(style_obj, NamedStyle):
        cell.font = style_obj.font
        cell.fill = style_obj.fill
        cell.border = style_obj.border
        cell.alignment = style_obj.alignment
        if getattr(style_obj, "number_format", None):
            cell.number_format = style_obj.number_format
    else:
        cell.style = 'Normal'


def _collect_template_placeholders(named_ranges, ws):
    """
    Собирает все плейсхолдеры вида {{ key }} из header/footer шаблона.
    Возвращает множество ключей (без фигурных скобок и пробелов).
    """
    placeholder_keys = set()
    placeholder_re = re.compile(r"\{\{\s*([^}]+?)\s*\}\}")

    for range_name in ("header", "footer"):
        if range_name not in named_ranges:
            continue
        defined_name = named_ranges.get(range_name)
        try:
            destinations = defined_name.destinations
        except AttributeError:
            continue

        for sheetname, coord in destinations:
            if sheetname != ws.title:
                continue

            if ":" in coord:
                cell_iter = ws[coord]
            else:
                cell_iter = [(ws[coord],)]

            for row in cell_iter:
                for cell in row:
                    val = cell.value
                    if not isinstance(val, str):
                        continue
                    for match in placeholder_re.finditer(val):
                        key = match.group(1).strip()
                        if key:
                            placeholder_keys.add(key)

    return placeholder_keys


def _resolve_dynamic_path(instance, path):
    """
    Резолвит ORM-путь вида 'customer_card__org_admin__email' от переданного инстанса.
    Первый сегмент считается полем модели отчёта; далее обычные атрибуты / related managers.
    """
    if instance is None or not path:
        return None

    current = instance
    parts = path.split("__")

    for part in parts:
        if current is None:
            return None
        try:
            current = getattr(current, part)
        except Exception:
            return None

    return current


def enrich_placeholders_from_dynamic_paths(placeholders, dynamic_keys, model_class, base_instance, named_ranges=None, ws=None):
    """
    Обогащает словарь placeholders значениями для динамических плейсхолдеров.
    dynamic_keys — множество ключей, которых ещё нет в placeholders.
    Резолв идёт от base_instance по ORM-пути.
    Если base_instance None или путь не разрешился, добавляет плейсхолдер со значением пустой строки.
    """
    if not dynamic_keys:
        return
    
    if base_instance is None:
        # Если base_instance None, добавляем все плейсхолдеры со значением пустой строки
        for key in dynamic_keys:
            placeholders[key] = ""
        return

    for key in dynamic_keys:
        try:
            value = _resolve_dynamic_path(base_instance, key)
            if value is None:
                # Если путь разрешился в None, используем пустую строку
                placeholders[key] = ""
                continue
        except Exception:
            # Если произошла ошибка при разрешении пути, используем пустую строку
            placeholders[key] = ""
            continue

        # Единая нормализация для всех мест (HTML/Excel/плейсхолдеры)
        normalized = _normalize_cell_value(try_repr(value))
        text = force_str(normalized)

        placeholders[key] = text


def _format_html_cell_link_or_text(val):
    """Если значение — dict с url/repr, отрисуем как ссылку, иначе как текст.
    Для списка таких dict (LinkListField) — несколько ссылок через запятую."""
    def _render_link(item):
        if not isinstance(item, dict) or not item.get("url"):
            return None
        url = item["url"]
        text = _normalize_cell_value(item.get("repr") or url)
        return f'<a href="{url}" target="_blank">{text}</a>'

    # Список ссылок (related_tasks и т.п.)
    def _render_links_list(links):
        if not links or not all(isinstance(item, dict) for item in links):
            return None
        parts = []
        repr_parts = []
        for item in links:
            rendered_link = _render_link(item)
            if rendered_link:
                parts.append(rendered_link)
            else:
                # Если ссылки не отрендерились (нет url), но есть repr — выводим repr через запятую.
                repr_text = item.get("repr")
                if repr_text:
                    repr_parts.append(force_str(_normalize_cell_value(repr_text)))
        if parts:
            return ", ".join(parts)
        if repr_parts:
            return ", ".join(repr_parts)
        return None

    link_list = _extract_link_list(val)
    if link_list is not None:
        rendered = _render_links_list(link_list)
        if rendered is not None:
            return rendered
        return ""
    link_dict = _extract_link_dict(val)
    if link_dict is not None:
        rendered = _render_link(link_dict)
        if rendered is not None:
            return rendered
    # Для прочих словарей пробуем вернуть их человекочитаемое представление
    if isinstance(val, dict):
        normalized = _normalize_cell_value(try_repr(val))
    else:
        normalized = _normalize_cell_value(val)
    return normalized

def try_repr(val):
    if isinstance(val, dict) and "repr" in val:
        return val["repr"]
    return val


def get_fields_dict(request):
    """Возвращает словарь fields из request'a (поддержка GET и POST)."""
    fields_param = request.query_params.get('fields') if request.method == 'GET' else request.data.get('fields')
    fields_dict = {}

    if fields_param:
        try:
            # В GET это может быть JSON-строка, в POST — уже словарь
            if isinstance(fields_param, str):
                fields_dict = json.loads(fields_param)
            elif isinstance(fields_param, dict):
                fields_dict = fields_param
        except json.JSONDecodeError:
            # fields_dict = {}
            raise ValidationError('Ошибка в fields')

    return fields_dict


 



def export_report_to_histogram(
    rows,
    group_fields,
    aggregate_fields,
):
    aggregate_names = [agg.get("name") for agg in aggregate_fields if agg.get("name")]

    data = []
    for row in rows:
        processed_row = {}
        for k, v in row.items():
            if k in aggregate_names:
                try:
                    processed_row[k] = float(v)
                except (ValueError, TypeError):
                    processed_row[k] = None  # или 0, если нужно
            else:
                processed_row[k] = try_repr(v)
        data.append(processed_row)

    if len(group_fields) < 1 or not aggregate_fields:
        return Response({"detail": "Нужен хотя бы 1 group_field и 1 aggregate_field"},
                        status=400)

    x_key = group_fields[0]
    group_key = group_fields[1] if len(group_fields) > 1 else None
    y_keys = [agg.get("name") for agg in aggregate_fields if agg.get("name")]

    df = pd.DataFrame(data)

    fig = go.Figure()

    if group_key:
        grouped = df.groupby([x_key, group_key], dropna=False).sum(numeric_only=True).reset_index()
        for group_val in grouped[group_key].unique():
            sub_df = grouped[grouped[group_key] == group_val]
            for y_key in y_keys:
                fig.add_trace(go.Bar(
                    name=f"{group_val} - {y_key}",
                    x=sub_df[x_key],
                    y=sub_df[y_key],
                ))
    else:
        for y_key in y_keys:
            fig.add_trace(go.Bar(
                name=y_key,
                x=df[x_key],
                y=df[y_key],
            ))

    fig.update_layout(
        title="Гистограмма",
        xaxis_title=x_key,
        yaxis_title="Значение",
        barmode='stack'
    )

    html = pio.to_html(fig, full_html=True, include_plotlyjs='cdn')
    return HttpResponse(html, content_type='text/html')


def parse_filter_value(value):
    """
    Преобразует строковое значение в соответствующий тип данных.
    - Дата в формате 'YYYY-MM-DD' → datetime.date
    - Дата и время в формате ISO 8601 (например, '2025-01-10T03:00:00Z') → datetime.datetime
    - Целое число (например, '42') → int
    - Число с плавающей точкой (например, '3.14') → float
    - Все остальные значения остаются строками
    """
    if not isinstance(value, str):
        return value

    value = value.strip()

    # Дата в формате YYYY-MM-DD
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", value):
        try:
            return datetime.strptime(value, "%Y-%m-%d").date()
        except ValueError:
            pass

    # Дата-время в формате ISO 8601
    try:
        return datetime.fromisoformat(value.replace("Z", "+00:00"))
    except ValueError:
        pass

    # Целое число
    if re.fullmatch(r"-?\d+", value):
        try:
            return int(value)
        except ValueError:
            pass

    # Десятичное число (float)
    if re.fullmatch(r"-?\d+\.\d+", value):
        try:
            return float(value)
        except ValueError:
            pass

    # Оставить строкой
    return value


def extract_field_names(raw_list):
    """На вход принимает список словарей: [{"name":"manager"},{"name":"manager_position"}],
    возвращает список значений ключей name: ["manager", "manager_position"]"""
    if not isinstance(raw_list, list):
        return []
    if all(isinstance(item, dict) and "name" in item for item in raw_list):
        return [item["name"] for item in raw_list]
    return raw_list  # предполагаем, что это уже список имён


def get_verbose_title(model_class, col, leveling_raw, group_fields_raw, aggregates, system_fields_raw):
    """Функция для получения названия поля с учётом title и обработки агрегированных полей"""
    # 0. Обработка агрегированных полей
    if col in aggregates:
        return (
            col
            .replace("__COLON", ":")
            .replace("__SEMICOLON", ";")
            .replace("__SPACE", " ")
            .replace("__COMMA", ",")
            .replace("__DOT", ".")
        )
    # 1. Поиск в переданных словарях
    sources = []
    for lst in (leveling_raw, group_fields_raw, system_fields_raw):
        if isinstance(lst, list):
            sources.extend(lst)

    for item in sources:
        if isinstance(item, dict) and item.get("name") == col and item.get("title"):
            return item.get("title") or col

    # 1.5. Вычисляемые поля (get_report_computed_fields_meta)
    if hasattr(model_class, 'get_report_computed_fields_meta'):
        parts = col.split("__")
        if len(parts) == 1:
            for item in model_class.get_report_computed_fields_meta():
                if item.get("name") == col:
                    verbose = item.get("verbose_name") or item.get("title")
                    return str(verbose) if verbose else col
        else:
            prefix_path, computed_name = "__".join(parts[:-1]), parts[-1]
            try:
                _, related_model = get_field_meta(prefix_path, model_class)
                if related_model and hasattr(related_model, 'get_report_computed_fields_meta'):
                    for item in related_model.get_report_computed_fields_meta():
                        if item.get("name") == computed_name:
                            verbose = item.get("verbose_name") or item.get("title")
                            return str(verbose) if verbose else col
            except Exception:
                pass

    # 2. Разбор по цепочке через модель
    try:
        parts = col.split("__")
        model = model_class
        field = None

        for i, part in enumerate(parts):
            # Осталась только часть даты (например, year), нужно проверить предыдущий field
            if i == len(parts) - 1 and field and isinstance(field, (models.DateField, models.DateTimeField)):
                if part in {"year", "month", "day", "week", "quarter", "date"}:
                    return f"{field.verbose_name}, {part}"

            field = model._meta.get_field(part)

            if isinstance(field, (models.ForeignKey, models.OneToOneField)):
                model = field.related_model
            elif i < len(parts) - 1:
                # Если это не последняя часть и не FK — путь обрывается
                return col

        return field.verbose_name if field else col

    except Exception:
        return col


def build_title_context(
    model_class,
    leveling_raw,
    group_fields_raw,
    aggregate_fields_raw,
    system_fields_raw,
):
    """
    Нормализует контекст для вычисления заголовков колонок отчета.
    """
    aggregate_names = []
    for aggregate_field in aggregate_fields_raw or []:
        if isinstance(aggregate_field, dict):
            aggregate_name = aggregate_field.get("name")
            if aggregate_name:
                aggregate_names.append(aggregate_name)
        elif isinstance(aggregate_field, str):
            aggregate_names.append(aggregate_field)

    return {
        "model_class": model_class,
        "leveling_raw": leveling_raw,
        "group_fields_raw": group_fields_raw,
        "aggregate_names": aggregate_names,
        "system_fields_raw": system_fields_raw,
    }


def build_display_title_map(columns, title_context):
    """
    Формирует единую карту человекочитаемых заголовков для колонок отчета.
    """
    title_map = {}
    for column_name in columns or []:
        title_map[column_name] = force_str(
            get_verbose_title(
                model_class=title_context["model_class"],
                col=column_name,
                leveling_raw=title_context["leveling_raw"],
                group_fields_raw=title_context["group_fields_raw"],
                aggregates=title_context["aggregate_names"],
                system_fields_raw=title_context["system_fields_raw"],
            )
        )
    return title_map


def build_json_display_name_map(
    model_class,
    visible_columns_order,
    leveling_raw,
    group_fields_raw,
    aggregate_fields_raw,
    system_fields_raw,
):
    """
    Формирует карту {техническое_имя: человекочитаемый_заголовок}
    для JSON-ответа отчёта по той же логике, что и HTML/Excel.
    """
    title_context = build_title_context(
        model_class=model_class,
        leveling_raw=leveling_raw,
        group_fields_raw=group_fields_raw,
        aggregate_fields_raw=aggregate_fields_raw,
        system_fields_raw=system_fields_raw,
    )
    return build_display_title_map(visible_columns_order, title_context)


def rename_rows_by_title(rows, title_map):
    """Переименовывает ключи строк по карте title_map."""
    if not rows or not title_map:
        return rows
    result = []
    for row in rows:
        renamed_row = {}
        for key, value in row.items():
            renamed_row[title_map.get(key, key)] = value
        result.append(renamed_row)
    return result
    

def _format_filter_value(field, value, model_class):
    """
    Единая функция приведения значения фильтра к человекочитаемому виду
    (как в таблице фильтров и в плейсхолдерах).
    """
    from .serializers import FastSerializerById

    model_field, model = get_field_meta(field, model_class)

    if not isinstance(value, list):
        values = [value]
    else:
        values = list(value)

    choices_map = None
    if model_field is not None and getattr(model_field, "choices", None):
        choices_map = {str(choice[0]): str(choice[1]) for choice in model_field.choices}

    is_relation = isinstance(
        model_field,
        (models.ForeignKey, models.OneToOneField, models.ManyToManyField, TreeForeignKey),
    )

    for i, each in enumerate(values):
        if isinstance(each, bool):
            values[i] = "Да" if each else "Нет"
        elif model and (is_relation or model_field is None):
            # обычный FK или вычисляемое поле отчёта типа ForeignKey (main_specialist и т.п.)
            if each is None or each == '':
                values[i] = ''
            else:
                serializer = FastSerializerById(model_class=model)
                try:
                    data = serializer.to_representation(each)
                    values[i] = (data or {}).get("repr", each)
                except Exception:
                    values[i] = each
        elif choices_map is not None:
            values[i] = choices_map.get(str(each), each)
        else:
            parsed = parse_filter_value(each)
            if isinstance(parsed, (datetime, date)):
                values[i] = parsed.strftime("%d.%m.%Y")
            else:
                values[i] = parsed

    return ", ".join(map(str, values))


def get_filters_table(filters_data, model_class, leveling_raw, group_fields_raw, aggregates, system_fields_raw):
    """Создает таблицу фильтров для экспорта отчетов с поддержкой вложенных групп."""
    COMPARISON_TYPES = {
        "=": _("Равно"),
        "!=": _("Не равно"),
        ">": _("Больше"),
        ">=": _("Больше или равно"),
        "<": _("Меньше"),
        "<=": _("Меньше или равно"),
        "icontains": _("Содержит"),
        "not icontains": _("Не содержит"),
        "isnull": _("Заполнено"),
        "not isnull": _("Не заполнено"),
        "in": _("В списке"),
        "not in": _("Не в списке"),
    }

    def process_filter_group(filter_group, level=0):
        """Рекурсивно обрабатывает группу фильтров."""
        filters_table = []

        if not isinstance(filter_group, dict):
            return filters_table

        filters = filter_group.get("filters", [])
        logic = filter_group.get("logic", "and")

        # Добавляем заголовок группы, если это не корневой уровень
        if level > 0 and filters:
            logic_display = "ИЛИ" if logic.lower() == "or" else "И"
            filters_table.append({
                "Поле": f"{'  ' * (level - 1)}{logic_display}",
                "Условие": "",
                "Значение": "",
                "__row_type__": "summary",
                "__indent__": level - 1
            })

        for f in filters:
            if "filters" in f:
                # Это вложенная группа
                nested_filters = process_filter_group(f, level + 1)
                filters_table.extend(nested_filters)
            else:
                # Это обычное условие фильтра
                field = f.get("field", "") or f.get("name", "")  # Поддержка как field, так и name
                comparison_type = f.get("comparison_type", "")
                raw_value = f.get("value", "")

                field_title = force_str(get_verbose_title(model_class, field, leveling_raw, group_fields_raw, aggregates, system_fields_raw))
                comparison_display = force_str(COMPARISON_TYPES.get(comparison_type, comparison_type))

                value = _format_filter_value(field, raw_value, model_class)

                filters_table.append({
                    "Поле": f"{'  ' * level}{field_title}",
                    "Условие": comparison_display,
                    "Значение": value,
                    "__row_type__": "data",
                    "__indent__": level
                })

        return filters_table

    filters_table = []
    if filters_data and isinstance(filters_data, dict):
        filters_table = process_filter_group(filters_data, level=0)

    # Всегда создаем DataFrame с заголовками, даже если фильтров нет
    if not filters_table:
        # Добавляем пустую строку, чтобы были видны заголовки
        filters_table.append({
            "Поле": "",
            "Условие": "",
            "Значение": "",
            "__row_type__": "data",
            "__indent__": 0
        })
    return filters_table


def get_filters_placeholders(filters_data, model_class, leveling_raw, group_fields_raw, aggregates, system_fields_raw):
    """
    Возвращает словарь плейсхолдеров для фильтров:
    filters__<field>__<lookup> -> человекочитаемое значение.
    """
    comparison_lookups = {
        "=": "exact",
        "!=": "not_exact",
        ">": "gt",
        ">=": "gte",
        "<": "lt",
        "<=": "lte",
        "icontains": "icontains",
        "not icontains": "not_icontains",
        "isnull": "isnull",
        "not isnull": "not_isnull",
        "in": "in",
        "not in": "not_in",
    }

    placeholders = {}

    def process_filter_group(filter_group):
        if not isinstance(filter_group, dict):
            return

        filters = filter_group.get("filters", [])
        for f in filters:
            if "filters" in f:
                process_filter_group(f)
            else:
                field = f.get("field", "") or f.get("name", "")
                comparison_type = f.get("comparison_type", "")
                raw_value = f.get("value", "")

                if not field:
                    continue

                lookup = comparison_lookups.get(comparison_type, comparison_type)
                display_value = _format_filter_value(field, raw_value, model_class)

                # Имя плейсхолдера: <field>__<lookup>
                key = f"{field}__{lookup}"
                if key in placeholders and placeholders[key]:
                    placeholders[key] = f"{placeholders[key]}, {display_value}"
                else:
                    placeholders[key] = display_value

    if filters_data and isinstance(filters_data, dict):
        process_filter_group(filters_data)

    return placeholders


def get_columns_order(group_fields_raw, aggregate_fields_raw, system_fields_raw, leveling=None, computed_fields_raw=None):
    """
    Возвращает список имён колонок, отсортированный по полю 'order'.
    Столбцы из leveling располагаются первыми в том порядке, в котором они перечислены в leveling.
    Остальные элементы без 'order' располагаются в конце, сохраняя исходный порядок.
    Автоматически исключает поля с is_visible=False.
    
    Args:
        group_fields_raw: Список групповых полей
        aggregate_fields_raw: Список агрегатных полей  
        system_fields_raw: Список системных полей
        leveling: Список полей для группировки
    """
    sources = []
    for field_list in (aggregate_fields_raw, group_fields_raw, system_fields_raw, computed_fields_raw or []):
        if isinstance(field_list, list):
            sources.extend(field_list)
    
    # Фильтруем по видимости (is_visible=False исключаются)
    sources = [item for item in sources if item.get('is_visible', True)]
    
    # Получаем все имена столбцов, отсортированные по order
    all_sorted_names = [
        item['name']
        for item in sorted(sources, key=lambda x: (x.get('order') is None, x.get('order', float('inf'))))
    ]
    
    # Если leveling не передан, возвращаем обычную сортировку
    if not leveling:
        return all_sorted_names
    
    # Создаем результирующий список, начиная с leveling
    result_names = list(leveling)
    
    # Добавляем остальные столбцы, исключая те, что уже добавлены из leveling
    for name in all_sorted_names:
        if name not in result_names:
            result_names.append(name)
    
    return result_names


def get_columns_order_for_calculations(group_fields_raw, aggregate_fields_raw, system_fields_raw, leveling=None, computed_fields_raw=None):
    """
    Возвращает список имён колонок для вычислений, включая невидимые поля.
    Используется для промежуточных вычислений, группировки и агрегации.
    
    Args:
        group_fields_raw: Список групповых полей
        aggregate_fields_raw: Список агрегатных полей  
        system_fields_raw: Список системных полей
        leveling: Список полей для группировки
    """
    sources = []
    for field_list in (aggregate_fields_raw, group_fields_raw, system_fields_raw, computed_fields_raw or []):
        if isinstance(field_list, list):
            sources.extend(field_list)
    
    # НЕ фильтруем по видимости - включаем все поля для вычислений
    
    # Получаем все имена столбцов, отсортированные по order
    all_sorted_names = [
        item['name']
        for item in sorted(sources, key=lambda x: (x.get('order') is None, x.get('order', float('inf'))))
    ]
    
    # Если leveling не передан, возвращаем обычную сортировку
    if not leveling:
        return all_sorted_names
    
    # Создаем результирующий список, начиная с leveling
    result_names = list(leveling)
    
    # Добавляем остальные столбцы, исключая те, что уже добавлены из leveling
    for name in all_sorted_names:
        if name not in result_names:
            result_names.append(name)
    
    return result_names

def export_report_to_excel(params, model_class, output_rows, final_df, df_filters, 
                           numeric_aggregate_fields, show_totals, fields_for_totals, totals_title, grand_totals_row,
                           leveling_raw, group_fields_raw, aggregate_fields_raw, system_fields_raw,
                           report_title, columns_order, template=None, placeholders=None, base_instance=None):
    # Берем плейсхолдеры из переданного словаря (формируются во view)
    placeholders = placeholders or {}
    
    # Получаем leveling из leveling_raw
    leveling = extract_field_names(leveling_raw)
    
    # Получаем aggregates из aggregate_fields_raw
    aggregates = [agg.get("name") for agg in aggregate_fields_raw if agg.get("name")]
    
    # Строим структуру таблицы и конфигурацию заголовков
    table_structure = build_table_structure(params, final_df, columns_order, leveling)
    headers_config = build_headers_config(params, leveling, leveling_raw, group_fields_raw, 
                                         aggregate_fields_raw, system_fields_raw, columns_order, 
                                         model_class, table_structure)
    all_columns = table_structure['all_columns']

    # === В НАЧАЛЕ ФУНКЦИИ: Определяем стили ===
    # Базовый набор стандартных стилей (fallback)
    named_styles_dict, filter_header_style, filter_row_style = _build_default_export_styles()
    # По умолчанию фильтры выводим
    write_filters = True
    # Стартовый столбец для таблицы данных (по умолчанию первый столбец)
    table_start_col_idx = 1

    if template:
        wb = load_workbook(template)
        ws = wb.worksheets[0]
        # Save column widths
        col_widths = {col: ws.column_dimensions[col].width for col in ws.column_dimensions}
        # Save row heights
        row_heights_template = {row: ws.row_dimensions[row].height for row in ws.row_dimensions}
        # Save named ranges for header/footer/table
        named_ranges = {n.name: n for n in wb.defined_names.definedName}
        # Собираем все плейсхолдеры, которые реально используются в header/footer
        template_placeholder_keys = _collect_template_placeholders(named_ranges, ws)
        # Динамические плейсхолдеры: есть в шаблоне, но ещё нет в словаре placeholders
        dynamic_keys = {key for key in template_placeholder_keys if key not in placeholders}
        # Разрешаем динамические плейсхолдеры через ORM-пути от base_instance
        # (если base_instance None или путь не разрешился, плейсхолдеры получат пустую строку)
        enrich_placeholders_from_dynamic_paths(placeholders, dynamic_keys, model_class, base_instance, named_ranges, ws)

        # === ЗАМЕНА ПЛЕЙСХОЛДЕРОВ В HEADER И FOOTER (в самом начале) ===
        # Используем регулярное выражение для замены всех плейсхолдеров
        # Если плейсхолдер не найден в словаре, заменяется на пустую строку (fallback)
        placeholder_re = re.compile(r"\{\{\s*([^}]+?)\s*\}\}")
        
        def replace_placeholders_in_cell(cell):
            """Заменяет все плейсхолдеры в ячейке на значения из словаря или пустую строку."""
            if not cell.value or not isinstance(cell.value, str):
                return
            original_value = cell.value
            new_value = placeholder_re.sub(
                lambda match: str(placeholders.get(match.group(1).strip(), "")),
                original_value
            )
            cell.value = new_value
        
        # Замена в header
        if 'header' in named_ranges:
            try:
                for dest in named_ranges['header'].destinations:
                    sheetname, coord = dest
                    if sheetname == ws.title:
                        if ':' in coord:
                            # Диапазон ячеек
                            for row in ws[coord]:
                                for cell in row:
                                    replace_placeholders_in_cell(cell)
                        else:
                            # Одиночная ячейка
                            if coord and coord[0].isalpha() and any(c.isdigit() for c in coord):
                                row, col = coordinate_to_tuple(coord)
                                cell = ws.cell(row=row, column=col)
                                replace_placeholders_in_cell(cell)
                            else:
                                print(f"ERROR: Invalid header coordinate: '{coord}' - skipping placeholder replacement")
            except AttributeError:
                print(f"ERROR: Invalid header named range definition - skipping placeholder replacement")
        
        # Замена в footer
        if 'footer' in named_ranges:
            try:
                for dest in named_ranges['footer'].destinations:
                    sheetname, coord = dest
                    if sheetname == ws.title:
                        if ':' in coord:
                            # Диапазон ячеек
                            for row in ws[coord]:
                                for cell in row:
                                    replace_placeholders_in_cell(cell)
                        else:
                            # Одиночная ячейка
                            if coord and coord[0].isalpha() and any(c.isdigit() for c in coord):
                                row, col = coordinate_to_tuple(coord)
                                cell = ws.cell(row=row, column=col)
                                replace_placeholders_in_cell(cell)
                            else:
                                print(f"ERROR: Invalid footer coordinate: '{coord}' - skipping placeholder replacement")
            except AttributeError:
                print(f"ERROR: Invalid footer named range definition - skipping placeholder replacement")
        # === КОНЕЦ ЗАМЕНЫ ПЛЕЙСХОЛДЕРОВ ===

        # === КОПИРОВАНИЕ СТИЛЕЙ ИЗ ШАБЛОНА В NamedStyle (с приоритетом над стандартными) ===
        def copy_style_from_cell(style_name, cell):
            ns = NamedStyle(name=style_name)
            ns.font = cell.font.copy() if cell.font else Font()
            ns.fill = cell.fill.copy() if cell.fill else PatternFill()
            ns.border = cell.border.copy() if cell.border else Border()
            ns.alignment = cell.alignment.copy() if cell.alignment else Alignment()
            return ns

        style_to_named_range = {
            "report_title": "report_title",
            "report_footer": "report_footer",
            "filter_header": "filter_header",
            "filter_row": "filter_row",
            "table_header": "table_header",
            "group_header_level_1": "group_header_level_1",
            "group_header_level_2": "group_header_level_2",
            "group_header_level_3": "group_header_level_3",
            "group_footer_level_1": "group_footer_level_1",
            "group_footer_level_2": "group_footer_level_2",
            "group_footer_level_3": "group_footer_level_3",
            "detail_rows": "detail_rows",
            "grand_totals": "grand_totals",
        }
        styles_from_template = set()
        for style_name, range_name in style_to_named_range.items():
            if range_name in named_ranges:
                for dest in named_ranges[range_name].destinations:
                    sheetname, coord = dest
                    if sheetname == ws.title:
                        if ':' in coord:
                            first_cell_coord = coord.split(':')[0].replace('$', '')
                        else:
                            first_cell_coord = coord.replace('$', '')
                        row, col = coordinate_to_tuple(first_cell_coord)
                        cell = ws.cell(row=row, column=col)
                        ns = copy_style_from_cell(style_name, cell)
                        named_styles_dict[style_name] = ns
                        styles_from_template.add(style_name)
                        break
        # === КОПИРОВАНИЕ СТИЛЕЙ ИЗ ШАБЛОНА В NamedStyle ===

        # Обновляем переменные стилей для фильтров
        filter_header_style = named_styles_dict.get('filter_header')
        filter_row_style = named_styles_dict.get('filter_row')

        # Get header/footer cell coordinates
        header_cells = set()
        footer_cells = set()
        header_max_row = 0
        if 'header' in named_ranges:
            for dest in named_ranges['header'].destinations:
                sheetname, coord = dest
                if sheetname == ws.title:
                    for row in ws[coord]:
                        for cell in row:
                            header_cells.add(cell.coordinate)
                            if cell.row and cell.row > header_max_row:
                                header_max_row = cell.row
        if 'footer' in named_ranges:
            for dest in named_ranges['footer'].destinations:
                sheetname, coord = dest
                if sheetname == ws.title:
                    for row in ws[coord]:
                        for cell in row:
                            footer_cells.add(cell.coordinate)
        # Стартовая строка по умолчанию — сразу после header (если он есть), иначе первая строка
        row_idx = header_max_row + 1 if header_max_row else 1
        # Clear all cells except header/footer
        for row in ws.iter_rows():
            for cell in row:
                if cell.coordinate not in header_cells and cell.coordinate not in footer_cells:
                    cell.value = None
                    cell.font = Font()
                    cell.fill = PatternFill()
                    cell.border = Border()
                    cell.alignment = Alignment()
                    cell.number_format = 'General'
                    cell.style = 'Normal'
                    from openpyxl.styles import Protection
                    cell.protection = Protection(locked=False)
        # Restore column widths
        for col, width in col_widths.items():
            if width:
                ws.column_dimensions[col].width = width
        # Restore row heights for header/footer
        for row, height in row_heights_template.items():
            if height:
                ws.row_dimensions[row].height = height
        # Определяем начальную позицию для вставки данных из именованного диапазона filter
        if 'filter' in named_ranges:
            filter_dest = None
            for dest in named_ranges['filter'].destinations:
                sheetname, coord = dest
                if sheetname == ws.title:
                    filter_dest = coord
                    break
            if filter_dest:
                if ':' in filter_dest:
                    filter_start_cell = filter_dest.split(':')[0]
                else:
                    filter_start_cell = filter_dest
                filter_start_row, filter_start_col = coordinate_to_tuple(filter_start_cell.replace('$', ''))
                row_idx = filter_start_row
        else:
            # В шаблоне нет именованного диапазона filter — фильтры не выводим
            write_filters = False

        # Определяем стартовый столбец для таблицы данных по именованному диапазону table (если есть)
        if 'table' in named_ranges:
            table_dest = None
            for dest in named_ranges['table'].destinations:
                sheetname, coord = dest
                if sheetname == ws.title:
                    table_dest = coord
                    break
            if table_dest:
                if ':' in table_dest:
                    table_start_cell = table_dest.split(':')[0]
                else:
                    table_start_cell = table_dest
                _, table_start_col = coordinate_to_tuple(table_start_cell.replace('$', ''))
                table_start_col_idx = table_start_col
    else:
        wb = Workbook()
        # Для ветки без шаблона используем те же дефолтные стили
        named_styles_dict, filter_header_style, filter_row_style = _build_default_export_styles()
        
        ws = wb.active
        ws.title = report_title[:31]
        ws.sheet_format = SheetFormatProperties(defaultRowHeight=15)
        ws.sheet_properties.outlinePr.summaryBelow = True
        row_idx = 1
        ws.cell(row=row_idx, column=1, value=report_title).font = Font(size=14, bold=True)
        row_idx += 1
        ws.cell(row=row_idx, column=1, value=placeholders.get("creation_date", ""))  # Дата создания
        ws.cell(row=row_idx, column=2, value=placeholders.get("timezone", ""))  # Часовой пояс
        row_idx += 2

        thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # 🔹 Фильтры с границами
    filter_row_idx_start = row_idx
    filter_rows_written = 0
    col_idx = 1  # Инициализируем col_idx значением по умолчанию
    # Если используем шаблон, но в нём нет именованного диапазона filter, фильтры не выводим
    if write_filters and not df_filters.empty:
        # Применяем стили к слову "Фильтры"
        cell = ws.cell(row=row_idx, column=col_idx, value="Фильтры")
        _apply_style(cell, 'filter_header', named_styles_dict)
        row_idx += 1
        filter_rows_written += 1

        # Исключаем служебные столбцы из фильтров
        filter_columns = [col for col in df_filters.columns if col not in ("__row_type__", "__indent__")]
        
        for cidx, col in enumerate(filter_columns, start=col_idx):
            # Применяем стили к заголовкам фильтров
            cell = ws.cell(row=row_idx, column=cidx, value=force_str(col))
            _apply_style(cell, 'filter_header', named_styles_dict)
        row_idx += 1
        filter_rows_written += 1

        # Группировка фильтров
        filter_prev_indent = None
        filter_prev_row_idx = None
        
        for row_index, row in df_filters.iterrows():
            indent = row.get("__indent__", 0)
            row_type = row.get("__row_type__", "data")
            
            # Настройка группировки строк
            ws.row_dimensions[row_idx].outlineLevel = indent
            
            # Группировки должны быть развёрнуты по умолчанию, поэтому не скрываем строки
            if indent > 0:
                ws.row_dimensions[row_idx].hidden = False
            
            # Оставляем родительские группы раскрытыми
            if filter_prev_indent is not None and indent > filter_prev_indent:
                ws.row_dimensions[filter_prev_row_idx].collapsed = False
            
            filter_prev_indent = indent
            filter_prev_row_idx = row_idx

            for cidx, col in enumerate(filter_columns, start=col_idx):
                # Применяем стили к данным фильтров
                if row_type == "summary":
                    # Группирующие строки используют стиль filter_header
                    cell = ws.cell(row=row_idx, column=cidx, value=force_str(row[col]))
                    _apply_style(cell, 'filter_header', named_styles_dict)
                else:
                    # Обычные строки используют стиль filter_row
                    cell = ws.cell(row=row_idx, column=cidx, value=force_str(row[col]))
                    _apply_style(cell, 'filter_row', named_styles_dict)
            row_idx += 1
            filter_rows_written += 1

        row_idx += 1
        filter_rows_written += 1
    filter_row_idx_end = row_idx - 1

    # --- Если шаблон и есть table, сдвигаем table вниз, если фильтров больше, чем в шаблоне ---
    if template and 'table' in named_ranges:
        # Определяем начальную строку table
        table_dest = None
        for dest in named_ranges['table'].destinations:
            sheetname, coord = dest
            if sheetname == ws.title:
                table_dest = coord
                break
        if table_dest:
            if ':' in table_dest:
                table_start_cell = table_dest.split(':')[0]
            else:
                table_start_cell = table_dest
            table_start_row, table_start_col = coordinate_to_tuple(table_start_cell.replace('$', ''))
            # Определяем сколько строк было в шаблоне под фильтры
            if 'filter' in named_ranges:
                filter_dest = None
                for dest in named_ranges['filter'].destinations:
                    sheetname, coord = dest
                    if sheetname == ws.title:
                        filter_dest = coord
                        break
                if filter_dest:
                    if ':' in filter_dest:
                        filter_start_cell, filter_end_cell = filter_dest.split(':')
                        filter_start_row, filter_start_col = coordinate_to_tuple(filter_start_cell.replace('$', ''))
                        filter_end_row, filter_end_col = coordinate_to_tuple(filter_end_cell.replace('$', ''))
                        template_filter_rows = filter_end_row - filter_start_row + 1
                    else:
                        template_filter_rows = 1
                else:
                    template_filter_rows = 0
            else:
                template_filter_rows = 0
            # --- Фикс: table всегда начинается сразу после всех реально выведенных фильтров ---
            if row_idx > table_start_row:
                # Фильтры заняли больше места, чем было в шаблоне — table сразу после фильтров
                pass  # row_idx уже правильный
            else:
                # Фильтров меньше или столько же — сдвигаем table вниз, если надо
                if filter_rows_written > template_filter_rows:
                    ws.insert_rows(table_start_row, amount=filter_rows_written - template_filter_rows)
                    table_start_row += filter_rows_written - template_filter_rows
                row_idx = table_start_row
        else:
            # Если не нашли table_dest, row_idx уже после фильтров
            pass
    # 🔹 Заголовки данных с границами
    row_idx = render_excel_headers(
        ws,
        row_idx,
        headers_config,
        table_structure,
        model_class,
        leveling_raw,
        group_fields_raw,
        aggregate_fields_raw,
        system_fields_raw,
        named_styles_dict,
        start_col_idx=table_start_col_idx,
    )

    # Итоги уже рассчитаны в prepare_report_data

    # === КОРРЕКТИРОВКА ПОЛОЖЕНИЯ FOOTER ДО ВЫВОДА ДАННЫХ ===
    if template and 'table' in named_ranges and 'footer' in named_ranges:
        table_dest = None
        footer_dest = None
        for dest in named_ranges['table'].destinations:
            sheetname, coord = dest
            if sheetname == ws.title:
                table_dest = coord
                break
        for dest in named_ranges['footer'].destinations:
            sheetname, coord = dest
            if sheetname == ws.title:
                footer_dest = coord
                break
        if table_dest and footer_dest:
            if ':' in table_dest:
                table_start_cell, table_end_cell = table_dest.split(':')
            else:
                table_start_cell = table_end_cell = table_dest
            table_start_row, table_start_col = coordinate_to_tuple(table_start_cell.replace('$', ''))
            table_end_row, table_end_col = coordinate_to_tuple(table_end_cell.replace('$', ''))
            table_rows_in_template = table_end_row - table_start_row + 1

            if ':' in footer_dest:
                footer_start_cell = footer_dest.split(':')[0]
            else:
                footer_start_cell = footer_dest
            footer_start_row, footer_start_col = coordinate_to_tuple(footer_start_cell.replace('$', ''))

            # Сколько строк между концом table и началом footer в шаблоне
            gap_in_template = footer_start_row - (table_start_row + table_rows_in_template)
            if gap_in_template < 0:
                gap_in_template = 0  # на всякий случай

            data_start_row = row_idx
            data_end_row = data_start_row + len(output_rows)
            # Если будет строка итогов — учесть её
            if show_totals and fields_for_totals and not final_df.empty:
                data_end_row += 1
            # diff теперь учитывает gap_in_template
            diff = (footer_start_row - data_end_row) - gap_in_template
            if diff > 0:
                ws.delete_rows(data_end_row, amount=diff)
            elif diff < 0:
                ws.insert_rows(footer_start_row, amount=abs(diff))
            # === Сдвигаем именованный диапазон footer ===
            if diff != 0:
                from openpyxl.utils.cell import coordinate_from_string
                old_footer_range = footer_dest
                if ':' in old_footer_range:
                    start_cell, end_cell = old_footer_range.split(':')
                else:
                    start_cell = end_cell = old_footer_range
                def shift_cell(cell, row_shift):
                    col, row = coordinate_from_string(cell.replace('$', ''))
                    return f"{col}{row + row_shift}"
                new_start = shift_cell(start_cell, diff)
                new_end = shift_cell(end_cell, diff)
                new_footer_range = f"{new_start}:{new_end}" if start_cell != end_cell else new_start
                for defined_name in wb.defined_names.definedName:
                    if defined_name.name == 'footer':
                        defined_name.attr_text = f"{ws.title}!{new_footer_range}"
                        break
    # === КОНЕЦ КОРРЕКТИРОВКИ ===

    row_heights = {}
    data_prev_indent = None
    data_prev_row_idx = None
    max_widths = {
        col: len(force_str(get_verbose_title(model_class, col, leveling_raw, group_fields_raw, aggregates, system_fields_raw)))
        for col in all_columns
    }

    # Теперь row_idx указывает на первую строку данных, выводим данные начиная с row_idx
    excel_output_rows = _prepare_excel_rows_for_repr_only_columns(output_rows, model_class)
    for row in excel_output_rows:
        indent = row.get("__indent__", 0)
        row_type = row.get("__row_type__", "data")
        ws.row_dimensions[row_idx].outlineLevel = indent

        # Группировки должны быть развёрнуты по умолчанию, поэтому не скрываем строки
        if indent > 0:
            ws.row_dimensions[row_idx].hidden = False

        # Оставляем родительские группы раскрытыми
        if data_prev_indent is not None and indent > data_prev_indent:
            ws.row_dimensions[data_prev_row_idx].collapsed = False

        data_prev_indent = indent
        data_prev_row_idx = row_idx

        for col_idx, col in enumerate(all_columns, start=table_start_col_idx):
            raw_val = row.get(col, "")
            # Универсальная нормализация значений (NaN/None/"nan"/"None")
            raw_val = _normalize_cell_value(raw_val)
            cell_value = None

            link_text = None
            link_url = None
            # Ссылка передаётся как dict {"repr","url"}; иногда прилетает строковый repr dict
            parsed_link = _extract_link_dict(raw_val)
            if parsed_link is not None:
                link_url = parsed_link.get("url")
                link_text = parsed_link.get("repr")

            # Любые фактически числовые значения пишем числом (если это не ссылка)
            if link_url is None and isinstance(raw_val, (int, float, np.integer, np.floating)):
                cell_value = raw_val
            else:
                if link_text is not None:
                    display = link_text
                else:
                    if raw_val is None:
                        display = ""
                    elif isinstance(raw_val, dict):
                        display = force_str(try_repr(raw_val))
                    elif isinstance(raw_val, (list, tuple, np.ndarray)) and raw_val:
                        items = raw_val
                        if isinstance(raw_val, np.ndarray):
                            try:
                                items = raw_val.tolist()
                            except Exception:
                                items = []
                        else:
                            items = list(raw_val)

                        if items and all(isinstance(item, dict) for item in items):
                            repr_texts = [
                                force_str(item.get("repr") or item.get("url") or "")
                                for item in items
                                if item.get("repr") or item.get("url")
                            ]
                        else:
                            repr_texts = []
                        display = ", ".join(repr_texts) if repr_texts else ""
                    else:
                        display = force_str(raw_val)
                cell_value = display

            cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
            # Для всех вещественных значений формат до 2 знаков после запятой
            if isinstance(cell_value, (float, np.floating)):
                cell.number_format = '0.00'

            # Если это ссылка — сделаем ячейку гиперссылкой
            if link_url:
                cell.hyperlink = link_url
                cell.font = Font(color="0000EE", underline="single")

            # Всегда применяем стили по типу строки, не учитывая, была ли ячейка в table
            if row_type == "summary":
                if indent == 0 and named_styles_dict.get('group_header_level_1'):
                    _apply_style(cell, 'group_header_level_1', named_styles_dict)
                elif indent == 1 and named_styles_dict.get('group_header_level_2'):
                    _apply_style(cell, 'group_header_level_2', named_styles_dict)
                elif indent == 2 and named_styles_dict.get('group_header_level_3'):
                    _apply_style(cell, 'group_header_level_3', named_styles_dict)
                else:
                    cell.style = 'Normal'
            else:
                if named_styles_dict.get('detail_rows'):
                    _apply_style(cell, 'detail_rows', named_styles_dict)
                else:
                    cell.style = 'Normal'

            val_str = force_str(cell.value if cell_value is not None else "")
            max_widths[col] = max(max_widths[col], len(val_str))

            # Высота строки: рассчитываем необходимое количество строк
            if len(val_str) > 50:
                cell.alignment = Alignment(wrap_text=True)
                lines = (len(val_str) // 50) + 1
                row_heights[row_idx] = max(row_heights.get(row_idx, 15), lines * 15)
            else:
                row_heights[row_idx] = max(row_heights.get(row_idx, 15), 15)

        row_idx += 1

    # 🔹 Обработка общих итогов (grand_totals) - используем готовые данные
    if show_totals and grand_totals_row:
        # При использовании шаблона строка итогов должна начинаться с того же столбца,
        # откуда начинается таблица (table_start_col_idx); без шаблона он равен 1.
        for col_idx, col in enumerate(all_columns, start=table_start_col_idx):
            val = grand_totals_row.get(col, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if isinstance(val, (float, np.floating)):
                cell.number_format = '0.00'
            # Всегда применяем стиль grand_totals_style, если он есть
            if named_styles_dict.get('grand_totals'):
                _apply_style(cell, 'grand_totals', named_styles_dict)
            else:
                cell.style = 'Normal'

            if isinstance(val, str) and len(val) > 50:
                cell.alignment = Alignment(wrap_text=True)
                lines = (len(val) // 50) + 1
                row_heights[row_idx] = max(row_heights.get(row_idx, 15), lines * 15)
            else:
                row_heights[row_idx] = max(row_heights.get(row_idx, 15), 15)

        row_idx += 1

    row_idx += 1
    if not template:
        ws.cell(row=row_idx, column=1, value=placeholders.get("creation_date", ""))

    # 🔹 Ширина колонок
    if template:
        # Use template widths – они уже восстановлены ранее; дополнительных сдвигов не делаем
        for col_idx, col in enumerate(all_columns, start=table_start_col_idx):
            col_letter = get_column_letter(col_idx)
            if col_letter in col_widths and col_widths[col_letter]:
                ws.column_dimensions[col_letter].width = col_widths[col_letter]
    else:
        for col_idx, col in enumerate(all_columns, start=1):
            col_letter = get_column_letter(col_idx)
            adjusted_width = min(max_widths[col] + 2, 50)
            ws.column_dimensions[col_letter].width = adjusted_width

    # 🔹 Высота строк
    for idx, height in row_heights.items():
        ws.row_dimensions[idx].height = min(height, 450)  # Максимальная высота строки

    # 🔹 Удаляем все именованные диапазоны перед сохранением
    wb.defined_names.definedName = []
    # 🔹 Сброс флага шаблона (важно для корректного открытия в Excel)
    wb.template = False
    # 🔹 Сохранение
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{report_title}.xlsx"'}
    )


def export_report_to_html(params, model_class, output_rows, final_df, df_filters, 
                         numeric_aggregate_fields, show_totals, fields_for_totals, totals_title, grand_totals_row,
                         leveling_raw, group_fields_raw, aggregate_fields_raw, system_fields_raw, 
                         report_title, columns_order, placeholders=None):
    # Используем переданные плейсхолдеры (формируются во view)
    placeholders = placeholders or {}
    
    # Получаем leveling из leveling_raw
    leveling = extract_field_names(leveling_raw)
    
    # Строим структуру таблицы и конфигурацию заголовков
    table_structure = build_table_structure(params, final_df, columns_order, leveling)
    headers_config = build_headers_config(params, leveling, leveling_raw, group_fields_raw, 
                                         aggregate_fields_raw, system_fields_raw, columns_order, 
                                         model_class, table_structure)
    all_columns = table_structure['all_columns']

    # Стили и скрипты
    html_output = """
        <style>
            .group-toggle { cursor: pointer; font-weight: bold; background-color: #f0f0f0; }
            .group-child { display: none; }
            .indent-0 td { padding-left: 0px; }
            .indent-1 td { padding-left: 20px; }
            .indent-2 td { padding-left: 40px; }
            .indent-3 td { padding-left: 60px; }
            .grand-totals { font-weight: bold; background-color: #f8f8f8; }
            .grand-totals td { text-align: right; }
            .numeric-field { text-align: right; }
        </style>
        <script>
            function toggleGroup(groupId) {
                const rows = document.querySelectorAll(`.group-child-of-${groupId}`);
                const icon = document.getElementById(`icon-${groupId}`);
                const isHidden = rows.length && getComputedStyle(rows[0]).display === "none";

                if (isHidden) {
                    // 🔓 Показываем только прямых потомков
                    rows.forEach(row => {
                        row.style.display = "table-row";
                    });
                    if (icon) icon.textContent = "−";
                } else {
                    // 🔒 Скрываем всех потомков рекурсивно
                    hideDescendants(groupId);
                    if (icon) icon.textContent = "+";
                }
            }

            function hideDescendants(groupId) {
                const descendants = document.querySelectorAll(`.group-child-of-${groupId}`);
                descendants.forEach(row => {
                    row.style.display = "none";

                    // 📍 Получаем вложенный groupId из class
                    const match = row.className.match(/group-id-(g\d+_\d+)/);
                    if (match) {
                        const childGroupId = match[1];
                        hideDescendants(childGroupId);
                        const icon = document.getElementById(`icon-${childGroupId}`);
                        if (icon) icon.textContent = "+";
                    }
                });
            }
        </script>
    """

    # 🔹 Генерация HTML
    html_output += f"<h1>{placeholders.get('report_title', '')}</h1>"
    html_output += f"<p class='creation-date'>{placeholders.get('creation_date', '')} | {placeholders.get('timezone', '')}</p>"

    if not df_filters.empty:
        html_output += ""
        html_output += f"""
            <h2>Фильтры</h2>
            <div class="table-wrapper">
            <table border="1">
            <thead>
            <tr>
        """
        
        # Заголовки таблицы фильтров (исключаем служебные поля)
        filter_columns = [col for col in df_filters.columns if not col.startswith('__')]
        for col in filter_columns:
            html_output += f"<th>{col}</th>"
        
        html_output += """
            </tr>
            </thead>
            <tbody>
        """
        
        # Вывод строк фильтров с группировкой
        filter_group_ids = {}
        filter_group_counters = {}
        
        for idx, row in enumerate(df_filters.to_dict('records')):
            row_type = row.get("__row_type__", "data")
            indent_level = row.get("__indent__", 0)
            
            if row_type == "summary":
                # Строка-заголовок группы
                filter_group_counters[indent_level] = filter_group_counters.get(indent_level, 0) + 1
                current_group_id = f"filter_g{indent_level}_{filter_group_counters[indent_level]}"
                filter_group_ids[indent_level] = current_group_id
                parent_group_id = filter_group_ids.get(indent_level - 1)
                parent_class = f"group-child group-child-of-{parent_group_id}" if parent_group_id else ""
                
                html_output += f'<tr class="group-toggle indent-{indent_level} group-id-{current_group_id} {parent_class}" onclick="toggleGroup(\'{current_group_id}\')">'
                for col_idx, col in enumerate(filter_columns):
                    val = row.get(col, "")
                    # Обрабатываем None и nan значения для HTML отображения
                    if val is None or val == "nan" or (isinstance(val, float) and str(val) == "nan"):
                        val = ""
                    # Добавляем знак "+" только для первого столбца (Поле) в заголовках групп
                    if col_idx == 0 and row_type == "summary":
                        icon = f'<span id="icon-{current_group_id}">+</span> '
                        html_output += f"<td>{icon}{val}</td>"
                    else:
                        html_output += f"<td>{val}</td>"
                html_output += "</tr>"
            else:
                # Обычная строка данных
                if indent_level > 0:
                    parent_group_id = filter_group_ids.get(indent_level - 1)
                    row_class = f'group-child group-child-of-{parent_group_id} indent-{indent_level}'
                else:
                    row_class = f'indent-{indent_level}'
                
                html_output += f'<tr class="{row_class}">'
                for col in filter_columns:
                    val = row.get(col, "")
                    # Обрабатываем None и nan значения для HTML отображения
                    if val is None or val == "nan" or (isinstance(val, float) and str(val) == "nan"):
                        val = ""
                    html_output += f"<td>{val}</td>"
                html_output += "</tr>"
        
        html_output += """
            </tbody>
            </table>
            </div>
        """
    else:
        # Пустая таблица фильтров
        html_output += """
            <h2>Фильтры</h2>
            <div class="table-wrapper">
            <table border="1">
            <thead>
            <tr>
            <th>Поле</th>
            <th>Условие</th>
            <th>Значение</th>
            </tr>
            </thead>
            <tbody>
            <tr>
            <td></td>
            <td></td>
            <td></td>
            </tr>
            </tbody>
            </table>
            </div>
        """

    html_output += """
    <h2>Данные</h2>
    <div class=\"table-wrapper\">
    <table border=\"1\">
    <thead>"""

    # Рендерим заголовки таблицы
    html_output += render_html_headers(headers_config, model_class, leveling_raw, group_fields_raw, 
                                      aggregate_fields_raw, system_fields_raw, columns_order)
    html_output += "</thead><tbody>"

    if not final_df.empty:
        data_group_ids = {}
        data_group_counters = {}
        for idx, row in enumerate(output_rows):
            row_type = row.get("__row_type__")
            indent_level = row.get("__indent__", 0)

            data_group_counters[indent_level] = data_group_counters.get(indent_level, 0) + 1
            current_group_id = f"data_g{indent_level}_{data_group_counters[indent_level]}"
            data_group_ids[indent_level] = current_group_id
            parent_group_id = data_group_ids.get(indent_level - 1)
            parent_class = f"group-child group-child-of-{parent_group_id}" if parent_group_id else ""

            if row_type == "summary":
                html_output += f'<tr class="group-toggle indent-{indent_level} group-id-{current_group_id} {parent_class}" onclick="toggleGroup(\'{current_group_id}\')">'
                for col in all_columns:
                    # При layout == "together" иконка добавляется для первого столбца
                    if table_structure['layout'] == "together":
                        icon = f'<span id="icon-{current_group_id}">+</span> ' if col == leveling[0] else ""
                    else:
                        icon = f'<span id="icon-{current_group_id}">+</span> ' if col == leveling[indent_level] else ""
                    val = row.get(col, "")
                    # Данные уже подготовлены в build_grouped_rows; нормализуем и округлим по правилам
                    val = _format_html_cell_link_or_text(val)
                    # Добавляем класс numeric-field для числовых агрегационных полей
                    numeric_class = ' class="numeric-field"' if col in numeric_aggregate_fields else ""
                    html_output += f"<td{numeric_class}>{icon}{val}</td>"
                html_output += "</tr>"

            elif row_type == "data":
                if not leveling:
                    row_class = f'indent-{indent_level}'
                else:
                    row_class = f'group-child group-child-of-{parent_group_id} indent-{indent_level}'

                html_output += f'<tr class="{row_class}">'
                for col in all_columns:
                    val = row.get(col)
                    # Данные уже подготовлены в build_grouped_rows; нормализуем и округлим по правилам
                    val = _format_html_cell_link_or_text(val)
                    # Добавляем класс numeric-field для числовых агрегационных полей
                    numeric_class = ' class="numeric-field"' if col in numeric_aggregate_fields else ""
                    html_output += f"<td{numeric_class}>{val}</td>"
                html_output += "</tr>"
    # Создаем пустую строку таблицы если данных нет
    else:
        html_output += "<tr>"
        for col in all_columns:
            html_output += f"<td></td>"
        html_output += "</tr>"

    html_output += "</tbody>"
    
    # Итоги уже рассчитаны в prepare_report_data

    # 🔹 Вывод строки общих итогов - используем готовые данные
    if show_totals and grand_totals_row:
        html_output += "<tr class='grand-totals'>"
        
        if table_structure['layout'] == "together" and leveling:
            # Для layout=together точно повторяем структуру заголовков
            if headers_config['type'] == 'multi_row':
                cols_before_leveling = headers_config['cols_before_leveling']
                cols_after_leveling = headers_config['cols_after_leveling']
                
                # Столбцы до leveling (с rowspan в заголовках)
                for col in cols_before_leveling:
                    val = grand_totals_row.get(col, "")
                    val = _normalize_cell_value(val)
                    html_output += f"<td>{val}</td>"
                
                # Leveling столбец (первый элемент стека)
                val = grand_totals_row.get(leveling[0], "")
                html_output += f"<td>{val}</td>"
                
                # Столбцы после leveling (с rowspan в заголовках)
                for col in cols_after_leveling:
                    val = grand_totals_row.get(col, "")
                    val = _format_html_cell_link_or_text(val)
                    html_output += f"<td>{val}</td>"
            else:
                # Если по какой-то причине не multi_row
                for col in all_columns:
                    val = grand_totals_row.get(col, "")
                    val = _format_html_cell_link_or_text(val)
                    html_output += f"<td>{val}</td>"
        else:
            # Обычная логика для других layout
            for col in all_columns:
                val = grand_totals_row.get(col, "")
                val = _format_html_cell_link_or_text(val)
                html_output += f"<td>{val}</td>"
        
        html_output += "</tr>"
    
    html_output += "</table></div>"
    return HttpResponse(html_output, content_type="text/html; charset=utf-8")


def build_grouped_rows(params, df, leveling, aggregates, aggregate_fields_raw, together_merged_column=None, level=0, parent_keys=(), duration_fields=None, aggregate_to_source_field=None):
    """
    Строит иерархически сгруппированные строки с подведением итогов на каждом уровне.

    Выполняет рекурсивную группировку DataFrame по полям из leveling.
    Для поддерживаемых агрегатов ('sum', 'min', 'max') вычисляет сумму по колонке `name`
    из aggregate_fields_raw. Остальные агрегаты пропускаются.

    Возвращает список словарей, представляющих строки данных и итогов с отступами.
    """
    layout = params.get("grouping_fields_layout", "together")
    
    # Примечание: замена None на "Не заполнено" уже выполнена в prepare_report_data
    # перед вызовом build_grouped_rows, поэтому здесь не дублируем
    


    def build_agg_func_map(aggregate_fields_raw):
        """Возвращает отображение: имя агрегатного поля → функция агрегации ('sum' для поддерживаемых, иначе None)."""
        agg_func_map = {}
        for field in aggregate_fields_raw:
            name = field["name"]
            if any(k in field for k in ("sum", "min", "max", "count", "distinct_count")):
                agg_func_map[name] = "sum"
            else:
                agg_func_map[name] = None  # неподдерживаемая агрегация
        return agg_func_map

    def data_row(row, level):
        """
        Преобразует строку DataFrame в словарь для экспорта с типом "data" и уровнем вложенности.

        - layout == 'separate':      Все поля строки добавляются без изменений.
        - layout == 'separate_in_totals_only':  В словарь не добавляются поля из leveling.
        - layout == 'together':      Первый не-leveling столбец кладётся под ключом первого из leveling,
                                 остальные не-leveling поля — под своими именами.
        """
        if layout == 'separate':
            r = row.to_dict()
        elif layout == 'separate_in_totals_only':
            r = {k: v for k, v in row.to_dict().items() if k not in leveling}
        elif layout == 'together':
            row_dict = row.to_dict()
            non_leveling_items = [(k, v) for k, v in row_dict.items() if k not in leveling]
            r = {}
            if non_leveling_items:
                indent = "  " * level  # два пробела на уровень
                # Ищем столбец, который должен быть объединён с leveling[0]
                if together_merged_column and together_merged_column in row_dict:
                    # Обрабатываем значение объединённого столбца - заменяем None на пустую строку
                    merged_value = row_dict[together_merged_column]
                    if pd.isna(merged_value) or merged_value == "nan" or merged_value == "None":
                        merged_value = ""
                    # Сохраняем гиперссылку, если значение — dict с url/repr
                    if isinstance(merged_value, dict) and merged_value.get("url"):
                        merged_text = merged_value.get("repr") or merged_value.get("url") or ""
                        r[leveling[0]] = {"repr": f"{indent}{merged_text}", "url": merged_value.get("url")}
                    else:
                        r[leveling[0]] = f"{indent}{merged_value}"
                    # Остальные не-leveling поля, кроме объединённого
                    for k, v in non_leveling_items:
                        if k != together_merged_column:
                            r[k] = v
                else:
                    # Fallback к старой логике, если объединённый столбец не найден
                    first_value = non_leveling_items[0][1]
                    if pd.isna(first_value) or first_value == "nan" or first_value == "None":
                        first_value = ""
                    if isinstance(first_value, dict) and first_value.get("url"):
                        first_text = first_value.get("repr") or first_value.get("url") or ""
                        r[leveling[0]] = {"repr": f"{indent}{first_text}", "url": first_value.get("url")}
                    else:
                        r[leveling[0]] = f"{indent}{first_value}"
                    for k, v in non_leveling_items[1:]:
                        r[k] = v
        
        r["__row_type__"] = "data"
        r["__indent__"] = level
        return r

    def summary_row(group_df, group_field):
        """
        Формирует строку-итог для группы с учётом layout:
        - 'separate': включает все предыдущие группировки (leveling[:level]) и текущую, плюс агрегации.
        - 'separate_in_totals_only': только текущая группировка и агрегации.
        - 'together': ключ группировки всегда первый из leveling, агрегации как обычно.
        """
        if layout == 'separate':
            # Собираем все предыдущие ключи группировки и их значения из parent_keys
            summary_row = {lev_field: parent_keys[i] for i, lev_field in enumerate(leveling[:level])}
            # Добавляем текущую группировку
            summary_row[group_field] = key
        elif layout == 'separate_in_totals_only':
            summary_row = {group_field: key}
        elif layout == 'together':
            # Ключ группировки всегда первый из leveling, значение с отступом по уровню
            indent = "  " * level  # два пробела на уровень
            summary_row = {leveling[0]: f"{indent}{key}"}
        else:
            summary_row = {group_field: key}

        for agg_field in aggregates:
            agg_func = agg_func_map.get(agg_field)
            if agg_func == "sum":
                raw_sum = pd.to_numeric(group_df[agg_field], errors='coerce').sum()
                if not pd.isna(raw_sum):
                    if _is_duration_field(agg_field, duration_fields, aggregate_fields_raw):
                        # Для duration полей округляем до целого числа и форматируем
                        summary_row[agg_field] = _format_duration_seconds(int(round(raw_sum)))
                    else:
                        # Для остальных полей округляем до 2 знаков после запятой
                        summary_row[agg_field] = round(raw_sum, 2)
                else:
                    summary_row[agg_field] = ""
            else:
                summary_row[agg_field] = ""

        summary_row["__row_type__"] = "summary"
        summary_row["__indent__"] = level
        return summary_row


    output_rows = []
    if df.empty:
        return output_rows
    if level >= len(leveling):
        for index, row in df.iterrows():
            output_rows.append(data_row(row, level))
        return output_rows

    group_field = leveling[level]
    grouped = df.groupby(group_field, dropna=False, sort=False)
    agg_func_map = build_agg_func_map(aggregate_fields_raw)

    for key, group_df in grouped:
        output_rows.append(summary_row(group_df, group_field))
        output_rows.extend(
            build_grouped_rows(params, group_df, leveling, aggregates, aggregate_fields_raw, together_merged_column, level + 1, parent_keys + (key,), duration_fields, aggregate_to_source_field)
            )

    return output_rows


def build_q_filter(filter_dict):
    """
    Строит Q-объект из словаря условий фильтрации.
    Поддерживает простые условия: {"field": value, "field__lookup": value}
    
    Args:
        filter_dict: Словарь с условиями фильтрации, например:
                     {"is_result": false} или {"created_at__year": 2025}
    
    Returns:
        Q-объект или None, если фильтр не задан
    """
    if not filter_dict or not isinstance(filter_dict, dict):
        return None
    
    q_filter = Q()
    for key, value in filter_dict.items():
        q_filter &= Q(**{key: value})
    
    return q_filter if q_filter.children else None


def current_month_date_range_strings():
    """
    Первая и последняя календарные даты текущего месяца (локальная дата Django) в формате YYYY-MM-DD.
    """
    today = timezone.localdate()
    first = date(today.year, today.month, 1)
    last_day = calendar.monthrange(today.year, today.month)[1]
    last = date(today.year, today.month, last_day)
    return first.isoformat(), last.isoformat()


def apply_filter_preset_to_report_metadata(metadata, filter_presets, preset_key):
    """
    Копия metadata с наложением filter_patch пресета на элементы metadata['filters'] по полю name.
    Ключи из patch (кроме name) мержатся в существующий dict фильтра.
    """
    if not preset_key:
        return copy.deepcopy(metadata)

    presets = filter_presets or {}
    preset = presets.get(preset_key)
    if not preset:
        raise ValidationError(
            {'filter_preset': [f'Пресет «{preset_key}» не найден в filter_presets этого отчёта.']}
        )

    patch_list = preset.get('filter_patch')
    if patch_list is None:
        raise ValidationError(
            {'filter_preset': [f'У пресета «{preset_key}» отсутствует filter_patch.']}
        )
    if not isinstance(patch_list, list):
        raise ValidationError(
            {'filter_preset': [f'filter_patch пресета «{preset_key}» должен быть списком.']}
        )

    out = copy.deepcopy(metadata)
    filters_list = out.get('filters') or []
    if not isinstance(filters_list, list):
        raise ValidationError({'metadata': ['metadata.filters должен быть списком.']})

    by_name = {}
    for idx, filt in enumerate(filters_list):
        if not isinstance(filt, dict):
            continue
        fname = filt.get('name')
        if fname is not None:
            by_name[fname] = idx

    for patch in patch_list:
        if not isinstance(patch, dict):
            continue
        fname = patch.get('name')
        if fname is None:
            raise ValidationError(
                {'filter_preset': ['Каждый элемент filter_patch должен содержать name.']}
            )
        idx = by_name.get(fname)
        if idx is None:
            raise ValidationError(
                {
                    'filter_preset': [
                        f'Фильтр «{fname}» из filter_patch не найден в metadata.filters этого отчёта.'
                    ]
                }
            )
        base = filters_list[idx]
        for key, val in patch.items():
            if key == 'name':
                continue
            base[key] = val

    out['filters'] = filters_list
    return out


def apply_default_active_date_filter_values(metadata):
    """
    Для ответа API (деталка шаблона): активные фильтры с типом даты/времени и value=null
    получают границы текущего месяца. Простой фильтр: value — пара [начало, конец], comparison "=".
    Сложный фильтр (complexFilter=true): в complexFilters для ">=" — начало месяца, для "<=" — конец (строка YYYY-MM-DD).
    Мутирует переданный dict (только payload ответа, не БД).
    """
    date_types = {
        'CustomDateField',
        'CustomDateTimeField',
        'DateField',
        'DateTimeField',
    }
    start_s, end_s = current_month_date_range_strings()

    last_30_fields = {'user__last_activity'}
    today = timezone.localdate()
    last_30_start = (today - timedelta(days=30)).isoformat()
    last_30_end = today.isoformat()

    for item in metadata.get('filters', []):
        if not item.get('active'):
            continue
        if item.get('type') not in date_types:
            continue
        if item.get('value') is not None:
            continue
        if item.get('name') in last_30_fields:
            item['value'] = [last_30_start, last_30_end]
        else:
            item['value'] = [start_s, end_s]

    if metadata.get('complexFilter'):
        for item in metadata.get('complexFilters', []):
            if not item.get('active'):
                continue
            if item.get('type') not in date_types:
                continue
            if item.get('value') is not None:
                continue
            if item.get('name') in last_30_fields:
                s_start, s_end = last_30_start, last_30_end
            else:
                s_start, s_end = start_s, end_s
            comparison = item.get('comparison_type')
            if comparison == '>=':
                item['value'] = s_start
            elif comparison == '<=':
                item['value'] = s_end

    return metadata


def resolve_base_report_settings(report_id):
    """Базовый ReportSettingsModel: user id → base_report, иначе сам отчёт; None без base_report; ValidationError если id не найден."""
    if not report_id:
        return None

    from . import models

    try:
        user_report = models.UserReportSettingsModel.objects.select_related('base_report').get(pk=report_id)
        return user_report.base_report
    except models.UserReportSettingsModel.DoesNotExist:
        pass
    try:
        return models.ReportSettingsModel.objects.get(pk=report_id)
    except models.ReportSettingsModel.DoesNotExist:
        raise ValidationError(f'Отчет с id={report_id} не найден')


def get_base_report_metadata(report_id):
    """
    Metadata только базового ReportSettingsModel (apexchart и прочее на уровне общего отчёта).
    Не подмешивает metadata пользовательского шаблона.
    """
    if not report_id:
        return {}
    report_settings = resolve_base_report_settings(report_id)
    if not report_settings:
        return {}
    return dict(report_settings.metadata or {})


def _apex_format_category(value, label_format):
    if value is None:
        return ''
    if isinstance(value, dict):
        if value.get('url'):
            value = value.get('repr', value.get('label', ''))
        else:
            value = value.get('repr', str(value))
    if label_format:
        dt_value = None
        if isinstance(value, datetime):
            dt_value = value
        elif isinstance(value, date):
            dt_value = datetime.combine(value, datetime.min.time())
        else:
            text = str(value).strip()
            dt_value = parse_datetime(text)
            if dt_value is None:
                date_part = text[:10] if len(text) >= 10 else text
                parsed_d = parse_date(date_part)
                if parsed_d is not None:
                    dt_value = datetime.combine(parsed_d, datetime.min.time())
        if dt_value is not None:
            try:
                return dt_value.strftime(label_format)
            except (ValueError, OSError):
                pass
    return try_repr(value)


def _apex_series_numeric(value):
    if value is None:
        return None
    if isinstance(value, float) and pd.isna(value):
        return None
    if isinstance(value, dict):
        if value.get('url'):
            return _apex_series_numeric(value.get('repr'))
        if not value:
            return None
        first_val = next(iter(value.values()), None)
        return _apex_series_numeric(first_val)
    try:
        return float(value)
    except (TypeError, ValueError):
        try:
            return float(str(value).replace(',', '.'))
        except (TypeError, ValueError):
            return value


def build_apexchart_payload(rows, apex_config, title_to_key=None):
    """series + chartOptions из metadata.apexchart; source_field/name задаются только человекочитаемыми title."""
    if not apex_config or not isinstance(apex_config, dict):
        raise ValidationError(
            'В metadata отчёта нет блока apexchart. Задайте его в настройках отчёта или проверьте id.'
        )
    xaxis_cfg = apex_config.get('xaxis') or {}
    x_source = xaxis_cfg.get('source_field')
    if not x_source:
        raise ValidationError('В apexchart.xaxis нужно указать source_field.')
    label_format = xaxis_cfg.get('label_format')
    series_cfg = apex_config.get('series')
    if not series_cfg or not isinstance(series_cfg, list):
        raise ValidationError('В apexchart.series ожидается непустой массив.')

    key_by_title = title_to_key or {}
    x_title = str(x_source)
    x_col = key_by_title.get(x_title)
    if x_col is None:
        raise ValidationError(
            f'Ось X «{x_title}» не найдена среди заголовков отчёта. '
            f'Укажите title поля группировки (как в groups[].title).'
        )
    series_names = []
    series_columns = []
    for series_item in series_cfg:
        if not isinstance(series_item, dict):
            continue
        if series_item.get('name') is None:
            raise ValidationError('Каждый элемент apexchart.series должен содержать name.')
        series_name = str(series_item['name'])
        series_names.append(series_name)
        series_col = key_by_title.get(series_name)
        if series_col is None:
            raise ValidationError(
                f'Серия «{series_name}» не найдена среди заголовков отчёта. '
                f'Укажите title агрегата (как в aggregates[].title).'
            )
        series_columns.append(series_col)
    if not series_names:
        raise ValidationError('В apexchart.series нет корректных элементов.')

    if rows:
        row_keys = set(rows[0])
        if x_col not in row_keys:
            raise ValidationError(
                f'Столбец оси X «{x_col}» не найден в данных отчёта. Проверьте xaxis.source_field и группировку.'
            )
        for series_name, series_col in zip(series_names, series_columns):
            if series_col not in row_keys:
                raise ValidationError(
                    f'Столбец серии «{series_name}» не найден в данных отчёта. Проверьте title в series.'
                )

    series_out = [
        {'name': series_name, 'data': [], 'source_col': series_col}
        for series_name, series_col in zip(series_names, series_columns)
    ]
    categories = []
    for row in rows:
        categories.append(_apex_format_category(row.get(x_col), label_format))
        for ser in series_out:
            ser['data'].append(_apex_series_numeric(row.get(ser['source_col'])))
    for ser in series_out:
        del ser['source_col']

    chart_options = copy.deepcopy(apex_config)
    chart_options.pop('series', None)

    chart_block = dict(chart_options.get('chart') or {})
    if not chart_block.get('type'):
        chart_block.setdefault('type', 'bar')
    chart_options['chart'] = chart_block

    xaxis_out = {
        key: val
        for key, val in (chart_options.get('xaxis') or {}).items()
        if key not in ('source_field', 'label_format')
    }
    xaxis_out['title'] = x_title
    xaxis_out['categories'] = categories
    chart_options['xaxis'] = xaxis_out
    yaxis_cfg = chart_options.get('yaxis')
    if isinstance(yaxis_cfg, dict):
        yo = dict(yaxis_cfg)
        if isinstance(yo.get('title'), str):
            yo['title'] = {'text': yo['title']}
        chart_options['yaxis'] = yo

    return {'series': series_out, 'chartOptions': chart_options}


def get_template_from_report(report_id):
    """
    Устарело: единственный источник правды — get_template_path_from_report.
    """
    return None


def get_template_path_from_report(report_id):
    """
    Возвращает путь к файлу шаблона Excel для отчёта (из report_templates).
    """
    if not report_id:
        return None

    import os
    from django.conf import settings

    try:
        report_settings = resolve_base_report_settings(report_id)
    except ValidationError:
        return None

    if not report_settings or not report_settings.template_path:
        return None
    path = os.path.join(settings.MEDIA_ROOT, report_settings.template_path)
    return path if os.path.exists(path) else None


def extract_annotations(model_class, aggregate_fields):
    annotations = {}
    for agg in aggregate_fields:
        name = agg.get("name")
        if not name:
            continue

        # Построение фильтра если указан
        q_filter = build_q_filter(agg.get("filter"))

        if "sum" in agg:
            sum_source = agg["sum"]
            field_type = get_report_field_type(model_class, sum_source)
            if field_type == "BooleanField":
                annotations[name] = Sum(
                    models.Case(
                        models.When(**{sum_source: True}, then=models.Value(1)),
                        default=models.Value(0),
                        output_field=models.IntegerField(),
                    ),
                    filter=q_filter,
                )
            else:
                annotations[name] = Sum(sum_source, filter=q_filter)
        elif "avg" in agg:
            annotations[name] = Avg(agg["avg"], filter=q_filter)
        elif "min" in agg:
            annotations[name] = Min(agg["min"], filter=q_filter)
        elif "max" in agg:
            annotations[name] = Max(agg["max"], filter=q_filter)
        elif "count" in agg:
            annotations[name] = Count(agg["count"], filter=q_filter)
        elif "distinct_count" in agg:
            annotations[name] = Count(agg["distinct_count"], distinct=True, filter=q_filter)
        elif "concatenate" in agg:
            field = agg["concatenate"]
            delimiter = agg.get("delimiter", ", ")
            has_created_at = any(f.name == 'created_at' for f in model_class._meta.fields)
            
            # Объединяем фильтр concatenate с пользовательским фильтром
            base_filter = ~Q(**{f"{field}__isnull": True}) & ~Q(**{f"{field}": ""})
            if q_filter:
                base_filter &= q_filter
            
            annotations[name] = StringAgg(
                F(field),
                delimiter=delimiter,
                ordering='created_at' if has_created_at else 'id',
                filter=base_filter
            )
    return annotations


def parse_grand_totals_config(params, numeric_aggregate_fields, df, aggregate_fields_raw=None):
    """
    Парсит конфигурацию grand_totals из параметров и возвращает настройки для отображения итогов.
    
    Args:
        params: Параметры отчета
        numeric_aggregate_fields: Множество числовых агрегатных полей
        df: DataFrame с данными
    
    Returns:
        tuple: (show_totals, fields_for_totals, totals_title)
    """
    totals_param = params.get("totals")
    grand_totals_config = None

    if totals_param:
        try:
            if isinstance(totals_param, str):
                totals_param = json.loads(totals_param)
            grand_totals_config = totals_param.get("grand_totals")
        except Exception:
            grand_totals_config = None

    show_totals = True
    # Базовые поля для итогов: все числовые агрегаты, КРОМЕ avg по умолчанию
    if aggregate_fields_raw is not None:
        avg_fields = {f.get('name') for f in aggregate_fields_raw if f.get('name') and ('avg' in f)}
        fields_for_totals = [f for f in numeric_aggregate_fields if f not in avg_fields]
    else:
        fields_for_totals = list(numeric_aggregate_fields)
    totals_title = "ВСЕГО"

    if grand_totals_config:
        if grand_totals_config.get("position") is None:
            show_totals = False
        elif grand_totals_config.get("position") != "bottom":
            show_totals = False
        elif grand_totals_config.get("by") != "columns":
            show_totals = False
        else:
            totals_title = grand_totals_config.get("title", totals_title)
            fields_for_totals = grand_totals_config.get("fields") or fields_for_totals
            fields_for_totals = [f for f in fields_for_totals if f in df.columns]

    return show_totals, fields_for_totals, totals_title


def prepare_report_data(params, model_class, df, leveling, aggregates, aggregate_fields_raw, 
                       leveling_raw, group_fields_raw, system_fields_raw, columns_order):
    """
    Универсальная функция для подготовки итоговых данных отчетов.
    
    Подготавливает данные с учетом группировок, промежуточных и общих итогов,
    возвращает готовую таблицу для экспорта в HTML или Excel.
    
    Args:
        params: Параметры отчета
        model_class: Класс модели Django
        df: Исходный DataFrame с данными
        leveling: Список полей для группировки
        aggregates: Список агрегатных полей
        aggregate_fields_raw: Сырые данные агрегатных полей
        leveling_raw: Сырые данные полей группировки
        group_fields_raw: Сырые данные групповых полей
        system_fields_raw: Сырые данные системных полей
        columns_order: Порядок колонок
    
    Returns:
        tuple: (output_rows, final_df, numeric_aggregate_fields, show_totals, fields_for_totals, totals_title)
    """
    # Определяем числовые агрегатные поля
    numeric_aggregate_fields = {
        field["name"] for field in aggregate_fields_raw
        if not any(k in field for k in ("concatenate",))
    }
    
    # Определяем duration поля из computed fields метаданных (основная и связанные модели)
    duration_fields = _get_duration_computed_field_names(model_class)
    for agg in aggregate_fields_raw:
        source = agg.get("sum") or agg.get("avg") or agg.get("min") or agg.get("max")
        if not source or "__" not in source:
            continue
        try:
            prefix_path, _ = source.rsplit("__", 1)
            _, related_model = get_field_meta(prefix_path, model_class)
            duration_fields |= _get_duration_computed_field_names(related_model, prefix_path)
        except Exception:
            continue
    
    # Числовые столбцы приводим к числам в df до to_dict/build_grouped_rows — тогда в output_rows уже числа
    numeric_columns = _excel_numeric_columns(columns_order, model_class, numeric_aggregate_fields)
    for col in numeric_columns:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")
    
    # Обрабатываем None значения в группировочных полях один раз в начале
    # Это нужно для корректной группировки: None и "Не заполнено" должны быть в одной группе
    replace_none_with_empty_for_leveling(df, leveling)
    
    # Сбор итогов и строк
    if not leveling:
        output_rows = df.to_dict("records")
        for row in output_rows:
            row["__row_type__"] = "data"
            row["__indent__"] = 0
    else:
        # Для layout == "together" определяем, какой столбец объединяется с leveling[0]
        layout = params.get("grouping_fields_layout", "together")
        together_merged_column = None
        if layout == "together" and leveling:
            leveling_index = columns_order.index(leveling[0])
            data_columns_after_leveling = [col for col in columns_order[leveling_index + 1:] if col not in leveling]
            if data_columns_after_leveling:
                together_merged_column = data_columns_after_leveling[0]
        output_rows = build_grouped_rows(params, df, leveling, aggregates, aggregate_fields_raw, together_merged_column, duration_fields=duration_fields, aggregate_to_source_field=None)
    
    # Форматируем duration поля в output_rows после суммирования
    for row in output_rows:
        for col in columns_order:
            if col in row and _is_duration_field(col, duration_fields, aggregate_fields_raw):
                row[col] = _format_duration_value(row[col])
    
    # Создаем финальный DataFrame после форматирования
    final_df = pd.DataFrame(output_rows).fillna("")
    if not final_df.empty:
        final_df = final_df[[col for col in columns_order if col in final_df.columns] + ["__row_type__", "__indent__"]]
    else:
        final_df = pd.DataFrame(columns=columns_order)  # Создаем пустой DataFrame с колонками
    
    # Парсим конфигурацию общих итогов
    show_totals, fields_for_totals, totals_title = parse_grand_totals_config(params, numeric_aggregate_fields, df, aggregate_fields_raw)
    
    # Рассчитываем общие итоги на основе исходного df
    grand_totals_row = None
    if show_totals and fields_for_totals and not df.empty:
        grand_totals_row = {col: "" for col in columns_order}
        total_row_label_written = False
        
        for col in columns_order:
            if col in fields_for_totals and col in df.columns:
                raw_total = pd.to_numeric(df[col], errors='coerce').sum()
                if not pd.isna(raw_total):
                    if _is_duration_field(col, duration_fields, aggregate_fields_raw):
                        grand_totals_row[col] = _format_duration_seconds(int(round(raw_total)))
                    else:
                        grand_totals_row[col] = round(raw_total, 2)
                else:
                    grand_totals_row[col] = ""
            elif not total_row_label_written:
                grand_totals_row[col] = totals_title
                total_row_label_written = True
    
    return output_rows, final_df, numeric_aggregate_fields, show_totals, fields_for_totals, totals_title, grand_totals_row


def build_table_structure(params, final_df, columns_order, leveling):
    """
    Определяет структуру таблицы в зависимости от layout.
    
    Args:
        params: Параметры отчета
        final_df: DataFrame с данными
        columns_order: Порядок колонок (только видимые поля)
        leveling: Список полей для группировки
    
    Returns:
        dict: Структура таблицы с полями all_columns, leveling_index, first_data_column
    """
    layout = params.get("grouping_fields_layout", "together")
    
    if layout == "together" and leveling:
        # Проверяем, есть ли leveling[0] в видимых колонках
        if leveling[0] in columns_order:
            leveling_index = columns_order.index(leveling[0])
            data_columns_after_leveling = [col for col in columns_order[leveling_index + 1:] if col not in leveling]
            first_data_column = data_columns_after_leveling[0] if data_columns_after_leveling else None
            
            # Исключаем ВСЕ поля leveling кроме первого, а также first_data_column (который объединяется с первым)
            all_columns = []
            for col in columns_order:
                if col == leveling[0]:
                    # Первый leveling столбец остается (в него "запихиваются" остальные)
                    all_columns.append(col)
                elif col in leveling:
                    # Остальные leveling столбцы исключаем (они показываются в первом столбце)
                    continue
                elif col == first_data_column:
                    # first_data_column тоже исключаем (он объединяется с первым leveling)
                    continue
                else:
                    # Остальные столбцы добавляем
                    all_columns.append(col)
        else:
            # Если первый leveling столбец невидимый, переходим к обычной логике
            all_columns = columns_order[:]
            leveling_index = None
            first_data_column = None
    else:
        all_columns = columns_order[:]
        leveling_index = None
        first_data_column = None
    
    return {
        'layout': layout,
        'all_columns': all_columns,
        'leveling_index': leveling_index,
        'first_data_column': first_data_column
    }


def build_headers_config(params, leveling, leveling_raw, group_fields_raw, aggregate_fields_raw, 
                        system_fields_raw, columns_order, model_class, table_structure):
    """
    Создает конфигурацию заголовков для таблицы.
    
    Args:
        params: Параметры отчета
        leveling: Список полей для группировки
        leveling_raw: Сырые данные полей группировки
        group_fields_raw: Сырые данные групповых полей
        aggregate_fields_raw: Сырые данные агрегатных полей
        system_fields_raw: Сырые данные системных полей
        columns_order: Порядок колонок
        model_class: Класс модели Django
        table_structure: Структура таблицы из build_table_structure
    
    Returns:
        dict: Конфигурация заголовков
    """
    layout = table_structure['layout']
    all_columns = table_structure['all_columns']
    first_data_column = table_structure['first_data_column']
    
    # Получаем aggregates из aggregate_fields_raw
    aggregates = [agg.get("name") for agg in aggregate_fields_raw if agg.get("name")]
    
    if layout == "together" and leveling and leveling[0] in all_columns:
        # Определяем структуру заголовков
        leveling_titles = [get_verbose_title(model_class, lev, leveling_raw, group_fields_raw, aggregates, system_fields_raw) for lev in leveling]
        
        # Находим первый не-групповой столбец, который идёт ПОСЛЕ первого leveling
        first_data_title = ""
        if first_data_column:
            first_data_title = get_verbose_title(model_class, first_data_column, leveling_raw, group_fields_raw, aggregates, system_fields_raw)
        
        # Высота стека заголовков
        stack_height = len(leveling_titles) + (1 if first_data_title else 0)
        
        # Формируем заголовки: до leveling, в leveling (стек), после leveling
        leveling_index = all_columns.index(leveling[0])
        cols_before_leveling = all_columns[:leveling_index]
        cols_after_leveling = all_columns[leveling_index + 1:]
        
        # Формируем header_stack для многострочных заголовков
        header_stack = leveling_titles + ([first_data_title] if first_data_title else [])
        
        return {
            'type': 'multi_row',
            'stack_height': stack_height,
            'header_stack': header_stack,
            'cols_before_leveling': cols_before_leveling,
            'cols_after_leveling': cols_after_leveling,
            'leveling_titles': leveling_titles,
            'first_data_title': first_data_title
        }
    else:
        # Обычные заголовки (одна строка)
        return {
            'type': 'single_row',
            'columns': all_columns
        }


def render_excel_headers(ws, row_idx, headers_config, table_structure, model_class, 
                        leveling_raw, group_fields_raw, aggregate_fields_raw, system_fields_raw,
                        named_styles_dict, start_col_idx=1):
    """
    Рендерит заголовки таблицы в Excel.
    
    Returns:
        int: Следующий номер строки после заголовков
    """
    aggregates = [agg.get("name") for agg in aggregate_fields_raw if agg.get("name")]
    
    if headers_config['type'] == 'multi_row':
        # Многострочные заголовки для layout=together
        stack_height = headers_config['stack_height']
        header_stack = headers_config['header_stack']
        cols_before_leveling = headers_config['cols_before_leveling']
        cols_after_leveling = headers_config['cols_after_leveling']
        
        col_idx = start_col_idx
        # Столбцы до leveling
        for col in cols_before_leveling:
            title = get_verbose_title(model_class, col, leveling_raw, group_fields_raw, aggregates, system_fields_raw)
            cell = ws.cell(row=row_idx, column=col_idx, value=force_str(title))
            
            # Для Excel: применяем границы ко всем ячейкам в диапазоне ДО объединения
            # Это гарантирует, что правая граница будет отображаться корректно
            table_header_style = named_styles_dict.get('table_header')
            if stack_height > 1 and table_header_style and hasattr(table_header_style, 'border') and table_header_style.border:
                for r in range(row_idx, row_idx + stack_height):
                    temp_cell = ws.cell(row=r, column=col_idx)
                    temp_cell.border = table_header_style.border
            
            # Объединяем ячейки по вертикали
            if stack_height > 1:
                ws.merge_cells(start_row=row_idx, start_column=col_idx, end_row=row_idx + stack_height - 1, end_column=col_idx)
            
            # Применяем стили
            if table_header_style:
                _apply_style(cell, 'table_header', named_styles_dict)
            else:
                cell.style = 'Normal'
            col_idx += 1
        
        # Leveling столбец: вертикальный стек
        for i, title in enumerate(header_stack):
            cell = ws.cell(row=row_idx + i, column=col_idx, value=force_str(title))
            # Применяем стили
            if named_styles_dict.get('table_header'):
                _apply_style(cell, 'table_header', named_styles_dict)
            else:
                cell.style = 'Normal'
        col_idx += 1
        
        # Столбцы после leveling: объединяем по вертикали
        for col in cols_after_leveling:
            title = get_verbose_title(model_class, col, leveling_raw, group_fields_raw, aggregates, system_fields_raw)
            cell = ws.cell(row=row_idx, column=col_idx, value=force_str(title))
            
            # Для Excel: применяем границы ко всем ячейкам в диапазоне ДО объединения
            # Это гарантирует, что правая граница будет отображаться корректно
            table_header_style = named_styles_dict.get('table_header')
            if stack_height > 1 and table_header_style and hasattr(table_header_style, 'border') and table_header_style.border:
                for r in range(row_idx, row_idx + stack_height):
                    temp_cell = ws.cell(row=r, column=col_idx)
                    temp_cell.border = table_header_style.border
            
            # Объединяем ячейки по вертикали
            if stack_height > 1:
                ws.merge_cells(start_row=row_idx, start_column=col_idx, end_row=row_idx + stack_height - 1, end_column=col_idx)
            
            # Применяем стили
            if table_header_style:
                _apply_style(cell, 'table_header', named_styles_dict)
            else:
                cell.style = 'Normal'
            col_idx += 1
        
        return row_idx + stack_height
        
    else:
        # Обычные заголовки (одна строка)
        columns = headers_config['columns']
        for col_idx, col in enumerate(columns, start=start_col_idx):
            title = get_verbose_title(model_class, col, leveling_raw, group_fields_raw, aggregates, system_fields_raw)
            cell = ws.cell(row=row_idx, column=col_idx, value=force_str(title))
            # Применяем стили
            if named_styles_dict.get('table_header'):
                _apply_style(cell, 'table_header', named_styles_dict)
            else:
                cell.style = 'Normal'
        return row_idx + 1


def render_html_headers(headers_config, model_class, leveling_raw, group_fields_raw, 
                       aggregate_fields_raw, system_fields_raw, columns_order):
    """
    Рендерит заголовки таблицы в HTML.
    
    Returns:
        str: HTML-строка с заголовками
    """
    aggregates = [agg.get("name") for agg in aggregate_fields_raw if agg.get("name")]
    html_output = ""
    
    if headers_config['type'] == 'multi_row':
        # Многострочные заголовки для layout=together
        stack_height = headers_config['stack_height']
        header_stack = headers_config['header_stack']
        cols_before_leveling = headers_config['cols_before_leveling']
        cols_after_leveling = headers_config['cols_after_leveling']
        
        # Первая строка: столбцы до leveling + leveling (первый элемент) + столбцы после leveling
        html_output += "<tr>"
        # Столбцы до leveling
        for col in cols_before_leveling:
            th_title = get_verbose_title(model_class, col, leveling_raw, group_fields_raw, aggregates, system_fields_raw)
            html_output += f'<th rowspan="{stack_height}">{th_title}</th>'
        # Leveling столбец (первый элемент стека)
        html_output += f'<th>{force_str(header_stack[0])}</th>'
        # Столбцы после leveling
        for col in cols_after_leveling:
            th_title = get_verbose_title(model_class, col, leveling_raw, group_fields_raw, aggregates, system_fields_raw)
            html_output += f'<th rowspan="{stack_height}">{th_title}</th>'
        html_output += "</tr>"
        # Остальные строки: только leveling столбец (продолжение header_stack)
        for i in range(1, len(header_stack)):
            html_output += f'<tr><th>{force_str(header_stack[i])}</th></tr>'
    else:
        # Обычные заголовки (одна строка)
        columns = headers_config['columns']
        html_output += "<tr>"
        for col in columns:
            title = get_verbose_title(model_class, col, leveling_raw, group_fields_raw, aggregates, system_fields_raw)
            html_output += f"<th>{title}</th>"
        html_output += "</tr>"
    
    return html_output


def merge_leveling_with_ordering(ordering_parts, leveling, model_class=None):
    """
    Объединяет ordering с leveling, гарантируя что поля группировки идут первыми.
    Сохраняет направление сортировки (asc/desc) из ordering.
    
    Args:
        ordering_parts: список полей сортировки из запроса
        leveling: список полей группировки
        model_class: класс модели (опционально, для валидации)
    
    Returns:
        список полей сортировки с учетом группировки
    """
    if not leveling:
        return ordering_parts
    
    # Создаем словарь: поле → направление сортировки из ordering
    ordering_directions = {}
    for ord_field in ordering_parts:
        is_desc = ord_field.startswith('-')
        ord_field_clean = ord_field.lstrip('-')
        ord_field_short = ord_field_clean.split('__')[-1]
        ordering_directions[ord_field_clean] = is_desc
        ordering_directions[ord_field_short] = is_desc
    
    # Строим новый ordering начиная с leveling (в правильном порядке)
    new_ordering = []
    for lev_field in leveling:
        lev_field_short = lev_field.split('__')[-1]
        # Проверяем указано ли направление сортировки
        if lev_field in ordering_directions:
            is_desc = ordering_directions[lev_field]
        elif lev_field_short in ordering_directions:
            is_desc = ordering_directions[lev_field_short]
        else:
            is_desc = False  # по умолчанию asc
        
        # Добавляем с правильным направлением
        new_ordering.append(f"-{lev_field}" if is_desc else lev_field)
    
    # Добавляем остальные поля из ordering (которых нет в leveling)
    for ord_field in ordering_parts:
        ord_field_clean = ord_field.lstrip('-')
        ord_field_short = ord_field_clean.split('__')[-1]
        
        # Проверяем есть ли в leveling
        found = False
        for lev_field in leveling:
            lev_field_short = lev_field.split('__')[-1]
            if ord_field_clean == lev_field or ord_field_short == lev_field_short:
                found = True
                break
        
        if not found:
            new_ordering.append(ord_field)
    
    return new_ordering


def get_computed_order_field(model_class, computed_field_name):
    """
    Возвращает поле для сортировки вычисляемого поля.
    - Если указан order_by_field в метаданных → возвращает его
    - Если order_by_field не указан → возвращает само имя поля (для сортировки по аннотации)
    - Если поле не вычисляемое → возвращает None
    """
    if not hasattr(model_class, 'get_report_computed_fields_meta'):
        return None
    try:
        for field_meta in model_class.get_report_computed_fields_meta():
            if field_meta.get('name') == computed_field_name:
                return field_meta.get('order_by_field') or computed_field_name
    except Exception:
        pass
    return None


def _replace_computed_field_for_ordering(field_name, model_class):
    """
    Заменяет вычисляемое поле на order_by_field для сортировки.
    Возвращает кортеж (replaced_field, should_continue), где:
    - replaced_field - замененное поле или исходное, если замена не нужна
    - should_continue - True если поле было обработано (вычисляемое), False иначе
    """
    if '__' in field_name:
        # Связанное вычисляемое поле
        relation_path, computed_field_name = field_name.rsplit('__', 1)
        try:
            _, related_model = get_field_meta(relation_path, model_class)
            if related_model:
                order_by_field = get_computed_order_field(related_model, computed_field_name)
                if order_by_field:
                    if order_by_field != computed_field_name:
                        return f"{relation_path}__{order_by_field}", True
                    return field_name, True
        except Exception:
            pass
    else:
        # Локальное вычисляемое поле
        order_by_field = get_computed_order_field(model_class, field_name)
        if order_by_field:
            if order_by_field != field_name:
                return order_by_field, True
            return field_name, True
    
    return field_name, False


def expand_ordering_fields(ordering_fields, model_class):
    """
    Расширяет поля сортировки, заменяя ForeignKey на соответствующие поля связанной модели.
    Чтобы при сортировке по ['author'] сортировка была по ФИО (get_order_param у модели).
    Также обрабатывает вычисляемые поля (computed fields), заменяя их на order_by_field из метаданных.
    Также добавляет стабильную сортировку (created_at/id) если её нет.
    """
    if not model_class:
        return ordering_fields if ordering_fields else []
    
    expanded = []
    
    # Расширяем переданные поля
    if ordering_fields:
        for part in ordering_fields:
            is_desc = part.startswith('-')
            field_name = part[1:] if is_desc else part
            sign = '-' if is_desc else ''
            
            # Обработка вычисляемых полей (включая связанные, например task__task_link)
            replaced_field, was_computed = _replace_computed_field_for_ordering(field_name, model_class)
            if was_computed:
                expanded.append(f"{sign}{replaced_field}")
                continue
            
            # Пытаемся расширить ForeignKey до полей сортировки связанной модели
            try:
                field = model_class._meta.get_field(field_name)
                if isinstance(field, (models.ForeignKey, models.OneToOneField)):
                    related_model = field.related_model
                    if hasattr(related_model, 'get_order_param'):
                        order_params = related_model.get_order_param()
                        if order_params:
                            for param in order_params:
                                param_str = str(param)
                                param_is_desc = param_str.startswith('-')
                                param_clean = param_str.lstrip('-')
                                final_is_desc = param_is_desc or is_desc
                                param_sign = '-' if final_is_desc else ''
                                expanded.append(f"{param_sign}{field_name}__{param_clean}")
                            continue
            except Exception:
                pass
            
            # Стандартная обработка - добавляем как есть
            expanded.append(part)
    
    # Добавляем стабильную сортировку по id/created_at для детерминированного порядка
    # ВСЕГДА (даже если ordering_fields был пустой)
    has_stable_sort = any(
        part.lstrip('-') in ('id', 'pk', 'created_at') 
        for part in expanded
    )
    if not has_stable_sort:
        # Добавляем created_at если есть, иначе id
        has_created_at = any(f.name == 'created_at' for f in model_class._meta.fields)
        if has_created_at:
            expanded.append('-created_at')
        else:
            expanded.append('id')
    
    return expanded


def compute_final_ordering(ordering_parts, leveling, model_class, group_fields=None, annotations=None):
    """
    Полный конвейер вычисления итогового order_by:
    - объединение с leveling
    - расширение FK до полей сортировки связанных моделей + стабильная сортировка
    - фильтрация при группировке, чтобы не разворачивать values()
    - детерминированный фолбэк при пустом order_by

    Args:
        ordering_parts: список полей сортировки из запроса
        leveling: список полей группировки (для приоритета в ordering)
        model_class: модель
        group_fields: фактические group_fields (values(...))
        annotations: доступные аннотации (dict), ключи которых можно использовать в order_by
    """
    ordering_parts = merge_leveling_with_ordering(ordering_parts or [], leveling, model_class)
    ordering_parts = expand_ordering_fields(ordering_parts, model_class)

    # Фильтруем ordering_parts, чтобы не разворачивать values() группировку
    # Разрешаем только поля, которые есть в values() или начинаются с них + "__"
    if group_fields:
        available_fields = set(group_fields)
        if annotations:
            available_fields.update(annotations.keys())
        
        ordering_parts = [
            f for f in ordering_parts 
            if (f.lstrip('-') in available_fields or 
                any(f.lstrip('-').startswith(f"{field}__") for field in available_fields))
        ]

    if not ordering_parts:
        if group_fields:
            ordering_parts = list(group_fields)
        else:
            has_created_at = bool(model_class) and any(f.name == 'created_at' for f in model_class._meta.fields)
            ordering_parts = ['-created_at'] if has_created_at else ['id']

    return ordering_parts


def replace_none_with_empty_for_leveling(df, leveling):
    """
    Заменяет None/NaN значения на "Не заполнено" для полей leveling в DataFrame.
    Это нужно для корректной группировки: None и "Не заполнено" должны быть в одной группе.
    
    Args:
        df: pandas DataFrame
        leveling: список полей группировки
    
    Returns:
        DataFrame с замененными значениями (изменяется in-place, но возвращается для удобства)
    """
    if leveling:
        for field in leveling:
            if field in df.columns:
                def normalize_leveling_value(value):
                    # Нормализуем NaN/None под одно значение для группировки
                    if value is None:
                        return "Не заполнено"
                    try:
                        if pd.isna(value):
                            return "Не заполнено"
                    except Exception:
                        pass
                    if isinstance(value, str) and value in ("nan", "None"):
                        return "Не заполнено"

                    # Если в leveling внезапно прилетает список (например list(dict) для FK-list),
                    # приводим к строке, чтобы pandas/groupby мог работать.
                    if isinstance(value, (list, tuple, set)):
                        if not value:
                            return ""
                        if all(isinstance(item, dict) for item in value):
                            parts = []
                            for item in value:
                                part = item.get("repr") or item.get("url") or ""
                                part = force_str(part)
                                if part:
                                    parts.append(part)
                            return ", ".join(parts) if parts else ""
                        parts = [force_str(item) for item in value if item not in (None, "")]
                        return ", ".join(parts)

                    return value

                df[field] = df[field].apply(normalize_leveling_value)
    return df


def add_index_fields_with_pandas(data, ordering_fields, leveling, system_fields_requested):
    """
    Добавляет нумерационные поля (index и group_index) через pandas.
    Выполняется после получения данных из БД и сортировки.
    
    Args:
        data: список словарей (результат serializer.data)
        ordering_fields: список полей сортировки (после compute_final_ordering)
        leveling: список полей группировки
        system_fields_requested: список запрошенных system_fields (например, ["index", "group_index"])
    
    Returns:
        список словарей с добавленными полями index и/или group_index
    """
    if not data or not system_fields_requested:
        return data
    
    # Создаем DataFrame из данных
    df = pd.DataFrame(data)
    
    if df.empty:
        return data
    
    # Обрабатываем None значения в группировочных полях перед группировкой
    # Это нужно для корректной группировки: None и "Не заполнено" должны быть в одной группе
    replace_none_with_empty_for_leveling(df, leveling)
    
    # Добавляем index (сквозная нумерация)
    if "index" in system_fields_requested:
        df['index'] = range(1, len(df) + 1)
    
    # Добавляем group_index (нумерация внутри групп leveling)
    if "group_index" in system_fields_requested:
        if leveling:
            # Проверяем, что все поля leveling присутствуют в DataFrame
            leveling_present = [field for field in leveling if field in df.columns]
            if leveling_present:
                # Группируем по полям leveling и нумеруем внутри каждой группы
                # dropna=False гарантирует, что "Не заполнено" будет правильно обработано
                df['group_index'] = df.groupby(leveling_present, dropna=False).cumcount() + 1
            else:
                # Если полей leveling нет в DataFrame, делаем простую нумерацию
                df['group_index'] = range(1, len(df) + 1)
        else:
            # Если leveling нет, делаем простую нумерацию
            df['group_index'] = range(1, len(df) + 1)
    
    # Преобразуем обратно в список словарей
    return df.to_dict('records')


