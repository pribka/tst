import secrets
import string
import json
import base64
import datetime
import email
from email.utils import make_msgid, parseaddr
from email.header import Header
import io
import logging
import math
import re
import uuid
from typing import Optional, Union, List, Tuple
from urllib.parse import urlparse, parse_qs, unquote, urlsplit

import openpyxl
import requests
import telebot
from bs4 import BeautifulSoup
from email import encoders
from email.message import EmailMessage
from email.mime.base import MIMEBase
from email.mime.message import MIMEMessage
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.utils import parsedate_tz, mktime_tz, formatdate
from imapclient import IMAPClient
from smtplib import SMTP, SMTP_SSL
from telebot import types as tg_types

from django.core.exceptions import ValidationError, ObjectDoesNotExist
from django.core.files import File as DjangoFile
from django.core.serializers.json import DjangoJSONEncoder
from django.db import transaction
from django.db.models import Sum, Q
from django.utils import timezone
from django_q.tasks import async_task

from rest_framework.response import Response
from rest_framework import exceptions as drf_exceptions

from bkz3.settings import FRONTEND_URL, BACKEND_URL, DOWNLOADER_PATH, SOCKETIO_SYSTEM_CHANNEL

from common.models import File, FileBaseModel
from common.catalogs.models import ContractorModel, ContractorProfileModel
from common.redis import socketio_redis
from common.utils import send_socketio_about_update_current_work

from tags.serializers import TagModelCreateSerializer
from users.models import ProfileModel
from users.utils import get_tree_departments_related_organizations

from contractor_permissions.utils import check_contractor_permission

from . import models


def handle_command_start(message, bot):
    telegram_id = message.from_user.id
    try:
        org_admin = ContractorModel.objects.get(help_desk_config__telegram_token=bot.token)
    except (ValidationError, ObjectDoesNotExist):
        # Ответить что мы вас не знаем
        return Response("ne ok")
    try:
        contact_person = models.ContactPersonModel.objects.get(
            customer_card__org_admin=org_admin,
            telegram_id=telegram_id,
        )
    except ObjectDoesNotExist:
        contractor_profile_id = message.text.split(' ')[-1]
        telegram_user = message.from_user.username
        full_name = message.from_user.full_name
        if contractor_profile_id and not contractor_profile_id == '/start':
            try:
                contractor_profile = ContractorProfileModel.objects.get(pk=contractor_profile_id)
            except (ValidationError, ObjectDoesNotExist):
                #  TODO вставить тут?
                bot.send_message(
                    chat_id=telegram_id,
                    text=f'❗️ Возникла ошибка при запуске бота.\n'
                         f'Пожалуйста, повторите попытку через несколько минут.\n'
                         f'Мы уже уведомили техническую команду и работаем над устранением проблемы.\n'
                         f'🔁 Чтобы перезапустить — нажмите /start'
                )
                return Response('contractor_profile not found')
            try:
                customer_card = models.CustomerCardModel.objects.get(
                    org_admin=org_admin,
                    customer=contractor_profile.contractor
                )
            except (ValidationError, ObjectDoesNotExist):
                # Отправляем что организация не найдена
                return Response('customer_card not found')
            contact_person, created = models.ContactPersonModel.objects.update_or_create(
                customer_card=customer_card,
                telegram_id=telegram_id,
                defaults={
                    'name': full_name,
                    'user': contractor_profile.user,
                    'telegram': telegram_user,
                }
            )
            bot.send_message(
                chat_id=telegram_id,
                text=f'Добро пожаловать в техподдержку {org_admin.name}! Вы успешно подключены. '
                     f'Напишите свой вопрос и опишите проблему. '
                     f'Мы поможем вам как можно скорее.'
            )
            return Response('ok')
        else:
            # TODO создаем лида и неизвестный контакт
            unknown_customer_card = get_or_create_unknown_customer_card(org_admin)
            contact_person = models.ContactPersonModel.objects.create(
                customer_card=unknown_customer_card,
                telegram_id=telegram_id,
                name=full_name,
                telegram=telegram_user,
                unknown=True,
            )
            bot.send_message(
                chat_id=telegram_id,
                text=f'Добро пожаловать в техподдержку {org_admin.name}! Вы успешно подключены. '
                     f'Напишите свой вопрос и опишите проблему. '
                     f'Мы поможем вам как можно скорее.'
            )
            # bot.send_message(
            #     text=f'⚠️ Вы открыли бота по ссылке, но не прошли авторизацию.\n'
            #          f'В целях безопасности, пожалуйста, выполните вход через {FRONTEND_URL}. \n'
            #          f'После этого вы сможете продолжить общение с техподдержкой.',
            #     chat_id=telegram_id
            # )
            return Response('ok')
    else:
        if contact_person.spam:
            return Response('ok')
        bot.send_message(
            text='Вы уже запустили бот. Добро пожаловать. Снова.',
            chat_id=telegram_id
        )
        return Response('ok')


def get_emails(config):
    """"
    Подключается к imap-серверу из конфигураци организации и возвращает список сообщений,
    полученных после последнего отправленного сообщения.
    config может быть объектом HelpDeskConfigModel или строкой с id записи HelpDeskConfigModel.
    """
    if isinstance(config, str):
        config = models.HelpDeskConfigModel.objects.get(pk=config)
    imap_server = config.imap_server
    imap_port = config.imap_port
    email_username = config.email_username
    email_pass = config.email_pass
    if not imap_server or not imap_port or not email_username or not email_pass:
        raise drf_exceptions.ValidationError('Отсутствуют настройки подключения')
    last_uid = config.email_last_uid
    if last_uid:
        lookup = (("UID", f"{last_uid}:*"), ("NOT", ("UID", last_uid)))
    else:
        lookup = ("SINCE", timezone.localdate())
    contact_persons = list(models.ContactPersonModel.objects.filter(
        is_active=True,
        customer_card__org_admin=config.contractor
    ).values('email', 'id', 'spam'))

    with IMAPClient(host=imap_server, port=imap_port) as imap:
        imap.login(email_username, email_pass)
        imap.select_folder('INBOX', readonly=True)
        messages = imap.search(lookup)
        last_letter_uid = None
        letter_count = 0
        for letter_uid in messages:
            email_message = imap.fetch(letter_uid, "RFC822")
            parsed_message = email.message_from_bytes(
                email_message[letter_uid][b"RFC822"],
                policy=email.policy.default,
                _class=EmailMessage
            )
            from_address = get_from_address(parsed_message)
            if not from_address:
                continue
            contact_person_id, created = get_or_create_contact_person_from_email_message(
                config,
                from_address,
                contact_persons
            )
            if not contact_person_id:
                continue
            if created:
                contact_persons.append({'email': from_address, 'id': contact_person_id, 'spam': False})
            body = parsed_message.get_body(preferencelist=('html', 'plain', 'related'))  # noqa
            body_string = body.get_payload(decode='utf-8')
            if body_string is None:
                body_string = ''
            try:
                body_string = body_string.decode('utf-8')
            except AttributeError:
                pass
            except UnicodeDecodeError:
                try:
                    body_string = body_string.decode('latin-1')
                except UnicodeDecodeError:
                    body_string = '*******Неизвестная кодировка********'
            soup = BeautifulSoup(body_string, 'lxml')
            image_urls = set(img_tag.get('src', '').replace('cid:', '') for img_tag in soup.find_all('img'))
            image_urls.discard('')

            attachments = list()
            for part in parsed_message.walk():
                content_maintype = part.get_content_maintype()
                content_disposition = part.get('Content-Disposition')
                if content_maintype == 'multipart' or content_disposition is None:
                    continue
                filename = part.get_filename()
                if filename:
                    content_type = part.get_content_type()
                    content_id = part.get('Content-ID')
                    if isinstance(content_id, str):
                        content_id = content_id.strip('<>')
                    if content_id in image_urls:
                        payload = part.get_payload(decode=False)
                        if payload:
                            body_string = body_string.replace(f'cid:{content_id}',
                                                              f'data:{content_type};base64,{payload}')
                    else:
                        try:
                            payload = part.get_payload(decode=True)
                        except AttributeError:
                            payload = part.get_payload()
                        if payload:
                            attachments.append((filename, content_type, payload, content_id))

            message_date = get_message_date(email_message)
            message_id = parsed_message.get('Message-ID')
            reply_message_id = parsed_message.get('In-Reply-To')
            if reply_message_id:
                reply = models.ContactPersonMessageModel.objects.filter(
                    message_id=reply_message_id
                ).order_by('-created_at').first()
            else:
                reply = None
            with transaction.atomic():
                contact_person_message = models.ContactPersonMessageModel.objects.create(
                    contact_person_id=contact_person_id,
                    text=body_string,
                    message_id=message_id,
                    channel_id='email',
                    message_date=message_date,
                    email_subject=parsed_message.get('Subject', ''),
                    reply=reply,
                )
                for attachment in attachments:
                    bytesio = io.BytesIO()
                    bytesio.write(attachment[2])
                    django_file = DjangoFile(bytesio, name=attachment[0])
                    attach_file = File()
                    attach_file.upload = django_file
                    attach_file.save()
                    FileBaseModel.objects.create(file=attach_file, related_object=contact_person_message)

            letter_count += 1
            last_letter_uid = letter_uid
        if last_letter_uid:
            config.email_last_uid = last_letter_uid
            config.save(update_fields=('email_last_uid',))
    return letter_count


def get_from_address(email_message):
    from_address_raw = email_message.get('from')
    if from_address_raw:
        match = re.search(r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}", from_address_raw)
        if match:
            from_address = match.group(0)
        else:
            return None
    else:
        return None
    return from_address


def send_email(config, email_address, message):
    smtp_server_name = config.smtp_server
    smtp_port = config.smtp_port
    email_username = config.email_username
    email_pass = config.email_pass
    if not smtp_server_name or not email_username or not email_pass:
        raise drf_exceptions.ValidationError('Отсутствуют настройки подключения')
    email_message = MIMEMultipart('alternative')
    email_message['From'] = email_username
    email_message['To'] = email_address
    email_message['Date'] = formatdate(localtime=True)
    email_message['Message-ID'] = message.message_id
    email_message['Subject'] = Header(message.email_subject, 'utf-8')
    reply = message.reply
    if reply:
        reply_message_id = reply.message_id
        email_message['In-Reply-To'] = reply_message_id
        references = get_references(message)
        email_message['References'] = references
    text = f"{message.text} <br> {config.email_signature}"
    soup = BeautifulSoup(text, 'lxml')
    img_urls = set(img_tag.get('src', '') for img_tag in soup.find_all('img'))
    for each in img_urls:
        match = re.match(
            re.escape(BACKEND_URL + DOWNLOADER_PATH) +
            r'/\?path=%3Fid%3D([0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12})%26target%3Dckeditor',
            each
        )
        try:
            file_instance_id = match.group(1)
        except AttributeError:
            continue
        else:
            try:
                file_instance = File.objects.get(pk=file_instance_id)
            except (ValidationError, ObjectDoesNotExist):
                continue

            text = text.replace(
                each,
                f"data:{file_instance.mime_type_id};base64,{base64.b64encode(file_instance.upload.read()).decode('utf-8')}"
            )

    email_message.attach(MIMEText(text, 'html', _charset='utf-8'))
    message_attachments = message.attachments.all().order_by('created_at')

    for message_attachment in message_attachments:
        part = MIMEBase('application', 'octet-stream')
        payload = message_attachment.upload.read()
        part.set_payload(payload)
        encoders.encode_base64(part)
        part.add_header(
            'Content-Disposition',
            f'attachment; filename={message_attachment.name + "." + message_attachment.extension}'
        )
        email_message.attach(part)

    with SMTP_SSL(host=smtp_server_name, port=smtp_port, ) as smtp_server:
        smtp_server.login(email_username, email_pass)
        smtp_server.sendmail(
            email_username,
            email_address,
            email_message.as_string(),
        )


def get_message_id(config: models.HelpDeskConfigModel):
    sender_email = config.email_username
    _, email_addr = parseaddr(sender_email)
    domain = email_addr.split('@')[-1]
    message_id = make_msgid(domain=domain)
    return message_id


def get_references(message: models.ContactPersonMessageModel):
    references = list()
    while True:
        reply = message.reply
        if not reply:
            references.reverse()
            return " ".join(references)
        message_id = reply.message_id
        if message_id:
            references.append(message_id)
        message = reply


def get_message_date(email_message: email.message.EmailMessage) -> datetime.datetime:
    """
    Возвращает дату email соолщения из объекта email.message.EmailMessage.
    Если не получается получить дату сообщения, возвращает timezone.now().
    """
    date_raw = email_message.get('date')
    if date_raw:
        try:
            date_tuple = parsedate_tz(date_raw)
            if date_tuple:
                message_date = datetime.datetime.fromtimestamp(mktime_tz(date_tuple))
            else:
                message_date = timezone.now()
        except Exception as e:
            message_date = timezone.now()
    else:
        message_date = timezone.now()
    return message_date


def get_or_create_contact_person_from_email_message(config, from_address, contact_persons):
    """
    Возвращает контактное лицо из email-адреса. Если контактного лица нет, создает "неизвестное" контактное лицо.
    """
    created = False
    contact_person = list(filter(lambda s: s['email'] == from_address, contact_persons))
    if contact_person:
        spam = contact_person[0]['spam']
        if spam:
            return None, False
        contact_person_id = contact_person[0]['id']
    else:
        contact_person = create_unknown_contact_person(config, from_address)
        created = True
        contact_person_id = contact_person.id
    return contact_person_id, created


def create_unknown_contact_person(config, from_address: str):
    customer_card = get_or_create_unknown_customer_card(config.contractor)
    contact_person = models.ContactPersonModel.objects.create(
        customer_card=customer_card,
        email=from_address,
        name='Новый контакт',
        unknown=True,
    )
    return contact_person


def get_or_create_unknown_customer_card(org_admin):
    customer_card, created = models.CustomerCardModel.objects.get_or_create(
        unknown=True,
        org_admin=org_admin,
        defaults={
            'name': 'Неизвестная организация',
            'full_name': 'Неизвестная организация',
        }
    )
    return customer_card


def get_tg_message_date(message):
    return datetime.datetime.fromtimestamp(message.date).astimezone(timezone.get_current_timezone())


logger = logging.getLogger(__name__)

YOUTUBE_HOSTS = set(["youtube.com", "www.youtube.com", "youtu.be", "m.youtube.com"])


def _is_youtube(url: str) -> bool:
    try:
        host = urlsplit(url).hostname or ""
        return host.lower() in YOUTUBE_HOSTS
    except Exception:
        return False


def _extract_uuid_from_backend_download(url: str) -> Optional[str]:
    """
    Ожидаемый пример:
      https://connect.delocloud.ru/download_file/?path=%3Fid%3Da41bb844-b00a-11f0-8287-0242ac11000e%26target%3Dckeditor
    В path лежит URL-encoded строка "?id=<uuid>&target=..."
    """
    try:
        u = urlparse(url)
        backend_host = BACKEND_URL.split("://", 1)[-1].split("/", 1)[0]
        if not u.netloc or backend_host not in u.netloc:
            return None
        q = parse_qs(u.query)
        raw_path = q.get("path", [None])[0]
        if not raw_path:
            return None
        # raw_path = "%3Fid%3D<uuid>%26target%3D..."
        un = unquote(raw_path)  # "?id=<uuid>&target=..."
        inner_q = parse_qs(urlparse(un).query)
        _id = inner_q.get("id", [None])[0]
        uuid.UUID(str(_id))  # валидация
        return _id
    except Exception:
        return None


def _fetch_photo_source(src: str) -> Union[str, Tuple[io.BufferedReader, str], Tuple[io.BytesIO, str]]:
    """
    Возвращает один из вариантов, пригодных для pytelegrambotapi:
    - (file_like, filename) для локальной передачи (если нашли File в БД или скачали удалённый)
    - str (url) — можно сразу отдавать в send_photo/media_group
    """
    # 1) Пробуем как внутренний файл по UUID
    fid = _extract_uuid_from_backend_download(src)
    if fid:
        f = File.objects.filter(id=fid, is_deleted=False).first()
        if f and getattr(f, "upload", None):
            try:
                fileobj = open(f.upload.path, "rb")  # Telegram сам закроет после отправки, но мы закроем ниже
                filename = f.upload.name.rsplit("/", 1)[-1]
                return (fileobj, filename)
            except Exception as e:
                logger.warning("Failed to open local File(id=%s): %s; fallback to URL", fid, e)

    # 2) Иначе: пытаемся скачать
    try:
        resp = requests.get(src, timeout=10)
        resp.raise_for_status()
        b = io.BytesIO(resp.content)
        # имя файла — хвост пути или дефолт
        tail = urlsplit(src).path.rsplit("/", 1)[-1]
        b.name = tail or "photo.jpg"
        return (b, b.name)  # BytesIO пойдёт как файл
    except Exception:
        # 3) Фолбэк — пусть телега сама тянет по URL
        return src


def _clean_text_from_html(html: str) -> List[str]:
    """
    Достаём осмысленные текстовые блоки (параграфы/строки) без форматирования.
    Пустые — отбрасываем.
    """
    soup = BeautifulSoup(html or "", "html.parser")

    # Уберём возможные figure/iframe (ссылки соберём отдельно)
    for fig in soup.select("figure, iframe"):
        fig.decompose()

    chunks = []
    for p in soup.find_all(["p", "div", "li", "h1", "h2", "h3", "h4", "h5"]):
        text = p.get_text(" ", strip=True)
        if text:
            chunks.append(text)

    # "голый" текст (на случай отсутствия параграфов)
    body_text = soup.get_text(" ", strip=True)
    if body_text:
        joined = " ".join(chunks)
        if body_text != joined and body_text not in chunks:
            chunks.append(body_text)

    cleaned = []
    for t in chunks:
        tt = re.sub(r"\s+", " ", t).strip()
        if tt:
            cleaned.append(tt)
    return cleaned


def _collect_media_and_links(html: str) -> Tuple[List[str], List[str], List[str]]:
    """
    Возвращает:
      photos: список источников картинок (src)
      links:  список обычных ссылок (включая youtube)
      videos: список явных видео-URL (<video src> / <source src>)
    """
    soup = BeautifulSoup(html or "", "html.parser")

    photos = []
    for imgtag in soup.find_all("img"):
        src = (imgtag.get("src") or "").strip()
        if src:
            photos.append(src)

    links = set()
    for a in soup.find_all("a"):
        href = (a.get("href") or "").strip()
        if href:
            links.add(href)

    # embed-ютуб
    for iframe in soup.find_all("iframe"):
        src = (iframe.get("src") or "").strip()
        if src and _is_youtube(src):
            links.add(src)

    for div in soup.find_all("div"):
        data_oe = (div.get("data-oembed-url") or "").strip()
        if data_oe and _is_youtube(data_oe):
            links.add(data_oe)

    videos = set()
    for v in soup.find_all("video"):
        src = (v.get("src") or "").strip()
        if src:
            videos.add(src)
        for s in v.find_all("source"):
            ss = (s.get("src") or "").strip()
            if ss:
                videos.add(ss)

    return photos, sorted(list(links)), sorted(list(videos))


def _send_media_group(
    bot: telebot.TeleBot,
    chat_id: Union[int, str],
    photo_sources: List[Union[str, Tuple[io.BufferedReader, str], Tuple[io.BytesIO, str]]],
    caption: Optional[str] = None
) -> List[telebot.types.Message]:
    """
    Отправка пачкой до 10 фото. Первый элемент может нести caption.
    """
    if not photo_sources:
        return []

    sent_msgs = []
    CHUNK = 10
    for i in range(0, len(photo_sources), CHUNK):
        chunk = photo_sources[i:i + CHUNK]
        media = []
        for j, src in enumerate(chunk):
            if isinstance(src, tuple):
                fileobj, _fname = src
                if j == 0 and caption:
                    media.append(tg_types.InputMediaPhoto(media=fileobj, caption=caption))
                else:
                    media.append(tg_types.InputMediaPhoto(media=fileobj))
            else:
                if j == 0 and caption:
                    media.append(tg_types.InputMediaPhoto(media=src, caption=caption))
                else:
                    media.append(tg_types.InputMediaPhoto(media=src))
        msgs = bot.send_media_group(chat_id, media=media)
        sent_msgs.extend(msgs)
    return sent_msgs


def _close_files(photo_sources: List[Union[str, Tuple[io.BufferedReader, str], Tuple[io.BytesIO, str]]]) -> None:
    for src in photo_sources:
        if isinstance(src, tuple):
            fileobj, _fn = src
            try:
                fileobj.close()
            except Exception:
                pass


def _html_to_telegram(
    bot: telebot.TeleBot,
    telegram_id: Union[int, str],
    html: str,
    reply_to_message_id: Optional[int]
) -> Optional[telebot.types.Message]:
    """
    Главный “рендер”: отправляет текстовые блоки, картинки (медиагруппы), ссылки/видео.
    Возвращает последний message (для фикса id/date), если удалось.
    """
    text_blocks = _clean_text_from_html(html)
    photos_raw, links, videos = _collect_media_and_links(html)

    photo_sources = [_fetch_photo_source(x) for x in photos_raw]

    last_tg = None  # type: Optional[telebot.types.Message]

    # 1) Тексты
    for i, block in enumerate(text_blocks):
        _reply = reply_to_message_id if (i == 0) else None
        last_tg = bot.send_message(telegram_id, text=block, reply_to_message_id=_reply)

    # 2) Картинки
    try:
        caption = text_blocks[-1] if text_blocks else None
        photos_only = [x for x in photo_sources if x]
        if photos_only:
            if len(photos_only) == 1:
                src = photos_only[0]
                if isinstance(src, tuple):
                    fobj, _fn = src
                    last_tg = bot.send_photo(telegram_id, fobj, caption=caption)
                else:
                    last_tg = bot.send_photo(telegram_id, src, caption=caption)
            else:
                msgs = _send_media_group(bot, telegram_id, photos_only, caption=caption)
                if msgs:
                    last_tg = msgs[-1]
    finally:
        _close_files(photo_sources)

    # 3) Видеоссылки и ссылки
    link_lines = []
    if videos:
        link_lines.append("Видео:")
        link_lines.extend(videos)
    if links:
        link_lines.append("Ссылки:" if videos else "Ссылки:")
        link_lines.extend(links)
    if link_lines:
        last_tg = bot.send_message(telegram_id, text="\n".join(link_lines))

    return last_tg


def send_tg_message(config, telegram_id: Union[int, str], message) -> None:
    """
    Обновлённая версия:
    - reply цепочкой
    - парсит HTML и рассылает текст/картинки/ссылки
    - отправляет ПОСЛЕ КОММИТА
    - отдельно отправляет attachments из message.attachments
    """
    bot_token = config.telegram_token
    bot = telebot.TeleBot(bot_token)

    reply = getattr(message, "reply", None)
    if reply and getattr(reply, "channel_id", None) == "telegram":
        try:
            reply_to_message_id = int(reply.message_id)
        except Exception:
            reply_to_message_id = None
    else:
        reply_to_message_id = None

    html_text = (getattr(message, "text", "") or "").strip()

    def _do_send():
        last_message = None

        # 1) HTML → Telegram
        if html_text:
            last_message = _html_to_telegram(bot, telegram_id, html_text, reply_to_message_id)

        # 2) Прикреплённые файлы (если у модели message есть related manager attachments)
        try:
            attachments_qs = message.attachments.all().order_by("created_at")
        except Exception:
            attachments_qs = []

        for each in attachments_qs:
            try:
                if getattr(each, "is_image", False):
                    tg = bot.send_photo(telegram_id, each.upload)
                elif getattr(each, "is_video", False):
                    tg = bot.send_video(telegram_id, each.upload)
                elif getattr(each, "is_audio", False):
                    tg = bot.send_audio(telegram_id, each.upload)
                else:
                    tg = bot.send_document(telegram_id, each.upload)
                last_message = tg
            except Exception as e:
                logger.error("Failed to send attachment id=%s: %s", getattr(each, "id", None), e)

        # 3) Сохраняем id/date по последнему отправленному сообщению
        if last_message is not None:
            try:
                message.message_id = last_message.message_id
                message.message_date = get_tg_message_date(last_message)
                message.save(update_fields=("message_id", "message_date",))
            except Exception as e:
                logger.warning("Failed to update message meta: %s", e)

    # Отправляем ПОСЛЕ КОММИТА транзакции
    transaction.on_commit(_do_send)

#
# def send_tg_message(config, telegram_id, message):
#     bot_token = config.telegram_token
#     bot = telebot.TeleBot(bot_token)
#     reply = message.reply
#     if reply and reply.channel_id == 'telegram':
#         reply_to_message_id = int(reply.message_id)
#     else:
#         reply_to_message_id = None
#     tg_message = bot.send_message(
#         telegram_id,
#         text=message.text,
#         reply_to_message_id=reply_to_message_id,
#     )
#     message.message_id = tg_message.message_id
#     message.message_date = get_tg_message_date(tg_message)
#     message.save(update_fields=('message_id', 'message_date',))
#
#     attachments = message.attachments.all().order_by('created_at')
#     for each in attachments:
#         if each.is_image:
#             bot.send_photo(
#                 telegram_id,
#                 each.upload
#             )
#         elif each.is_video:
#             bot.send_video(
#                 telegram_id,
#                 each.upload
#             )
#         elif each.is_audio:
#             bot.send_audio(
#                 telegram_id,
#                 each.upload
#             )
#         else:
#             bot.send_document(
#                 telegram_id,
#                 each.upload
#             )
#

from bpms.chat.models import ChatModel, MessageModel
from bpms.chat.serializers import MessageListSerializer


def send_internal_chat_message(
        chat: ChatModel,
        message: models.ContactPersonMessageModel,
        message_reply: MessageModel,
        ticket: models.HelpDeskTicketModel
):
    if not chat.is_public:
        return
    notify_message = MessageModel()
    notify_message.is_system = True
    notify_message.message_reply = message_reply
    notify_message.text = message.text
    # if not chat.is_public:
    #     notify_message.share = ticket
    notify_message.chat = chat
    notify_message.created = timezone.now()
    notify_message.share = ticket
    notify_message.save()
    message.message_id = notify_message.message_uid
    message.save(update_fields=('message_id',))
    message_data = MessageListSerializer(notify_message).data
    message_data['chat_uid'] = str(chat.chat_uid)
    message_data['chat_name'] = chat.name
    message_data['is_public'] = chat.is_public
    message_data['is_new'] = True
    data = json.dumps(
        {
            "event": "chat_message",
            "data": message_data
        },
        cls=DjangoJSONEncoder,
    )
    socketio_redis.publish(SOCKETIO_SYSTEM_CHANNEL, data)


def get_work_log_duration(user, ticket):
    work_logs = models.HelpDeskWorkLogModel.objects.filter(
        is_active=True,
        user=user,
        ticket=ticket,
    )
    complete_duration = work_logs.filter(
        is_current=False
    ).aggregate(complete_duration=Sum('duration'))['complete_duration']
    if complete_duration is None:
        complete_duration = 0
    incomplete_log = work_logs.filter(is_current=True).first()
    if incomplete_log:
        now = timezone.now()
        incomplete_duration = get_incomplete_duration(now, incomplete_log.created_at)
        is_current = True
    else:
        incomplete_duration = 0
        is_current = False
    duration = complete_duration + incomplete_duration
    return duration, is_current, incomplete_duration


def get_work_log_duration_sum(ticket):
    work_logs = models.HelpDeskWorkLogModel.objects.filter(
        is_active=True,
        ticket=ticket,
        is_current=False
    )
    duration_sum = work_logs.aggregate(duration_sum=Sum('duration'))['duration_sum']
    if not duration_sum:
        duration_sum = 0
    return duration_sum


def get_incomplete_duration(now, created_at):
    return (now - created_at).seconds


def save_telegram_file(bytes_string, filename, message):
    stream = io.BytesIO()
    stream.write(bytes_string)
    stream.seek(0)
    django_file = DjangoFile(
        file=stream,
        name=filename,
    )
    attach_file = File()
    attach_file.upload = django_file
    attach_file.is_confined = True
    attach_file.save()
    message.attachments.set((attach_file,))


def handle_upload_customer_cards(org_admin, file):
    workbook = openpyxl.load_workbook(filename=file)
    sheet = workbook.active
    row_count = sheet.max_row + 1
    contractors_tree_ids = get_tree_departments_related_organizations((org_admin.pk,))

    with transaction.atomic():
        for each in range(2, row_count):
            full_name = get_cell_value(sheet, each, 1)[:511]
            name = full_name[:255]

            # создаем карточку клиента:
            customer_bin = get_cell_value(sheet, each, 2)
            legal_address = get_cell_value(sheet, each, 3)
            if not customer_bin:
                break
            if customer_bin:
                customer_card, created = models.CustomerCardModel.objects.update_or_create(
                    inn=customer_bin,
                    org_admin=org_admin,
                    defaults={
                        'name': name,
                        'full_name': full_name,
                        'legal_address': legal_address,
                    }
                )

            # Добавляем админа
            customer_admin_name = get_cell_value(sheet, each, 4)
            customer_admin_bin = get_cell_value(sheet, each, 5)
            if customer_admin_bin:
                customer_admin, admin_created = models.CustomerCardAdminModel.objects.update_or_create(
                    org_admin=org_admin,
                    bin=customer_admin_bin,
                    defaults={
                        'name': customer_admin_name,
                    }
                )
                models.CustomerCardAdminThroughModel.objects.filter(customer_card=customer_card).delete()
                models.CustomerCardAdminThroughModel.objects.create(
                    customer_card=customer_card,
                    admin=customer_admin,
                )

            # Добавляем тэг
            tag_name = get_cell_value(sheet, each, 6)
            if tag_name:
                tag_serializer = TagModelCreateSerializer(
                    data={
                        'name': tag_name,
                        'related_object': customer_card.pk,
                        'contractor': org_admin.pk,
                    }
                )
                tag_serializer.is_valid(raise_exception=True)
                tag_serializer.save()

            # Добавляем контакт
            contact_email = get_cell_value(sheet, each, 10)
            if contact_email:
                contact_name = get_cell_value(sheet, each, 7)
                contact_post = get_cell_value(sheet, each, 8)
                contact_phone = get_cell_value(sheet, each, 9)
                contact_comment = get_cell_value(sheet, each, 11)
                # Получаем или создаем должность
                if contact_post:
                    try:
                        contact_post_uuid = uuid.UUID(contact_post)
                    except ValueError:
                        post_inst = models.ContactPersonPostModel.objects.filter(
                            contractor_id__in=contractors_tree_ids,
                            name=contact_post
                        ).first()
                        if not post_inst:
                            post_inst = models.ContactPersonPostModel.objects.create(
                                contractor=org_admin,
                                name_ru=contact_post,
                            )
                    else:
                        try:
                            post_inst = models.ContactPersonPostModel.objects.filter(
                                contractor_id__in=contractors_tree_ids
                            ).get(
                                pk=contact_post,
                            )
                        except (ValidationError, ObjectDoesNotExist):
                            raise ValidationError(f"line {each}: должность не найдена: {contact_post}")

                else:
                    post_inst = None
                contact_person = models.ContactPersonModel.objects.filter(
                    is_active=True,
                    email=contact_email,
                    customer_card=customer_card
                ).order_by('-created_at').first()
                if contact_person:
                    contact_person.name = contact_name
                    contact_person.post_inst = post_inst
                    contact_person.phone = contact_phone
                    contact_person.comment = contact_comment
                    contact_person.save()
                else:
                    models.ContactPersonModel.objects.create(
                        customer_card=customer_card,
                        name=contact_name,
                        post_inst=post_inst,
                        phone=contact_phone,
                        comment=contact_comment,
                        email=contact_email,
                    )
            # Добавляем ответственных
            specialist_email = get_cell_value(sheet, each, 12)
            if specialist_email:
                try:
                    specialist_profile = ProfileModel.objects.get(user__email=specialist_email)
                except (ValidationError, ObjectDoesNotExist):
                    continue
                reserve_email = get_cell_value(sheet, each, 14)
                if not reserve_email:
                    reserve_profile = None
                try:
                    reserve_profile = ProfileModel.objects.get(user__email=reserve_email)
                except (ValidationError, ObjectDoesNotExist):
                    reserve_profile = None
                models.CustomerSupportSpecialistModel.objects.get_or_create(
                    customer_card=customer_card,
                    user=specialist_profile,
                    reserve=reserve_profile,
                )
    return 'ok'


def get_cell_value(sheet, row, column):
    value = sheet.cell(row=row, column=column).value
    if not value:
        value = ''
    else:
        value = str(value).strip()
    return value


def update_customer_card_index(card_id):
    """
    Функция для асинхронной переиндексации CustomerCardModel.
    Вызывается асинхронно через django-q при изменении связанных объектов.
    """
    try:
        card = models.CustomerCardModel.objects.get(pk=card_id)
        from .search_indexes import CustomerCardIndex
        CustomerCardIndex().update_object(card)
        return 'done.'
    except models.CustomerCardModel.DoesNotExist:
        pass


def get_my_actual_specialists(user):
    customer_cards = user.contact_persons.filter(is_active=True, spam=False).values_list('customer_card', flat=True)
    local_date = timezone.localdate()
    actual_specialist_users = models.CustomerSupportSpecialistModel.objects.filter(
        Q(start_date__lte=local_date, end_date__gte=local_date, ) |
        Q(start_date__isnull=True, end_date__gte=local_date) |
        Q(start_date__lte=local_date, end_date__isnull=True) |
        Q(start_date__isnull=True, end_date__isnull=True),
        customer_card__in=customer_cards,
    ).exclude(
        vacation_dates__start_date__lte=local_date,
        vacation_dates__end_date__gte=local_date,
    ).values_list('user', 'customer_card__org_admin__name', 'is_reserve')
    return actual_specialist_users


def get_actual_specialist_from_user(client_user, specialist_user):
    customer_cards = client_user.contact_persons.filter(
        is_active=True,
        spam=False
    ).values_list('customer_card', flat=True)
    local_date = timezone.localdate()
    actual_specialists = models.CustomerSupportSpecialistModel.objects.filter(
        Q(start_date__lte=local_date, end_date__gte=local_date, ) |
        Q(start_date__isnull=True, end_date__gte=local_date) |
        Q(start_date__lte=local_date, end_date__isnull=True) |
        Q(start_date__isnull=True, end_date__isnull=True),
        customer_card__in=customer_cards,
        user=specialist_user,
    ).exclude(
        vacation_dates__start_date__lte=local_date,
        vacation_dates__end_date__gte=local_date,
    )
    return actual_specialists


def get_help_desk_admin_contractors(user):
    from contractor_permissions.utils import contractors_where_user_has_permission
    contractors_id = contractors_where_user_has_permission(
        user.pk, 'help_desk_admin', None
    )
    return contractors_id


def get_help_desk_admin_manager_contractors(user):
    from contractor_permissions.utils import contractors_where_user_has_permission
    contractors_id = contractors_where_user_has_permission(
        user.pk, ('help_desk_admin', 'help_desk_manager', 'help_desk_supervisor'), None
    )
    return contractors_id


def start_work_log_timer(user, ticket):
    """Запускает таймер учета времени для пользователя и тикета."""
    if not user == ticket.specialist:
        raise drf_exceptions.PermissionDenied('Вы не можете запускать учет времени для этого тикета')
    from bpms.tasks.models import TaskExecutionTimeModel
    now = timezone.now()
    if not models.HelpDeskWorkLogModel.objects.filter(user=user, ticket=ticket, is_current=True).exists():
        with transaction.atomic():
            completed_statuses = get_completed_statuses_id()
            work_logs = (
                models.HelpDeskWorkLogModel.objects
                .select_related('ticket')
                .filter(user=user, is_current=True)
            )
            for each in work_logs:
                paused_ticket = each.ticket
                each.duration = get_incomplete_duration(now, each.created_at)
                each.is_current = False
                each.finished_date = now
                each.save()
                if paused_ticket and paused_ticket.status_id not in completed_statuses and paused_ticket.status_id != 'on_pause':
                    paused_ticket.status_id = 'on_pause'
                    paused_ticket.save(update_fields=('status_id', 'updated_at'))
            task_work_logs = (
                TaskExecutionTimeModel.objects
                    .select_related('task')
                    .filter(user=user, is_current=True)
            )
            for each in task_work_logs:
                # paused_task = each.task
                each.duration = get_incomplete_duration(now, each.created_at)
                each.is_current = False
                each.save()
                # if paused_task and paused_task.status_id not in completed_statuses and paused_task.status_id != 'on_pause':
                #     paused_task.status_id = 'on_pause'
                #     paused_task.save(update_fields=('status_id', 'updated_at'))
            current_log = models.HelpDeskWorkLogModel()
            current_log.user = user
            current_log.ticket = ticket
            current_log.is_current = True
            current_log.duration = 0
            current_log.save()
        ProfileModel.objects.filter(pk=user.pk).update(current_work=current_log.ticket)
        transaction.on_commit(lambda: send_socketio_about_update_current_work([str(user.pk), ]))
        transaction.on_commit(lambda: send_socketio_about_update_ticket(ticket))
    duration, is_current, incomplete_duration = get_work_log_duration(user, ticket)
    return duration, is_current


def stop_work_log_timer(user, ticket, provided_duration=None, description='', is_result=False):
    """Останавливает таймер учета времени для пользователя и тикета."""
    now = timezone.now()
    incomplete_log = models.HelpDeskWorkLogModel.objects.filter(user=user, ticket=ticket, is_current=True).first()
    if not isinstance(description, str):
        description = ''
    else:
        description = description[:1024]
    if incomplete_log:
        if provided_duration is not None:
            # жёстко доверяем фронту, но слегка валидируем
            try:
                provided_duration = int(provided_duration)
            except (TypeError, ValueError):
                provided_duration = 0
            if provided_duration < 0:
                provided_duration = 0
            duration_to_save = provided_duration
        else:
            # старое поведение
            duration_to_save = get_incomplete_duration(now, incomplete_log.created_at)

        incomplete_log.duration = duration_to_save
        incomplete_log.is_current = False
        incomplete_log.finished_date = now
        incomplete_log.description = description
        incomplete_log.is_result = is_result
        incomplete_log.save()

    # как и раньше: возвращаем duration/is_current, без ломающей смены сигнатуры
    duration, is_current, incomplete_duration = get_work_log_duration(user, ticket)
    transaction.on_commit(lambda: send_socketio_about_update_current_work([str(user.pk), ]))
    transaction.on_commit(lambda: send_socketio_about_update_ticket(ticket))
    return duration, is_current


def change_ticket_status(ticket, user, status_code, assign_specialist=False):
    """Изменяет статус тикета и запускает/останавливает таймер если нужно.
    Для телеграм-бота."""

    with transaction.atomic():
        if ticket.status_id == status_code:
            raise drf_exceptions.ValidationError(f"Обращение уже находится в статусе '{ticket.status.name}'")

        from .serializers import UpdateTicketStatusSerializer
        serializer_data = {"status": status_code}
        serializer = UpdateTicketStatusSerializer(
            ticket,
            data=serializer_data,
            partial=True,
            context={'empty_return': True}
        )
        serializer.is_valid(raise_exception=True)
        ticket = serializer.save()

        if ticket.specialist is None and assign_specialist:
            ticket.specialist = user
            ticket.save()
        completed_statuses = get_completed_statuses_id()
        if status_code == 'in_work':
                start_work_log_timer(user, ticket)
        elif status_code in ['on_pause', *completed_statuses]:
                stop_work_log_timer(user, ticket)

        from .notifications import notify_about_new_status
        transaction.on_commit(
            lambda: async_task(notify_about_new_status, str(user.pk), str(ticket.pk), str(ticket.status_id))
        )



def get_completed_statuses_id():
    return ['completed', 'rejected', ]


def check_ticket_category_create_permission(user, contractor_id: uuid.UUID):
    contractors_id = get_help_desk_admin_contractors(user)
    from users.utils import get_ancestor_departments_related_organizations
    contractors_id = get_ancestor_departments_related_organizations(contractors_id, include_self=True)
    if contractor_id in contractors_id:
        return True
    return False


def check_contact_person_create_permission(user, contractor_id: uuid.UUID):
    try:
        check_contractor_permission(user.pk, contractor_id, ('help_desk_manager', 'help_desk_admin'), None)
    except drf_exceptions.PermissionDenied:
        return False
    return True


def get_create_message_for_client_permission(ticket):
    if ticket.channel_id not in ('internal', 'internal_chat',) or ticket.status_id in get_completed_statuses_id():
        return False
    else:
        return True


def get_contact_persons_queryset(user_id):
    """Возвращает queryset доступных контактных лиц для пользователя с указанным user_id."""
    from contractor_permissions.utils import contractors_where_user_has_permission, contractors_where_im_director
    from .models import ContactPersonModel
    
    user = ProfileModel.objects.get(pk=user_id)
    qs = ContactPersonModel.objects.filter(is_active=True)
    available_contractors = set(contractors_where_user_has_permission(
        user_id,
        ('help_desk_admin', 'help_desk_manager', 'admin',),
        None,
    ))
    available_contractors.update(set(contractors_where_im_director(user)))
    qs = qs.filter(customer_card__org_admin__in=available_contractors)
    qs = qs.order_by('name')
    return qs.select_related('post_inst', 'customer_card',)


def get_ticket_first_chat_message(ticket):
    first_message = ticket.messages.all().order_by('created_at').first()
    if not first_message:
        return None
    message_uid = first_message.message_id
    try:
        chat_message = MessageModel.objects.get(message_uid=message_uid)
    except (ValidationError, ObjectDoesNotExist):
        return None
    return chat_message


def send_socketio_about_update_ticket(ticket: models.HelpDeskTicketModel):
    from .serializers import HelpDeskTicketDetailSerializer
    s_data = HelpDeskTicketDetailSerializer(instance=ticket).data
    data = json.dumps(
        {
            'event': 'ticket_update',
            'data': s_data,
        },
        cls=DjangoJSONEncoder
    )
    socketio_redis.publish(SOCKETIO_SYSTEM_CHANNEL, data)


def generate_telegram_webhook_token():
    """
    Генерирует криптографически стойкую строку,
    подходящую для X-Telegram-Bot-Api-Secret-Token.
    """
    alphabet = string.ascii_letters + string.digits + "_-"
    token = ''.join(secrets.choice(alphabet) for _ in range(64))
    return token


def get_priority_emoji(priority_code: str) -> str:
    """Возвращает эмоджи для приоритета тикета"""
    priority_emojis = {
        '0': '⏳',  # Очень низкий - песочные часы
        '1': '🕐',  # Низкий - часы
        '2': '⚠️',  # Обычный - восклицательный знак
        '3': '⚡',  # Высокий - молния
        '4': '🔥',  # Очень высокий - огонь
    }
    return priority_emojis.get(priority_code, '⚠️')


def status_color_to_emoji(color_name: str) -> str:
    mapping = {
        'blue': '🔵',
        'cyan': '🟢',
        'default': '⚪',
        'grey': '⚫',
        'geekblue': '🔵',
        'green': '🟢',
        'orange': '🟠',
        'brown': '🟠',
        'pink': '🟣',
        'purple': '🟣',
        'red': '🔴',
        'yellow': '🟡',
        'black': '⚫',
    }

    return mapping.get((color_name or '').lower(), '🔵')