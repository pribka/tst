from collections import Counter
import pytz
import json
import pyexcelerate
from datetime import datetime, timedelta, time
from bs4 import BeautifulSoup
from math import ceil

from django.db import transaction
from django.db.models import Q, Count, Case, Value, IntegerField, When, OuterRef, Subquery, Sum, Avg, BooleanField, \
    Min, F, DecimalField, Prefetch, Exists
from django.utils import timezone, dateparse
from django.core.cache import cache
from django.core.exceptions import ObjectDoesNotExist, ValidationError
from django.core.serializers.json import DjangoJSONEncoder

from rest_framework import exceptions as drf_exceptions

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Border, Side, PatternFill

from django_q.tasks import async_task

from bkz3.settings import TIME_ZONE, SOCKETIO_SYSTEM_CHANNEL, DEFAULT_CURRENCY_CODE, FILTER_BY_ORGANIZATIONS
from common.redis import socketio_redis
from common.models import BaseModel, FiltersStore
from common.serializers import BaseCatalogRetrieveSerializer
from common.utils import get_filter_queryset, get_field_keys, order_queryset_from_get_param
from common.utils import send_socketio_about_update_current_work
from common.catalogs.models import CurrencyModel
from common.catalogs.serializers import AppCurrencySerializer
from bpms.chat.models import MessageModel
from bpms.chat.serializers import MessageListSerializer
from bpms.workgroups.models import WorkgroupModel, WorkgroupMembersModel
from tags.models import TagModel
from . import models
from django.db import connection
from django.db.models.sql.compiler import SQLCompiler  # будем из queryset-ов доставать тексты запросов
from bkz3.settings import TEMPDB_TABLESPACE as TABLESPACE

from users.models import ProfileModel

from contractor_permissions.utils import contractors_where_user_has_permission

from bpms.event_calendar.models import EventCalendarModel, EventCalendarTypeModel, EventCalendarMemberModel
from bpms.event_calendar.utils import get_or_create_related_calendar
from bpms.favorites.models import FavoriteModel


def get_user_today_from_iso(iso_str):
    """
    Возвращает текущую дату в часовом поясе пользователя.
    iso_str: ISO строка с датой/временем и часовым поясом (например, '2026-01-13T00:00:00.000+03:00')
    """
    # Извлекаем часовой пояс из параметра (например, +03:00)
    dt = dateparse.parse_datetime(iso_str)
    if dt and dt.tzinfo:
        # Конвертируем текущее время в часовой пояс пользователя
        user_now = timezone.now().astimezone(dt.tzinfo)
        return user_now.date()
    else:
        # Если не удалось определить часовой пояс, используем локальное время сервера
        return timezone.localdate()


def clear_my_day_grouped_cache(profile_id):
    """Сбрасывает кэш сгруппированных задач «Мой день» для профиля."""
    pattern = f"my_day_grouped_{profile_id}_*"
    cache.delete_pattern(pattern)


def create_temp_bases():
    with connection.cursor() as cursor:
        cursor.execute('''
            CREATE TEMPORARY TABLE IF NOT EXISTS task_temporary_table_by_permissions_step1 
            (task_id UUID PRIMARY KEY, 
             task_children_id UUID,
             task_children_children_id UUID,
             task_children_children_children_id UUID)   ''' + TABLESPACE + ';')
    with connection.cursor() as cursor:
        cursor.execute('''
            CREATE TEMPORARY TABLE IF NOT EXISTS task_temporary_table_by_permissions 
            (task_id UUID PRIMARY KEY );
        ''')


def get_children_count(instance: models.TaskModel):
    children = instance.children.all()
    children_count = 0
    for each in children:
        if each.is_active is True:
            children_count += 1
    return children_count


def get_task_sprint_queryset(request):
    tasks_qs = models.TaskModel.objects.filter(is_active=True).exclude(sprint__isnull=True)
    user = request.user.profile
    only_participant = request.query_params.get('only_participant', '')

    my_projects_id = user.workgroupmembersmodel_set.filter(
            membership_request_status__code='APPROVED',
            membership_role__code__in=('FOUNDER', 'MODERATOR',)
        ).values_list('work_group', flat=True)
    if only_participant == '1':
        my_tasks_list = tasks_qs.filter(
            Q(owner=user) |
            Q(operator=user) |
            Q(visors=user) |
            Q(cooperators=user)
        ).distinct()
        my_tasks_sprint_ids = set(my_tasks_list.values_list('sprint_id', flat=True))
        queryset = models.TaskSprintModel.objects.filter(
            Q(pk__in=my_tasks_sprint_ids) | Q(author=user),
            is_active=True
        ).distinct()
    else:
        my_tasks_list = filter_by_permissions(tasks_qs, user)
        my_tasks_sprint_ids = set(my_tasks_list.values_list('sprint_id', flat=True))
        queryset = models.TaskSprintModel.objects.filter(
            Q(pk__in=my_tasks_sprint_ids) | Q(projects__in=my_projects_id) | Q(author=user),
            is_active=True
        ).distinct()
    query_params = getattr(request, 'query_params', request.GET)
    dead_line_after = query_params.get('dead_line_after', None)
    dead_line_before = query_params.get('dead_line_before', None)
    dead_line_is_null = query_params.get('dead_line_is_null', None)
    only_no_dead_line = query_params.get('only_no_dead_line', None)
    exclude_id = query_params.get('exclude')
    for_set_task_id = query_params.get('for_set_task')
    filtered_queryset = get_filter_queryset(request, models.TaskSprintModel, queryset)
    queryset = filtered_queryset
    if only_no_dead_line == "1":
        queryset = queryset.filter(dead_line__isnull=True)
    base_isnull_deadline_lu = Q(dead_line__isnull=True)
    if dead_line_after:
        dl_af_lu = Q(dead_line__gt=dead_line_after)
        if dead_line_is_null == "1":
            dl_af_lu = dl_af_lu | Q(base_isnull_deadline_lu)
        queryset = queryset.filter(dl_af_lu)
    if dead_line_before:
        dl_bef_lu = Q(dead_line__lt=dead_line_before)
        if dead_line_is_null == "1":
            dl_bef_lu = dl_bef_lu | Q(base_isnull_deadline_lu)
        queryset = queryset.filter(dl_bef_lu)
    if exclude_id:
        queryset = queryset.exclude(pk=exclude_id)
    if for_set_task_id:
        try:
            task = models.TaskModel.objects.get(is_active=True, pk=for_set_task_id)
        except ObjectDoesNotExist:
            raise drf_exceptions.ValidationError(f'Задача {for_set_task_id} не найдена')
        if task.sprint:
            return queryset.none()
        project = task.project
        if project and project.get_update_permission(request):
            queryset = queryset.filter(projects=project).exclude(status='completed')
        else:
            return queryset.none()
    if query_params.get('active_sprints', '') == '1':
        queryset = queryset.exclude(status__in=('new', 'completed',))
    return queryset.distinct()


def get_task_queryset(request, queryset, list_type: str = 'default'):
    try:
        query_params = request.query_params
    except AttributeError:
        query_params = request.GET

    if isinstance(query_params.get('task_type'), str):
        task_type = query_params.get('task_type').split(',')
    else:
        task_type = None
    dead_line_after = query_params.get('dead_line_after', None)
    dead_line_before = query_params.get('dead_line_before', None)
    dead_line_is_null = query_params.get('dead_line_is_null', None)
    only_no_dead_line = query_params.get('only_no_dead_line', None)
    or_project = query_params.get('or_project', None)
    or_workgroup = query_params.get('or_workgroup', None)
    or_group_or_project_lookup = Q()
    or_group_or_project_lookup_active = False
    if or_project:
        or_group_or_project_lookup_active = True
        or_group_or_project_lookup |= Q(project=or_project)
    if or_workgroup:
        or_group_or_project_lookup_active = True
        or_group_or_project_lookup |= Q(workgroup=or_workgroup)
    queryset = FavoriteModel.annotate_favorites(queryset)

    #-------------------------------------------------------------
    # если есть хранимые фильтры, возвращаем всегда плоский список:
    has_stored_filters = False
    try:
        filter_store = FiltersStore.objects.get(
            author=request.user.profile,
            model='tasks.TaskModel',
            page_name=query_params.get('page_name', '')
        )
    except FiltersStore.DoesNotExist:
        filter_store = None
    if filter_store:
        filter_fields = filter_store.filters.get('values', dict())
        if isinstance(filter_fields, dict):
            has_stored_filters = any(_.get('active') for _ in filter_fields.values())

    if has_stored_filters:
        if 'filters' in list(query_params.keys()):
            try:
                filters_dict = json.loads(query_params.get('filters'))
            except (json.JSONDecodeError, TypeError):
                filters_dict = dict()
            if filters_dict and 'parent' in filters_dict:
                filters_dict.pop('parent', None)
                request.query_params._mutable = True
                request.query_params['filters'] = json.dumps(filters_dict)
                request.query_params._mutable = False
    #-------------------------------------------------------------

    filtered_queryset = get_filter_queryset(request, models.TaskModel, queryset)
    if isinstance(query_params.get('parent'), str):
        parent = query_params.get('parent').split(',')
        if parent == ['all']:
            queryset = filtered_queryset
        elif parent == ['null']:
            queryset = filtered_queryset.filter(parent=None)
        else:
            queryset = filtered_queryset.filter(parent__in=parent)
    else:
        queryset = filtered_queryset

    user = request.user.profile

    exclude_id = query_params.get('exclude')
    if exclude_id:
        try:
            exclude_task = models.TaskModel.objects.get(is_active=True, pk=exclude_id)
        except ObjectDoesNotExist:
            exclude_task = None
        if exclude_task:
            exclude_tree = exclude_task.get_descendants(
                include_self=True).filter(is_active=True).values_list('id', flat=True)
            queryset = queryset.exclude(pk__in=exclude_tree)
    if query_params.get('dead_line__month'):
        try:
            dead_line__month = dateparse.parse_date(query_params.get('dead_line__month'))
        except ValueError:
            dead_line__month = None
        if dead_line__month:
            dead_line__gte = pytz.timezone(TIME_ZONE).localize(datetime(year=dead_line__month.year,
                                                                        month=dead_line__month.month,
                                                                        day=1,
                                                                        hour=0,
                                                                        minute=0,
                                                                        second=0,
                                                                        microsecond=0
                                                                        ), is_dst=None)
            dead_line__lt = pytz.timezone(TIME_ZONE).localize(datetime(year=dead_line__month.year,
                                                                       month=dead_line__month.month + 1,
                                                                       day=1,
                                                                       hour=0,
                                                                       minute=0,
                                                                       second=0,
                                                                       microsecond=0), is_dst=None)
            queryset = queryset.filter(dead_line__gte=dead_line__gte, dead_line__lt=dead_line__lt)
    if only_no_dead_line == "1":
        queryset = queryset.filter(dead_line__isnull=True)
    if or_group_or_project_lookup_active:
        queryset = queryset.filter(or_group_or_project_lookup)
    base_isnull_deadline_lu = Q(dead_line__isnull=True)
    if dead_line_after:
        dl_af_lu = Q(dead_line__gt=dead_line_after)
        if dead_line_is_null == "1":
            dl_af_lu = dl_af_lu | Q(base_isnull_deadline_lu)
        queryset = queryset.filter(dl_af_lu)
    if dead_line_before:

        dl_bef_lu = Q(dead_line__lt=dead_line_before)
        if dead_line_is_null == "1":
            dl_bef_lu = dl_bef_lu | Q(base_isnull_deadline_lu)
        queryset = queryset.filter(dl_bef_lu)
    if query_params.get('only_participant') == '1':
        queryset = queryset.filter(
            Q(owner=user) |
            Q(operator=user) |
            Q(visors=user) |
            Q(cooperators=user)
        ).distinct()
    if query_params.get('im_executor', '') == '1':
        queryset = queryset.filter(
            Q(operator=user) |
            Q(cooperators=user)
        )

    if query_params.get('is_overdue', '') == '1':
        _, not_complete_statuses, _ = get_cached_statuses()
        queryset = queryset.filter(
            dead_line__isnull=False, dead_line__lte=timezone.now(), status__in=not_complete_statuses
        )
    if query_params.get('is_risk', '') == '1':
        _, not_complete_statuses, _ = get_cached_statuses()
        queryset = queryset.filter(
            Q(
                dead_line__isnull=False,
                dead_line__lte=timezone.now(),
                status__in=not_complete_statuses
            ) |
            Q(object_tags__isnull=False),
        )
    status_code_str = query_params.get('status')
    if status_code_str:
        status_code_list = status_code_str.split(',')
        if status_code_list:
            queryset = queryset.filter(status_id__in=status_code_list)
    # prices = GoodsPriceModel.objects.filter(is_active=True, goods__shopping_carts=OuterRef('pk'),
    #                                         price_type_id=price_type_id)
    # queryset = cls.objects.prefetch_related(
    #     Prefetch('goods__gallery', queryset=GalleryModel.objects.filter(is_active=True, is_main=True))
    # ).select_related('goods', 'warehouse__manager__user').filter(user=user).annotate(
    #     # price=Subquery(prices.values('price'), output_field=models.DecimalField(max_digits=15, decimal_places=2)),
    #     price=F('goods__price_by_catalog'),
    #     currency_name=Subquery(prices.values('price_type__currency__name')),
    #     currency_icon=Subquery(prices.values('price_type__currency__icon')),
    #     remnant_sum=Sum('goods__remnants__quantity'),
    #     amount=F('price') * F('quantity'),
    # )
    if task_type:
        if task_type == ['interest'] and isinstance(query_params.get('parent'), str):
            queryset = queryset.filter(task_type_id='task')
        else:
            queryset = queryset.filter(task_type_id__in=task_type)
    queryset = filter_by_permissions(queryset, user)
    # Закомментировано 14.10.2025 @beluza_n. Не понятно, зачем нужна эта аннотация. В сериализаторах задач этих полей нет.
        # task_status_types = models.TaskStatusTypeModel.objects.filter(
        #     is_active=True,
        #     task_status__tasks=OuterRef('pk'),
        #     task_type_id__in=task_type)
        # queryset = queryset.annotate(
        #     is_open=Subquery(task_status_types.values('is_open')),
        #     is_complete=Subquery(task_status_types.values('is_complete')),
        #     show_btn=Subquery(task_status_types.values('show_btn'))
        # ).distinct()
    return queryset.distinct()


def order_tasks_queryset_from_get_param(request, queryset):
    model = models.TaskModel
    field_keys = get_field_keys(request, model)
    if field_keys:
        queryset = queryset.order_by(*field_keys, '-priority', '-created_at')
    else:
        order_by = queryset.query.order_by
        queryset = queryset.annotate(annotate_is_finished=Case(
            When(status_id='completed',
                 then=Value(1)),
            default=Value(0),
            output_field=IntegerField()))

        if request.query_params.get('task_type', 'task') == 'logistic' \
                and (
                request.user.profile.is_driver
                or request.query_params.get('page_name', '') in (
                        'page_list_logistic_task.TaskModel', 'page_list_logistic_task.TaskModel_user_tab'
                )
        ):
            queryset = queryset.annotate(
                annotate_order_exists=Case(
                    When(
                        task_delivery_points=None,
                        then=Value(0)
                    ),
                    default=Value(1),
                    output_field=IntegerField()
                ),
                annotate_order_date_start_plan=Min("task_delivery_points__start_goods_orders__delivery_date_plan_gte"),
            )
            queryset = queryset.order_by(
                *order_by,
                '-is_auction',
                'annotate_is_finished',
                'annotate_order_exists',
                'annotate_order_date_start_plan',
                'date_start_plan',
                '-priority',
                '-created_at')
        else:
            queryset = queryset.order_by(
                *order_by,
                '-is_auction',
                'annotate_is_finished',
                'date_start_plan',
                '-priority',
                '-created_at'
            )
    return queryset


def get_gantt_chart_task_queryset(request, queryset):
    """Возвращает список задач для диаграммы Ганта: убирает задачи без даты начала и конца,
    убирает завершенные задачи, фильтрует по дате начала и конца задач."""

    # Получаем стандартный кверисет задач
    queryset = order_queryset_from_get_param(
        request, models.TaskModel, get_task_queryset(request, queryset)
    )

    # Убираем задачи без даты начала и конца
    queryset = queryset.filter(
        is_active=True,
        date_start_plan__isnull=False,
        dead_line__isnull=False,
        )

    # # Убираем завершенные задачи.
    # try:
    #     query_params = request.query_params
    # except AttributeError:
    #     query_params = request.GET
    # if isinstance(query_params.get('task_type'), str):
    #     task_type = query_params.get('task_type').split(',')
    # else:
    #     task_type = ['task']
    # complete_statuses = models.TaskStatusTypeModel.objects.filter(
    #     task_type__in=task_type,
    #     is_complete=True,
    # ).values_list('task_status_id', flat=True)
    # queryset = queryset.exclude(
    #     status_id__in=complete_statuses
    # )

    # # Проверяем, что в пользовательских фильтрах была дата начала и конца. Если нет, то фильтруем даты по умолчанию.
    has_filter_store = False
    try:
        filter_store = FiltersStore.objects.get(author=request.user.profile,
                                                model=request.query_params.get('model', ''),
                                                page_name=request.query_params.get('page_name', '')
                                                )
    except FiltersStore.DoesNotExist:
        filter_store = None
    if filter_store:
        date_start_plan_active = filter_store.filters.get('values', dict()).get('date_start_plan', dict()).get(
            'active', False)
        dead_line_active = filter_store.filters.get('values', dict()).get('dead_line', dict()).get('active', False)
        if date_start_plan_active or dead_line_active:
            has_filter_store = True
    # if not has_filter_store:
    #     queryset = filter_gantt_chart_task_queryset_by_default_dates(request, queryset)

    queryset = queryset.order_by('level', 'date_start_plan',)
    return queryset


def filter_gantt_chart_task_queryset_by_default_dates(request, queryset):
    """Фильтрует кверисет задач для диаграммы Ганта по датам начала и конца проекта, либо по текущему месяцу."""

    filters = request.query_params.get('filters', '')
    try:
        filters_dict = json.loads(filters)
    except json.JSONDecodeError:
        filters_dict = dict()
    project_id = filters_dict.get('project')
    if project_id:
        try:
            project = WorkgroupModel.objects.get(pk=project_id)
        except WorkgroupModel.DoesNotExist:
            project = None
        if project:
            project_date_start_plan = project.date_start_plan if project.date_start_plan else project.created_at
            project_dead_line = project.dead_line
            if project_date_start_plan and project_dead_line:
                return queryset.filter(
                    date_start_plan__gte=project_date_start_plan, dead_line__lte=project_dead_line
                )
    now_date = timezone.now()
    now_date = now_date.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    try:
        now_date_plus_month = now_date.replace(now_date.year, now_date.month + 1, 1)
    except ValueError:
        now_date_plus_month = now_date.replace(now_date.year + 1, 1, 1)
    return queryset.filter(date_start_plan__gte=now_date, dead_line__lt=now_date_plus_month)


def check_need_set_calendar_event_by_task(some_task_id):
    some_task = models.TaskModel.objects.get(pk=some_task_id)

    if some_task.dead_line and (
            some_task.is_need_to_make_event or some_task.author.is_make_events_in_task_automatically):

        calendar = get_or_create_related_calendar('НЕ ТРЕБУЕТСЯ', some_task.id, False)

        event = EventCalendarModel.objects.filter(calendar=calendar,
                                                  event_type='task_deadline',
                                                  is_active=True).first()
        if not event:
            event = EventCalendarModel()
            event.calendar = calendar
            event.color = 'red'
            event.event_type, is_new = EventCalendarTypeModel.objects.get_or_create(code='task_deadline',
                                                                                    defaults={
                                                                                        'name': 'Сдача задачи',
                                                                                        'is_active': False,
                                                                                    }
                                                                                    )
        event.start_at = some_task.dead_line
        event.end_at = some_task.dead_line
        event.name = 'Сдать задачу ' + str(some_task)
        event.notify_at = some_task.dead_line - timedelta(days=1)
        event.description = some_task.description
        has_complete_status = some_task.status.task_status_type.first()

        event.is_finished = False
        if has_complete_status:
            event.is_finished = has_complete_status.is_complete

        event.save()

        allowed_user_ids = set()
        if some_task.is_need_to_make_event or some_task.operator.is_make_events_in_task_automatically:
            allowed_user_ids.add(some_task.operator_id)
        if some_task.is_need_to_make_event or some_task.owner.is_make_events_in_task_automatically:
            allowed_user_ids.add(some_task.owner_id)

        EventCalendarMemberModel.objects.filter(event=event).exclude(
            user_id__in=allowed_user_ids
        ).delete()

        for user_id in allowed_user_ids:
            EventCalendarMemberModel.objects.get_or_create(event=event, user_id=user_id)


def filter_by_permissions(queryset, user):
    """Фильтрует queryset задач по правам доступа пользователя."""
    touch_logistic = queryset.values_list('task_type', flat=True).distinct()
    if touch_logistic.count() == 1 and touch_logistic[0] == 'logistic':
        # Если в списке только логистические задачи, то проводим проверку и уходим
        if user.is_storekeeper or user.me_logistic_manager_only or user.can_create_logistic_task:
            return queryset

    # Получаем SQL из queryset для tree_id
    tree_ids_qs = queryset.values_list('tree_id', flat=True).order_by()
    compiler = SQLCompiler(tree_ids_qs.query, connection, tree_ids_qs.db)
    sql_tree_ids, params_tree_ids = compiler.as_sql()

    with connection.cursor() as cursor:
        # Шаг 1: Создаем временную таблицу для tree_id из исходного queryset
        cursor.execute('''
            CREATE TEMPORARY TABLE IF NOT EXISTS task_source_tree_ids 
            (tree_id BIGINT PRIMARY KEY) ''' + TABLESPACE + ';')
        cursor.execute('TRUNCATE TABLE task_source_tree_ids')
        
        cursor.execute('''
            INSERT INTO task_source_tree_ids (tree_id)
            SELECT DISTINCT tree_id FROM (''' + sql_tree_ids + ''') as src
        ''', params_tree_ids)
        
        # Шаг 2: Получаем ВСЕ задачи из этих деревьев для проверки прав
        cursor.execute('''
            CREATE TEMPORARY TABLE IF NOT EXISTS task_temporary_table_by_permissions_step1 
            (task_id UUID PRIMARY KEY, tree_id BIGINT) ''' + TABLESPACE + ';')
        cursor.execute('TRUNCATE TABLE task_temporary_table_by_permissions_step1')
        
        cursor.execute('''
            INSERT INTO task_temporary_table_by_permissions_step1 (task_id, tree_id)
            SELECT t.basemodel_ptr_id, t.tree_id
            FROM tasks_taskmodel t
            INNER JOIN task_source_tree_ids src ON src.tree_id = t.tree_id
            WHERE t.is_active_custom = true
        ''')

    # Шаг 3: Проверяем права на задачи через ORM
    tmp_objects = models.TmpSortedTableByPermissionStep1.objects.all()
    
    # Проверяем права доступа к задачам
    lookup = Q(task__visors=user)
    lookup = lookup | Q(task__cooperators=user)
    lookup = lookup | Q(task__operator=user)
    lookup = lookup | Q(task__owner=user)

    # Супервизор:
    visor_orgs = contractors_where_user_has_permission(user.pk, 'tasks_supervisor', None)
    if visor_orgs:
        lookup = lookup | Q(task__organization_id__in=visor_orgs)

    qs_full = tmp_objects.filter(lookup).values_list('task')

    # Директор организации
    lookup = Q(
        task__contractor__contractor_profile__user=user,
        task__contractor__contractor_profile__director=True,
    )
    qs_full = qs_full.union(tmp_objects.filter(lookup).values_list('task'))

    # Участник проекта/команды
    lookup = Q(
        task__workgroup__workgroupmembersmodel__is_active=True,
        task__workgroup__workgroupmembersmodel__member=user,
        task__workgroup__workgroupmembersmodel__membership_request_status__code='APPROVED',
    )
    qs_full = qs_full.union(tmp_objects.filter(lookup).values_list('task'))

    lookup = Q(
        task__project__workgroupmembersmodel__is_active=True,
        task__project__workgroupmembersmodel__member=user,
        task__project__workgroupmembersmodel__membership_request_status__code='APPROVED',
    )
    qs_full = qs_full.union(tmp_objects.filter(lookup).values_list('task'))
    
    # Поиск лидов по контрагенту
    lookup = Q(task__message_share__chat__member__user=user)
    qs_full = qs_full.union(tmp_objects.filter(lookup).values_list('task'))

    lookup = Q(Q(task__is_auction=True, ) & Q(task__task_type__code='task'))
    qs_full = qs_full.union(tmp_objects.filter(lookup).values_list('task').distinct())

    if user.is_auctioneer:
        lookup = Q(Q(task__is_auction=True, ) & Q(task__task_type__code='interest'))
        qs_full = qs_full.union(tmp_objects.filter(lookup).values_list('task').distinct())

    if user.is_auction_moderator:
        lookup = Q(task__task_type__code='interest')
        qs_full = qs_full.union(tmp_objects.filter(lookup).values_list('task').distinct())

    # Шаг 4: Извлекаем уникальные tree_id для задач с правами
    compiler = SQLCompiler(qs_full.query, connection, qs_full.db)
    sql_tasks_with_rights, params_tasks_with_rights = compiler.as_sql()
    
    with connection.cursor() as cursor:
        # Создаём временную таблицу для разрешенных tree_id
        cursor.execute('''
            CREATE TEMPORARY TABLE IF NOT EXISTS task_allowed_tree_ids 
            (tree_id BIGINT PRIMARY KEY) ''' + TABLESPACE + ';')
        cursor.execute('TRUNCATE TABLE task_allowed_tree_ids')
        
        # Получаем tree_id из задач с правами напрямую через SQL
        cursor.execute('''
            INSERT INTO task_allowed_tree_ids (tree_id)
            SELECT DISTINCT step1.tree_id
            FROM (''' + sql_tasks_with_rights + ''') as tasks_with_rights(task_id)
            INNER JOIN task_temporary_table_by_permissions_step1 step1 
                ON step1.task_id = tasks_with_rights.task_id
        ''', params_tasks_with_rights)

    # Шаг 5: Фильтруем исходный queryset по разрешенным tree_id через raw SQL
    # Получаем список pk из исходного queryset, которые есть в разрешенных деревьях
    filtered_qs = queryset.extra(
        tables=['task_allowed_tree_ids'],
        where=['tasks_taskmodel.tree_id = task_allowed_tree_ids.tree_id']
    )
    return filtered_qs


def get_reason_object(reason_id):
    if not reason_id:
        return None
    try:
        reason_obj = BaseModel.objects.super_get(pk=reason_id)
    except (ValidationError, ValueError, TypeError):
        reason_obj = None
    if not reason_obj:
        try:
            reason_obj = MessageModel.objects.get(message_uid=reason_id)
        except (MessageModel.DoesNotExist, ValidationError, ValueError, TypeError):
            return None
    return reason_obj


def create_chat_message_about_task_reason(task):
    """Если задача создана на основе сообщения чата, в этот чат отправляется системное сообщение."""
    reason_id = task.reason
    if not reason_id:
        return 'Task has no reason.'
    try:
        reason_obj = MessageModel.objects.get(message_uid=reason_id)
    except MessageModel.DoesNotExist:
        return 'Reason is not chat message.'
    message = MessageModel()
    message.is_system = True
    message.text = f'На основе этого сообщения {task.author.full_name} создал задачу "#{task}".'
    message.share = task
    message.message_reply = reason_obj
    reason_chat = reason_obj.chat
    message.chat = reason_chat
    message.created = timezone.now()
    message.save()
    message_data = MessageListSerializer(message).data
    message_data['chat_uid'] = str(message.chat.chat_uid)
    message_data['chat_name'] = reason_chat.name
    message_data['is_public'] = reason_chat.is_public
    message_data['is_new'] = True
    data = json.dumps(
        {
            "event": "chat_message",
            "data": message_data
        },
        cls=DjangoJSONEncoder,
    )
    socketio_redis.publish(SOCKETIO_SYSTEM_CHANNEL, data)
    return 'done.'


def send_socketio_about_new_task(task):
    from .serializers import ListTaskSerializer
    s_data = ListTaskSerializer(instance=task).data
    if task.task_type_id == 'logistic':
        s_data['has_order'] = task.task_delivery_points.all().exists()
    data = json.dumps({
        'event': 'task_create',
        'data': s_data,
    }, cls=DjangoJSONEncoder)
    socketio_redis.publish(SOCKETIO_SYSTEM_CHANNEL, data)
    return 'message sent'


def send_socketio_about_update_task(task):
    from .serializers import DetailTaskSerializer
    s_data = DetailTaskSerializer(instance=task).data
    if task.task_type_id == 'logistic':
        s_data['has_order'] = task.task_delivery_points.all().exists()
    data = json.dumps(
        {
            'event': 'task_update',
            'data': s_data,
        },
        cls=DjangoJSONEncoder
    )
    socketio_redis.publish(SOCKETIO_SYSTEM_CHANNEL, data)
    return 'message_sent'


def send_socketio_about_delete_task(task):
    data = json.dumps(
        {
            'event': 'task_delete',
            'data': str(task.pk),
        },
        cls=DjangoJSONEncoder
    )
    socketio_redis.publish(SOCKETIO_SYSTEM_CHANNEL, data)
    return 'message_sent'


def send_socketio_about_task_touch(task):
    if task.operator:
        data = json.dumps({
            'event': 'system_notify',
            'data': {
                'user': str(task.operator.id),
                'count': task.operator.operator_tasks.filter(status__task_status_type__is_complete=False).values_list(
                    'pk', flat=True).distinct().count()
            },
        }, cls=DjangoJSONEncoder)

        socketio_redis.publish(SOCKETIO_SYSTEM_CHANNEL, data)
    return 'message sent'


def get_status_data(task: models.TaskModel) -> dict:
    """Возвращает сериализованный статус задачи."""
    if task.task_type_id == 'milestone':
        task_status_code = task.milestone_status
    elif task.task_type_id == 'stage':
        task_status_code = task.stage_status
    else:
        task_status_code = task.status_id

    from .serializers import CachedTaskStatusTypeModelSerializer
    task_status_type = str(task.task_type_id) + '__' + task_status_code
    data = CachedTaskStatusTypeModelSerializer(task_status_type).data
    return data


def get_cooperator_status_data(task_cooperator: models.TaskCooperator) -> dict:
    """Возвращает сериализованный статус выполнения задачи соисполнителем."""
    task_status = task_cooperator.status
    from .serializers import TaskStatusModelSerializer
    data = TaskStatusModelSerializer(task_status).data
    task_type = task_cooperator.task.task_type # тип задачи соисполнителя определяется типом основной задачи
    try:
        task_status_type = task_status.task_status_type.get(task_type=task_type)
    except ObjectDoesNotExist:
        return data
    data['is_open'] = task_status_type.is_open
    data['is_complete'] = task_status_type.is_complete
    data['show_btn'] = task_status_type.show_btn
    return data


def get_budget_aggregate_qs(obj, request):
    qs = models.TaskBudgetModel.objects.filter(is_active=True)
    if isinstance(obj, models.TaskModel):
        if not obj.get_detail_permission(request):
            return qs.none()
        return qs.filter(task=obj, )

    elif isinstance(obj, models.TaskSprintModel):
        user = request.user.profile
        if not obj.author == user and not filter_by_permissions(obj.tasks.filter(is_active=True), user).exists():
            return qs.none()
        tasks = obj.tasks.filter(is_active=True)
        return qs.filter(task__in=tasks, )
    elif isinstance(obj, WorkgroupModel):
        if not obj.get_detail_permission(request):
            return qs.none()
        if obj.is_project:
            tasks = obj.project_tasks.filter(is_active=True)
            return qs.filter(task__in=tasks)
        else:
            tasks = obj.workgroup_tasks.filter(is_active=True)
            return qs.filter(task__in=tasks)
    else:
        return qs.none()


def get_budget_aggregate_data(qs):
    total_sum = qs.aggregate(total_sum=Sum('amount_sum'))
    if total_sum['total_sum'] is None:
        return None
    detail = tuple(qs)
    for each in detail:
        each['cost_item'] = BaseCatalogRetrieveSerializer(
            instance=models.CostItemTaskModel.objects.get(pk=each['cost_item'])).data
    data = {
        'detail_sum': detail,
        'currency': AppCurrencySerializer(CurrencyModel.objects.get(code=DEFAULT_CURRENCY_CODE)).data,
    }
    data.update(total_sum)
    return data


def get_difficulty_aggregate_qs(obj, request):
    qs = models.TaskDifficulty.objects.filter(is_active=True)
    if isinstance(obj, models.TaskModel):
        if not obj.get_detail_permission(request):
            return qs.none()
        return qs.filter(task=obj, )

    elif isinstance(obj, models.TaskSprintModel):
        user = request.user.profile
        if not obj.author == user and not filter_by_permissions(obj.tasks.filter(is_active=True), user).exists():
            return qs.none()
        tasks = obj.tasks.filter(is_active=True)
        return qs.filter(task__in=tasks, )
    elif isinstance(obj, WorkgroupModel):
        if not obj.get_detail_permission(request):
            return qs.none()
        if obj.is_project:
            return qs.filter(task__project=obj, )
        else:
            return qs.filter(task__workgroup=obj)
    else:
        return qs.none()


def get_difficulty_aggregate_data(qs):
    total_avg = qs.aggregate(total_avg=Avg('score'))
    if total_avg['total_avg'] is None:
        return None
    qs = qs.values(
        'criterion',
    ).annotate(score_avg=Avg('score'))
    detail = tuple(qs)
    for each in detail:
        each['criterion'] = BaseCatalogRetrieveSerializer(
            instance=models.TaskDifficultyCriterion.objects.get(pk=each['criterion'])).data
        try:
            each['score_avg'] = round(each['score_avg'], 2)
        except TypeError:
            each['score_avg'] = 0
    data = {
        'max_value': 10,
        'detail_avg': detail,
        'total_avg': round(total_avg['total_avg'], 2)
    }
    return data


def create_workbook_task_analytics(queryset) -> Workbook:
    """Создает рабочую книгу Excel по аналитике задач. Возвращает объект WorkBook."""
    wb = load_workbook('task_analytics.xlsx', )
    sheet = wb.active
    row_counter = 4
    est = pytz.timezone(TIME_ZONE)
    bd = Side(style='hair', color='000000')
    border = Border(left=bd, right=bd, bottom=bd, top=bd)
    for each in queryset:
        sheet.cell(row_counter, 1, each.counter).border = border
        sheet.cell(row_counter, 2, each.name).border = border
        sheet.cell(row_counter, 3, get_description_text(each)).border = border
        sheet.cell(
            row_counter, 4, each.created_at.astimezone(est).replace(tzinfo=None) if each.created_at else ""
        ).border = border
        sheet.cell(
            row_counter, 5, each.finished_date.astimezone(est).replace(tzinfo=None) if each.finished_date else ""
        ).border = border
        sheet.cell(row_counter, 6, each.budget_sum if each.budget_sum is not None else 0).border = border
        sheet.cell(row_counter, 7, each.difficulty_avg if each.difficulty_avg is not None else 0).border = border
        sheet.cell(
            row_counter, 8, each.execution_time_sum if each.execution_time_sum is not None else 0
        ).border = border
        row_counter += 1
    return wb


def create_workbook_task_list(queryset) -> Workbook:
    """Создает рабочую книгу Excel по аналитике задач. Возвращает объект WorkBook."""
    wb = load_workbook('task_report.xlsx', )
    sheet = wb.active
    row_counter = 4
    fill_color = 'FFFFFF'
    est = pytz.timezone(TIME_ZONE)
    bd = Side(border_style='hair', color='000000')
    border = Border(left=bd, right=bd, bottom=bd, top=bd)
    for each in queryset:
        task_steps = each.task_steps
        row = []
        if task_steps:
            for step in task_steps:
                row.append(sheet.cell(row_counter, 1, each.counter))  # Номер задачи
                row.append(sheet.cell(row_counter, 2, each.name))  # Название задачи
                row.append(sheet.cell(row_counter, 3, get_description_text(each)))  # Описание задачи
                row.append(sheet.cell(
                    row_counter, 4,
                    each.finished_date.astimezone(est).replace(
                        tzinfo=None) if each.finished_date else ''))  # Дата сдачи
                row.append(sheet.cell(row_counter, 5, step.hours))  # Затрачено
                row.append(sheet.cell(row_counter, 6, step.measure_unit.name_short))  # Единица измерения
                row.append(sheet.cell(row_counter, 7, step.work_type.name))  # Вид работ
                row.append(sheet.cell(row_counter, 8, step.description))  # Описание работ
                row.append(sheet.cell(row_counter, 9, step.author.user.get_full_name()))  # Исполнитель
                row_counter += 1
        else:
            row.append(sheet.cell(row_counter, 1, each.counter))  # Номер задачи
            row.append(sheet.cell(row_counter, 2, each.name))  # Название задачи
            row.append(sheet.cell(row_counter, 3, get_description_text(each)))  # Описание задачи
            row.append(sheet.cell(
                row_counter, 4,
                each.finished_date.astimezone(est).replace(tzinfo=None) if each.finished_date else ''))  # Дата сдачи
            row.append(sheet.cell(row_counter, 5, ''))  # Потраченное время
            row.append(sheet.cell(row_counter, 6, ''))  # Единица измерения
            row.append(sheet.cell(row_counter, 7, ''))  # Вид работ
            row.append(sheet.cell(row_counter, 8, ''))  # Описание работ
            row.append(sheet.cell(
                row_counter, 9,
                each.operator.user.get_full_name() if each.operator is not None else ''))  # Ответственный исполнитель
            row_counter += 1
        for cell in row:
            cell.border = border
            cell.fill = PatternFill(fill_type='solid', fgColor=fill_color)
        fill_color = 'F8F8FF' if fill_color == 'FFFFFF' else 'FFFFFF'
    return wb


def create_workbook_interest_list(queryset) -> Workbook:
    """Создает рабочую книгу Excel по аналитике задач. Возвращает объект WorkBook."""
    wb = load_workbook('interest_report.xlsx', )
    sheet = wb.active
    row_counter = 4
    fill_color = 'FFFFFF'
    est = pytz.timezone(TIME_ZONE)
    bd = Side(border_style='hair', color='000000')
    border = Border(left=bd, right=bd, bottom=bd, top=bd)
    for each in queryset:
        task_steps = each.task_steps
        row = []
        if task_steps:
            for step in task_steps:
                row.append(sheet.cell(row_counter, 1, each.counter))  # Номер задачи
                row.append(sheet.cell(row_counter, 2, each.name))  # Название задачи
                row.append(sheet.cell(row_counter, 3, get_description_text(each)))  # Описание задачи
                row.append(sheet.cell(row_counter, 4,
                                      each.potential_contractor.name if each.potential_contractor else ''))  # Клиент
                row.append(sheet.cell(row_counter, 5,
                                      each.potential_contractor.company_name if each.potential_contractor else ''))  # Организация
                row.append(sheet.cell(row_counter, 6,
                                      each.potential_contractor.phone if each.potential_contractor else ''))  # Телефон
                row.append(sheet.cell(row_counter, 7,
                                      each.potential_contractor.email if each.potential_contractor else ''))  # Электронная почта
                row.append(
                    sheet.cell(row_counter, 8, each.created_at.astimezone(est).replace(tzinfo=None)))  # Дата создания
                row.append(sheet.cell(row_counter, 9, each.finished_date.astimezone(est).replace(
                    tzinfo=None) if each.finished_date else ''))  # Дата завершения
                row.append(sheet.cell(row_counter, 10, each.lead_source.name if each.lead_source else ''))  # Источник
                row.append(sheet.cell(row_counter, 11, step.work_type.name))  # Вид работ
                row.append(sheet.cell(row_counter, 12, step.date if step.date else ''))  # Дата проведения работы
                row.append(sheet.cell(row_counter, 13, each.status.name if each.status.name else ''))  # Статус
                row.append(sheet.cell(row_counter, 14,
                                      each.rejection_reason.name if each.rejection_reason else ''))  # Причина отказа
                row.append(sheet.cell(row_counter, 15,
                                      each.operator.user.get_full_name() if each.operator is not None else ''))  # Ответственный
                row_counter += 1
        else:
            row.append(sheet.cell(row_counter, 1, each.counter))  # Номер задачи
            row.append(sheet.cell(row_counter, 2, each.name))  # Название задачи
            row.append(sheet.cell(row_counter, 3, get_description_text(each)))  # Описание задачи
            row.append(sheet.cell(row_counter, 4,
                                  each.potential_contractor.name if each.potential_contractor else ''))  # Клиент
            row.append(sheet.cell(row_counter, 5,
                                  each.potential_contractor.company_name if each.potential_contractor else ''))  # Организация
            row.append(sheet.cell(row_counter, 6,
                                  each.potential_contractor.phone if each.potential_contractor else ''))  # Телефон
            row.append(sheet.cell(row_counter, 7,
                                  each.potential_contractor.email if each.potential_contractor else ''))  # Электронная почта
            row.append(
                sheet.cell(row_counter, 8, each.created_at.astimezone(est).replace(tzinfo=None)))  # Дата создания
            row.append(sheet.cell(row_counter, 9, each.finished_date.astimezone(est).replace(
                tzinfo=None) if each.finished_date else ''))  # Дата завершения
            row.append(sheet.cell(row_counter, 10, each.lead_source.name if each.lead_source else ''))  # Источник
            row.append(sheet.cell(row_counter, 11, ''))  # Вид работ
            row.append(sheet.cell(row_counter, 12, ''))  # Дата проведения работы
            row.append(sheet.cell(row_counter, 13, each.status.name if each.status.name else ''))  # Статус
            row.append(sheet.cell(row_counter, 14,
                                  each.rejection_reason.name if each.rejection_reason else ''))  # Причина отказа
            row.append(sheet.cell(row_counter, 15,
                                  each.operator.user.get_full_name() if each.operator is not None else ''))  # Ответственный
            row_counter += 1
        for cell in row:
            cell.border = border
            cell.fill = PatternFill(fill_type='solid', fgColor=fill_color)
        fill_color = 'F8F8FF' if fill_color == 'FFFFFF' else 'FFFFFF'
    return wb


def get_description_text(instance):
    description = instance.description
    if description:
        return BeautifulSoup(description, 'lxml').get_text(separator=" ").strip()
    else:
        return ''


def plus_one_day_logistic_tasks():
    qs = models.TaskModel.objects.filter(
        is_active=True,
        task_type_id='logistic'
    ).exclude(status_id='completed').order_by('created_at')
    count = ceil(qs.count() / 100)
    # print(f"\nstart.")
    delta = timedelta(days=1)
    for each in range(count):
        tasks = qs[each * 100:each * 100 + 100]
        for task in tasks:
            date_start_plan = task.date_start_plan
            if date_start_plan:
                task.date_start_plan = date_start_plan + delta
            date_start_fact = task.date_start_fact
            if date_start_fact:
                task.date_start_fact = date_start_fact + delta
            task.save(update_fields=('date_start_plan', 'date_start_fact'))
            for delivery_point in task.task_delivery_points.all():
                delivery_date = delivery_point.delivery_date
                if delivery_date:
                    delivery_point.delivery_date = delivery_date + delta


def delete_task_delivery_points(task_delivery_points):
    for task_delivery_point in task_delivery_points:
        if task_delivery_point.is_start:
            orders = task_delivery_point.start_goods_orders.all()
            for order in orders:
                order_task_delivery_point = order.task_delivery_point
                order.task_delivery_point = None
                order.start_task_delivery_point = None
                order.save(update_fields=('task_delivery_point', 'start_task_delivery_point',))
                if order_task_delivery_point:
                    if not order_task_delivery_point.goods_orders.filter(is_active=True).exists():
                        order_task_delivery_point.delete()
        else:
            orders = task_delivery_point.goods_orders.all()
            for order in orders:
                start_order_task_delivery_point = order.start_task_delivery_point
                order.task_delivery_point = None
                order.start_task_delivery_point = None
                order.save(update_fields=('task_delivery_point', 'start_task_delivery_point',))
                if start_order_task_delivery_point:
                    if not start_order_task_delivery_point.start_goods_orders.filter(is_active=True).exists():
                        start_order_task_delivery_point.delete()
        task_delivery_point.delete()


def can_update_status(profile_id, task):
    if task.get_available_statuses(profile_id):
        return True
    else:
        return False
    # if task.operator_id and task.operator_id == user_id:
    #     return True
    # if task.owner_id and task.owner_id == user_id:
    #     return True
    # if getattr(getattr(task, 'project', None), 'author_id', '') == user_id:
    #     return True
    # if getattr(getattr(task, 'workgroup', None), 'author_id', '') == user_id:
    #     return True
    # if task.sprint and task.sprint.author_id and task.sprint.author_id == user_id:
    #     return True
    # if task.contractor and task.contractor.profiles.filter(pk=user_id).exists():
    #     return True
    # return False


def set_delivery_status_in_orders(task):
    from .notifications import notify_about_order_in_transit
    from crm.models import GoodsOrderModel
    task_delivery_points = task.task_delivery_points.filter(is_active=True, is_start=False).values_list('pk', flat=True)
    orders = GoodsOrderModel.objects.filter(is_active=True, task_delivery_point_id__in=task_delivery_points).exclude(
        Q(delivery_status_id__in=('in_transit', 'delivered', 'partially_delivered',))
        | Q(execute_status_id__in=('completed', 'canceled', 'partially_canceled',))
    )

    for order in orders:
        order.delivery_status_id = 'in_transit'
        order.save(update_fields=('delivery_status_id',))
        async_task(notify_about_order_in_transit, task, order)
    return


def get_goods_by_warehouses(task, warehouses=None):
    from crm.models import GoodsOrderModel, TPGoodsOrderModel
    delivery_points = task.task_delivery_points.filter(is_active=True, is_start=False).values_list('pk', flat=True)
    orders = GoodsOrderModel.objects.filter(
        is_active=True,
        task_delivery_point_id__in=delivery_points
    ).order_by('warehouse__name')
    if warehouses:
        orders = orders.filter(warehouse__in=warehouses)
    warehouses = orders.values('warehouse').annotate(
        id=F('warehouse_id'),
        name=F('warehouse__name'),
        need_amount_pay=Sum(
            'start_task_delivery_point__need_amount_pay',
            filter=Q(
                start_task_delivery_point__is_active=True,
                start_task_delivery_point__task__is_active=True,
            ),
        )
    )
    for warehouse in warehouses:
        goods = TPGoodsOrderModel.objects.filter(
            is_active=True,
            owner__in=orders,
            owner__warehouse_id=warehouse['id']
        ).order_by(
            'goods__name',
        ).values(
            'goods',
        ).annotate(
            id=F('goods_id'),
            name=F('goods__name'),
            code=F('goods__code'),
            count=Count('goods'),
            quantity=Sum('quantity'),
            quantity_success=Sum('quantity_success'),
            amount=Sum('amount'),
            measure_unit_name_short=F('goods__base_measure_unit__name_short')
        )
        for each in goods:
            quantity_loaded = task.task_loading_goods.filter(
                warehouse_id=warehouse['id'], goods_id=each['id']
            ).aggregate(Sum('quantity'))['quantity__sum']
            if quantity_loaded is None:
                quantity_loaded = 0
            each['quantity_loaded'] = quantity_loaded
        warehouse['goods'] = goods
        amount_paid = task.task_loading_goods.filter(
            is_active=True, warehouse_id=warehouse['id']).aggregate(Sum('amount_paid'))['amount_paid__sum']
        if amount_paid is None:
            amount_paid = 0
        warehouse['amount_paid'] = amount_paid
    return warehouses


def get_cached_statuses():
    """Возвращает кортеж (statuses, not_complete_statuses, complete_statuses) из кеша."""
    cache_key = 'task_status_type_model_cache'
    result = cache.get(cache_key)
    if result is None:
        rows = list(
            models.TaskStatusTypeModel.objects.filter(
                is_active=True,
                task_status__is_active=True,
                task_type='task',
            ).order_by('sort', 'task_status__name').values_list('task_status', 'is_complete')
        )
        statuses = [row[0] for row in rows]
        not_complete_statuses = [row[0] for row in rows if not row[1]]
        complete_statuses = [row[0] for row in rows if row[1]]
        result = (statuses, not_complete_statuses, complete_statuses)
        cache.set(cache_key, result, timeout=86400)  # 1 день
    return result


def get_tasks_status_count(queryset):
    """Принимает queryset объектов TaskModel. Считает количество задач в каждом статусе.
    Добавляет вычисляемый статус overdue - просроченные."""
    statuses, not_complete_statuses, complete_statuses = get_cached_statuses()
    aggregate_lookup = {
        status_code: Count('status', filter=Q(status=status_code)) for status_code in statuses
    }
    data = queryset.aggregate(
        **aggregate_lookup,
        overdue=Count(
            'pk',
            filter=Q(dead_line__isnull=False, dead_line__lte=timezone.now(), status__in=not_complete_statuses)
        )
    )
    return data


def get_my_tasks_count(request, queryset):
    """Принимает queryset незавершённых задач. Считает количество задач,
    где я исполнитель, я постановщик, я наблюдатель, а также количество просроченных."""
    user = request.user.profile
    cooperator_task_ids = models.TaskCooperator.objects.filter(
        user_id=user.pk
    ).values('task_id')
    visor_task_ids = models.TaskVisor.objects.filter(
        user_id=user.pk
    ).values('task_id')
    data = {
        'im_operator': queryset.filter(
            Q(operator_id=user.pk) | Q(pk__in=cooperator_task_ids)
        ).count(),
        'im_owner': queryset.filter(owner_id=user.pk).count(),
        'im_visor': queryset.filter(pk__in=visor_task_ids).count(),
        'im_participant': queryset.filter(
            Q(operator_id=user.pk) |
            Q(owner_id=user.pk) |
            Q(pk__in=cooperator_task_ids) |
            Q(pk__in=visor_task_ids)
        ).count(),
    }
    overdue_qs = queryset.filter(
        dead_line__isnull=False, dead_line__lt=timezone.now(),
    )
    overdue_qs = filter_by_permissions(overdue_qs, user)
    data['overdue'] = overdue_qs.count()
    return data


def get_tasks_for_sprint_qs(queryset, request,):
    from common import utils as common_utils
    qs = common_utils.order_queryset_from_get_param(
        request, models.TaskModel, get_task_queryset(request, queryset, list_type='sprint_list')
    )
    sprint_id = request.query_params.get('sprint', None)
    if not sprint_id:
        sprint_id = request.data.get('sprint')
    if sprint_id:
        sprint = models.TaskSprintModel.objects.get(pk=sprint_id)
        projects = sprint.projects.all().values_list('pk', flat=True)
        if projects.exists():
            qs = qs.filter(project__in=projects)
        else:
            qs = qs.none()
    else:
        user = request.user.profile
        projects_id = user.workgroupmembersmodel_set.filter(
            membership_request_status__code='APPROVED',
            membership_role__code__in=('FOUNDER', 'MODERATOR',)
        ).values_list('work_group', flat=True)
        qs = qs.filter(project__in=projects_id)
    return qs.filter(sprint__isnull=True).exclude(status_id='completed',)


align_center = pyexcelerate.Alignment(horizontal='center', vertical='center', wrap_text=True)
align_left = pyexcelerate.Alignment(horizontal='left', vertical='center', wrap_text=True)

normal_9 = pyexcelerate.Font(family='Times new roman', size=9,)
bold_9 = pyexcelerate.Font(bold=True, family='Times new roman', size=9,)

bold_9_center_style = pyexcelerate.Style(font=bold_9, alignment=align_center,)
normal_9_left_style = pyexcelerate.Style(font=normal_9, alignment=align_left)
normal_9_center_style = pyexcelerate.Style(font=normal_9, alignment=align_center)


def get_sprint_report_file_act(sprint):

    wb = pyexcelerate.Workbook()
    ws = wb.new_sheet('Акт выполненных работ')

    ws.set_col_style(1, pyexcelerate.Style(size=5,))
    ws.set_col_style(2, pyexcelerate.Style(size=15,))
    ws.set_col_style(3, pyexcelerate.Style(size=50,))
    ws.set_col_style(4, pyexcelerate.Style(size=30,))
    ws.set_col_style(5, pyexcelerate.Style(size=15,))
    ws.set_col_style(6, pyexcelerate.Style(size=15,))
    ws.set_col_style(7, pyexcelerate.Style(size=20,))

    row = 1
    first_col = 1
    last_col = 7
    localdate = timezone.localdate()
    datetime_begin = sprint.begin_date
    if not datetime_begin:
        return wb
    date_begin = datetime_begin.strftime("%d.%m.%Y")
    datetime_end = sprint.finished_date if sprint.finished_date else localdate
    date_end = datetime_end.strftime("%d.%m.%Y")
    sprint_tasks_id = get_all_sprint_tasks(sprint)
    sprint_tasks_qs = models.TaskModel.objects.filter(pk__in=sprint_tasks_id)
    sprint_tasks_time_executed = get_sprint_time_tracking(sprint)
    time_executed_sum = sprint_tasks_time_executed.aggregate(hours_sum=Sum('hours'))['hours_sum']

    title_data = (
        ('АКТ ВЫПОЛНЕННЫХ РАБОТ',),
        (localdate.strftime("%d.%m.%Y"),),
        (f'Общая информация',),
        (f'Период работ: {date_begin} - {date_end}',),
        (f'Общий объем трудозатрат: {time_executed_sum} ч.',),
    )
    ws.range((row, first_col), (row + 4, first_col)).value = title_data
    ws.set_cell_style(row, first_col, normal_9_center_style)
    ws.set_cell_style(row + 1, first_col, normal_9_center_style)
    ws.set_cell_style(row + 2, first_col, normal_9_center_style)
    ws.set_cell_style(row + 3, first_col, normal_9_left_style)
    ws.set_cell_style(row + 4, first_col, normal_9_left_style)

    merged_cells = [
        ((row, first_col), (row, last_col)),
        ((row + 1, first_col), (row + 1, last_col)),
        ((row + 2, first_col), (row + 2, last_col)),
        ((row + 3, first_col), (row + 3, last_col)),
        ((row + 4, first_col), (row + 4, last_col)),
    ]

    table_header_data = (
        (
            '№',
            '№ задачи',
            'Задача',
            'Исполнители',
            'Плановые трудозатраты (ч)',
            'Фактические трудозатраты (ч)',
            'Статус',
        ),
    )
    row += 7
    ws.range((row, first_col), (row, last_col)).value = table_header_data
    ws.range((row, first_col), (row, last_col)).style.font = bold_9
    ws.range((row, first_col), (row, last_col)).style.alignment = align_center
    row += 1
    projects_id = set(sprint_tasks_qs.values_list('project', flat=True))
    has_no_project = False
    if None in projects_id:
        has_no_project = True
        projects_id.discard(None)
    projects = list(WorkgroupModel.objects.filter(is_active=True, pk__in=projects_id))
    if has_no_project:
        projects.append(None)
    total_time_plan = 0
    total_time_fact = 0
    if projects:
        for project in projects:
            project_tasks = sprint_tasks_qs.filter(project=project)
            if project_tasks:
                project_time_fact_qs = sprint_tasks_time_executed.filter(task__project=project)
                project_time_fact = project_time_fact_qs.aggregate(project_time_fact=Sum('hours'))['project_time_fact']
                if not project_time_fact:
                    continue
                if project:
                    ws.set_cell_value(row, first_col, f'Проект "{project.name}"')
                else:
                    ws.set_cell_value(row, first_col, f'Задачи без проекта')
                ws.set_cell_style(row, first_col, bold_9_center_style)
                merged_cells.append(((row, first_col), (row, last_col)))
                task_counter = 1
                row += 1
                project_data = list()
                project_time_plan = 0
                for project_task in project_tasks:
                    time_qs = project_time_fact_qs.filter(task=project_task)
                    execution_time_fact = time_qs.aggregate(hours_sum=Sum('hours'))['hours_sum']
                    if not execution_time_fact:
                        continue
                    execution_time_plan = project_task.execution_time_plan \
                        if project_task.execution_time_plan is not None else 0
                    operators_qs = time_qs.values_list('user__user__first_name', 'user__user__last_name').distinct('author')
                    operators_str = ', '.join([f"{_[0]} {_[1]}" for _ in operators_qs])
                    project_data.append(
                        (
                            task_counter,
                            project_task.counter,
                            project_task.name,
                            operators_str,
                            execution_time_plan,
                            execution_time_fact,
                            project_task.status.name
                        )
                    )
                    project_time_plan += execution_time_plan
                    task_counter += 1
                ws.range((row, first_col), (row + task_counter - 1, last_col)).value = project_data
                ws.range((row, first_col), (row + task_counter - 1, last_col)).style.font = normal_9
                ws.range((row, first_col), (row + task_counter - 1, last_col)).style.alignment = align_left
                row = row + task_counter - 1
                ws.range((row, 2), (row, 6)).value = (
                    ('Итого по проекту', None, None, project_time_plan, project_time_fact, ),
                )
                ws.range((row, 2), (row, 6)).style.font = bold_9
                ws.range((row, 2), (row, 6)).style.alignment = align_left
                merged_cells.append(((row, 2), (row, 4)))
                total_time_plan += project_time_plan
                total_time_fact += project_time_fact
                row += 1
            else:
                row += 1
        ws.range((row, 2), (row, 6)).value = (('Итого', None, None, total_time_plan, total_time_fact, ),)
        ws.range((row, 2), (row, 6)).style.font = bold_9
        ws.range((row, 2), (row, 6)).style.alignment = align_left
        merged_cells.append(((row, 2), (row, 4)))

    # Границы ячеек:
    ws.range((8, first_col), (row, last_col)).style.borders.top.style = '_'
    ws.range((8, first_col), (row, last_col)).style.borders.top.color = pyexcelerate.Color(0, 0, 0)
    ws.range((8, first_col), (row, last_col)).style.borders.left.style = '_'
    ws.range((8, first_col), (row, last_col)).style.borders.left.color = pyexcelerate.Color(0, 0, 0)
    ws.range((8, first_col), (row, last_col)).style.borders.bottom.style = '_'
    ws.range((8, first_col), (row, last_col)).style.borders.bottom.color = pyexcelerate.Color(0, 0, 0)
    ws.range((8, first_col), (row, last_col)).style.borders.right.style = '_'
    ws.range((8, first_col), (row, last_col)).style.borders.right.color = pyexcelerate.Color(0, 0, 0)
    # Объединяем ячейки в самом конце, иначе глючат границы:
    for merged_cell in merged_cells:
        ws.range(*merged_cell).merge()
    return wb


def get_sprint_report_file_report(sprint):
    wb = pyexcelerate.Workbook()
    ws = wb.new_sheet('Отчет по проектам')

    ws.set_col_style(1, pyexcelerate.Style(size=20, ))
    ws.set_col_style(2, pyexcelerate.Style(size=20, ))
    ws.set_col_style(3, pyexcelerate.Style(size=20, ))
    ws.set_col_style(4, pyexcelerate.Style(size=20, ))
    ws.set_col_style(5, pyexcelerate.Style(size=20, ))
    ws.set_col_style(6, pyexcelerate.Style(size=20, ))
    ws.set_col_style(7, pyexcelerate.Style(size=20, ))
    ws.set_col_style(8, pyexcelerate.Style(size=20, ))
    ws.set_col_style(9, pyexcelerate.Style(size=20, ))
    ws.set_col_style(10, pyexcelerate.Style(size=20, ))
    ws.set_col_style(11, pyexcelerate.Style(size=20, ))
    ws.set_col_style(12, pyexcelerate.Style(size=20, ))
    ws.set_col_style(13, pyexcelerate.Style(size=20, ))
    row = 1

    sprint_tasks_time_executed = get_sprint_time_tracking(sprint)
    sprint_tasks_id = get_all_sprint_tasks(sprint)
    sprint_tasks_qs = models.TaskModel.objects.filter(is_active=True, pk__in=sprint_tasks_id)
    sprint_tasks_budget = models.TaskBudgetModel.objects.filter(
        is_active=True,
        task_id__in=sprint_tasks_id,
    )

    merged_cells = []
    first_col = 1
    last_col = 13
    ws.set_cell_value(1, first_col, f"Отчет по проектам спринта \"{sprint.name}\"")
    ws.set_cell_style(1, first_col, bold_9_center_style)
    merged_cells.append(((1, 1), (1, last_col)))

    table_header_data = (
        (
            'Статус',
            'Дата начала',
            'Крайний срок',
            'Продолжительность (в днях)',
            '№ задачи',
            'Название задачи',
            'Исполнители',
            'Виды работ',
            'Результат',
            'Стоимость плановая',
            'Стоимость фактическая',
            'Плановые (в часах)',
            'Фактические (в часах)',
        ),
    )
    row += 2
    ws.range((row, first_col), (row, last_col)).value = table_header_data
    ws.range((row, first_col), (row, last_col)).style.font = bold_9
    ws.range((row, first_col), (row, last_col)).style.alignment = align_center
    row += 1
    projects_id = set(sprint_tasks_qs.values_list('project', flat=True))
    has_no_project = False
    if None in projects_id:
        has_no_project = True
        projects_id.discard(None)
    projects = list(WorkgroupModel.objects.filter(is_active=True, pk__in=projects_id))
    if has_no_project:
        projects.append(None)
    if projects:
        for project in projects:
            project_tasks = sprint_tasks_qs.filter(project=project, is_active=True)
            if project_tasks:
                project_tasks_time_fact_qs = sprint_tasks_time_executed.filter(
                    task__project=project
                )
                project_time_fact_sum = project_tasks_time_fact_qs.aggregate(project_time_fact=Sum('hours'))['project_time_fact']
                if not project_time_fact_sum:
                    project_time_fact_sum = 0
                project_budget_fact = sprint_tasks_budget.filter(
                    task__project=project
                ).aggregate(project_budget_fact=Sum('amount'))['project_budget_fact']
                project_tasks_aggr = project_tasks.aggregate(
                    project_tasks_funds=Sum('funds'),
                    project_tasks_time_plane=Sum('execution_time_plan')
                )
                project_tasks_funds = project_tasks_aggr['project_tasks_funds'] if project_tasks_aggr[
                    'project_tasks_funds'] else 0
                project_tasks_time_plane = project_tasks_aggr['project_tasks_time_plane'] if project_tasks_aggr[
                    'project_tasks_time_plane'] else 0
                project_aggr_data = (
                    (
                        f'Проект "{project.name}"' if project else 'Задачи без проекта',
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        sprint.progress,
                        project_tasks_funds,
                        project_budget_fact,
                        project_tasks_time_plane,
                        project_time_fact_sum,
                    ),
                )
                ws.range((row, first_col), (row, last_col)).value = project_aggr_data
                ws.range((row, first_col), (row, last_col)).style.font = bold_9
                ws.range((row, first_col), (row, last_col)).style.alignment = align_left
                merged_cells.append(((row, first_col), (row, 8)))
                task_counter = 1
                row += 1
                project_data = list()
                project_time_plan = 0
                correction = 0
                time_by_work_type_len_acc = 0
                task_merge_row_start = row
                for project_task in project_tasks:
                    time_qs = project_tasks_time_fact_qs.filter(task=project_task)
                    execution_time_fact = time_qs.aggregate(hours_sum=Sum('hours'))['hours_sum']
                    if not execution_time_fact:
                        execution_time_fact = 0
                    execution_time_plan = project_task.execution_time_plan \
                        if project_task.execution_time_plan is not None else 0
                    budget_fact = project_task.task_budgets.filter(
                        is_active=True
                    ).aggregate(budget_fact=Sum('amount'))['budget_fact']
                    if not budget_fact:
                        budget_fact = 0
                    operators_qs = time_qs.values_list('user__user__first_name', 'user__user__last_name').distinct(
                        'author')
                    operators_str = ', '.join([f"{_[0]} {_[1]}" for _ in operators_qs])
                    # work_types_qs = time_qs.values_list('work_type__name', flat=True).distinct('work_type')
                    # work_types_str = ', '.join([f"{_}" for _ in work_types_qs])
                    date_begin = project_task.date_start_plan
                    if date_begin:
                        date_begin_str = date_begin.strftime('%d.%m.%Y')
                    else:
                        date_begin_str = '-'
                    date_end = project_task.finished_date
                    if date_end:
                        date_end_str = date_end.strftime('%d.%m.%Y')
                    else:
                        date_end_str = '-'
                    if date_begin and date_end:
                        duration = date_end - date_begin
                        duration_days = duration.days
                    else:
                        duration_days = '-'
                    time_author_counter = 0
                    if not execution_time_fact:
                        project_data.append(
                            (
                                project_task.status.name,
                                date_begin_str,
                                date_end_str,
                                duration_days,
                                project_task.counter,
                                project_task.name,
                                operators_str,
                                '-',
                                project_task.result if project_task.result else '-',
                                project_task.funds,
                                budget_fact,
                                execution_time_plan,
                                execution_time_fact,
                            )
                        )
                        task_merge_row_start += 1
                    else:
                        time_authors = time_qs.values(
                            'author',
                        ).annotate(
                            author_first_name=F('user__user__first_name'),
                            author_last_name=F('user__user__last_name'),
                            author_middle_name=F('user__user__middle_name'),
                        ).values('author', 'author_first_name', 'author_last_name', 'author_middle_name',).distinct('author')

                        is_first = True

                        for time_author in time_authors:
                            time_by_work_type = time_qs.filter(author_id=time_author['author']).values(
                                'work_type'
                            ).annotate(
                                work_type_name=F('work_type__name'),
                                hours_sum=Sum('hours'),
                            ).values(
                                'work_type_name',
                                'hours_sum',
                            )
                            is_author_first = True
                            author_full_name = f"{time_author['author_last_name']} {time_author['author_first_name']} {time_author['author_middle_name']}"
                            for each in time_by_work_type:
                                if is_first:
                                    project_data.append(
                                        (
                                            project_task.status.name,
                                            date_begin_str,
                                            date_end_str,
                                            duration_days,
                                            project_task.counter,
                                            project_task.name,
                                            author_full_name,
                                            each['work_type_name'],
                                            project_task.result if project_task.result else '-',
                                            project_task.funds,
                                            budget_fact,
                                            execution_time_plan,
                                            each['hours_sum'],
                                        )
                                    )
                                    is_first = False
                                else:
                                    if is_author_first:
                                        project_data.append(
                                            (
                                                None,
                                                None,
                                                None,
                                                None,
                                                None,
                                                None,
                                                author_full_name,
                                                each['work_type_name'],
                                                None,
                                                None,
                                                None,
                                                None,
                                                each['hours_sum'],
                                            )
                                        )
                                        correction += 1
                                        time_author_counter += 1
                                        is_author_first = False
                                    else:
                                        project_data.append(
                                            (
                                                None,
                                                None,
                                                None,
                                                None,
                                                None,
                                                None,
                                                None,
                                                each['work_type_name'],
                                                None,
                                                None,
                                                None,
                                                None,
                                                each['hours_sum'],
                                            )
                                        )
                                        correction += 1

                            time_by_work_type_len = len(time_by_work_type)

                            if time_by_work_type_len > 1:
                                merge_row_start = row + task_counter + time_author_counter + time_by_work_type_len_acc - 1
                                merge_row_end = row + task_counter + time_author_counter + time_by_work_type_len_acc - 2 + time_by_work_type_len
                                # merged_cells.append(((merge_row_start, 6), (merge_row_end, 6)))
                                time_by_work_type_len_acc += time_by_work_type_len - 1

                    task_merge_row_end = row + task_counter + time_author_counter + time_by_work_type_len_acc - 1
                    # if task_merge_row_start < task_merge_row_end:
                    #     merged_cells.append(((task_merge_row_start, 1), (task_merge_row_end, 1)))
                    #     merged_cells.append(((task_merge_row_start, 2), (task_merge_row_end, 2)))
                    #     merged_cells.append(((task_merge_row_start, 3), (task_merge_row_end, 3)))
                    #     merged_cells.append(((task_merge_row_start, 4), (task_merge_row_end, 4)))
                    #     merged_cells.append(((task_merge_row_start, 5), (task_merge_row_end, 5)))
                    #     merged_cells.append(((task_merge_row_start, 8), (task_merge_row_end, 8)))
                    #     merged_cells.append(((task_merge_row_start, 9), (task_merge_row_end, 9)))
                    #     merged_cells.append(((task_merge_row_start, 10), (task_merge_row_end, 10)))
                    #     merged_cells.append(((task_merge_row_start, 11), (task_merge_row_end, 11)))
                    project_time_plan += execution_time_plan
                    task_counter += 1
                    task_merge_row_start = task_merge_row_end
                ws.range((row, first_col), (row + correction + task_counter - 1, last_col)).value = project_data
                ws.range((row, first_col), (row + correction + task_counter - 1, last_col)).style.font = normal_9
                ws.range((row, first_col), (row + correction + task_counter - 1, last_col)).style.alignment = align_left
                row = row + task_counter + correction - 1

    # Границы ячеек:
    row -= 1
    ws.range((3, first_col), (row, last_col)).style.borders.top.style = '_'
    ws.range((3, first_col), (row, last_col)).style.borders.top.color = pyexcelerate.Color(0, 0, 0)
    ws.range((3, first_col), (row, last_col)).style.borders.left.style = '_'
    ws.range((3, first_col), (row, last_col)).style.borders.left.color = pyexcelerate.Color(0, 0, 0)
    ws.range((3, first_col), (row, last_col)).style.borders.bottom.style = '_'
    ws.range((3, first_col), (row, last_col)).style.borders.bottom.color = pyexcelerate.Color(0, 0, 0)
    ws.range((3, first_col), (row, last_col)).style.borders.right.style = '_'
    ws.range((3, first_col), (row, last_col)).style.borders.right.color = pyexcelerate.Color(0, 0, 0)
    # Объединяем ячейки в самом конце, иначе глючат границы:
    for merged_cell in merged_cells:
        ws.range(*merged_cell).merge()
    return wb


def get_all_sprint_tasks(sprint) -> set:
    included_tasks = set(sprint.tasks.filter(is_active=True).values_list('pk', flat=True))
    excluded_tasks = set(sprint.task_sprint_history.all().values_list('task', flat=True))
    all_tasks = included_tasks | excluded_tasks
    return all_tasks


def get_sprint_time_tracking(sprint):
    sprint_tasks_id = get_all_sprint_tasks(sprint)

    sprint_tasks_time_executed = models.TaskExecutionTimeModel.objects.filter(is_active=True,
                                                                              task_id__in=sprint_tasks_id,
                                                                              sprint=sprint)
    return sprint_tasks_time_executed


def prepare_list_task_queryset(qs, request):
    """Подготавливает queryset к передачи в сериализатор ListTaskSerializer.
    Делает необходимые select_related, prefetch_related и аннотации для сокращения количества запросов в базу."""

    now = timezone.now()
    if isinstance(request.query_params.get('task_type'), str):
        task_type = request.query_params.get('task_type').split(',')
    else:
        task_type = ['task']
    qs = qs.select_related(
        'parent',
        'workgroup',
        'project',
        'sprint',
        'contractor__contact_person',
        'potential_contractor',
        'rejection_reason',
        'lead_source',

        # следующие поля теперь проходят через кеширование
        # 'workgroup__workgroup_logo',
        # 'workgroup__workgroup_logo__mime_type',
        # 'workgroup__workgroup_logo__mime_type__file_type',
        # 'project__workgroup_logo',
        # 'project__workgroup_logo__mime_type',
        # 'project__workgroup_logo__mime_type__file_type',
        # 'organization',
        # 'contractor',

        # 'author__user',
        # 'author__avatar',
        # 'owner__user',
        # 'owner__avatar',
        # 'operator__user',
        # 'operator__avatar',
        # 'workgroup__author',
        # 'workgroup__author__avatar',
        # 'workgroup__author__user',
        # 'project__author',
        # 'project__author__avatar',
        # 'project__author__user',
        # 'contractor__curator',
        # 'contractor__curator__user',
        # 'contractor__curator__avatar',

    ).prefetch_related(
        'prerequisites',
        'attachments',
        'children',
        'visors',
        'cooperators',
        'task_delivery_points',
        'cooperator_tasks',
        Prefetch(
            'object_tags',
            queryset=TagModel.objects.prefetch_related('tag_object_through')
        ),
        Prefetch(
            'task_points',
            queryset=models.TaskPointModel.objects.filter(is_active=True)
        ),
        Prefetch(
            'project__workgroupmembersmodel_set',
            queryset=WorkgroupMembersModel.objects.select_related(
                'membership_request_status', 'membership_role'
            ),
            to_attr='prefetched_members'
        ),
        # Prefetch('execution_time',
        #             queryset=models.TaskExecutionTimeModel.objects.filter(is_active=True).select_related('work_type').order_by('-date'),
        #             to_attr='prefetched_exec_times',
        #             ),
        # Prefetch(
        #     'event_calendars__events',
        #     queryset=EventCalendarModel.objects.filter(is_active=True, is_finished=False, start_at__gte=now,).order_by('start_at'),
        #     to_attr='prefetched_future_events'
        # )
    )

    # Аннотация participants_count - убрана для оптимизации производительности 14.10.2025
    # participants_count теперь возвращает -1 в сериализаторе
    # qs = qs.annotate(
    #     participants_count=Count(
    #         'owner',
    #         filter=Q(owner__is_active=True),
    #         distinct=True
    #     ) + Count(
    #         'operator',
    #         filter=Q(operator__is_active=True) & ~Q(operator_id=F('owner_id')),
    #         distinct=True
    #     ) + Count(
    #         'visors',
    #         filter=Q(visors__is_active=True) & ~Q(visors__id=F('owner_id')) & ~Q(visors__id=F('operator_id')),
    #         distinct=True
    #     )
    # )

    if task_type == ['logistic']:
        qs = qs.prefetch_related(
            Prefetch(
                'task_delivery_points',
                queryset=models.TaskDeliveryPointModel.objects.select_related('delivery_point').filter(
                    is_start=False,
                    delivery_date__isnull=True
                ).order_by('sort'),
                to_attr='next_delivery_point'
            ),
        )

    if not set(task_type).isdisjoint(('task', 'interest',)):
        qs = qs.annotate(
            completed_children_count=Count(
                'children',
                filter=Q(
                    children__is_active=True,
                    children__status__code__in=(
                        'completed',
                        'successfully'
                    )
                ),
                distinct=True
            ),
            comment_count=Count(
                'comments',
                filter=Q(
                    comments__is_active=True
                ),
                distinct=True
            ),
            attachments_count=Count(
                'files',
                filter=Q(
                    files__is_active=True
                ),
                distinct=True
            ),
            has_description=Case(
                When(
                    description__gt='',
                    then=Value(True)
                ),
                default=Value(False),
                output_field=BooleanField()
            )
        )
        # с этим тоже, что-то надо решать. Так как дублируем, по сути, код
    if task_type == ['interest']:
        # CRM: у интереса факт "заказ уже оформлен" определяется не через
        # точки доставки задачи, а через GoodsOrderModel.reason = id интереса.
        from crm.models import GoodsOrderModel
        order_exists_expression = Exists(
            GoodsOrderModel.objects.filter(is_active=True, reason_id=OuterRef('pk'))
        )
    else:
        order_exists_expression = Case(
            When(
                task_delivery_points=None,
                then=Value(0)
            ),
            default=Value(1),
            output_field=IntegerField()
        )

    qs = qs.annotate(annotate_is_finished=Case(
            When(status_id='completed',
                    then=Value(1)),
            default=Value(0),
        output_field=IntegerField()),

        annotate_order_exists=order_exists_expression,
        annotate_order_date_start_plan=Min("task_delivery_points__start_goods_orders__delivery_date_plan_gte"),
    )
    return qs


def prepare_my_day_task_queryset(qs, request, start, end, profile_ids=None):
    """Подготавливает queryset к передаче в сериализатор MyDayTaskSerializer.
    Делает необходимые select_related, prefetch_related и аннотации для сокращения количества запросов в базу."""

    profile = request.user.profile
    profile_id = profile.pk
    
    # Преобразуем start и end в даты для использования в аннотациях
    # Извлекаем только дату из ISO строки (игнорируя время и часовой пояс)
    # (пользователь выбирает только дату)
    start_date = dateparse.parse_date(start.split('T')[0])
    end_date = dateparse.parse_date(end.split('T')[0])

    # Используем Exists - это подзапрос, но он НЕ создает JOIN и не приводит к дубликатам
    # Любое обращение к ManyToMany полю через аннотацию (Count, JOIN) создаст JOIN и дубликаты
    # Exists - единственный способ проверить наличие связи без JOIN
    
    qs = qs.annotate(
        is_executor=Case(
            When(
                Q(operator_id=profile_id) | Exists(
                    models.TaskCooperator.objects.filter(
                        task_id=OuterRef('pk'),
                        user_id=profile_id
                    )
                ),
                then=Value(True),
            ),
            default=Value(False),
            output_field=BooleanField(),
        ),
        is_owner=Case(
            When(owner_id=profile_id, then=Value(True)),
            default=Value(False),
            output_field=BooleanField(),
        ),
        is_visor=Case(
            When(
                Exists(
                    models.TaskVisor.objects.filter(
                        task_id=OuterRef('pk'),
                        user_id=profile_id
                    )
                ),
                then=Value(True)
            ),
            default=Value(False),
            output_field=BooleanField(),
        ),
    )

    # Добавляем аннотацию для проверки наличия трудозатрат в периоде.
    # Используем Exists вместо Count: Count создаёт LEFT JOIN на execution_time,
    # что в сочетании с другими JOIN-ами приводит к умножению строк и необходимости
    # GROUP BY + DISTINCT. Exists — подзапрос без JOIN, нет умножения.
    if start_date and end_date and profile_ids:
        qs = qs.annotate(
            has_execution_time=Case(
                When(
                    Exists(
                        models.TaskExecutionTimeModel.objects.filter(
                            task_id=OuterRef('pk'),
                            user_id__in=profile_ids,
                            date__gte=start_date,
                            date__lte=end_date,
                            is_active=True,
                        )
                    ),
                    then=Value(1),
                ),
                default=Value(0),
                output_field=IntegerField(),
            )
        )
    else:
        qs = qs.annotate(
            has_execution_time=Value(0, output_field=IntegerField())
        )
    
    # Добавляем prefetch_related для ManyToMany полей
    # Формируем queryset для pinned_by с учетом временного промежутка
    # Для СВОЕГО рабочего дня за СЕГОДНЯ показываем только активные pinned задачи
    profile_id = str(request.user.profile.pk)
    
    # Для отображения флажка "Фокус дня" всегда учитываем только активные pinned на сегодняшний день
    pinned_by_queryset = models.TaskPinnedModel.objects.filter(
        user_id=profile_id,
        is_active=True
    )
    
    qs = qs.prefetch_related(
        "visors",
        "cooperators",
        Prefetch(
            'object_tags',
            queryset=TagModel.objects.prefetch_related('tag_object_through')
        ),
        Prefetch(
            'pinned_by',
            queryset=pinned_by_queryset
        ),
    )
    
    # Сортируем: сначала задачи с трудозатратами, потом по дедлайну, затем по статусу
    qs = qs.select_related("status").order_by("-has_execution_time", "dead_line", "status__sort")

    return qs


def prepare_list_kanban_task_queryset(qs, request):
    """Подготавливает queryset к передачи в сериализатор ListKanbanTaskSerializer.
    Делает необходимые select_related, prefetch_related и аннотации для сокращения количества запросов в базу."""
    qs = qs.select_related(
        'parent',
        'project',
        'project__workgroup_logo',
        'project__workgroup_logo__mime_type',
        'project__workgroup_logo__mime_type__file_type',
        'status',
    ).prefetch_related(
        'visors',
        'cooperators',
        'cooperator_tasks',
        Prefetch(
            'object_tags',
            queryset=TagModel.objects.prefetch_related('tag_object_through')
        ),
        Prefetch(
            'project__workgroupmembersmodel_set',
            queryset=WorkgroupMembersModel.objects.select_related(
                'membership_request_status', 'membership_role'
            ),
            to_attr='prefetched_members'
        ),
    )
    return qs


def get_sprint_task_count_data(sprint):
    tasks_qs = sprint.tasks.filter(is_active=True)
    sprint_tasks = sprint.tasks.filter(is_active=True).prefetch_related('status').values('status')
    count_of_tasks_by_status = Counter(task['status'] for task in sprint_tasks)
    data = {
        'new_task_count': count_of_tasks_by_status.__getitem__('new'),
        'completed_task_count': count_of_tasks_by_status.__getitem__('completed'),
        'in_work_task_count': tasks_qs.exclude(status_id__in=('new', 'completed',)).count(),
        'overdue_task_count': tasks_qs.filter(
            dead_line__isnull=False, dead_line__lt=timezone.now()).exclude(status_id='completed').count()
    }
    return data


def create_onboarding_data(profile_id):
    """
    Создает обучающие задачи для нового пользователя на основе данных из Excel-файла.
    """
    import os
    from openpyxl import load_workbook
    from datetime import timedelta
    
    # Путь к Excel-файлу с данными
    excel_file_path = os.path.join(
        os.path.dirname(__file__), 
        'data', 
        'Onboarding_data.xlsx'
    )
    
    if not os.path.exists(excel_file_path):
        raise FileNotFoundError(f"Файл {excel_file_path} не найден")
    
    try:
        # Получаем профиль пользователя
        from users.models import ProfileModel
        profile = ProfileModel.objects.get(pk=profile_id)
        
        # Проверяем, есть ли уже обучающие задачи
        existing_onboarding_tasks = profile.owner_tasks.filter(
            is_onboarding=True,
            is_active=True
        )
        
        if existing_onboarding_tasks.exists():
            return {
                'success': False,
                'message': 'У пользователя уже есть обучающие задачи',
                'existing_tasks_count': existing_onboarding_tasks.count()
            }
        
        # Ищем контрагент: сначала current_contractor профиля, если нет - первый активный ContractorProfileModel
        contractor = None
        if profile.current_contractor:
            contractor = profile.current_contractor
        else:
            from common.catalogs.models import ContractorProfileModel
            contractor_profile = ContractorProfileModel.objects.filter(
                user=profile,
                is_active=True
            ).first()
            
            if contractor_profile:
                contractor = contractor_profile.contractor
        
        if not contractor:
            return {
                'success': False,
                'message': 'У пользователя нет активного контрагента'
            }
        
        # Загружаем Excel-файл
        workbook = load_workbook(excel_file_path, data_only=True)
        worksheet = workbook.active
        
        now = timezone.now()
        created_tasks = []
        parent_task = None
        
        # Обрабатываем строки начиная со второй (индекс 1, так как первая - заголовки)
        for row_idx in range(1, worksheet.max_row + 1):
            row = row_idx + 1  # Excel строки начинаются с 1, но мы пропускаем заголовок
            
            # Получаем данные из строки
            name = worksheet.cell(row=row, column=1).value
            description = worksheet.cell(row=row, column=2).value or ""
            result = worksheet.cell(row=row, column=3).value or ""
            duration_days = worksheet.cell(row=row, column=4).value

            # Проверяем, что все необходимые данные есть
            if not name or duration_days is None:
                continue
            
            try:
                duration_days = int(duration_days)
            except (ValueError, TypeError):
                duration_days = 1  # По умолчанию 1 день
            
            # Создаем задачу
            task = models.TaskModel()
            task.name = str(name)
            task.description = str(description)
            task.result = str(result)
            task.author = profile
            task.owner = profile
            task.operator = profile
            task.organization = contractor
            task.task_type_id = 'task'
            task.status_id = 'new'
            task.priority = 2
            task.is_onboarding = True
            
            # Устанавливаем даты
            if row_idx == 1:  # Первая задача (родительская)
                task.date_start_plan = now
                task.dead_line = now + timedelta(days=duration_days)
                parent_task = task
            elif row_idx == 2:  # Первая подзадача
                task.date_start_plan = now
                task.dead_line = now + timedelta(days=duration_days)
                task.parent = parent_task
            else:  # Остальные подзадачи
                # date_start_plan = dead_line предыдущей задачи
                previous_task = created_tasks[-1]
                task.date_start_plan = previous_task.dead_line
                task.dead_line = task.date_start_plan + timedelta(days=duration_days)
                task.parent = parent_task
            
            task.save()
            created_tasks.append(task)

        workbook.close()
        
        return {
            'success': True,
            'message': f'Создано {len(created_tasks)} обучающих задач'
        }
        
    except Exception as e:
        # Закрываем workbook в случае ошибки
        try:
            workbook.close()
        except:
            pass
        
        return {
            'success': False,
            'message': f'Ошибка при создании обучающих задач: {str(e)}',
            'error': str(e)
        }


def delete_onboarding_data(profile_id):
    """
    Удаляет (деактивирует) обучающие задачи пользователя
    """
    try:
        # Получаем профиль пользователя
        from users.models import ProfileModel
        profile = ProfileModel.objects.get(pk=profile_id)
        
        # Находим все активные обучающие задачи пользователя
        onboarding_tasks = profile.owner_tasks.filter(
            is_onboarding=True,
            is_active=True
        )
        
        if not onboarding_tasks.exists():
            return {
                'success': False,
                'message': 'У пользователя нет активных обучающих задач'
            }
        
        # Деактивируем задачи
        deleted_count = onboarding_tasks.count()
        onboarding_tasks.update(is_active=False)
        
        return {
            'success': True,
            'message': f'Удалено {deleted_count} обучающих задач'
        }
        
    except ProfileModel.DoesNotExist:
        return {
            'success': False,
            'message': 'Пользователь не найден'
        }
    except Exception as e:
        return {
            'success': False,
            'message': f'Ошибка при удалении обучающих задач: {str(e)}',
            'error': str(e)
        }


def get_sprint_members_set(sprint: models.TaskSprintModel):
    """Возвращает set с id профилей участников спринта sprint: кол-во в задачах спринта и кол-во в исключенных из спринта"""
    included_tasks = models.TaskModel.objects.filter(sprint=sprint, is_active=True)
    included_tasks_id = set(included_tasks.filter(
        sprint=sprint, is_active=True
    ).values_list('pk', flat=True))
    excluded_tasks = sprint.task_sprint_history.filter(
        for_completed=True,
    )
    excluded_tasks_id = set(excluded_tasks.values_list('task', flat=True))
    included_tasks_operators_id = set(included_tasks.filter(
        operator__isnull=False
    ).values_list('operator', flat=True))
    included_tasks_owners_id = set(included_tasks.filter(owner__isnull=False).values_list('owner', flat=True))
    included_tasks_cooperators_id = set(models.TaskCooperator.objects.filter(
        task__in=included_tasks_id
    ).values_list('user', flat=True))
    included_tasks_visors_id = set(models.TaskVisor.objects.filter(
        task__in=included_tasks_id
    ).values_list('user', flat=True))
    included_tasks_members_id = (
            included_tasks_operators_id |
            included_tasks_owners_id |
            included_tasks_cooperators_id |
            included_tasks_visors_id
    )
    excluded_tasks_operators_id = set(
        excluded_tasks.filter(task__operator__isnull=False).values_list('task__operator', flat=True))
    excluded_tasks_owners_id = set(
        excluded_tasks.filter(task__owner__isnull=False).values_list('task__owner', flat=True)
    )
    excluded_tasks_cooperators_id = set(
        models.TaskCooperator.objects.filter(task__in=excluded_tasks_id).values_list('user', flat=True))
    excluded_tasks_visors_id = set(
        models.TaskVisor.objects.filter(task__in=excluded_tasks_id).values_list('user', flat=True))
    excluded_tasks_members_id = (
            excluded_tasks_operators_id |
            excluded_tasks_owners_id |
            excluded_tasks_cooperators_id |
            excluded_tasks_visors_id
    )
    return included_tasks_members_id, excluded_tasks_members_id


def get_tasks_after_start_sprint_id(sprint):
    begin_date = sprint.begin_date
    if begin_date:
        moved_tasks_id = set(sprint.task_sprint_history.filter(
                                for_completed=True,
                                add_sprint_date__isnull=False,
                                add_sprint_date__gte=begin_date
                            ).values_list('task', flat=True))
        sprint_tasks_id = set(sprint.tasks.filter(
            is_active=True,
            add_sprint_date__isnull=False,
            add_sprint_date__gte=begin_date
        ).values_list('pk', flat=True))
        tasks_id = moved_tasks_id | sprint_tasks_id
    else:
        tasks_id = set()
    return tasks_id


def get_expected_results_qs(sprint):
    expected_results = models.SprintExpectedResultModel.objects.filter(
        is_active=True, sprint=sprint
    ).exclude(
        Q(task__result='') |
        Q(task_id__in=sprint.task_sprint_history.filter(for_completed=False).values_list('task', flat=True))
    )

    return expected_results


def get_incomplete_duration(now, created_at):
    return (now - created_at).seconds


def get_work_log_duration(user, task):
    work_logs = models.TaskExecutionTimeModel.objects.filter(
        is_active=True,
        user=user,
        task=task,
    )
    complete_duration = work_logs.filter(
        is_current=False
    ).aggregate(complete_duration=Sum('duration'))['complete_duration']
    if complete_duration is None:
        complete_duration = 0
    incomplete_log = work_logs.filter(is_current=True).first()
    if incomplete_log:
        now = timezone.now()
        incomplete_duration = get_incomplete_duration(now, incomplete_log.created_at)
        is_current = True
    else:
        incomplete_duration = 0
        is_current = False
    duration = complete_duration + incomplete_duration
    return duration, is_current, incomplete_duration


def start_work_log_timer(user, task):
    """Запускает таймер учета времени для пользователя и задачи."""
    if user not in (*list(task.visors.all()), *list(task.cooperators.all()), task.owner, task.operator):
        raise drf_exceptions.PermissionDenied('Вы не являетесь оператором задачи')
    from help_desk.models import HelpDeskWorkLogModel
    from help_desk import utils as help_desk_utils
    now = timezone.now()
    if not models.TaskExecutionTimeModel.objects.filter(user=user, task=task, is_current=True).exists():
        with transaction.atomic():
            _, _, completed_statuses = get_cached_statuses()
            work_logs = (
                models.TaskExecutionTimeModel.objects
                .select_related('task')
                .filter(user=user, is_current=True)
            )
            for each in work_logs:
                # paused_task = each.task
                each.duration = get_incomplete_duration(now, each.created_at)
                each.is_current = False
                each.save()
                # if paused_task and paused_task.status_id not in completed_statuses and paused_task.status_id != 'on_pause':
                #     paused_task.status_id = 'on_pause'
                #     paused_task.save(update_fields=('status_id', 'updated_at'))

            help_desk_work_logs = HelpDeskWorkLogModel.objects.select_related('ticket').filter(
                user=user,
                is_current=True
            )

            for each in help_desk_work_logs:
                # paused_ticket = each.ticket
                each.duration = help_desk_utils.get_incomplete_duration(now, each.created_at)
                each.is_current = False
                each.finished_date = now
                each.save()
                # if paused_ticket and paused_ticket.status_id not in completed_statuses and paused_ticket.status_id != 'on_pause':
                #     paused_ticket.status_id = 'on_pause'
                #     paused_ticket.save(update_fields=('status_id', 'updated_at'))
            current_log = models.TaskExecutionTimeModel()
            current_log.user = user
            current_log.task = task
            current_log.is_current = True
            current_log.duration = 0
            current_log.save()
        ProfileModel.objects.filter(pk=user.pk).update(current_work=current_log.task)
        transaction.on_commit(lambda: send_socketio_about_update_current_work([str(user.pk), ]))

    duration, is_current, incomplete_duration = get_work_log_duration(user, task)
    return duration, is_current


def stop_work_log_timer(user, task, provided_duration=None, description=''):
    """Останавливает таймер учета времени для пользователя и тикета."""
    now = timezone.now()
    incomplete_log = models.TaskExecutionTimeModel.objects.filter(user=user, task=task, is_current=True).first()
    if not isinstance(description, str):
        description = ''
    else:
        description = description[:1024]
    if incomplete_log:
        if provided_duration is not None:
            # жёстко доверяем фронту, но слегка валидируем
            try:
                provided_duration = int(provided_duration)
            except (TypeError, ValueError):
                provided_duration = 0
            if provided_duration < 0:
                provided_duration = 0
            duration_to_save = provided_duration
        else:
            # старое поведение
            duration_to_save = get_incomplete_duration(now, incomplete_log.created_at)

        incomplete_log.duration = duration_to_save
        incomplete_log.is_current = False
        incomplete_log.description = description
        incomplete_log.save()

        # как и раньше: возвращаем duration/is_current, без ломающей смены сигнатуры
    duration, is_current, incomplete_duration = get_work_log_duration(user, task)
    transaction.on_commit(lambda: send_socketio_about_update_current_work([str(user.pk), ]))
    return duration, is_current
