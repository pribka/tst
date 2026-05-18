from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from django.urls import reverse_lazy
from django.db.models import Q
from openpyxl import load_workbook

from rest_framework import exceptions

from bkz3.settings import TIME_ZONE
from bpms.workgroups import models as wg_models
from bpms.workload import utils
from common.catalogs import models as c_models
from common.accounting_catalogs.models import KATOCodesModel
from invest_projects_info import models as invest_models


@login_required(redirect_field_name=None, login_url=reverse_lazy('admin:index'))
def project_generator_view(request):
    """Генерация проектов из .xlsx файла. Столбцы таблицы:
        1 - № п/п
        2 - Название проекта (str)
        3 - Описание проекта (str)
        4 - Начало (datetime)
        5 - Завершение (datetime)
        6 - id организации
    """
    if request.method == 'POST' and request.FILES.get('file'):
        filename = request.FILES['file']
        sheet = load_workbook(filename).active
        fst_row, lst_row = (sheet.min_row + 1), (sheet.max_row + 1)
        for row in range(fst_row, lst_row):
            cells = [cell.value for cell in sheet[row][1:]]
            try:
                organization = c_models.ContractorModel.objects.get(pk=cells[4])
            except c_models.ContractorModel.DoesNotExist:
                raise exceptions.NotFound('Организация не найдена')
            
            data = dict()
            data['organization'] = organization
            data['name'] = cells[0]
            data['description'] = cells[1]
            data['date_start_plan'] = utils.get_correct_timezone_date(cells[2], TIME_ZONE)
            data['dead_line'] = utils.get_correct_timezone_date(cells[3], TIME_ZONE)
            data['is_project'] = True
            create_workgroup(organization, data, request)

            # здесь может быть логика добавления задач
        # return render(request, 'management/project_generator.html')
    return render(request, 'management/project_generator.html')


def create_workgroup(organization, data, request):
    workgroup, _ = wg_models.WorkgroupModel.objects.get_or_create(**data)

    status = wg_models.WorkgroupMembershipStatus.objects.get(code='APPROVED')
    founder_role = wg_models.WorkgroupMembershipRole.objects.get(code='FOUNDER')
    moderator_role = wg_models.WorkgroupMembershipRole.objects.get(code='MODERATOR')
    # Основатель проекта - директор организации, тот кто генерирует - модератор
    author, _ = wg_models.WorkgroupMembersModel.objects.get_or_create(
        work_group=workgroup,
        member=request.user.profile,
        membership_request_status=status,
        membership_role=moderator_role,
        member_visible = False
    )
    director, _ = wg_models.WorkgroupMembersModel.objects.get_or_create(
        work_group=workgroup,
        member=organization.director,
        membership_request_status=status,
        membership_role=founder_role,
        member_visible = True
    )
    return workgroup


@login_required(redirect_field_name=None, login_url=reverse_lazy('admin:index'))
def invest_project_generator_view(request):
    """Генерация инвест-проектов из .xlsx файла. Данные начинаются с 3 строки.
    Столбцы таблицы:
        0 - id организации
        1 - генерировать связанный проект. 1 - да, 0 - нет.
        2 - Название 1 уровня локации из справочника КАТО
        3 - Название 2 уровня локации из справочника КАТО (не обязательно)
        4 - Столбец не используется
        5 - Категория
        6 - Подкатегория
        7 - Столбец не используется
        8 - Наименование инвест-проекта
        9 - Комментарий к инвест-проекту
        10 - Столбец не используется
        11 - Начало (дата)
        12 - Завершение (дата)
    ФИНАНСИРОВАНИЕ:
        13 - Комментарий к финансированию
        14 - РБ – республиканский бюджет
        15 - МБ – местный бюджет
        16 - НФ – Национальный Фонд
        17 - Планируется
        18 - За счет инвестора
        19 - ДИ – дополнительные инвестиции
    """
    if request.method == 'POST' and request.FILES.get('file'):
        filename = request.FILES['file']
        sheet = load_workbook(filename).active
        fst_row, lst_row = (sheet.min_row + 2), (sheet.max_row + 1)
        for row in range(fst_row, lst_row):
            cells = [cell.value for cell in sheet[row][0:]]
            try:
                organization = c_models.ContractorModel.objects.get(pk=cells[0])
            except c_models.ContractorModel.DoesNotExist:
                raise exceptions.NotFound('Организация не найдена')
            # при необходимости сначала генерируем проект
            if cells[1] == str(1):
                workgroup_data = dict()
                workgroup_data['organization'] = organization
                workgroup_data['name'] = cells[8]
                workgroup_data['description'] = f'Инвестпроект. {cells[9]}'
                workgroup_data['date_start_plan'] = utils.get_correct_timezone_date(cells[11], TIME_ZONE)
                workgroup_data['dead_line'] = utils.get_correct_timezone_date(cells[12], TIME_ZONE)
                workgroup_data['is_project'] = True
                workgroup = create_workgroup(organization, workgroup_data, request)
            
            # получаем местоположение по справочнику KATO
            location = get_KATO_location(cells[2], cells[3])
            # получаем категорию и подкатегорию
            try:
                category = invest_models.InvestProjectCategoryModel.objects.get(name=cells[5])
            except invest_models.InvestProjectCategoryModel.DoesNotExist:
                raise exceptions.NotFound('Категория не найдена')
            try:
                subcategory = invest_models.InvestProjectSubcategoryModel.objects.get(name=cells[6])
            except invest_models.InvestProjectSubcategoryModel.DoesNotExist:
                raise exceptions.NotFound('Подкатегория не найдена')

            # создаем сам инвест-проект
            invest_project_data = dict()
            invest_project_data['organization'] = organization
            invest_project_data['location'] = location
            invest_project_data['project_name'] = cells[8]
            invest_project_data['category'] = category
            invest_project_data['subcategory'] = subcategory
            invest_project_data['comment'] = cells[9]
            invest_project_data['date_start'] = utils.get_correct_timezone_date(cells[11], TIME_ZONE)
            invest_project_data['dead_line'] = utils.get_correct_timezone_date(cells[12], TIME_ZONE)
            amount_list = [cells[14], cells[15], cells[16], cells[17], cells[18], cells[19]]
            invest_project_data['funds'] = sum(filter(None, amount_list))
            if workgroup:
                invest_project_data['project'] = workgroup

            invest_project, _ = invest_models.InvestProjectInfoModel.objects.get_or_create(**invest_project_data)

            # создаем связанную модель финансирования
            funding_source_names = [
                'РБ – республиканский бюджет',
                'МБ – местный бюджет',
                'НФ – Национальный Фонд',
                'Планируется',
                'За счет инвестора',
                'ДИ – дополнительные инвестиции'
                ]
            for i in range(len(funding_source_names)):
                funding_source = invest_models.InvestProjectFundingSourceModel.objects.get(name=funding_source_names[i])
                amount = cells[i+14]
                if not amount:
                    continue
                invest_models.FundingSourceAndAmountModel.objects.create(
                    invest_project_info = invest_project,
                    funding_source = funding_source,
                    amount = amount
                )
            print(f'Сгенерирован инвест проект № {row} - {cells[8]}')

    return render(request, 'management/invest_project_generator.html')


def get_KATO_location(first_level, second_level=None):
    """Возвращает локацию из справочника KATO"""
    def get_first_level(first_level):
        try:
            first_level_location = KATOCodesModel.objects.get(
                cd='00', ef='00', hij='000',
                name=first_level)
        except KATOCodesModel.DoesNotExist:
            raise exceptions.NotFound(f'Локация 1 уровня не найдена: {first_level}')
        return first_level_location
    
    def get_second_level(first_level, second_level):
        first_level_location = get_first_level(first_level)
        try:
            second_level_location = KATOCodesModel.objects.get(
                ~Q(cd='00'),
                ab=first_level_location.ab,
                ef='00', hij='000',
                name=second_level)
        except KATOCodesModel.DoesNotExist:
            raise exceptions.NotFound(f'Локация 2 уровня не найдена: {second_level}')
        return second_level_location

    if second_level is None: # если указан только 1 уровень
        location = get_first_level(first_level)
    else: # если указан 2 уровень
        location = get_second_level(first_level, second_level)
    return location
        

