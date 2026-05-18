import datetime

import openpyxl
from django.apps import apps
from django.db import transaction
from django.db.models import Q
from django.utils import timezone
from django.utils.module_loading import import_string
from django_q.tasks import async_task
from rest_framework import exceptions as drf_exceptions

from common.current_profile.middleware import get_current_authenticated_profile
from common.models import File
from common.utils import use_access_groups
from contractor_permissions.utils import check_contractor_permission

from contractor_permissions.models import ContractorPermissionModel

from .. import models, notifications, utils


class ReportFormBaseClass():
    """
    Базовый класс форм отчетности консолидаций.
    """

    _subclasses = dict()

    def __init__(self) -> None:
        self._name = ''
        self._description = ''
        self._add_button_icon = 'cloud-upload'
        self._report_widget = 'FileUploadWidget'
        self._extra_fields_model = None

    def __init_subclass__(cls, *args, **kwargs):
        cls._subclasses[cls._code] = cls

    def get_code(self):
        return self._code

    def get_name(self):
        return self._name

    def __str__(self) -> str:
        return f'Форма отчетности - {self._name}'

    def get_report_form(self):
        return models.ReportFormModel.objects.filter(
            is_active=True,
            code=self._code
        ).first()

    def get_extra_fields(self, consolidation, data: dict) -> None:
        if not self._extra_fields_model:
            return
        serializer = self._extra_fields_model.get_serializer_class(action='retrieve')
        extra_fields_obj = getattr(consolidation, f'{self._code}_extra', None)
        extra_data = serializer(extra_fields_obj).data
        data[f'{self._code}_extra'] = extra_data

    def set_extra_fields(self, *args, **kwargs): ...

    def report_update_is_disabled(self, *args, **kwargs) -> bool:
        is_disabled = False
        message = ''
        return is_disabled, message

    def upload_report(self, consolidation, report, request, *args, **kwargs): ...

    def add_specific_fields_in_report(self, data, report, *args, **kwargs):
        data['add_button_icon'] = self._add_button_icon
        data['widget'] = self._report_widget

    def report_rejected(self, report, *args, **kwargs): ...

    def report_approved(self, report, *args, **kwargs): ...

    def all_uploaded_files_validate(self, request, *args, **kwargs): ...

    def before_approve(self, report): ...

    def before_create_scheduled_consolidation(self, *args, **kwargs): ...


class F2GO(ReportFormBaseClass):
    """
    Класс для работы с формой отчетности 'Обращения граждан'.
    """

    _code = 'f2go'

    def __init__(self) -> None:
        super().__init__()
        self._name = 'Обращения граждан'
        self._report_widget = 'F2GOWidget'

    def add_specific_fields_in_report(self, data, report, *args, **kwargs):
        super().add_specific_fields_in_report(data, report, *args, **kwargs)
        disintegration = getattr(report, 'disintegration', None)
        data['revoked_without_routing'] = disintegration.revoked_without_routing if disintegration is not None else None
        data['transferring_to_another_system'] = disintegration.transferring_to_another_system if disintegration is not None else None

    def report_update_is_disabled(self, *args, **kwargs):
        consolidation = kwargs.get('consolidation')
        user = get_current_authenticated_profile()
        current_date = datetime.datetime.now().date()
        if current_date <= consolidation.end:
            is_disabled = True
            message = ('Отчетный период еще не завершен. Загрузка отчетов '
                       'запрещена до окончания отчетного периода.')
            return is_disabled, message
        report = kwargs.get('report')
        get_descendants = import_string(
            'users.utils.get_descendants_departments_related_organizations'
        )
        descendants = get_descendants((report.contractor,), include_self=False)
        if use_access_groups(user.pk):
            try:
                check_contractor_permission(
                    user.pk, report.contractor_id, 'create_consolidation', None
                )
            except drf_exceptions.PermissionDenied:
                consolidation_create_permission = False
            else:
                consolidation_create_permission = True
        else:
            consolidation_create_permission = ContractorPermissionModel.objects.filter(
                Q(aux_conditions=self.get_report_form()) | Q(
                    aux_conditions__isnull=True),
                is_active=True,
                contractor_permission_role__is_active=True,
                permission_type_id='create_consolidation',
                contractor_permission_role__contractor_profiles__user=user,
                contractor_permission_role__contractor_id=report.contractor
            ).exists()
        if report.contractor != consolidation.org_administrator and descendants and consolidation_create_permission:
            is_disabled = True
            message = ('Загрузка сводного отчёта доступна только через '
                       'кнопку "Отправить" в консолидации Вашей организации.')
            return is_disabled, message
        return super().report_update_is_disabled(*args, **kwargs)

    def before_create_scheduled_consolidation(self, *args, **kwargs):
        super().before_create_scheduled_consolidation(*args, **kwargs)
        consolidation = kwargs.get('consolidation')
        if consolidation.dead_line <= consolidation.end:
            raise drf_exceptions.ValidationError(
                'Срок подачи отчетов не может быть раньше, '
                'чем дата окончания отчетного периода.'
            )

    def upload_report(self, consolidation, report, request, *args, **kwargs):
        # Проверка загруженных данных
        self.all_uploaded_files_validate(request)

        # Сохранение загруженных данных
        self.upload(consolidation, report, request, *args, **kwargs)

    def upload(self, consolidation, report, request, *args, **kwargs):
        uploaded_report_files = request.data.get('report_files', None)
        no_inquiries = request.data.get('no_inquiries', None)
        revoked_without_routing = request.data.get('revoked_without_routing', None)
        transferring_to_another_system = request.data.get('transferring_to_another_system', None)

        report_files = report.report_files.filter(
            is_active=True
        )
        for each in uploaded_report_files:
            original_file = each.get('original_file', None)
            code = each.get('code', None)
            if not code:
                raise drf_exceptions.ValidationError('Не удалось получить код.')
            try:
                rf = report_files.get(id=each.get('id', None))
            except models.ReportFileModel.DoesNotExist:
                raise drf_exceptions.ValidationError('Не удалось найти файл отчета.')
            old_original_file = rf.original_file
            file = None
            if original_file and original_file.get('id'):
                try:
                    file = File.objects.get(
                        is_active=True,
                        id=original_file.get('id')
                    )
                except File.DoesNotExist:
                    raise drf_exceptions.ValidationError('Не удалось найти файл.')

                if no_inquiries is not None:
                    utils.report_file_validation(file.id, consolidation, code, no_inquiries=no_inquiries) # TODO Перенести в класс
                else:
                    utils.report_file_validation(file.id, consolidation, code) # TODO Перенести в класс

            if file:
                utils.save_report_data(file.id, report, code) # TODO Перенести в класс

            is_generated = each.get('is_generated', False)
            if file != old_original_file:
                rf.original_file = file
                rf.upload_date = timezone.now() if file else None
                rf.uploaded_by = request.user.profile if file else None
                rf.is_generated = is_generated
                rf.save(update_fields=(
                    'original_file',
                    'pdf_file',
                    'upload_date',
                    'uploaded_by',
                    'is_generated',
                ))
                utils.handle_generate_sources(report, rf) # TODO Перенести в класс
        if no_inquiries is not None:
            report.no_inquiries = no_inquiries
            report.save(update_fields=(
                'no_inquiries',
            ))

        if revoked_without_routing is not None and transferring_to_another_system is not None:
            disintegration, created = models.DisintegrationModel.objects.get_or_create(
                report=report,
                defaults={
                    'revoked_without_routing': abs(int(revoked_without_routing)),
                    'transferring_to_another_system': abs(int(transferring_to_another_system))
                }
            )
            if not created:
                disintegration.revoked_without_routing = abs(int(revoked_without_routing))
                disintegration.transferring_to_another_system = abs(int(transferring_to_another_system))
                disintegration.save(update_fields=(
                    'revoked_without_routing',
                    'transferring_to_another_system'
                ))

        report.status_id = 'approved' if consolidation.auto_approve else 'on_review'
        report.save(update_fields=('status_id',))
        consolidation.status_id = 'in_progress'
        consolidation.save(update_fields=('status_id', ))

        if not consolidation.source_reports.filter(
            is_active=True,
            status__code__in=['new', 'not_loaded']
        ).exists():
            async_task(
                notifications.notify_all_reports_are_uploaded,
                str(consolidation.id),
                str(request.user.profile.id)
            )

    def all_uploaded_files_validate(self, request):
        uploaded_report_files = request.data.get('report_files', None)
        revoked_without_routing = request.data.get('revoked_without_routing', None)
        transferring_to_another_system = request.data.get('transferring_to_another_system', None)
        if revoked_without_routing is None:
            raise drf_exceptions.ValidationError(
                'Значение поля "Отозвано без маршрутизации" не может быть пустым.'
            )
        if transferring_to_another_system is None:
            raise drf_exceptions.ValidationError(
                'Значение поля "Перенос в другую систему" не может быть пустым.'
            )
        f2go, f2go_id = None, None
        risk_matrix, risk_matrix_id = None, None
        for rf in uploaded_report_files:
            code = rf.get('code', None)
            if code == 'f2go' and rf['original_file']:
                f2go_id = rf['original_file'].get('id', None)
            if code == 'risk_matrix' and rf['original_file']:
                risk_matrix_id = rf['original_file'].get('id', None)
        if not f2go_id or not risk_matrix_id:
            raise drf_exceptions.ValidationError(
                'Загружены не все файлы, проверка невозможна.'
            )
        try:
            f2go = File.objects.get(id=f2go_id)
        except File.DoesNotExist:
            raise drf_exceptions.ValidationError(
                'Не удалось получить файл отчета Ф2ГО.'
            )
        try:
            risk_matrix = File.objects.get(id=risk_matrix_id)
        except File.DoesNotExist:
            raise drf_exceptions.ValidationError(
                'Не удалось получить файл с картой рисков.'
            )
        try:
            f2go_workbook = openpyxl.load_workbook(filename=f2go.upload.url[1:])
        except:
            raise drf_exceptions.ValidationError(
                'Не удалось прочитать данные из файла Ф2ГО.'
            )
        try:
            risk_matrix_workbook = openpyxl.load_workbook(filename=risk_matrix.upload.url[1:])
        except:
            raise drf_exceptions.ValidationError(
                'Не удалось прочитать данные из файла с картой риска.'
            )

        f2go_sheet = f2go_workbook.worksheets[0]
        risk_matrix_sheet = risk_matrix_workbook.worksheets[0]
        entry_counter = 0

        for row in risk_matrix_sheet.iter_rows(min_row=4):
            values_counter = 0
            for cell in row:
                if cell.value is not None:
                    values_counter += 1
            if values_counter == 16:
                entry_counter += 1

        if int(f2go_sheet['B2'].value) != entry_counter:
            difference = 'больше' if int(f2go_sheet['B2'].value) < entry_counter else 'меньше'
            raise drf_exceptions.ValidationError(
                f'Количество записей в карте рисков {difference} '
                'количества зарегистрированных обращений.'
            )
        return True

    def before_approve(self, report):
        is_all_files_upload = not report.report_files.filter(
            is_active=True,
            file_type_id__in=['f2go', 'risk_matrix'],
            original_file__isnull=True
        ).exists()
        is_disintegration_set = False
        disintegration = getattr(report, 'disintegration', None)
        if disintegration is not None:
            is_disintegration_set = (disintegration.revoked_without_routing is not None) and (disintegration.transferring_to_another_system is not None)
        if not is_all_files_upload or not is_disintegration_set:
            raise drf_exceptions.ValidationError(
                'Данные в отчет внесены не полностью'
            )


class F2GOWithVerificationAct(ReportFormBaseClass):
    """
    Класс для работы с формой отчетности 'Ф2ГО (отчет с нарастающим итогом)'.
    """

    _code = 'f2go_with_verification_act'

    def __init__(self) -> None:
        super().__init__()
        self._name = 'Ф2ГО (отчет с нарастающим итогом)'
        self._report_widget = 'F2GOWidget'

    def add_specific_fields_in_report(self, data, report, *args, **kwargs):
        super().add_specific_fields_in_report(data, report, *args, **kwargs)
        disintegration = getattr(report, 'disintegration', None)
        data['revoked_without_routing'] = disintegration.revoked_without_routing if disintegration is not None else None
        data['transferring_to_another_system'] = disintegration.transferring_to_another_system if disintegration is not None else None

    def report_update_is_disabled(self, *args, **kwargs):
        consolidation = kwargs.get('consolidation')
        user = get_current_authenticated_profile()
        current_date = datetime.datetime.now().date()
        if current_date <= consolidation.end:
            is_disabled = True
            message = ('Отчетный период еще не завершен. Загрузка отчетов '
                       'запрещена до окончания отчетного периода.')
            return is_disabled, message
        report = kwargs.get('report')
        get_descendants = import_string(
            'users.utils.get_descendants_departments_related_organizations'
        )
        descendants = get_descendants((report.contractor,), include_self=False)
        if use_access_groups(user.pk):
            try:
                check_contractor_permission(user.pk, report.contractor_id, 'create_consolidation', None)
            except drf_exceptions.PermissionDenied:
                consolidation_create_permission = False
            else:
                consolidation_create_permission = True
        else:
            consolidation_create_permission = ContractorPermissionModel.objects.filter(
                Q(aux_conditions=self.get_report_form()) | Q(
                    aux_conditions__isnull=True),
                is_active=True,
                contractor_permission_role__is_active=True,
                permission_type_id='create_consolidation',
                contractor_permission_role__contractor_profiles__user=user,
                contractor_permission_role__contractor_id=report.contractor
            ).exists()
        if report.contractor != consolidation.org_administrator and descendants and consolidation_create_permission:
            is_disabled = True
            message = ('Загрузка сводного отчёта доступна только через '
                       'кнопку "Отправить" в консолидации Вашей организации.')
            return is_disabled, message
        return super().report_update_is_disabled(*args, **kwargs)

    def before_create_scheduled_consolidation(self, *args, **kwargs):
        super().before_create_scheduled_consolidation(*args, **kwargs)
        consolidation = kwargs.get('consolidation')
        if consolidation.dead_line <= consolidation.end:
            raise drf_exceptions.ValidationError(
                'Срок подачи отчетов не может быть раньше, '
                'чем дата окончания отчетного периода.'
            )

    def disintegration_check_validity(self, consolidation, request, revoked_without_routing, transferring_to_another_system):
        previous_report = utils.get_previous_report(consolidation, request)
        if previous_report is None:
            return

        errors = dict()
        disintegration = getattr(previous_report, 'disintegration', None)
        if disintegration is not None:
            rwr_comparison_value = getattr(disintegration, 'revoked_without_routing', 0)
            ttas_comparison_value = getattr(disintegration, 'transferring_to_another_system', 0)
            if revoked_without_routing < rwr_comparison_value:
                errors['rwr'] = ('Значение "Отозвано без маршрутизации" '
                                 'меньше предыдущего. Предыдущее значение: '
                                 f'{rwr_comparison_value}')
            if transferring_to_another_system < ttas_comparison_value:
                errors['ttas'] = ('Значение "Перенос в другую систему" меньше '
                                  'предыдущего. Предыдущее значение: '
                                  f'{ttas_comparison_value}')
        if errors:
            raise drf_exceptions.ValidationError(errors)
        return

    def upload_report(self, consolidation, report, request, *args, **kwargs):
        self.all_uploaded_files_validate(request)

        uploaded_report_files = request.data.get('report_files', None)
        no_inquiries = request.data.get('no_inquiries', None)
        revoked_without_routing = request.data.get('revoked_without_routing', None)
        transferring_to_another_system = request.data.get('transferring_to_another_system', None)

        # self.disintegration_check_validity(consolidation, request,
        #                                    revoked_without_routing,
        #                                    transferring_to_another_system)

        report_files = report.report_files.filter(
            is_active=True
        )
        for each in uploaded_report_files:
            original_file = each.get('original_file', None)
            code = each.get('code', None)
            if not code:
                raise drf_exceptions.ValidationError('Не удалось получить код.')
            try:
                rf = report_files.get(id=each.get('id', None))
            except models.ReportFileModel.DoesNotExist:
                raise drf_exceptions.ValidationError('Не удалось найти файл отчета.')
            old_original_file = rf.original_file
            file = None
            if original_file and original_file.get('id'):
                try:
                    file = File.objects.get(
                        is_active=True,
                        id=original_file.get('id')
                    )
                except File.DoesNotExist:
                    raise drf_exceptions.ValidationError('Не удалось найти файл.')

                if no_inquiries is not None:
                    utils.report_file_validation(file.id, consolidation, code, no_inquiries=no_inquiries) # TODO Перенести в класс
                else:
                    utils.report_file_validation(file.id, consolidation, code) # TODO Перенести в класс

            if file:
                utils.save_report_data(file.id, report, code) # TODO Перенести в класс

            is_generated = each.get('is_generated', False)
            if file != old_original_file:
                rf.original_file = file
                rf.upload_date = timezone.now() if file else None
                rf.uploaded_by = request.user.profile if file else None
                rf.is_generated = is_generated
                rf.save(update_fields=(
                    'original_file',
                    'pdf_file',
                    'upload_date',
                    'uploaded_by',
                    'is_generated',
                ))
                utils.handle_generate_sources(report, rf) # TODO Перенести в класс
        if no_inquiries is not None:
            report.no_inquiries = no_inquiries
            report.save(update_fields=(
                'no_inquiries',
            ))

        if revoked_without_routing is not None and transferring_to_another_system is not None:
            disintegration, created = models.DisintegrationModel.objects.get_or_create(
                report=report,
                defaults={
                    'revoked_without_routing': abs(int(revoked_without_routing)),
                    'transferring_to_another_system': abs(int(transferring_to_another_system))
                }
            )
            if not created:
                disintegration.revoked_without_routing = abs(int(revoked_without_routing))
                disintegration.transferring_to_another_system = abs(int(transferring_to_another_system))
                disintegration.save(update_fields=(
                    'revoked_without_routing',
                    'transferring_to_another_system'
                ))

        report.status_id = 'approved' if consolidation.auto_approve else 'on_review'
        report.save(update_fields=('status_id',))
        consolidation.status_id = 'in_progress'
        consolidation.save(update_fields=('status_id', ))

        if not consolidation.source_reports.filter(
            is_active=True,
            status__code__in=['new', 'not_loaded']
        ).exists():
            async_task(
                notifications.notify_all_reports_are_uploaded,
                str(consolidation.id),
                str(request.user.profile.id)
            )

    def all_uploaded_files_validate(self, request):
        uploaded_report_files = request.data.get('report_files', None)
        revoked_without_routing = request.data.get('revoked_without_routing', None)
        transferring_to_another_system = request.data.get('transferring_to_another_system', None)
        if revoked_without_routing is None:
            raise drf_exceptions.ValidationError(
                'Значение поля "Отозвано без маршрутизации" не может быть пустым.'
            )
        if transferring_to_another_system is None:
            raise drf_exceptions.ValidationError(
                'Значение поля "Перенос в другую систему" не может быть пустым.'
            )
        f2go_report_file = None
        for rf in uploaded_report_files:
            if rf['code'] == 'f2go':
                f2go_report_file = rf
                break
        if f2go_report_file is None or f2go_report_file['original_file'] is None:
            raise drf_exceptions.ValidationError('Файл с отчетом Ф2ГО не загружен')

    def before_approve(self, report):
        is_all_files_upload = not report.report_files.filter(
            is_active=True,
            file_type_id__in=['f2go',],
            original_file__isnull=True
        ).exists()
        is_disintegration_set = False
        disintegration = getattr(report, 'disintegration', None)
        if disintegration is not None:
            is_disintegration_set = (disintegration.revoked_without_routing is not None) and (disintegration.transferring_to_another_system is not None)
        if not is_all_files_upload or not is_disintegration_set:
            raise drf_exceptions.ValidationError(
                'Данные в отчет внесены не полностью'
            )


class IPFProposal(ReportFormBaseClass):
    """
    Класс для работы с формой отчетности 'Заявка на ИПФ'.
    """

    _code = 'ipf_proposal'

    def __init__(self) -> None:
        super().__init__()
        self._name = 'Заявка на ИПФ'
        self._description = 'Заявка на изменение плана финансирования'
        self._add_button_icon = 'plus-circle'
        self._report_widget = 'IpsProposalWidget'
        self._extra_fields_model = apps.get_model(
            'accounting_reports',
            'IpfProposalConsolidationExtraModel'
        )

    def set_extra_fields(self, clear=False, *args, **kwargs):
        """
        Метод создает/изменяет объект в модели с дополнительными
        полями консолидации формы отчетности 'Заявка на ИПФ'.

        Аргументы:
        clear: bool - если равен True, запись с дополнительными
                      полями консолидации будет удалена
        extra_fields: dict - словарь значений полей date, number, subtype
        consolidation - объект консолидации для которого указываются
                        значения дополнительных полей
        """
        if clear:
            consolidation = kwargs.get('consolidation')
            try:
                obj = self._extra_fields_model.objects.get(
                    consolidation=consolidation
                )
            except self._extra_fields_model.DoesNotExist:
                pass
            else:
                obj.delete()
        else:
            consolidation = kwargs.get('consolidation')
            extra_fields = kwargs.get('extra_fields')
            if not consolidation or not isinstance(extra_fields, dict):
                raise drf_exceptions.ValidationError(
                    'Не удалось получить консолидацию или значения дополнительных полей.'
                )
            date = extra_fields.get('date')
            number = extra_fields.get('number')
            subtype = extra_fields.get('subtype')
            self._extra_fields_model.objects.update_or_create(
                consolidation=consolidation,
                defaults={
                    'date': date,
                    'number': number,
                    'subtype_id': subtype.code
                }
            )

    def upload_report(self, consolidation, report, request, *args, **kwargs):
        from consolidation import notifications
        ipf_proposals = request.data.get('ipf_proposals', None)
        without_attachments = request.data.get('without_attachments', False)
        if not isinstance(ipf_proposals, list):
            raise drf_exceptions.ValidationError(
                'Неподдерживаемый тип данных.'
            )
        if not without_attachments and not ipf_proposals:
            raise drf_exceptions.ValidationError(
                'Выберите заявки на ИПФ или укажите что заявки отсутствуют'
            )

        with transaction.atomic():
            # Имеющимся в отчете заявкам присвоим статус "Новый"
            for proposal in report.ipf_proposal_reports.filter(is_active=True):
                proposal.status_id = 'new'
                proposal.save(update_fields=('status_id',))
            # Удаляем все связи отчёты
            report.ipf_proposal_reports.clear()
            # Создаем новые связи
            report.ipf_proposal_reports.set(ipf_proposals)
            # Новым заявкам присвоим статус "Обработан"
            for proposal in report.ipf_proposal_reports.filter(is_active=True):
                proposal.status_id = 'processed'
                proposal.save(update_fields=('status_id',))

            report.without_attachments = without_attachments

            # Присвоение статуса
            if consolidation.auto_approve and (report.ipf_proposal_reports.count() > 0 or without_attachments):
                report.status_id = 'approved'
            else:
                report.status_id = 'on_review'
            report.save(update_fields=(
                    'status_id',
                    'without_attachments'
                ))
            consolidation.status_id = 'in_progress'
            consolidation.save(update_fields=('status_id', ))

        # Проверка все ли отчеты загружены
        if not consolidation.source_reports.filter(
            is_active=True,
            status__code__in=['new', 'not_loaded']
        ).exists():
            async_task(
                notifications.notify_all_reports_are_uploaded,
                str(consolidation.id),
                str(request.user.profile.id)
            )

    def report_rejected(self, report, *args, **kwargs):
        for proposal in report.ipf_proposal_reports.filter(is_active=True):
            proposal.status_id = 'rejected'
            proposal.save(update_fields=('status_id',))

    def report_approved(self, report, *args, **kwargs):
        for proposal in report.ipf_proposal_reports.filter(is_active=True):
            proposal.status_id = 'processed'
            proposal.save(update_fields=('status_id',))

    def add_specific_fields_in_report(self, data, report, *args, **kwargs):
        from accounting_reports.serializers import \
            FPCReportModelWidgetSerializer

        super().add_specific_fields_in_report(data, report, *args, **kwargs)
        ipf_proposals = report.ipf_proposal_reports.filter(is_active=True)
        data['ipf_proposals'] = FPCReportModelWidgetSerializer(
            ipf_proposals,
            many=True
        ).data


class ChangeCalculation(ReportFormBaseClass):
    """
    Класс для работы с формой отчетности 'Расчет на внесение изменений в ИПФ'.
    """

    _code = 'change_calculation'

    def __init__(self) -> None:
        super().__init__()
        self._name = 'Расчет на внесение изменений в ИПФ'
        self._description = 'Сводный расчет на внесение изменений в индивидуальный план финансирования по платежам'
        self._add_button_icon = 'plus-circle'
        self._report_widget = 'ChangeCalculationWidget'

    def upload_report(self, consolidation, report, request, *args, **kwargs):
        from consolidation import notifications
        calculations = request.data.get('calculations', None)
        without_attachments = request.data.get('without_attachments', False)
        if not isinstance(calculations, list):
            raise drf_exceptions.ValidationError(
                'Неподдерживаемый тип данных.'
            )
        if not without_attachments and not calculations:
            raise drf_exceptions.ValidationError(
                'Выберите расчеты на внесение изменений в ИПФ '
                'или укажите что изменения в ИПФ отсутствуют'
            )
        with transaction.atomic():
            # Имеющимся в отчете заявкам присвоим статус "Новый"
            for proposal in report.change_calculation_reports.filter(is_active=True):
                proposal.status_id = 'new'
                proposal.save(update_fields=('status_id',))
            # Удаляем все связи отчёты
            report.change_calculation_reports.clear()
            # Создаем новые связи
            report.change_calculation_reports.set(calculations)
            # Новым заявкам присвоим статус "Обработан"
            for proposal in report.change_calculation_reports.filter(is_active=True):
                proposal.status_id = 'processed'
                proposal.save(update_fields=('status_id',))

            report.without_attachments = without_attachments

            # Присвоение статуса
            if consolidation.auto_approve and (report.change_calculation_reports.count() > 0 or without_attachments):
                report.status_id = 'approved'
            else:
                report.status_id = 'on_review'
            report.save(update_fields=(
                    'status_id',
                    'without_attachments'
                ))
            consolidation.status_id = 'in_progress'
            consolidation.save(update_fields=('status_id', ))

        # Проверка все ли отчеты загружены
        if not consolidation.source_reports.filter(
            is_active=True,
            change_calculation_reports__isnull=True
        ).exists():
            async_task(
                notifications.notify_all_reports_are_uploaded,
                str(consolidation.id),
                str(request.user.profile.id)
            )

    def report_rejected(self, report, *args, **kwargs):
        for proposal in report.change_calculation_reports.filter(is_active=True):
            proposal.status_id = 'rejected'
            proposal.save(update_fields=('status_id',))

    def report_approved(self, report, *args, **kwargs):
        for proposal in report.change_calculation_reports.filter(is_active=True):
            proposal.status_id = 'processed'
            proposal.save(update_fields=('status_id',))

    def add_specific_fields_in_report(self, data, report, *args, **kwargs):
        from accounting_reports.serializers import \
            ChangeCalculationReportModelWidgetSerializer

        super().add_specific_fields_in_report(data, report, *args, **kwargs)
        calculations = report.change_calculation_reports.filter(is_active=True)
        data['calculations'] = ChangeCalculationReportModelWidgetSerializer(
            calculations,
            many=True
        ).data


class RiskMapWithPersonalReception(F2GO):
    """
    Класс для работы с формой отчетности 'Карта риска + Личный прием'.
    """

    _code = 'risk_map_with_personal_reception'

    def __init__(self) -> None:
        super().__init__()
        self._name = 'Карта риска + Личный прием'
        self._report_widget = 'F2GOWidget'

    def add_specific_fields_in_report(self, data, report, *args, **kwargs):
        super().add_specific_fields_in_report(data, report, *args, **kwargs)
        personal_reception = getattr(report, 'personal_reception', None)
        data['personal_reception_is_edit'] = personal_reception is not None
        data['personal_reception_quantity'] = personal_reception.quantity if personal_reception is not None else None
        data['no_personal_reception'] = personal_reception.no_personal_reception if personal_reception is not None else False
        data['is_personal_reception_not_loaded'] = personal_reception is None
        data['personal_reception_issues'] = personal_reception.issues if personal_reception is not None else list()

    def all_uploaded_files_validate(self, request):
        uploaded_report_files = request.data.get('report_files', None)
        revoked_without_routing = request.data.get('revoked_without_routing', None)
        transferring_to_another_system = request.data.get('transferring_to_another_system', None)
        personal_reception_issues = request.data.get('personal_reception_issues', None)
        no_personal_reception = request.data.get('no_personal_reception', None)

        if revoked_without_routing is None:
            raise drf_exceptions.ValidationError(
                'Значение поля "Отозвано без маршрутизации" не может быть пустым.'
            )
        if transferring_to_another_system is None:
            raise drf_exceptions.ValidationError(
                'Значение поля "Перенос в другую систему" не может быть пустым.'
            )
        if personal_reception_issues is None:
            raise drf_exceptions.ValidationError(
                'Список обращений при личном приеме не передан.'
            )
        f2go, f2go_id = None, None
        risk_matrix, risk_matrix_id = None, None
        for rf in uploaded_report_files:
            code = rf.get('code', None)
            if code == 'f2go' and rf['original_file']:
                f2go_id = rf['original_file'].get('id', None)
            if code == 'risk_matrix' and rf['original_file']:
                risk_matrix_id = rf['original_file'].get('id', None)
        if not f2go_id or not risk_matrix_id:
            raise drf_exceptions.ValidationError(
                'Загружены не все файлы, проверка невозможна.'
            )
        try:
            f2go = File.objects.get(id=f2go_id)
        except File.DoesNotExist:
            raise drf_exceptions.ValidationError(
                'Не удалось получить файл отчета Ф2ГО.'
            )
        try:
            risk_matrix = File.objects.get(id=risk_matrix_id)
        except File.DoesNotExist:
            raise drf_exceptions.ValidationError(
                'Не удалось получить файл с картой рисков.'
            )
        try:
            f2go_workbook = openpyxl.load_workbook(filename=f2go.upload.url[1:])
        except:
            raise drf_exceptions.ValidationError(
                'Не удалось прочитать данные из файла Ф2ГО.'
            )
        try:
            risk_matrix_workbook = openpyxl.load_workbook(filename=risk_matrix.upload.url[1:])
        except:
            raise drf_exceptions.ValidationError(
                'Не удалось прочитать данные из файла с картой риска.'
            )

        f2go_sheet = f2go_workbook.worksheets[0]
        risk_matrix_sheet = risk_matrix_workbook.worksheets[0]
        entry_counter = 0

        for row in risk_matrix_sheet.iter_rows(min_row=4):
            values_counter = 0
            for cell in row:
                if cell.value is not None:
                    values_counter += 1
            if values_counter == 16:
                entry_counter += 1

        if int(f2go_sheet['B2'].value) != entry_counter:
            difference = 'больше' if int(f2go_sheet['B2'].value) < entry_counter else 'меньше'
            raise drf_exceptions.ValidationError(
                f'Количество записей в карте рисков {difference} '
                'количества зарегистрированных обращений.'
            )

        pri_count = len(personal_reception_issues)
        b9_value = int(f2go_sheet['B9'].value)
        if b9_value != pri_count:
            difference = 'меньше' if b9_value < pri_count else 'больше'
            raise drf_exceptions.ValidationError(
                'Количество обращений на личном приеме в отчете Ф2ГО '
                f'({int(f2go_sheet["B9"].value)}, ячейка B9) {difference} '
                f'количества переданных обращений ({pri_count}).'
            )
        if no_personal_reception is False and pri_count == 0:
            raise drf_exceptions.ValidationError(
                'Список обращений с категорией вопроса "Личный прием '
                'заявителей" пуст. Если приём не проводился, включите '
                'переключатель "Личный приём не проводился в отчётном '
                'периоде".'
            )
        return True

    def upload(self, consolidation, report, request, *args, **kwargs):
        from risk_assessment.models import PersonalReceptionModel

        no_personal_reception = request.data.get('no_personal_reception', None)

        if no_personal_reception:
            quantity = 0
            issues = list()
        else:
            quantity = request.data.get('personal_reception_quantity', None)
            issues = request.data.get('personal_reception_issues', [])

        if not isinstance(issues, list):
            raise drf_exceptions.ValidationError(
                'Обращения должны быть переданы списком.'
            )
        # Удалим лишние ключи перед сохранением обращений
        for i in issues:
            del i['personal_reception']['status']

        with transaction.atomic():
            personal_reception, created = models.ReportPersonalReceptionModel.objects.get_or_create(
                report=report,
                defaults={
                    'quantity': quantity,
                    'no_personal_reception': no_personal_reception,
                    'issues': [{
                        str(report.contractor.id): {
                            'org_name': report.contractor.name,
                            'issues': issues,
                            'personal_reception_quantity': quantity
                        }
                    }]
                }
            )
            if not created:
                personal_reception.issues = [{
                    str(report.contractor.id): {
                            'org_name': report.contractor.name,
                            'issues': issues,
                            'personal_reception_quantity': quantity
                        }
                }]
                personal_reception.no_personal_reception = no_personal_reception
                personal_reception.quantity = quantity
                personal_reception.save(update_fields=(
                    'issues',
                    'no_personal_reception',
                    'quantity'
                ))

            for issue in issues:
                PersonalReceptionModel.objects.filter(
                    is_active=True,
                    issue_id=issue['id']
                ).update(
                    days_in_queue=issue['personal_reception']['days_in_queue']
                )

            transaction.on_commit(
                lambda: super(RiskMapWithPersonalReception, self).upload(consolidation, report, request, *args, **kwargs)
            )
