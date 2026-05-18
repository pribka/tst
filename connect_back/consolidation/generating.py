import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, NamedStyle
from io import BytesIO

import pyexcelerate

from django.core.files import File as DjangoFile
from django.db import transaction
from django.utils import timezone

from rest_framework import exceptions as drf_exceptions

from common.models import File
from common.openpyxl_utils import get_height_for_row
from common.current_profile.middleware import get_current_authenticated_profile

from users.utils import get_descendants_departments_related_organizations
from common.catalogs.models import ContractorModel

from risk_assessment.models import RiskAssessmentModel

from .utils import (convert_xls_to_pdf,
                    save_report_data,
                    get_risk_matrix_template_workbook_pyexcelerate)
from . import models


# def generate_reports(consolidation: models.ConsolidationModel):
#     issue_date_gte = consolidation.start
#     issue_date_lte = consolidation.end
#     organization = consolidation.org_administrator
#     title_font = Font(
#         name='Times New Roman',
#         size=14,
#         bold=True,
#         italic=False,
#         vertAlign=None,
#         underline='none',
#         strike=False,
#         color='00000000'
#     )
#     center_alignment = Alignment(
#         horizontal='center',
#         vertical='center',
#         wrap_text=True,
#         shrink_to_fit=True,
#         indent=0
#     )
#     reports = consolidation.source_reports.filter(is_active=True)
#     xls_template = consolidation.report_form.attachments.first()
#
#     for each in reports:
#         risk_assessments = RiskAssessmentModel.objects.filter(
#             is_active=True,
#             organization_id=each.contractor_id,
#             issue__issue_date__gte=issue_date_gte,
#             issue__issue_date__lte=issue_date_lte,
#         ).order_by('issue__issue_date', 'created_at', )
#         if risk_assessments:
#             # Генерируем файл отчета из оценок обращений:
#             workbook = openpyxl.load_workbook(filename=xls_template.file.upload.url[1:])
#             sheet = workbook.active
#             sheet.cell(1, 1, f'Форма первичной оценки или Карта риска {each.contractor.name}')
#             sheet.cell(
#                 2,
#                 1,
#                 f'за период с {consolidation.start.strftime("%d.%m.%Y")} '
#                 f'по {consolidation.end.strftime("%d.%m.%Y")}'
#             )
#             sheet.merge_cells(start_column=1, end_column=16, start_row=2, end_row=2)
#             sheet.cell(2, 1).font = title_font
#             sheet.cell(2, 1).alignment = center_alignment
#             row_number = 4
#             i = 1
#             issue_date = risk_assessments[0].issue.issue_date
#             sheet.cell(row_number, 1, issue_date.strftime("%d.%m.%Y"))
#             sheet.merge_cells(start_row=row_number, end_row=row_number, start_column=1, end_column=16)
#             row_number += 1
#             for risk_assessment in risk_assessments:
#                 issue = risk_assessment.issue
#                 if issue.issue_date > issue_date:
#                     issue_date = issue.issue_date
#                     sheet.cell(row_number, 1, issue_date.strftime("%d.%m.%Y"))
#                     sheet.merge_cells(start_row=row_number, end_row=row_number, start_column=1, end_column=16)
#                     row_number += 1
#                 criteria = list(risk_assessment.risk_assessment_criteria.all().order_by('criteria__sort'))
#                 sheet.cell(row_number, 1, i)
#                 sheet.cell(row_number, 2, issue.number)
#                 sheet.cell(row_number, 3, issue.issue_type.name)
#                 sheet.cell(row_number, 4, criteria[0].value)
#                 sheet.cell(row_number, 5, criteria[1].value)
#                 sheet.cell(row_number, 6, criteria[2].value)
#                 sheet.cell(row_number, 7, criteria[3].value)
#                 sheet.cell(row_number, 8, criteria[4].value)
#                 sheet.cell(row_number, 9, criteria[5].value)
#                 sheet.cell(row_number, 10, criteria[6].value)
#                 sheet.cell(row_number, 11, criteria[7].value)
#                 sheet.cell(row_number, 12, criteria[8].value)
#                 sheet.cell(row_number, 13, criteria[9].value)
#                 sheet.cell(row_number, 14, risk_assessment.total_value)
#                 sheet.cell(row_number, 15, risk_assessment.sent_for)
#                 sheet.cell(row_number, 16, issue.summary)
#                 row_number += 1
#                 i += 1
#                 risk_assessment.status_id = 'processed'
#                 risk_assessment.save(update_fields=('status_id',))
#             sheet.delete_rows(row_number, 300)
#             workbook.remove(workbook['Лист2'])
#             workbook.template = False
#             stream = BytesIO()
#             workbook.save(stream)
#             django_file = DjangoFile(
#                 file=stream,
#                 name=f'Карта рисков за период с {consolidation.start.strftime("%d.%m.%Y")} по '
#                      f'{consolidation.end.strftime("%d.%m.%Y")}.xlsx',
#             )
#             report_file = File()
#             report_file.upload = django_file
#             report_file.is_confined = True
#             report_file.save()
#             report_file_instance = each.report_files.get(file_type_id='risk_matrix')
#             report_file_instance.original_file = report_file
#             report_file_instance.upload_date = timezone.now()
#             report_file_instance.is_generated = True
#             report_file_instance.save()
#
#             save_report_data(report_file.pk, each, 'risk_matrix')


def generate_report(report: models.ReportModel, file_code: str, no_inquiries: bool):
    consolidation = report.parent
    issue_date_gte = consolidation.start
    issue_date_lte = consolidation.end
    organization = report.contractor
    if report.parent.report_form.code in set(['f2go', 'risk_map_with_personal_reception']) and file_code == 'risk_matrix':
        if no_inquiries:
            risk_assessments = RiskAssessmentModel.objects.none()
        else:
            risk_assessments = RiskAssessmentModel.objects.filter(
                is_active=True,
                organization=organization,
                issue__issue_date__gte=issue_date_gte,
                issue__issue_date__lte=issue_date_lte,
            ).order_by('issue__issue_date', 'created_at', )
            if not risk_assessments:
                return None, 0
        if risk_assessments or no_inquiries:
            font_bold_14 = pyexcelerate.Font(bold=True, family='Times', size=14,)
            font_bold_11 = pyexcelerate.Font(bold=True, family='Times', size=11,)
            font_normal_11 = pyexcelerate.Font(family='Times', size=11,)
            center_alignment = pyexcelerate.Alignment(horizontal='center', vertical='center', wrap_text=True, )
            title_style = pyexcelerate.Style(font=font_bold_14, alignment=center_alignment)
            date_style = pyexcelerate.Style(font=font_bold_11, alignment=center_alignment)
            workbook = get_risk_matrix_template_workbook_pyexcelerate()
            ws = workbook._worksheets[0]
            ws.set_cell_value(1, 1, f'Форма первичной оценки или Карта риска {organization.name}')
            ws.set_cell_style(1, 1, title_style)
            ws.range("A1", "P1").merge()
            ws.set_cell_value(
                2,
                1,
                f'за период с {consolidation.start.strftime("%d.%m.%Y")} '
                f'по {consolidation.end.strftime("%d.%m.%Y")}'
            )
            ws.set_cell_style(2, 1, title_style)
            ws.range("A2", "P2").merge()
            row_number = 4
            if risk_assessments:
                issue_date = risk_assessments[0].issue.issue_date
                risk_assessments_data = [[issue_date.strftime("%d.%m.%Y")]]
                row_number += 1
                i = 1
                date_rows = [4]
                for risk_assessment in risk_assessments:
                    issue = risk_assessment.issue
                    if issue.issue_date > issue_date:
                        issue_date = issue.issue_date
                        risk_assessments_data.append([issue_date.strftime("%d.%m.%Y")])
                        date_rows.append(row_number)
                        row_number += 1
                    criteria = list(risk_assessment.risk_assessment_criteria.all().order_by('criteria__sort').values_list('value', flat=True))
                    if issue.summary != '':
                        issue_category = issue.summary
                    elif issue.issue_category:
                        issue_category = issue.issue_category.name
                    else:
                        issue_category = 'Не заполнено'
                    risk_assessments_data.append(
                        [
                            i,
                            issue.number,
                            issue.issue_type.name,
                            criteria[0],
                            criteria[1],
                            criteria[2],
                            criteria[3],
                            criteria[4],
                            criteria[5],
                            criteria[6],
                            criteria[7],
                            criteria[8],
                            criteria[9],
                            risk_assessment.total_value,
                            1 if risk_assessment.sent_for == 1 else 0,
                            issue_category,
                        ]
                    )
                    row_number += 1
                    i += 1
                row_number -= 1
                ws.range((4, 1), (row_number, 16)).value = risk_assessments_data
                ws.range((4, 1), (row_number, 16)).style.font = font_normal_11
                ws.range((4, 1), (row_number, 16)).style.alignment = center_alignment

                # Сначала прописываем borders, потом объединяем ячейки! Причина тут:
                # https://github.com/kz26/PyExcelerate/issues/74
                # Спасибо Анастасии за наводку.
                ws.range((4, 1), (row_number, 16)).style.borders.top.style = '_'
                ws.range((4, 1), (row_number, 16)).style.borders.top.color = pyexcelerate.Color(0, 0, 0)
                ws.range((4, 1), (row_number, 16)).style.borders.right.style = '_'
                ws.range((4, 1), (row_number, 16)).style.borders.right.color = pyexcelerate.Color(0, 0, 0)
                ws.range((4, 1), (row_number, 16)).style.borders.bottom.style = '_'
                ws.range((4, 1), (row_number, 16)).style.borders.bottom.color = pyexcelerate.Color(0, 0, 0)
                ws.range((4, 1), (row_number, 16)).style.borders.left.style = '_'
                ws.range((4, 1), (row_number, 16)).style.borders.left.color = pyexcelerate.Color(0, 0, 0)

                for each in date_rows:
                    ws.set_cell_style(each, 1, date_style)
                    ws.range((each, 1), (each, 16)).merge()
            stream = BytesIO()
            workbook.save(stream)
            openpyxl_workbook = openpyxl.load_workbook(stream)
            sheet = openpyxl_workbook.active
            sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
            sheet.sheet_properties.pageSetUpPr.fitToPage = True
            current_profile = get_current_authenticated_profile()
            if current_profile:
                openpyxl_workbook.properties.creator = current_profile.full_name

            stream = BytesIO()
            openpyxl_workbook.save(stream)
            openpyxl_workbook.close()
            django_file = DjangoFile(
                file=stream,
                name=f'Карта рисков за период с {consolidation.start.strftime("%d.%m.%Y")} по '
                     f'{consolidation.end.strftime("%d.%m.%Y")}.xlsx',
            )
            report_file = File()
            report_file.upload = django_file
            report_file.is_confined = True
            report_file.save()
            return report_file, risk_assessments.count()


# def generate_consolidation(self, request, *args, **kwargs):
#     organization_id = request.data.get('organization')
#     # TODO проверить права на эту организацию
#     try:
#         organization = ContractorModel.objects.get(pk=organization_id)
#     except ContractorModel.DoesNotExist:
#         raise drf_exceptions.ValidationError('Organization not found.')
#
#
#     issue_date_gte = request.data.get('issue_date_gte')
#     issue_date_lte = request.data.get('issue_date_lte')
#     descendants_id = get_descendants_departments_related_organizations(
#         contractors_seed=(organization_id,),
#         include_self=False
#     )
#
#     with transaction.atomic():
#
#         # Создание консолидации
#         consolidation = models.ConsolidationModel()
#         consolidation.org_administrator = organization
#         consolidation.report_form = models.ReportFormModel.objects.get(code='f2go')
#
#         # TODO определиться откуда брать даты
#         consolidation.start = datetime.datetime.strptime(issue_date_gte, '%Y-%m-%d')
#         consolidation.end = datetime.datetime.strptime(issue_date_lte, '%Y-%m-%d')
#         consolidation.dead_line = datetime.datetime.strptime(issue_date_lte, '%Y-%m-%d')
#
#         consolidation.consolidator = request.user.profile
#
#         consolidation.add_org_administrator_in_members = False
#         consolidation.save()
#         title_font = Font(
#             name='Times New Roman',
#             size=14,
#             bold=True,
#             italic=False,
#             vertAlign=None,
#             underline='none',
#             strike=False,
#             color='00000000'
#         )
#         center_alignment = Alignment(
#             horizontal='center',
#             vertical='center',
#             wrap_text=True,
#             shrink_to_fit=True,
#             indent=0
#         )
#         for each in descendants_id:
#             member = consolidation.consolidation_members.create(organization_id=each)
#             risk_assessments = RiskAssessmentModel.objects.filter(
#                 is_active=True,
#                 organization_id=each,
#                 issue__issue_date__gte=issue_date_gte,
#                 issue__issue_date__lte=issue_date_lte,
#             ).order_by('issue__issue_date', 'created_at', )
#             organization = ContractorModel.objects.get(pk=each)
#             if risk_assessments:
#                 # Генерируем файл отчета из оценок обращений:
#                 xls_template = consolidation.report_form.attachments.first()
#                 workbook = openpyxl.load_workbook(filename=xls_template.file.upload.url[1:])
#                 sheet = workbook.active
#                 sheet.cell(1, 1, f'Форма первичной оценки или Карта риска {organization.name}')
#                 sheet.cell(
#                     2,
#                     1,
#                     f'за период с {consolidation.start.strftime("%d.%m.%Y")} '
#                     f'по {consolidation.end.strftime("%d.%m.%Y")}'
#                 )
#                 sheet.merge_cells(start_column=1, end_column=16, start_row=2, end_row=2)
#                 sheet.cell(2, 1).font = title_font
#                 sheet.cell(2, 1).alignment = center_alignment
#                 row_number = 4
#                 i = 1
#                 for risk_assessment in risk_assessments:
#                     issue = risk_assessment.issue
#                     criteria = list(risk_assessment.risk_assessment_criteria.all().order_by('criteria__sort'))
#                     sheet.cell(row_number, 1, i)
#                     sheet.cell(row_number, 2, issue.number)
#                     sheet.cell(row_number, 3, issue.issue_type.name)
#                     sheet.cell(row_number, 4, criteria[0].value)
#                     sheet.cell(row_number, 5, criteria[1].value)
#                     sheet.cell(row_number, 6, criteria[2].value)
#                     sheet.cell(row_number, 7, criteria[3].value)
#                     sheet.cell(row_number, 8, criteria[4].value)
#                     sheet.cell(row_number, 9, criteria[5].value)
#                     sheet.cell(row_number, 10, criteria[6].value)
#                     sheet.cell(row_number, 11, criteria[7].value)
#                     sheet.cell(row_number, 12, criteria[8].value)
#                     sheet.cell(row_number, 13, criteria[9].value)
#                     sheet.cell(row_number, 14, risk_assessment.total_value)
#                     sheet.cell(row_number, 15, risk_assessment.sent_for)
#                     sheet.cell(row_number, 16, issue.summary)
#                     row_number += 1
#                     i += 1
#                 sheet.delete_rows(row_number, 300)
#                 workbook.remove(workbook['Лист2'])
#                 workbook.template = False
#                 stream = BytesIO()
#                 workbook.save(stream)
#                 django_file = DjangoFile(
#                     file=stream,
#                     name=f'Ф2ГО за период {consolidation.start.strftime("%d.%m.%Y")}-'
#                          f'{consolidation.end.strftime("%d.%m.%Y")}.xlsx',
#                 )
#                 report_file = File()
#                 report_file.upload = django_file
#                 report_file.is_confined = True
#                 report_file.save()
#                 report = models.ReportModel.objects.create(
#                     consolidator=None,
#                     parent=consolidation,
#                     contractor_id=each,
#                     status_id='on_review',
#                 )
#                 report_file_instance = models.ReportFileModel.objects.create(
#                     original_file=report_file,
#                     upload_date=timezone.now(),
#                     description='Файл, заполненный на основе шаблона',
#                     name='Карта рисков',
#                     code='risk_matrix',
#                 )
#                 report.report_files.set((report_file_instance,))
#                 # TODO картам рисков поменять статус на 'processed'
#             else:
#                 report = models.ReportModel.objects.create(
#                     consolidator=None,  # TODO Кого вписывать?
#                     parent=consolidation,
#                     contractor_id=each,
#                     status_id='not_loaded',
#                 )

