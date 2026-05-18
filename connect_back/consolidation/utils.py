import json
import os
import shutil
import subprocess
import time
from collections import OrderedDict
from datetime import date, timedelta
from io import BytesIO
from tempfile import TemporaryDirectory
from urllib.parse import quote
from django.db.models import Q

import openpyxl
import pyexcelerate
from dateutil.relativedelta import relativedelta
from django.core.files import File as DjangoFile
from openpyxl.styles import (Alignment, Border, Font, NamedStyle, PatternFill,
                             Side)
from openpyxl.worksheet.page import PageMargins
from rest_framework import exceptions as drf_exceptions

from bkz3.settings import BACKEND_URL, DOWNLOADER_PATH
from bpms.event_calendar.models import (EventCalendarMemberModel,
                                        EventCalendarModel,
                                        EventCalendarTypeModel)
from bpms.event_calendar.utils import get_or_create_related_calendar
from common.current_profile.middleware import get_current_authenticated_profile
from common.humanize import get_humanized_month
from common.models import File
from common.utils import use_access_groups
from common.openpyxl_utils import get_height_for_cell, get_height_for_row
from contractor_permissions.models import ContractorPermissionModel

from . import models, notifications


def get_serialized_report_original_file(instance):
    from common.serializers import AppFileSerializer
    file = instance.original_file
    if not file:
        return None
    s_data = AppFileSerializer(file).data
    if DOWNLOADER_PATH is not None:
        parent_path = quote(f"?obj={instance.pk}&id={s_data.get('id')}&target=original_file")
        s_data['path'] = f'{BACKEND_URL}{DOWNLOADER_PATH}/?path={parent_path}'
    return s_data


def get_serialized_report_pdf_file(instance):
    from common.serializers import AppFileSerializer
    file = instance.pdf_file
    if not file:
        return None
    s_data = AppFileSerializer(file).data
    if DOWNLOADER_PATH is not None:
        parent_path = quote(f"?obj={instance.pk}&id={s_data.get('id')}&target=pdf_file")
        s_data['path'] = f'{BACKEND_URL}{DOWNLOADER_PATH}/?path={parent_path}'
    return s_data


def get_serialized_consolidation_original_file(instance):
    from common.serializers import AppFileSerializer
    file = instance.original_file
    if not file:
        return None
    s_data = AppFileSerializer(file).data
    if DOWNLOADER_PATH is not None:
        parent_path = quote(f"?obj={instance.pk}&id={s_data.get('id')}&target=original_file")
        s_data['path'] = f'{BACKEND_URL}{DOWNLOADER_PATH}/?path={parent_path}'
    return s_data


def get_serialized_consolidation_pdf_file(instance):
    from common.serializers import AppFileSerializer
    file = instance.pdf_file
    if not file:
        return None
    s_data = AppFileSerializer(file).data
    if DOWNLOADER_PATH is not None:
        parent_path = quote(f"?obj={instance.pk}&id={s_data.get('id')}&target=pdf_file")
        s_data['path'] = f'{BACKEND_URL}{DOWNLOADER_PATH}/?path={parent_path}'
    return s_data


def get_data_from_f2go(file) -> list:
    result = list()
    try:
        workbook = openpyxl.load_workbook(filename=file.upload.url[1:])
        sheet = workbook.worksheets[0]
        for index, cell in enumerate(sheet['B']):
            if index != 0:
                result.append(cell.value)
    except Exception:
        raise drf_exceptions.ValidationError(
            'Не удалось прочитать данные из файла'
        )
    return result


def get_data_from_risk_matrix(file) -> list:
    result = list()
    try:
        workbook = openpyxl.load_workbook(filename=file.upload.url[1:], read_only=True)
    except Exception as ex:
        raise drf_exceptions.ValidationError(
            f'Не удалось прочитать данные из файла: {ex.__str__()}'
        )
    else:
        sheet = workbook.worksheets[0]
        for row in sheet.iter_rows(min_row=4, max_row=sheet.max_row, max_col=16, values_only=True):
            if row[0] is None:
                break
            result.append(row)
            # try:
            #     datetime.strptime(row[0], "%d.%m.%Y")
            # except (ValueError, TypeError):
            #     if row[13] is not None:
            #         result.append(row)
            # else:
            #     result.append(row)


        workbook.close()
    return result


def get_contractor_personal_reception(pr_obj) -> list:
    result = list()
    for org in pr_obj.issues:
        (_, data) = org.items().__iter__().__next__()
        summary_row = list()
        issues_rows = list()
        in_queue_issues = list()
        completed_issues = list()
        pensioners_count = 0
        large_families_count = 0
        vulnerable_groups_count = 0
        entrepreneurs_count = 0
        public_and_private_workers_count = 0
        for i in data['issues']:
            status_code = i.get('personal_reception', {}).get('status_code', '')
            social_status_code = i.get('personal_reception', {}).get('social_status_code', '')
            if status_code == 'in_queue':
                in_queue_issues.append(i)
            elif status_code == 'completed':
                completed_issues.append(i)
            if social_status_code == 'pensioners':
                pensioners_count += 1
            if social_status_code == 'large_families':
                large_families_count += 1
            if social_status_code == 'vulnerable_groups':
                vulnerable_groups_count += 1
            if social_status_code == 'entrepreneurs':
                entrepreneurs_count += 1 
            if social_status_code == 'public_and_private_workers':
                public_and_private_workers_count += 1

        # Сводная строка
        summary_row.append(data['org_name'])  # Колонка 1: Название организации
        summary_row.append(data['personal_reception_quantity'])  # Колонка 2: Количество проведенных приемов
        summary_row.append(len(data['issues']))  # Колонка 3: Количество обращений
        summary_row.append(sum(1 for i in data['issues'] if i.get('sent_for', None) == 0))  # Колонка 4: Количество принятых граждан на личном приеме первых руководителей
        summary_row.append(sum(1 for i in data['issues'] if i.get('sent_for', None) == 2))  # Колонка 4: Количество принятых граждан на личном приеме земестителей первых руководителей
        summary_row.append(len(in_queue_issues))  # Колонка 5: Количество обращений в статусе "В очереди"
        summary_row.append(None)  # Колонка 6: Количество дней в очереди
        summary_row.append(len(completed_issues))  # Колонка 7: Количество обращений в статусе "Завершено"
        summary_row.append(pensioners_count)  # Колонка 8: Статус граждан "Пенсионеры"
        summary_row.append(large_families_count)  # Колонка 9: Статус граждан "Многодетные"
        summary_row.append(vulnerable_groups_count)  # Колонка 10: Статус граждан "Инвалиды, самозанятые, безработные"
        summary_row.append(entrepreneurs_count)  # Колонка 11: Статус граждан "Предприниматели"
        summary_row.append(public_and_private_workers_count)  # Колонка 12: Статус граждан "Работники бюдж. и внебюдж. сферы"

        # Деталлизация обращений
        max_length = max(len(in_queue_issues), len(completed_issues))
        in_queue_issues_extended = in_queue_issues + [None] * (max_length - len(in_queue_issues))
        completed_issues_extended = completed_issues + [None] * (max_length - len(completed_issues))

        for i in range(max_length):
            row = [None] * 5
            if in_queue_issues_extended[i] is not None:
                row += [in_queue_issues[i]['number'], in_queue_issues[i]['personal_reception']['days_in_queue']]
            else:
                row += [None] * 2
            if completed_issues_extended[i] is not None:
                row += [completed_issues_extended[i]['number']]
            else:
                row += [None]
            row += [None] * 5
            issues_rows.append(row)

        result.append(summary_row)
        result.extend(issues_rows)
    return result


def f2go_get_risk_matrix_workbook(data_from_reports):

    def get_bg_color(value):
        if value == 0:
            return None
        elif 1 <= value <= 2:
            return pyexcelerate.Color(255, 255, 0, 0)
        elif 3 <= value <= 5:
            return pyexcelerate.Color(255, 153, 0, 0)
        elif 6 <= value <= 10:
            return pyexcelerate.Color(255, 0, 0, 0)
        else:
            return pyexcelerate.Color(255, 0, 255, 0)

    workbook = get_risk_matrix_template_workbook_pyexcelerate()
    ws = workbook._worksheets[0]
    merged_cells = []

    font_bold_14 = pyexcelerate.Font(bold=True, family='Times', size=14, )
    font_bold_11 = pyexcelerate.Font(bold=True, family='Times', size=11, )
    font_normal_11 = pyexcelerate.Font(family='Times', size=11, )
    center_alignment = pyexcelerate.Alignment(horizontal='center', vertical='center', wrap_text=True, )
    left_alignment = pyexcelerate.Alignment(horizontal='left', vertical='center', wrap_text=True,)
    right_alignment = pyexcelerate.Alignment(horizontal='right', vertical='center', wrap_text=True,)
    title_style = pyexcelerate.Style(font=font_bold_14, alignment=center_alignment)
    date_style = pyexcelerate.Style(font=font_bold_11, alignment=center_alignment)

    ws.set_cell_value(1, 1, f'Форма первичной оценки или Карта риска {data_from_reports.get("org_administrator", "")}')
    ws.set_cell_style(1, 1, title_style)
    merged_cells.append(1)
    ws.set_cell_value(2, 1, f'за период {data_from_reports.get("range")}')
    ws.set_cell_style(2, 1, title_style)
    merged_cells.append(2)

    row = 4
    count_1_2, count_3_5, count_6_10 = 0, 0, 0
    correction = 3
    for organization, entries in data_from_reports['risk_matrix'].items():

        # Указываем организацию
        if entries:
            ws.set_cell_value(row, 1, organization)
            ws.set_cell_style(row, 1, date_style)
            merged_cells.append(row)
            row += 1
            correction += 1
        for entry in entries:
            entry_data = []
            if entry[0] == organization:
                continue
            if entry[1] is None:
                ws.set_cell_value(row, 1, entry[0])
                ws.set_cell_style(row, 1, date_style)
                merged_cells.append(row)
                correction += 1
                row += 1
                continue
            else:
                entry_data.append(row - correction)
            # try:
            #     datetime.strptime(entry[0], "%d.%m.%Y")
            # except (TypeError, ValueError):
            #     entry_data.append(row-correction)
            # else:
            #     ws.set_cell_value(row, 1, entry[0])
            #     ws.set_cell_style(row, 1, date_style)
            #     merged_cells.append(row)
            #     correction += 1
            #     row += 1
            #     continue
            entry_data = entry_data + [entry[_] for _ in range(1, 16)]

            entry_total = entry[13]
            if 1 <= entry_total <= 2:
                count_1_2 += 1
            if 3 <= entry_total <= 5:
                count_3_5 += 1
            if 6 <= entry_total <= 10:
                count_6_10 += 1
            ws.range((row, 1), (row, 16)).value = [entry_data]
            ws.range((row, 1), (row, 16)).style.font = font_normal_11
            ws.range((row, 1), (row, 16)).style.alignment = center_alignment
            ws[row][14].style.fill.background = get_bg_color(entry_total)
            row += 1

    total_data = [['1-2 балла', count_1_2], ['3-5 баллов', count_3_5], ['6-10 балла', count_6_10]]
    ws.range((row+2, 2), (row+4, 3)).value = total_data
    ws.range((row+2, 2), (row+4, 2)).style.alignment = left_alignment
    ws.range((row+2, 3), (row+4, 3)).style.alignment = right_alignment
    ws.range((row+2, 2), (row+2, 3)).style.fill.background = get_bg_color(1)
    ws.range((row+3, 2), (row+3, 3)).style.fill.background = get_bg_color(3)
    ws.range((row+4, 2), (row+4, 3)).style.fill.background = get_bg_color(6)
    row -= 1
    ws.range((4, 1), (row, 16)).style.borders.top.style = '_'
    ws.range((4, 1), (row, 16)).style.borders.top.color = pyexcelerate.Color(0, 0, 0)
    ws.range((4, 1), (row, 16)).style.borders.right.style = '_'
    ws.range((4, 1), (row, 16)).style.borders.right.color = pyexcelerate.Color(0, 0, 0)
    ws.range((4, 1), (row, 16)).style.borders.bottom.style = '_'
    ws.range((4, 1), (row, 16)).style.borders.bottom.color = pyexcelerate.Color(0, 0, 0)
    ws.range((4, 1), (row, 16)).style.borders.left.style = '_'

    for each in merged_cells:
        ws.range((each, 1), (each, 16)).merge()

    return workbook


def f2go_get_verification_act_workbook(data_from_reports):
    workbook = get_template_verification_act()
    sheet = workbook.active
    member_name_cell_style = NamedStyle(name="member_name_cell_style")
    value_cell_style = NamedStyle(name="value_cell_style")
    range_cell_style = NamedStyle(name="range_cell_style")
    footer_style = NamedStyle(name="footer_style")

    times_new_roman_16_bold_font = Font(
        name='Times New Roman',
        size=16,
        bold=True,
        italic=False,
        vertAlign=None,
        underline='none',
        strike=False,
        color='00000000'
    )
    times_new_roman_16_normal_font = Font(
        name='Times New Roman',
        size=16,
        bold=False,
        italic=False,
        vertAlign=None,
        underline='none',
        strike=False,
        color='00000000'
    )
    thin_border = Border(
        left=Side(border_style='thin', color='00000000'),
        right=Side(border_style='thin', color='00000000'),
        top=Side(border_style='thin', color='00000000'),
        bottom=Side(border_style='thin', color='00000000'),
    )
    center_alignment = Alignment(
        horizontal='center',
        vertical='center',
        text_rotation=0,
        wrap_text=True,
        shrink_to_fit=False,
        indent=0
    )
    left_alignment = Alignment(
        horizontal='left',
        vertical='center',
        text_rotation=0,
        wrap_text=True,
        shrink_to_fit=False,
        indent=0
    )
    left_alignment_no_wrap_text = Alignment(
        horizontal='left',
        vertical='center',
        text_rotation=0,
        wrap_text=False,
        shrink_to_fit=False,
        indent=0
    )
    range_cell_style.alignment = center_alignment
    range_cell_style.font = times_new_roman_16_bold_font
    member_name_cell_style.alignment = left_alignment
    member_name_cell_style.border = thin_border
    member_name_cell_style.font = times_new_roman_16_normal_font
    value_cell_style.alignment = center_alignment
    value_cell_style.border = thin_border
    value_cell_style.font = times_new_roman_16_bold_font
    footer_style.alignment = left_alignment_no_wrap_text
    footer_style.font = times_new_roman_16_bold_font

    rows = list(enumerate(list(data_from_reports['data'].items()), 13))
    sum_2, sum_3, sum_4, sum_5, sum_6, sum_7 = 0, 0, 0, 0, 0, 0
    revoked_without_routing, transferring_to_another_system, duplicates, stopped = 0, 0, 0, 0
    sheet.cell(4, 1, data_from_reports['org_administrator']).style = range_cell_style
    sheet.row_dimensions[4].height = 48
    sheet.merge_cells(start_column=1, end_column=7, start_row=4, end_row=4)
    sheet.cell(5, 1, f'за период {data_from_reports["range"]}').style = range_cell_style
    sheet.merge_cells(start_column=1, end_column=7, start_row=5, end_row=5)
    for table_row in rows:
        row = table_row[0]
        value = [table_row[1][1]['name'], table_row[1][1]['data']]
        sheet.cell(row, 1, value[0]).style = member_name_cell_style
        sheet.cell(row, 2, int(value[1][0])).style = value_cell_style
        sum_2 += int(value[1][0])
        sheet.cell(row, 3, int(value[1][33])).style = value_cell_style
        sum_3 += int(value[1][33])
        sheet.cell(row, 4, int(value[1][52])).style = value_cell_style
        sum_4 += int(value[1][52])
        sheet.cell(row, 5, int(value[1][44])).style = value_cell_style
        sum_5 += int(value[1][44])
        sheet.cell(row, 6, int(value[1][61])).style = value_cell_style
        sum_6 += int(value[1][61])
        sheet.cell(row, 7, int(value[1][50])).style = value_cell_style
        sum_7 += int(value[1][50])
        revoked_without_routing += table_row[1][1]['revoked_without_routing']
        transferring_to_another_system += table_row[1][1]['transferring_to_another_system']
        duplicates += int(value[1][2])
        stopped += int(value[1][40])
        sheet.row_dimensions[row].height = get_height_for_row(sheet, row-1, font_size=16)
    sheet.cell(12, 2, sum_2)
    sheet.cell(12, 3, sum_3)
    sheet.cell(12, 4, sum_4)
    sheet.cell(12, 5, sum_5)
    sheet.cell(12, 6, sum_6)
    sheet.cell(12, 7, sum_7)

    sheet.cell(row+2, 1, 'Отозвано без маршрутизации:').style = footer_style
    sheet.cell(row+2, 2, revoked_without_routing).style = footer_style
    sheet.cell(row+3, 1, 'Перенос в другую систему:').style = footer_style
    sheet.cell(row+3, 2, transferring_to_another_system).style = footer_style
    sheet.cell(row+4, 1, 'Дубликаты:').style = footer_style
    sheet.cell(row+4, 2, duplicates).style = footer_style
    sheet.cell(row+5, 1, 'Прекращено:').style = footer_style
    sheet.cell(row+5, 2, stopped).style = footer_style
    sheet.cell(row+7, 1, 'Проверили:').style = footer_style
    sheet.cell(row+8, 1, 'сотрудник ДКПСиСУ ГП РК').style = footer_style
    sheet.cell(row+11, 1, 'сотрудник государственного органа').style = footer_style

    return workbook


def get_personal_reception_workbook(data_from_reports):
    workbook = get_template_personal_reception()
    sheet = workbook.active

    member_name_cell_style = NamedStyle(name="member_name_cell_style")
    value_cell_style = NamedStyle(name="value_cell_style")
    range_cell_style = NamedStyle(name="range_cell_style")
    footer_style_title = NamedStyle(name="footer_style_title")
    footer_style_value = NamedStyle(name="footer_style_value")

    times_new_roman_16_bold_font = Font(
        name='Times New Roman',
        size=16,
        bold=True,
        italic=False,
        vertAlign=None,
        underline='none',
        strike=False,
        color='00000000'
    )
    times_new_roman_16_normal_font = Font(
        name='Times New Roman',
        size=16,
        bold=False,
        italic=False,
        vertAlign=None,
        underline='none',
        strike=False,
        color='00000000'
    )
    thin_border = Border(
        left=Side(border_style='thin', color='00000000'),
        right=Side(border_style='thin', color='00000000'),
        top=Side(border_style='thin', color='00000000'),
        bottom=Side(border_style='thin', color='00000000'),
    )
    center_alignment = Alignment(
        horizontal='center',
        vertical='center',
        text_rotation=0,
        wrap_text=True,
        shrink_to_fit=False,
        indent=0
    )
    left_alignment = Alignment(
        horizontal='left',
        vertical='center',
        text_rotation=0,
        wrap_text=True,
        shrink_to_fit=False,
        indent=0
    )

    range_cell_style.alignment = center_alignment
    range_cell_style.font = times_new_roman_16_bold_font

    member_name_cell_style.alignment = left_alignment
    member_name_cell_style.border = thin_border
    member_name_cell_style.font = times_new_roman_16_normal_font

    value_cell_style.alignment = center_alignment
    value_cell_style.border = thin_border
    value_cell_style.font = times_new_roman_16_normal_font

    footer_style_title.alignment = left_alignment
    footer_style_title.border = thin_border
    footer_style_title.font = times_new_roman_16_bold_font
    footer_style_title.fill = PatternFill(
        fill_type='solid',
        start_color='DCDCDC'
    )

    footer_style_value.alignment = center_alignment
    footer_style_value.border = thin_border
    footer_style_value.font = times_new_roman_16_bold_font
    footer_style_value.fill = PatternFill(
        fill_type='solid',
        start_color='DCDCDC'
    )

    # Шапка
    sheet.cell(4, 1, data_from_reports['org_administrator']).style = range_cell_style
    sheet.row_dimensions[4].height = 48
    sheet.merge_cells(start_column=1, end_column=13, start_row=4, end_row=4)
    sheet.cell(5, 1, f'за период {data_from_reports["range"]}').style = range_cell_style
    sheet.merge_cells(start_column=1, end_column=13, start_row=5, end_row=5)

    # Таблица
    personal_reception = data_from_reports.get('personal_reception', {})
    sum_2, sum_3, sum_4, sum_5, sum_6, sum_8, sum_9, sum_10, sum_11, sum_12, sum_13 = 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0
    organizations = list(enumerate(personal_reception.values(), 13))

    for _, rows in organizations:
        for row in rows:
            sheet.append(row)
            row_number = sheet.max_row
            if row[0]:
                sum_2 += sheet.cell(row_number, 2).value
                sum_3 += sheet.cell(row_number, 3).value
                sum_4 += sheet.cell(row_number, 4).value
                sum_5 += sheet.cell(row_number, 5).value
                sum_6 += sheet.cell(row_number, 6).value
                sum_8 += sheet.cell(row_number, 8).value
                sum_9 += sheet.cell(row_number, 9).value
                sum_10 += sheet.cell(row_number, 10).value
                sum_11 += sheet.cell(row_number, 11).value
                sum_12 += sheet.cell(row_number, 12).value
                sum_13 += sheet.cell(row_number, 13).value
                sheet.row_dimensions[row_number].height = get_height_for_row(sheet, row_number-1, font_size=16)
            for i in range(1, 14):
                sheet.cell(row_number, i).style = member_name_cell_style if i == 1 else value_cell_style

    # Итоговая строка
    sheet.cell(row_number+1, 1, 'Итого').style = footer_style_title
    sheet.cell(row_number+1, 2, sum_2).style = footer_style_value
    sheet.cell(row_number+1, 3, sum_3).style = footer_style_value
    sheet.cell(row_number+1, 4, sum_4).style = footer_style_value
    sheet.cell(row_number+1, 5, sum_5).style = footer_style_value
    sheet.cell(row_number+1, 6, sum_6).style = footer_style_value
    sheet.cell(row_number+1, 7).style = footer_style_value
    sheet.cell(row_number+1, 8, sum_8).style = footer_style_value
    sheet.cell(row_number+1, 9, sum_9).style = footer_style_value
    sheet.cell(row_number+1, 10, sum_10).style = footer_style_value
    sheet.cell(row_number+1, 11, sum_11).style = footer_style_value
    sheet.cell(row_number+1, 12, sum_12).style = footer_style_value
    sheet.cell(row_number+1, 13, sum_13).style = footer_style_value

    return workbook


def get_bg_color(row):
    if row in (2, 6, 11, 17, 35,):
        return '00C5D9F1'
    elif row in (16, 26, 47, 54, 64,):
        return '00B5CCE4'
    elif row in (3, 4, 5, 36, 37, 53,):
        return '00FFFFFF'
    elif row in (45, 52,):
        return '00DDD9C4'
    elif row in (46, 63,):
        return '00FCD5B4'
    elif row in (73, 63,):
        return '00538DD5'
    else:
        return '00DAEEF3'


def get_font_color(row):
    if row in (46, 47, 53, 63,):
        return '00FF0000'
    else:
        return '00000000'


def f2go_get_f2go_workbook(data_from_reports):

    workbook = get_template_citizen_inquiries()
    sheet = workbook.active
    summary_header_cell_style = NamedStyle(name="summary_header_cell_style")
    contractor_header_cell_style = NamedStyle(name="contractor_header_cell_style")
    row_cell_style = NamedStyle(name="row_cell_style")
    calibri_bold_8_font = Font(
        name='Calibri',
        size=8,
        bold=True,
        italic=False,
        vertAlign=None,
        underline='none',
        strike=False,
        color='00000000'
    )
    thin_border = Border(
        left=Side(border_style='thin', color='00000000'),
        right=Side(border_style='thin', color='00000000'),
        top=Side(border_style='thin', color='00000000'),
        bottom=Side(border_style='thin', color='00000000'),
    )
    center_alignment = Alignment(
        horizontal='center',
        vertical='center',
        text_rotation=0,
        wrap_text=True,
        shrink_to_fit=False,
        indent=0
    )
    top_center_alignment = Alignment(
        horizontal='center',
        vertical='top',
        text_rotation=0,
        wrap_text=True,
        shrink_to_fit=False,
        indent=0
    )
    summary_header_cell_style.alignment = center_alignment
    summary_header_cell_style.border = thin_border
    summary_header_cell_style.font = calibri_bold_8_font
    contractor_header_cell_style.alignment = top_center_alignment
    contractor_header_cell_style.border = thin_border
    contractor_header_cell_style.font = calibri_bold_8_font
    row_cell_style.alignment = center_alignment
    row_cell_style.border = thin_border

    sheet.cell(1, 2, data_from_reports['summary_column_title']).style = summary_header_cell_style

    for row in range(1, 73):
        column = 3
        sum = 0
        row_cell_style.fill = PatternFill(
            fill_type='solid',
            start_color=get_bg_color(row)
        )
        row_cell_style.font = Font(
            name='Calibri',
            size=12,
            bold=False,
            italic=False,
            vertAlign=None,
            underline='none',
            strike=False,
            color=get_font_color(row)
        )
        for each in data_from_reports['data'].items():
            if row == 1:
                sheet.cell(row, column, each[1]['name']).style = contractor_header_cell_style
                sheet.row_dimensions[1].height = get_height_for_row(sheet, 0, font_size=8)
            else:
                try:
                    value = int(each[1]['data'][row - 2])
                except (ValueError, TypeError):
                    value = each[1]['data'][row - 2]
                sheet.cell(row, column, value).style = row_cell_style
                sum = sum + int(each[1]['data'][row - 2]) if each[1]['data'][row - 2] is not None else None
            column += 1
        if row != 1:
            sheet.cell(row, 2, sum).style = row_cell_style
    return workbook


def f2go_get_data_from_reports(consolidation):
    reports = consolidation.source_reports.filter(
        is_active=True,
        report_files__original_file__isnull=False,
        status_id__in=['approved', 'consolidated']
    ).order_by(
        'sort', 'contractor__sort', 'contractor__name'
    ).distinct()
    form_info = consolidation.report_form.form_info
    report_form_code = consolidation.report_form.code
    if form_info:
        try:
            form_info = json.loads(form_info)
        except:
            pass
    data = dict()
    data['summary_column_title'] = (
        'Всего поступило за период с '
        f'{str(consolidation.start.day).zfill(2)}.{str(consolidation.start.month).zfill(2)}.{consolidation.start.year} '
        f'по {str(consolidation.end.day).zfill(2)}.{str(consolidation.end.month).zfill(2)}.{consolidation.end.year}'
    )
    data['range'] = (
        'с '
        f'{str(consolidation.start.day).zfill(2)}.{str(consolidation.start.month).zfill(2)}.{consolidation.start.year} '
        f'по {str(consolidation.end.day).zfill(2)}.{str(consolidation.end.month).zfill(2)}.{consolidation.end.year}'
    )
    data['org_administrator'] = consolidation.org_administrator.name
    data['data'] = dict()
    data['risk_matrix'] = dict()
    if report_form_code == 'risk_map_with_personal_reception':
        data['personal_reception'] = dict()

    for report in reports:
        try:
            f2go_file = report.report_files.get(
                file_type_id='f2go'
            ).original_file
        except models.ReportFileModel.DoesNotExist:
            raise drf_exceptions.ValidationError(
                'Не удалось получить файл с отчетом Ф2ГО.'
            )
        try:
            risk_matrix_file = report.report_files.get(
                file_type_id='risk_matrix'
            ).original_file
        except models.ReportFileModel.DoesNotExist:
            raise drf_exceptions.ValidationError(
                'Не удалось получить файл с картой рисков.'
            )

        data['data'][report.contractor_id] = dict()
        data['data'][report.contractor_id]['name'] = report.contractor.name
        data['data'][report.contractor_id]['data'] = get_data_from_f2go(f2go_file)

        disintegration = getattr(report, 'disintegration', None)
        if disintegration is not None:
            data['data'][report.contractor_id]['revoked_without_routing'] = disintegration.revoked_without_routing if disintegration.revoked_without_routing else 0
            data['data'][report.contractor_id]['transferring_to_another_system'] = disintegration.transferring_to_another_system if disintegration.transferring_to_another_system is not None else 0
        else:
            data['data'][report.contractor_id]['revoked_without_routing'] = 0
            data['data'][report.contractor_id]['transferring_to_another_system'] = 0

        if report.contractor == consolidation.org_administrator:
            data['data'][report.contractor_id]['balance'] = get_balance(
                contractor=report.contractor_id,
                year=consolidation.start.year-1
                )
        else:
            data['data'][report.contractor_id]['balance'] = get_balance(
                contractor=report.contractor_id,
                year=consolidation.start.year-1,
                with_descendants=True
                )
        data['risk_matrix'][report.contractor.name] = get_data_from_risk_matrix(risk_matrix_file)

        if report_form_code == 'risk_map_with_personal_reception':
            data['personal_reception'][report.contractor_id] = dict()
            personal_reception = getattr(report, 'personal_reception', None)
            if personal_reception is not None:
                data['personal_reception'][report.contractor_id] = get_contractor_personal_reception(personal_reception)

    return data


def f2go_with_verification_act_get_data_from_reports(consolidation):
    reports = consolidation.source_reports.filter(
        is_active=True,
        report_files__original_file__isnull=False,
        status_id__in=['approved', 'consolidated']
    ).order_by('sort', 'contractor__sort', 'contractor__name')
    form_info = consolidation.report_form.form_info
    if form_info:
        try:
            form_info = json.loads(form_info)
        except:
            pass
    data = dict()
    data['summary_column_title'] = (
        'Всего поступило за период с '
        f'{str(consolidation.start.day).zfill(2)}.{str(consolidation.start.month).zfill(2)}.{consolidation.start.year} '
        f'по {str(consolidation.end.day).zfill(2)}.{str(consolidation.end.month).zfill(2)}.{consolidation.end.year}'
    )
    data['range'] = (
        'с '
        f'{str(consolidation.start.day).zfill(2)}.{str(consolidation.start.month).zfill(2)}.{consolidation.start.year} '
        f'по {str(consolidation.end.day).zfill(2)}.{str(consolidation.end.month).zfill(2)}.{consolidation.end.year}'
    )
    data['org_administrator'] = consolidation.org_administrator.name
    data['data'] = dict()

    for report in reports:
        try:
            f2go_file = report.report_files.get(
                file_type_id='f2go'
            ).original_file
        except models.ReportFileModel.DoesNotExist:
            raise drf_exceptions.ValidationError(
                'Не удалось получить файл с отчетом Ф2ГО.'
            )

        data['data'][report.contractor_id] = dict()
        data['data'][report.contractor_id]['name'] = report.contractor.name
        data['data'][report.contractor_id]['data'] = get_data_from_f2go(f2go_file)
        disintegration = getattr(report, 'disintegration', None)
        if disintegration is not None:
            data['data'][report.contractor_id]['revoked_without_routing'] = disintegration.revoked_without_routing if disintegration.revoked_without_routing else 0
            data['data'][report.contractor_id]['transferring_to_another_system'] = disintegration.transferring_to_another_system if disintegration.transferring_to_another_system is not None else 0
        else:
            data['data'][report.contractor_id]['revoked_without_routing'] = 0
            data['data'][report.contractor_id]['transferring_to_another_system'] = 0
        if report.contractor == consolidation.org_administrator:
            data['data'][report.contractor_id]['balance'] = get_balance(
                contractor=report.contractor_id,
                year=consolidation.start.year-1
                )
        else:
            data['data'][report.contractor_id]['balance'] = get_balance(
                contractor=report.contractor_id,
                year=consolidation.start.year-1,
                with_descendants=True
                )

    return data


def get_f2go_report(data, consolidation):
    f2go_workbook = f2go_get_f2go_workbook(data)
    f2go_workbook.properties.identifier = str(consolidation.org_administrator_id)
    stream = BytesIO()
    f2go_workbook.save(stream)
    django_file = DjangoFile(
        file=stream,
        name=f"Ф2ГО за период {data['range']}.xlsx"
    )
    f2go_file = File()
    f2go_file.upload = django_file
    f2go_file.is_confined = True
    f2go_file.save()
    return models.ConsolidationFileModel.objects.create(
            name=f"Ф2ГО за период {data['range']}",
            original_file=f2go_file,
            file_type_id='f2go',
            sort=100
        )


def get_personal_reception_report(data, consolidation):
    personal_reception_workbook = get_personal_reception_workbook(data)
    personal_reception_workbook.properties.identifier = str(consolidation.org_administrator_id)
    stream = BytesIO()
    personal_reception_workbook.save(stream)
    django_file = DjangoFile(
        file=stream,
        name=f"Сведения о проведении личного приема за период {data['range']}.xlsx"
    )
    personal_reception_file = File()
    personal_reception_file.upload = django_file
    personal_reception_file.is_confined = True
    personal_reception_file.save()
    return models.ConsolidationFileModel.objects.create(
            name=f"Сведения о проведении личного приема за период {data['range']}",
            original_file=personal_reception_file,
            file_type_id='personal_reception',
            sort=400
        )


def get_verification_act_report(data, consolidation):
    verification_act_workbook = f2go_get_verification_act_workbook(data)
    verification_act_workbook.properties.identifier = str(consolidation.org_administrator_id)
    stream = BytesIO()
    verification_act_workbook.save(stream)
    django_file = DjangoFile(file=stream, name=f"Акт сверки за период {data['range']}.xlsx")
    verification_act_file = File()
    verification_act_file.upload = django_file
    verification_act_file.is_confined = True
    verification_act_file.save()
    return models.ConsolidationFileModel.objects.create(
            name=f"Акт сверки за период {data['range']}",
            original_file=verification_act_file,
            file_type_id='verification_act',
            sort=300
        )


def get_risk_matrix(data, consolidation):
    verification_act_workbook = f2go_get_risk_matrix_workbook(data)
    stream = BytesIO()
    verification_act_workbook.save(stream)
    openpyxl_workbook = openpyxl.load_workbook(stream)
    sheet = openpyxl_workbook.active
    sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    current_profile = get_current_authenticated_profile()
    if current_profile:
        openpyxl_workbook.properties.creator = current_profile.full_name
    openpyxl_workbook.properties.identifier = str(consolidation.org_administrator_id)
    stream = BytesIO()
    openpyxl_workbook.save(stream)
    openpyxl_workbook.close()
    django_file = DjangoFile(file=stream, name=f"Карта рисков за период {data['range']}.xlsx")
    verification_act_file = File()
    verification_act_file.upload = django_file
    verification_act_file.is_confined = True
    verification_act_file.save()
    return models.ConsolidationFileModel.objects.create(
            name=f"Карта рисков {data['range']}",
            original_file=verification_act_file,
            file_type_id='risk_matrix',
            sort=200
        )


def f2go_m_create_consolidated_report(consolidation) -> list:
    data = f2go_get_data_from_reports(consolidation)
    risk_matrix = get_risk_matrix(data, consolidation)
    f2go_report = get_f2go_report(data, consolidation)
    verification_act_report = get_verification_act_report(data, consolidation)

    return [f2go_report, verification_act_report, risk_matrix]


def rmwpr_create_consolidated_report(consolidation) -> list:
    data = f2go_get_data_from_reports(consolidation)
    risk_matrix = get_risk_matrix(data, consolidation)
    f2go_report = get_f2go_report(data, consolidation)
    verification_act_report = get_verification_act_report(data, consolidation)
    personal_reception_report = get_personal_reception_report(data, consolidation)

    return [f2go_report, verification_act_report, risk_matrix, personal_reception_report]


def f2go_with_verification_act_create_consolidated_report(consolidation) -> list:
    data = f2go_with_verification_act_get_data_from_reports(consolidation)
    f2go_report = get_f2go_report(data, consolidation)
    verification_act_report = get_verification_act_report(data, consolidation)

    return [f2go_report, verification_act_report]


def create_consolidated_reports(consolidation=None):
    if not consolidation:
        return None

    report_form = consolidation.report_form

    report_form_functions = report_form.report_form_info.get(
        report_form.code, None
    )
    if report_form_functions is None:
        raise drf_exceptions.ValidationError(
            'Не удалось получить инструменты для формы '
            f'отчетности "{report_form.name}"'
        )

    create_consolidated_report_function = report_form_functions.get(
        'create_consolidated_report', None
    )
    if create_consolidated_report_function is None:
        raise drf_exceptions.ValidationError(
            'Не удалось получить функцию создания консолидированного отчета '
            f'для формы "{report_form.name}"'
        )

    c_files = create_consolidated_report_function(consolidation)

    return c_files


def create_consolidated_report(consolidation=None):
    if not consolidation:
        return None

    report_form = consolidation.report_form

    report_form_functions = report_form.report_form_info.get(
        report_form.code, None
    )
    if report_form_functions is None:
        raise drf_exceptions.ValidationError(
            'Не удалось получить инструменты для формы '
            f'отчетности "{report_form.name}"'
        )

    create_consolidated_report_function = report_form_functions.get(
        'create_consolidated_report', None
    )
    if create_consolidated_report_function is None:
        raise drf_exceptions.ValidationError(
            'Не удалось получить функцию создания консолидированного отчета '
            f'для формы "{report_form.name}"'
        )

    file = create_consolidated_report_function(consolidation)

    return file


def all_reports_validate(consolidation=None):
    if not consolidation:
        return None

    report_form = consolidation.report_form

    report_form_functions = report_form.report_form_info.get(
        report_form.code, None
    )
    if report_form_functions is None:
        raise drf_exceptions.ValidationError(
            'Не удалось получить инструменты для формы '
            f'отчетности "{report_form.name}"'
        )

    all_reports_validate_function = report_form_functions.get(
        'all_reports_validate', None
    )
    if not all_reports_validate_function:
        return None

    validate = all_reports_validate_function(consolidation)

    return validate


def consolidation_validate(consolidation=None):
    if not consolidation:
        return None

    report_form = consolidation.report_form

    report_form_functions = report_form.report_form_info.get(
        report_form.code, None
    )
    if report_form_functions is None:
        raise drf_exceptions.ValidationError(
            'Не удалось получить инструменты для формы '
            f'отчетности "{report_form.name}"'
        )

    consolidation_validate_function = report_form_functions.get(
        'consolidation_validate', None
    )
    if not consolidation_validate_function:
        return None

    validate = consolidation_validate_function(consolidation)

    return validate


def report_file_validation(file_id, consolidation, code, *args, **kwargs):
    if not consolidation:
        return None

    report_form = consolidation.report_form

    report_form_info = report_form.report_form_info.get(
        report_form.code, None
    )
    if report_form_info is None:
        raise drf_exceptions.ValidationError(
            'Не удалось получить инструменты для формы '
            f'отчетности "{report_form.name}"'
        )
    file_info = None
    if code:
        report_form_files = report_form_info.get('files_info', None)
        if not report_form_files:
            raise drf_exceptions.ValidationError(
                'Информация о файлах не найдена.'
            )
        file_info = next((item for item in report_form_files if item.get('code')==code), None)
    validate_function = None
    if file_info:
        validate_function = file_info.get('validate', None)
    if not validate_function:
        return True

    try:
        file = File.objects.get(
            is_active=True,
            id=file_id
        )
    except File.DoesNotExist:
        raise drf_exceptions.ValidationError(
            'Файл не найден.'
        )

    validate = validate_function(file, *args, **kwargs)

    if not validate:
        file.is_orphaned = True
        file.save(update_fields=('is_orphaned',))
        raise drf_exceptions.ValidationError('Загружаемый файл не прошел проверку.')
    return True


def save_report_data(file_id, report, code):
    if not report:
        return None

    report_form = report.parent.report_form

    report_form_info = report_form.report_form_info.get(
        report_form.code, None
    )
    if report_form_info is None:
        raise drf_exceptions.ValidationError(
            'Не удалось получить инструменты для формы '
            f'отчетности "{report_form.name}"'
        )
    file_info = None
    if code:
        report_form_files = report_form_info.get('files_info', None)
        if not report_form_files:
            raise drf_exceptions.ValidationError(
                'Информация о файлах не найдена.'
            )
        file_info = next((item for item in report_form_files if item.get('code')==code), None)
    save_data_function = None
    if file_info:
        save_data_function = file_info.get('save_data', None)
    if not save_data_function:
        return None

    try:
        file = File.objects.get(
            is_active=True,
            id=file_id
        )
    except File.DoesNotExist:
        raise drf_exceptions.ValidationError(
            'Файл не найден.'
        )
    obj = save_data_function(file, report)
    return obj


def consolidation_is_available(reports, consolidation):
    all_reports_is_approved = True
    all_files_is_uploaded = True

    for report in reports:
        all_reports_is_approved = (all_reports_is_approved and
                                   report.status.code in [
                                       'approved',
                                       'consolidated'])
        all_files_is_uploaded = (all_files_is_uploaded and
                                 not report.report_files.filter(
                                    file_type_id__in=('f2go', 'risk_matrix'),
                                    original_file__isnull=True
                                 ).exists())
    return all_files_is_uploaded and all_reports_is_approved


def f2go_validate(file, *args, **kwargs):

    def get_difference(sum, value) -> str:
        return 'больше' if sum > value else 'меньше'

    def get_sum(lst, indexes):
        return sum(lst[i] for i in indexes)

    if file.extension != 'xlsx':
        raise drf_exceptions.ValidationError(
            'Файл должен иметь расширение xlsx'
        )
    try:
        workbook = openpyxl.load_workbook(filename=file.upload.url[1:])
    except Exception:
        raise drf_exceptions.ValidationError(
            'Не удалось прочитать данные из файла'
        )
    else:
        sheet = workbook.worksheets[0]
        column_b_values = sheet.iter_rows(
            min_row=2,
            min_col=2,
            max_col=2,
            values_only=True
        )
        raw_data = [item[0] for item in column_b_values]

    none_elements = (1, 4, 9, 15, 24, 34)
    data = list()
    for index, value in enumerate(raw_data):
        if index in none_elements:
            if raw_data[index] is not None:
                continue
            else:
                data.append(value)
        else:
            try:
                data.append(int(value))
            except (ValueError, TypeError):
                continue
    if len(data) != 71:
        raise drf_exceptions.ValidationError(
            'Неверный формат данных'
        )

    errors = dict()

    # Тест 1
    indexes = (10, 11, 12, 13)
    total = get_sum(data, indexes)
    if total != data[0]:
        difference = get_difference(total, data[0])
        errors['test_1'] = ('Ошибка в одной из ячеек (В12:В15) отчета Ф2ГО. '
                            'Количество зарегистрированных обращений по видам '
                            f'{difference}, чем всего зарегистрировано '
                            'обращений (ячейка В2).')

    # Тест 2
    indexes = (14, 16, 17, 18, 19, 20, 21, 22, 23, 25, 26, 27, 28, 29, 30, 31, 32)
    total = get_sum(data, indexes)
    if total != data[0]:
        difference = get_difference(total, data[0])
        errors['test_2'] = ('Ошибка в одной из ячеек (В16; В18:В25; В27:В34) '
                            'отчета Ф2ГО. Количество обращений, '
                            'перенаправленных и поступивших непосредственно в '
                            f'ГО, {difference}, чем всего зарегистрировано '
                            'обращений (ячейка В2).')

    # Тест 3
    indexes = (36, 37)
    total = get_sum(data, indexes)
    if total != data[35]:
        difference = get_difference(total, data[35])
        errors['test_3'] = ('Ошибка в одной из ячеек (В38:В39) отчета Ф2ГО. '
                            'Суммарное значение благоприятных и '
                            'обременительных административных актов не '
                            'совпадает с количеством принятых актов '
                            f'({difference}) (ячейка В37).')

    # Тест 4
    indexes = (35, 39, 40, 41, 42)
    total = get_sum(data, indexes)
    if total != data[33]:
        difference = get_difference(total, data[33])
        errors['test_4'] = ('Ошибка в одной из ячеек (В37; В41:В44) отчета '
                            'Ф2ГО. Количество рассмотренных обращений в ГО '
                            f'(ячейка В35) {difference}, чем суммарное '
                            'значение принятых решений.')

    # Тест 5
    indexes = (53, 54, 55, 56, 57, 58, 59, 60)
    total = get_sum(data, indexes)
    if total != data[52]:
        difference = get_difference(total, data[52])
        errors['test_5'] = ('Ошибка в одной из ячеек (В55:В62) отчета Ф2ГО. '
                            'Общее количество перенаправленных обращений из ГО '
                            f'(ячейка В54) {difference}, чем суммарное '
                            'количество перенаправленных обращений.')

    # Тест 6
    indexes = (63, 64, 65, 66, 67, 68, 69, 70)
    total = get_sum(data, indexes)
    if total != data[62]:
        difference = get_difference(total, data[62])
        errors['test_6'] = ('Ошибка в одной из ячеек (В65:В72) отчета Ф2ГО. '
                            'Общее количество частично перенаправленных '
                            f'обращений из ГО (ячейка В64) {difference}, '
                            'чем суммарное количество перенаправленных '
                            'обращений из ГО.')

    if errors:
        raise drf_exceptions.ValidationError(errors)
    return True


def risk_matrix_validate(file, *args, **kwargs):
    if file.extension != 'xlsx':
        raise drf_exceptions.ValidationError(
            'Файл должен иметь расширение xlsx'
        )
    try:
        workbook = openpyxl.load_workbook(filename=file.upload.url[1:])
    except Exception:
        raise drf_exceptions.ValidationError(
            'Не удалось прочитать данные из файла'
        )
    no_inquiries = kwargs.get('no_inquiries', False)
    max_row = 3 if no_inquiries else 4
    sheet = workbook.worksheets[0]
    if sheet.max_row < max_row or sheet.max_column != 16:
        raise drf_exceptions.ValidationError(
            'Неверный формат данных'
        )
    return True


def f2go_all_uploaded_files_validate(consolidation, uploaded_report_files):
    f2go, f2go_id = None, None
    risk_matrix, risk_matrix_id = None, None
    for rf in uploaded_report_files:
        code = rf.get('code', None)
        if code == 'f2go' and rf['original_file']:
            f2go_id = rf['original_file'].get('id', None)
        if code == 'risk_matrix' and rf['original_file']:
            risk_matrix_id = rf['original_file'].get('id', None)
    if not f2go_id or not risk_matrix_id:
        raise drf_exceptions.ValidationError(
            'Загружены не все файлы, проверка невозможна.'
        )
    try:
        f2go = File.objects.get(id=f2go_id)
    except File.DoesNotExist:
        raise drf_exceptions.ValidationError(
            'Не удалось получить файл отчета Ф2ГО.'
        )
    try:
        risk_matrix = File.objects.get(id=risk_matrix_id)
    except File.DoesNotExist:
        raise drf_exceptions.ValidationError(
            'Не удалось получить файл с картой рисков.'
        )
    try:
        f2go_workbook = openpyxl.load_workbook(filename=f2go.upload.url[1:])
    except:
        raise drf_exceptions.ValidationError(
            'Не удалось прочитать данные из файла Ф2ГО.'
        )
    try:
        risk_matrix_workbook = openpyxl.load_workbook(filename=risk_matrix.upload.url[1:])
    except:
        raise drf_exceptions.ValidationError(
            'Не удалось прочитать данные из файла с картой риска.'
        )

    f2go_sheet = f2go_workbook.worksheets[0]
    risk_matrix_sheet = risk_matrix_workbook.worksheets[0]
    entry_counter = 0

    for row in risk_matrix_sheet.iter_rows(min_row=4):
        values_counter = 0
        for cell in row:
            if cell.value is not None:
                values_counter += 1
        if values_counter == 16:
            entry_counter += 1

    if int(f2go_sheet['B2'].value) != entry_counter:
        difference = 'больше' if int(f2go_sheet['B2'].value) < entry_counter else 'меньше'
        raise drf_exceptions.ValidationError(
            f'Количество записей в карте рисков {difference} '
            'количества зарегистрированных обращений.'
        )

    return True


def f2go_with_verification_act_all_uploaded_files_validate(consolidation, uploaded_report_files):
    all_files_uploaded = True
    for each in uploaded_report_files:
        all_files_uploaded = all_files_uploaded and each['original_file'] is not None

    if not all_files_uploaded:
        raise drf_exceptions.ValidationError('Файл не загружен')


def convert_xls_to_pdf(file, file_type='pdf'):
    file_name, file_extension = os.path.splitext(
        os.path.basename(file.upload.file.name)
    )
    with TemporaryDirectory() as tmpdir:
        try:
            result = subprocess.check_output(
                ['soffice', '--headless', '--convert-to', file_type, '--outdir', tmpdir, file.upload.file.name],
                stderr=subprocess.STDOUT
            )
        except subprocess.CalledProcessError:
            raise drf_exceptions.ValidationError()
        except OSError:
            raise drf_exceptions.ValidationError()
        if result.startswith(b'Error'):
            raise drf_exceptions.ValidationError()
        tmp_file = os.path.join(tmpdir, f'{file_name}.{file_type}')
        with open(tmp_file, 'rb') as target_file:
            stream = BytesIO(target_file.read())
            django_file = DjangoFile(file=stream, name=f"{file.name}.{file_type}")
            pdf_file = File()
            pdf_file.upload = django_file
            pdf_file.is_confined = True
            pdf_file.save()
    return pdf_file


def get_pdf(file, file_type='pdf', max_attempts=10):
    file_name, file_extension = os.path.splitext(
        os.path.basename(file.upload.file.name)
    )
    with TemporaryDirectory() as tmpdir:
        temp_file_name = f"temp_{file_name}"
        temp_file_path = os.path.join(
            tmpdir,
            f'{temp_file_name}{file_extension}'
        )
        shutil.copy2(file.upload.file.name, temp_file_path)
        command = ['soffice', '--headless', '--convert-to', file_type, '--outdir', tmpdir, temp_file_path]
        attempts = 0
        error = 'Ошибка конвертирования'
        while attempts < max_attempts:
            try:
                result = subprocess.check_output(
                    command,
                    stderr=subprocess.STDOUT
                )
            except (subprocess.CalledProcessError, OSError) as e:
                attempts += 1
                error = e
            else:
                if result.startswith(b'Error'):
                    attempts += 1
                else:
                    tmp_file = os.path.join(tmpdir, f'{temp_file_name}.{file_type}')
                    with open(tmp_file, 'rb') as target_file:
                        pdf_file = BytesIO(target_file.read())
                    return pdf_file
            if attempts < max_attempts:
                time.sleep(0.1)
            else:
                raise drf_exceptions.ValidationError(error)


def check_need_set_calendar_event_by_consolidation(some_obj_id):
    from .models import ConsolidationModel
    some_obj = ConsolidationModel.objects.get(pk=some_obj_id)

    if some_obj.dead_line:

        calendar = get_or_create_related_calendar('НЕ ТРЕБУЕТСЯ', some_obj.id, False)

        event = EventCalendarModel.objects.filter(calendar=calendar,
                                                  event_type='task_deadline',
                                                  is_active=True).first()
        if not event:
            event = EventCalendarModel()
            event.calendar = calendar
            event.color = 'red'
            event.event_type, is_new = EventCalendarTypeModel.objects.get_or_create(
                code='task_deadline',
                defaults={
                    'name': 'Сдача задачи',
                    'is_active': False,
                }
            )
        event.start_at = some_obj.dead_line
        event.end_at = some_obj.dead_line
        event.name = 'Сдать документы по консолидации ' + str(some_obj)
        event.notify_at = some_obj.dead_line - timedelta(days=7)
        event.description = "<H2>ТЕКСТ ОПИСАНИЯ ЗАДАЧИ</H2><br>" + some_obj.description

        event.is_finished = False
        if some_obj.status.code == 'completed':
            event.is_finished = True

        event.save()
        members = some_obj.members.all()
        report_form = some_obj.report_form
        if use_access_groups(None):
            from contractor_permissions.utils import users_that_have_permission_in_contractors
            recipients = set(
                users_that_have_permission_in_contractors(
                    list(members.values_list('pk', flat=True)),
                    ('create_consolidation', 'send_report'),
                    None,
                )
            )
        else:
            recipients = set(ContractorPermissionModel.objects.filter(
                (Q(aux_conditions=report_form) | Q(aux_conditions__isnull=True)),
                contractor_permission_role__is_active=True,
                contractor_permission_role__contractor__in=members,
                permission_type_id__in=['create_consolidation', 'send_report'],
            ).values_list(
                'contractor_permission_role__contractor_profiles__user',
                flat=True
            ))
        recipients.discard(None)
        for each in members:
            try:
                EventCalendarMemberModel.objects.create(user=each.profile, event=event)
            except:
                pass


def save_f2go_report(file, report):

    def get_data(workbook):
        result = {
            'map': {},
            'data': {}
        }
        sheet = workbook.worksheets[0]
        for index, row in enumerate(sheet.iter_rows()):
            if index:
                result['map'][index] = row[0].value
                result['data'][index] = row[1].value
        return result

    try:
        workbook = openpyxl.load_workbook(filename=file.upload.url[1:])
    except Exception:
        raise drf_exceptions.ValidationError(
            'Не удалось прочитать данные из файла'
        )
    else:
        consolidation = report.parent
        identifier = getattr(workbook.properties, 'identifier', '')
        if identifier == str(report.contractor_id):
            # Если у загруженного файла указан ID организации в качестве
            # идентификатора, значит этот файл был создан в модуле "Консолидация".
            # Данные из такого файла не сохраняем.
            return

        is_month_start = consolidation.start.month != (consolidation.start - timedelta(days=1)).month
        is_month_end = consolidation.end.month != (consolidation.end + timedelta(days=1)).month
        are_within_same_month = consolidation.start.month == consolidation.end.month
        if is_month_start and is_month_end and are_within_same_month:
            data = get_data(workbook)
            obj, created = models.F2GOReportModel.objects.get_or_create(
                organization=report.contractor,
                report_form=consolidation.report_form,
                start=consolidation.start,
                end=consolidation.end,
                defaults={
                    'data': json.dumps(data)
                }
            )
            if not created:
                obj.data = json.dumps(data)
                obj.save(update_fields=('data', ))
    return


def get_f2go_analytics(organizations=[], request=None, data=None):
    if not request:
        return []

    start = request.query_params.get('start')
    end = request.query_params.get('end')
    if not organizations or not start or not end:
        return []

    row_numbers = data
    rows = data
    if not row_numbers or not isinstance(row_numbers, list):
        return []

    queryset = models.F2GOReportModel.objects.filter(
            start__gte=start,
            end__lte=end,
            organization_id__in=organizations
        )
    if not queryset:
        return []

    result_dict = OrderedDict()
    for report in queryset:
        report_data = json.loads(report.data)
        values = report_data.get('data')
        for row in rows:
            value = values.get(str(row['row_number']))
            key = row['label']
            if not key:
                continue
            if key in result_dict:
                if value:
                    result_dict[key] += int(value)
            else:
                if value:
                    result_dict[key] = int(value)
                else:
                    result_dict[key] = 0

    result = [{'name': key, 'value': value} for key, value in result_dict.items()]
    return result


def create_scheduled_consolidation(source_object, repeat_to, extra_fields):
    from dateutil.relativedelta import MO, relativedelta
    from django.db import transaction
    from django.utils import timezone

    report_form_instance = source_object.get_report_form_instance()

    report_form_instance.before_create_scheduled_consolidation(consolidation=source_object)

    def is_month_end(date):
        return date.month != (date + timedelta(days=1)).month

    TIMEDELTAS = {
        'WEEKLY': relativedelta(weeks=+1),
        'MONTHLY': relativedelta(months=+1),
        'YEARLY': relativedelta(years=+1),
    }

    repeat_period = source_object.repeat_period
    time_delta = TIMEDELTAS.get(repeat_period)
    if not time_delta:
        raise drf_exceptions.ValidationError(
            'Периодичность указана некорректно'
        )

    if repeat_period == 'WEEKLY':
        next_creation_date = timezone.now() + time_delta + relativedelta(weekday=MO(-1), hour=0, minute=0, second=0, microsecond=0)
    elif repeat_period == 'MONTHLY':
        next_creation_date = timezone.now() + time_delta + relativedelta(day=1, hour=0, minute=0, second=0, microsecond=0)
    elif repeat_period == 'YEARLY':
        next_creation_date = timezone.now() + time_delta + relativedelta(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
    else:
        next_creation_date = None

    if next_creation_date and repeat_to and next_creation_date.date() >= repeat_to:
        raise drf_exceptions.ValidationError(
            'Дата окончания создания консолидаций указана некорректно'
        )

    if repeat_period in ['MONTHLY', 'YEARLY'] and is_month_end(source_object.dead_line):
        next_dead_line = source_object.dead_line + time_delta + relativedelta(day=31) # , hour=0, minute=0, second=0, microsecond=0)
    else:
        next_dead_line = source_object.dead_line + time_delta # + relativedelta(hour=0, minute=0, second=0, microsecond=0)

    next_start = source_object.start + time_delta # + relativedelta(hour=0, minute=0, second=0, microsecond=0)

    if repeat_period in ['MONTHLY', 'YEARLY'] and is_month_end(source_object.end):
        next_end = source_object.end + time_delta + relativedelta(day=31) #, hour=23, minute=59, second=59, microsecond=999999)
    else:
        next_end = source_object.end + time_delta # + relativedelta(hour=23, minute=59, second=59, microsecond=999999)

    report_form = source_object.report_form
    report_form_info = report_form.report_form_info.get(report_form.code, None)
    files_info = report_form_info.get('files_info', []) if report_form_info else []

    with transaction.atomic():
        template = models.ConsolidationModel.objects.create(
            add_org_administrator_in_members=source_object.add_org_administrator_in_members,
            auto_approve=source_object.auto_approve,
            dead_line=source_object.dead_line,
            description=source_object.description,
            end=source_object.end,
            is_scheduled=True,
            name=source_object.name,
            next_creation_date=next_creation_date,
            next_dead_line=next_dead_line,
            next_end=next_end,
            next_start=next_start,
            org_administrator=source_object.org_administrator,
            repeat_period=repeat_period,
            repeat_to=repeat_to,
            report_form=source_object.report_form,
            start=source_object.start,
            generate_report_files=source_object.generate_report_files
        )
        members = source_object.members.filter(is_active=True)
        for member in members:
            models.ConsolidationMemberModel.objects.create(
                        organization=member,
                        consolidation=template
                    )
            report = models.ReportModel.objects.create(
                    consolidator=source_object.author,
                    parent=template,
                    contractor=member,
                    status_id='not_loaded'
                )
            if files_info:
                report_files = [
                    models.ReportFileModel.objects.create(
                        file_type_id=file_info.get('code', 'default'),
                        sort=file_info.get('sort', 500),
                    ) for file_info in files_info
                ]
                report.report_files.set(report_files)

        attachments = source_object.attachments.filter(is_active=True)
        template.attachments.set(attachments)
        source_object.template = template
        source_object.save(update_fields=('template',))
        template.get_report_form_instance().set_extra_fields(
            consolidation=template,
            extra_fields=extra_fields
        )
    return


def handle_generate_sources(report, report_file):
    if report_file.file_type_id == 'risk_matrix' and report_file.is_generated is True:
        from risk_assessment.models import RiskAssessmentModel
        consolidation = report.parent
        start = consolidation.start
        end = consolidation.end
        RiskAssessmentModel.objects.filter(
            is_active=True,
            organization=report.contractor,
            issue__issue_date__gte=start,
            issue__issue_date__lte=end
        ).update(status_id='processed')


def send_documents(instance, recipient_reports, user):
    if not instance or not recipient_reports:
        return None

    report_form = instance.report_form

    report_form_functions = report_form.report_form_info.get(
        report_form.code, None
    )
    if report_form_functions is None:
        raise drf_exceptions.ValidationError(
            'Не удалось получить инструменты для формы '
            f'отчетности "{report_form.name}"'
        )

    send_documents_function = report_form_functions.get(
        'send_documents', None
    )
    if not send_documents_function:
        return None

    return send_documents_function(instance, recipient_reports, user)


def get_copy(file_name, source_file):
    original_file, pdf_file = None, None
    if source_file.original_file:
        with open(source_file.original_file.upload.url[1:], 'rb') as target_file:
            stream = BytesIO(target_file.read())
            django_file = DjangoFile(file=stream, name=f"{file_name}.xlsx")
            original_file = File()
            original_file.upload = django_file
            original_file.is_confined = True
            original_file.save()
    if source_file.pdf_file:
        with open(source_file.pdf_file.upload.url[1:], 'rb') as target_file:
            stream = BytesIO(target_file.read())
            django_file = DjangoFile(file=stream, name=f"{file_name}.pdf")
            pdf_file = File()
            pdf_file.upload = django_file
            pdf_file.is_confined = True
            pdf_file.save()
    return original_file, pdf_file


def rmwpr_send_documents(instance, recipient_reports, user):
    from django.db import transaction

    with transaction.atomic():
        instance_reports = instance.source_reports.filter(
            is_active=True,
            status_id__in=['approved', 'consolidated']
        ).order_by('sort', 'contractor__sort', 'contractor__name')
        issues = []
        quantity = 0
        no_personal_reception = False
        for each in instance_reports:
            pr = getattr(each, 'personal_reception', None)
            if pr:
                issues.extend(pr.issues)
                if each.contractor == instance.org_administrator:
                    quantity = pr.quantity
                    no_personal_reception = pr.no_personal_reception
        for item in recipient_reports:
            report, _ = item
            report_recipient_report, created = models.ReportPersonalReceptionModel.objects.get_or_create(
                is_active=True,
                report=report,
                defaults={
                    'issues': issues,
                    'no_personal_reception': no_personal_reception,
                    'quantity': quantity,
                }
            )
            if not created:
                report_recipient_report.issues = issues
                report_recipient_report.no_personal_reception = no_personal_reception
                report_recipient_report.quantity = quantity
                report_recipient_report.save(update_fields=(
                    'issues',
                    'no_personal_reception',
                    'quantity',
                ))
        try:
            f2go_send_documents(instance, recipient_reports, user)
        except Exception:
            raise
        else:
            return True


def f2go_send_documents(instance, recipient_reports, user):
    from django.db import transaction
    from django.utils import timezone
    from django_q.tasks import async_task

    consolidation_files = instance.consolidation_files.filter(
        is_active=True
    )
    if not consolidation_files:
        return None
    range = (
        'с '
        f'{str(instance.start.day).zfill(2)}.{str(instance.start.month).zfill(2)}.{instance.start.year} '
        f'по {str(instance.end.day).zfill(2)}.{str(instance.end.month).zfill(2)}.{instance.end.year}'
    )

    try:
        f2go_source_file = consolidation_files.get(file_type_id='f2go')
    except models.ConsolidationFileModel.DoesNotExist:
        raise drf_exceptions.ValidationError(
            'Ошибка при получении файла отчета Ф2ГО'
        )
    try:
        risk_matrix_source_file = consolidation_files.get(file_type_id='risk_matrix')
    except models.ConsolidationFileModel.DoesNotExist:
        raise drf_exceptions.ValidationError(
            'Ошибка при получении Карты рисков'
        )

    with transaction.atomic():
        instance_reports = instance.source_reports.filter(
                is_active=True
            )
        transferring_sum, revoked_sum = 0, 0
        for each in instance_reports:
            disintegration = getattr(each, 'disintegration', None)
            if disintegration is not None:
                revoked_sum += disintegration.revoked_without_routing if disintegration.revoked_without_routing else 0
                transferring_sum += disintegration.transferring_to_another_system if disintegration.transferring_to_another_system else 0

        for item in recipient_reports:
            report, auto_approve = item
            report_files = report.report_files.filter(
                is_active=True
            )
            f2go_report_file = report_files.get(is_active=True, file_type_id='f2go')
            # Если в отчет уже загружен файл присвоим ему признак 'Без родителя'
            if f2go_report_file.original_file:
                f2go_report_file.original_file.is_orphaned = True
                f2go_report_file.original_file.save(update_fields=('is_orphaned',))
            if f2go_report_file.pdf_file:
                f2go_report_file.pdf_file.is_orphaned = True
                f2go_report_file.pdf_file.save(update_fields=('is_orphaned',))

            risk_matrix_report_file = report_files.get(is_active=True, file_type_id='risk_matrix')
            # Если в отчет уже загружен файл присвоим ему признак 'Без родителя'
            if risk_matrix_report_file.original_file:
                risk_matrix_report_file.original_file.is_orphaned = True
                risk_matrix_report_file.original_file.save(update_fields=('is_orphaned',))
            if risk_matrix_report_file.pdf_file:
                risk_matrix_report_file.pdf_file.is_orphaned = True
                risk_matrix_report_file.pdf_file.save(update_fields=('is_orphaned',))

            # Создаем копии файлов источника отчета Ф2ГО
            file_name = f'Отчет Ф2ГО за период {range}'
            original_file, pdf_file = get_copy(
                file_name,
                f2go_source_file
                )
            if original_file:
                f2go_report_file.file_type_id = 'f2go'
                f2go_report_file.name = file_name
                f2go_report_file.original_file = original_file
                f2go_report_file.sort = 100
                f2go_report_file.upload_date = timezone.now()
                f2go_report_file.uploaded_by = user
                f2go_report_file.save(update_fields=(
                    'file_type_id',
                    'name',
                    'original_file',
                    'sort',
                    'upload_date',
                    'uploaded_by'
                ))
            else:
                raise drf_exceptions.ValidationError(
                    'Ошибка при копировании файлов отчета Ф2ГО'
                )
            # Создаем копии файлов источника Карты рисков
            file_name = f'Карта рисков за период {range}'
            original_file, pdf_file = get_copy(
                file_name,
                risk_matrix_source_file
                )
            if original_file:
                risk_matrix_report_file.file_type_id = 'risk_matrix'
                risk_matrix_report_file.name = file_name
                risk_matrix_report_file.original_file = original_file
                risk_matrix_report_file.sort = 200
                risk_matrix_report_file.upload_date = timezone.now()
                risk_matrix_report_file.uploaded_by = user
                risk_matrix_report_file.save(update_fields=(
                    'file_type_id',
                    'name',
                    'original_file',
                    'sort',
                    'upload_date',
                    'uploaded_by'
                ))
            else:
                raise drf_exceptions.ValidationError(
                    'Ошибка при копировании файлов Карты рисков'
                )
            # Копируем значения полей Отозвано без маршрутизации и Перенос в другую систему
            report_disintegration, created = models.DisintegrationModel.objects.get_or_create(
                is_active=True,
                report=report,
                defaults={
                    'revoked_without_routing': revoked_sum,
                    'transferring_to_another_system': transferring_sum
                }
            )
            if not created:
                report_disintegration.revoked_without_routing = revoked_sum
                report_disintegration.transferring_to_another_system = transferring_sum
                report_disintegration.save(update_fields=(
                    'revoked_without_routing',
                    'transferring_to_another_system'
                ))

            # Присвоим отчету новый статус
            new_status = 'approved' if auto_approve else 'on_review'
            report.status_id = new_status
            report.save(update_fields=(
                'status_id',
            ))
            # Присвоим новый статус консолидации
            parent = report.parent
            new_status = 'in_progress'
            if not parent.source_reports.filter(
                is_active=True,
                status__code__in=['new', 'not_loaded']
            ).exists():
                if parent.all_reports_is_approved():
                    new_status = 'ready_to_consolidate'
                async_task(
                    notifications.notify_all_reports_are_uploaded,
                    str(parent.id),
                    str(user.id)
                )
            parent.status_id = new_status
            parent.save(update_fields=(
                'status_id',
            ))

    return True


def f2go_with_verification_act_send_documents(instance, recipient_reports, user):
    from django.db import transaction
    from django.utils import timezone
    from django_q.tasks import async_task

    consolidation_files = instance.consolidation_files.filter(
        is_active=True
    )
    if not consolidation_files:
        return None

    range = (
        'с '
        f'{str(instance.start.day).zfill(2)}.{str(instance.start.month).zfill(2)}.{instance.start.year} '
        f'по {str(instance.end.day).zfill(2)}.{str(instance.end.month).zfill(2)}.{instance.end.year}'
    )
    try:
        f2go_source_file = consolidation_files.get(file_type_id='f2go')
    except models.ConsolidationFileModel.DoesNotExist:
        raise drf_exceptions.ValidationError(
            'Ошибка при получении файла отчета Ф2ГО'
        )

    with transaction.atomic():
        instance_reports = instance.source_reports.filter(
                is_active=True
            )
        transferring_sum, revoked_sum = 0, 0
        for each in instance_reports:
            disintegration = getattr(each, 'disintegration', None)
            if disintegration is not None:
                revoked_sum += disintegration.revoked_without_routing if disintegration.revoked_without_routing else 0
                transferring_sum += disintegration.transferring_to_another_system if disintegration.transferring_to_another_system else 0
        for item in recipient_reports:
            report, auto_approve = item
            report_files = report.report_files.filter(
                is_active=True
            )
            f2go_report_file = report_files.get(is_active=True, file_type_id='f2go')
            # Если в отчет уже загружен файл присвоим ему признак 'Без родителя'
            if f2go_report_file.original_file:
                f2go_report_file.original_file.is_orphaned = True
                f2go_report_file.original_file.save(update_fields=('is_orphaned',))
            if f2go_report_file.pdf_file:
                f2go_report_file.pdf_file.is_orphaned = True
                f2go_report_file.pdf_file.save(update_fields=('is_orphaned',))
            # Создаем копии файлов источника отчета Ф2ГО
            file_name = f'Отчет Ф2ГО за период {range}'
            original_file, pdf_file = get_copy(
                file_name,
                f2go_source_file
                )
            if original_file:
                f2go_report_file.file_type_id = 'f2go'
                f2go_report_file.name = file_name
                f2go_report_file.original_file = original_file
                f2go_report_file.sort = 100
                f2go_report_file.upload_date = timezone.now()
                f2go_report_file.uploaded_by = user
                f2go_report_file.save(update_fields=(
                    'file_type_id',
                    'name',
                    'original_file',
                    'sort',
                    'upload_date',
                    'uploaded_by'
                ))
            else:
                raise drf_exceptions.ValidationError(
                    'Ошибка при копировании файлов отчета Ф2ГО'
                )
            # Копируем значения полей Отозвано без маршрутизации и Перенос в другую систему
            report_disintegration, created = models.DisintegrationModel.objects.get_or_create(
                is_active=True,
                report=report,
                defaults={
                    'revoked_without_routing': revoked_sum,
                    'transferring_to_another_system': transferring_sum
                }
            )
            if not created:
                report_disintegration.revoked_without_routing = revoked_sum
                report_disintegration.transferring_to_another_system = transferring_sum
                report_disintegration.save(update_fields=(
                    'revoked_without_routing',
                    'transferring_to_another_system'
                ))
            # Присвоим отчету новый статус
            new_status = 'approved' if auto_approve else 'on_review'
            report.status_id = new_status
            report.save(update_fields=(
                'status_id',
            ))
            # Присвоим новый статус консолидации
            parent = report.parent
            new_status = 'in_progress'
            if not parent.source_reports.filter(
                is_active=True,
                report_files__original_file__isnull=True
            ).exists():
                if parent.all_reports_is_approved():
                    new_status = 'ready_to_consolidate'
                async_task(
                    notifications.notify_all_reports_are_uploaded,
                    str(parent.id),
                    str(user.id)
                )
            parent.status_id = new_status
            parent.save(update_fields=(
                'status_id',
            ))

    return True


def rollback_consolidation(consolidation):
    if consolidation.auto_approve:
        report_status = 'approved'
        consolidation_status = 'ready_to_consolidate'
    else:
        report_status = 'on_review'
        consolidation_status = 'in_progress'

    consolidation.status_id = consolidation_status
    consolidation.consolidated_at = None
    consolidation.consolidator = None
    consolidation.save(update_fields=('status_id', 'consolidated_at', 'consolidator',),)
    reports = consolidation.source_reports.filter(
        is_active=True,
    )
    for report in reports:
        report.status_id = report_status
        report.save(update_fields=('status_id',),)
    consolidation.consolidation_files.clear()
    # TODO Осиротить файлы консолидированных отчетов
    consolidation.refresh_from_db()
    return consolidation


def ipf_create_consolidated_report(consolidation) -> list:

    from accounting_reports.models import FPCReportModel, ProposalItemModel
    from accounting_reports.utils import get_upload_for_1C
    from common.accounting_catalogs.models import (
        BudgetFunctionalGroupModel, BudgetProgramAdministratorModel,
        BudgetProgramModel)
    from common.catalogs.models import ContractorModel
    reports = consolidation.source_reports.filter(is_active=True, without_attachments=False)
    if not reports:
        raise drf_exceptions.ValidationError({'message': 'Консолидация не содержит отчетов'})
    accounting_reports = FPCReportModel.objects.filter(is_active=True, consolidation_reports__in=reports)
    proposal_items = ProposalItemModel.objects.filter(
        is_active=True, report__in=accounting_reports).order_by('functional_group__code')
    functional_groups = BudgetFunctionalGroupModel.objects.filter(
        is_active=True, proposal_items__in=proposal_items).order_by('code').distinct()
    workbook = openpyxl.load_workbook('accounting_reports.xlsx')

    title_font = Font(
        name='Times New Roman',
        size=8,
        bold=True,
        italic=False,
        vertAlign=None,
        underline='none',
        strike=False,
        color='00000000'
    )
    regular_font = Font(
        name='Times New Roman',
        size=8,
        bold=False,
        italic=False,
        vertAlign=None,
        underline='none',
        strike=False,
        color='00000000'
    )

    thin_border = Border(
        left=Side(border_style='thin', color='00000000'),
        right=Side(border_style='thin', color='00000000'),
        top=Side(border_style='thin', color='00000000'),
        bottom=Side(border_style='thin', color='00000000'),
    )
    thick_border = Border(
        left=Side(border_style='medium', color='00000000'),
        right=Side(border_style='medium', color='00000000'),
        top=Side(border_style='medium', color='00000000'),
        bottom=Side(border_style='medium', color='00000000'),
    )
    center_alignment = Alignment(
        horizontal='center',
        vertical='center',
        wrap_text=True,
        shrink_to_fit=True,
        indent=0
    )

    page_header_style = NamedStyle(name='page_header_style')
    page_header_style.font = regular_font

    common_style = NamedStyle(name='common_style')
    common_style.font = regular_font
    common_style.border = thin_border

    title_style = NamedStyle(name='title_style')
    title_style.font = title_font

    document_title_style = NamedStyle(name='document_title_style')
    document_title_style.font = title_font
    document_title_style.alignment = center_alignment

    footer_style = NamedStyle(name='footer_style')
    footer_style.font = title_font
    footer_style.border = thick_border

    # Сводная таблица:
    sheet = workbook.worksheets[0]

    org_administrator_name = consolidation.org_administrator.name

    sheet.cell(row=4, column=1, value=f'"__" _______ ____ г. {org_administrator_name}').style = page_header_style
    sheet.merge_cells(start_row=4, end_row=4, start_column=1, end_column=17)

    sheet.cell(row=9, column=1, value=f'на {consolidation.end.year} год').style = document_title_style
    sheet.merge_cells(start_row=9, end_row=9, start_column=1, end_column=17)

    sheet.cell(row=10, column=1, value=org_administrator_name).style = document_title_style
    sheet.merge_cells(start_row=10, end_row=10, start_column=1, end_column=17)

    row = 20
    total_sum = {
        'january__sum': 0,
        'february__sum': 0,
        'march__sum': 0,
        'april__sum': 0,
        'may__sum': 0,
        'june__sum': 0,
        'july__sum': 0,
        'august__sum': 0,
        'september__sum': 0,
        'october__sum': 0,
        'november__sum': 0,
        'december__sum': 0,
        'year__sum': 0,
    }
    for functional_group in functional_groups:

        admins = BudgetProgramAdministratorModel.objects.filter(
            is_active=True,
            proposal_items__in=proposal_items,
            functional_subgroup__functional_group=functional_group,
        ).order_by('code').distinct()
        for admin in admins:
            programs = BudgetProgramModel.objects.filter(
                is_active=True,
                proposal_items__in=proposal_items,
                budget_program_administrator=admin,
            ).order_by('code').distinct()
            for program in programs:
                row += 1
                sheet.cell(row=row, column=1, value=functional_group.code).style = common_style
                sheet.cell(row=row, column=2, value=admin.code).style = common_style
                sheet.cell(row=row, column=3, value=program.code).style = common_style
                sheet.cell(row=row, column=4, value=program.name).style = common_style
                aggregate_sum = proposal_items.filter(
                    functional_group=functional_group,
                    budget_program_administrator=admin,
                    program=program,
                ).aggregate(
                    Sum('january'),
                    Sum('february'),
                    Sum('march'),
                    Sum('april'),
                    Sum('may'),
                    Sum('june'),
                    Sum('july'),
                    Sum('august'),
                    Sum('september'),
                    Sum('october'),
                    Sum('november'),
                    Sum('december'),
                )
                year_sum = 0
                for key, value in aggregate_sum.items():
                    year_sum += value
                aggregate_sum['year__sum'] = year_sum

                for key, value in aggregate_sum.items():
                    total_sum[key] += value

                sheet.cell(row=row, column=5, value=year_sum).style = common_style
                sheet.cell(row=row, column=6, value=aggregate_sum['january__sum']).style = common_style
                sheet.cell(row=row, column=7, value=aggregate_sum['february__sum']).style = common_style
                sheet.cell(row=row, column=8, value=aggregate_sum['march__sum']).style = common_style
                sheet.cell(row=row, column=9, value=aggregate_sum['april__sum']).style = common_style
                sheet.cell(row=row, column=10, value=aggregate_sum['may__sum']).style = common_style
                sheet.cell(row=row, column=11, value=aggregate_sum['june__sum']).style = common_style
                sheet.cell(row=row, column=12, value=aggregate_sum['july__sum']).style = common_style
                sheet.cell(row=row, column=13, value=aggregate_sum['august__sum']).style = common_style
                sheet.cell(row=row, column=14, value=aggregate_sum['september__sum']).style = common_style
                sheet.cell(row=row, column=15, value=aggregate_sum['october__sum']).style = common_style
                sheet.cell(row=row, column=16, value=aggregate_sum['november__sum']).style = common_style
                sheet.cell(row=row, column=17, value=aggregate_sum['december__sum']).style = common_style
    row += 1
    sheet.cell(row=row, column=1,).style = footer_style
    sheet.cell(row=row, column=2,).style = footer_style
    sheet.cell(row=row, column=3, value='ИТОГО').style = footer_style
    sheet.cell(row=row, column=4,).style = footer_style
    sheet.cell(row=row, column=5, value=total_sum['year__sum']).style = footer_style
    sheet.cell(row=row, column=6, value=total_sum['january__sum']).style = footer_style
    sheet.cell(row=row, column=7, value=total_sum['february__sum']).style = footer_style
    sheet.cell(row=row, column=8, value=total_sum['march__sum']).style = footer_style
    sheet.cell(row=row, column=9, value=total_sum['april__sum']).style = footer_style
    sheet.cell(row=row, column=10, value=total_sum['may__sum']).style = footer_style
    sheet.cell(row=row, column=11, value=total_sum['june__sum']).style = footer_style
    sheet.cell(row=row, column=12, value=total_sum['july__sum']).style = footer_style
    sheet.cell(row=row, column=13, value=total_sum['august__sum']).style = footer_style
    sheet.cell(row=row, column=14, value=total_sum['september__sum']).style = footer_style
    sheet.cell(row=row, column=15, value=total_sum['october__sum']).style = footer_style
    sheet.cell(row=row, column=16, value=total_sum['november__sum']).style = footer_style
    sheet.cell(row=row, column=17, value=total_sum['december__sum']).style = footer_style

    # Таблица в разрезе БП:
    sheet = workbook.worksheets[1]

    sheet.cell(row=4, column=1, value=f'"__" _______ ____ г. {org_administrator_name}').style = page_header_style
    sheet.merge_cells(start_row=4, end_row=4, start_column=1, end_column=19)

    sheet.cell(row=9, column=1, value=f'на {consolidation.end.year} год').style = document_title_style
    sheet.merge_cells(start_row=9, end_row=9, start_column=1, end_column=19)

    sheet.cell(
        row=11, column=1, value=f'{org_administrator_name} просит внести изменения в'
    ).style = document_title_style
    sheet.merge_cells(start_row=11, end_row=11, start_column=1, end_column=19)

    row = 20
    for functional_group in functional_groups:
        admins = BudgetProgramAdministratorModel.objects.filter(
            is_active=True,
            proposal_items__in=proposal_items,
            functional_subgroup__functional_group=functional_group,
        ).order_by('code').distinct()
        for admin in admins:
            programs = BudgetProgramModel.objects.filter(
                is_active=True,
                proposal_items__in=proposal_items,
                budget_program_administrator=admin,
            ).order_by('code').distinct()
            for program in programs:
                sub_items = proposal_items.filter(
                        functional_group=functional_group,
                        budget_program_administrator=admin,
                        program=program,
                    ).order_by('subprogram__code', 'specificity__code')
                organizations = ContractorModel.objects.filter(
                    is_active=True,
                    accounting_reports__in=sub_items.values_list('report', flat=True).distinct()
                ).order_by('name').distinct()
                aggregate_program = sub_items.aggregate(
                    Sum('january'),
                    Sum('february'),
                    Sum('march'),
                    Sum('april'),
                    Sum('may'),
                    Sum('june'),
                    Sum('july'),
                    Sum('august'),
                    Sum('september'),
                    Sum('october'),
                    Sum('november'),
                    Sum('december'),
                )
                program_sum = 0
                for key, value in aggregate_program.items():
                    program_sum += value
                for organization in organizations:
                    row += 1
                    sheet.cell(row=row, column=1, value=organization.name).style = title_style
                    sheet.merge_cells(start_row=row, end_row=row, start_column=1, end_column=19)
                    org_items = sub_items.filter(report__organization=organization).order_by(
                        'subprogram__code', 'specificity__code',)
                    aggregate_org = org_items.aggregate(
                        Sum('january'),
                        Sum('february'),
                        Sum('march'),
                        Sum('april'),
                        Sum('may'),
                        Sum('june'),
                        Sum('july'),
                        Sum('august'),
                        Sum('september'),
                        Sum('october'),
                        Sum('november'),
                        Sum('december'),
                    )
                    year_sum = 0
                    for org_item in org_items:
                        row += 1
                        year = org_item.january + org_item.february + org_item.march + org_item.april + org_item.may + \
                               org_item.june + org_item.july + org_item.august + org_item.september + \
                               org_item.october + org_item.november + org_item.december
                        year_sum += year
                        sheet.cell(row=row, column=1, value=functional_group.code).style = common_style
                        sheet.cell(row=row, column=2, value=admin.code).style = common_style
                        sheet.cell(row=row, column=3, value=program.code).style = common_style
                        sheet.cell(row=row, column=4, value=org_item.subprogram.code if org_item.subprogram else ''
                                   ).style = common_style
                        sheet.cell(row=row, column=5, value=org_item.specificity.code).style = common_style
                        sheet.cell(row=row, column=6, value=org_item.specificity.name).style = common_style
                        sheet.cell(row=row, column=7, value=year).style = common_style
                        sheet.cell(row=row, column=8, value=org_item.january).style = common_style
                        sheet.cell(row=row, column=9, value=org_item.february).style = common_style
                        sheet.cell(row=row, column=10, value=org_item.march).style = common_style
                        sheet.cell(row=row, column=11, value=org_item.april).style = common_style
                        sheet.cell(row=row, column=12, value=org_item.may).style = common_style
                        sheet.cell(row=row, column=13, value=org_item.june).style = common_style
                        sheet.cell(row=row, column=14, value=org_item.july).style = common_style
                        sheet.cell(row=row, column=15, value=org_item.august).style = common_style
                        sheet.cell(row=row, column=16, value=org_item.september).style = common_style
                        sheet.cell(row=row, column=17, value=org_item.october).style = common_style
                        sheet.cell(row=row, column=18, value=org_item.november).style = common_style
                        sheet.cell(row=row, column=19, value=org_item.december).style = common_style
                    row += 1
                    sheet.cell(row=row, column=1).style = footer_style
                    sheet.merge_cells(start_row=row, end_row=row, start_column=1, end_column=5)
                    sheet.cell(row=row, column=6, value='Итого по бюджетополучателю:').style = footer_style
                    sheet.cell(row=row, column=7, value=year_sum).style = footer_style
                    sheet.cell(row=row, column=8, value=aggregate_org['january__sum']).style = footer_style
                    sheet.cell(row=row, column=9, value=aggregate_org['february__sum']).style = footer_style
                    sheet.cell(row=row, column=10, value=aggregate_org['march__sum']).style = footer_style
                    sheet.cell(row=row, column=11, value=aggregate_org['april__sum']).style = footer_style
                    sheet.cell(row=row, column=12, value=aggregate_org['may__sum']).style = footer_style
                    sheet.cell(row=row, column=13, value=aggregate_org['june__sum']).style = footer_style
                    sheet.cell(row=row, column=14, value=aggregate_org['july__sum']).style = footer_style
                    sheet.cell(row=row, column=15, value=aggregate_org['august__sum']).style = footer_style
                    sheet.cell(row=row, column=16, value=aggregate_org['september__sum']).style = footer_style
                    sheet.cell(row=row, column=17, value=aggregate_org['october__sum']).style = footer_style
                    sheet.cell(row=row, column=18, value=aggregate_org['november__sum']).style = footer_style
                    sheet.cell(row=row, column=19, value=aggregate_org['december__sum']).style = footer_style
                row += 1
                sheet.cell(row=row, column=1).style = footer_style
                sheet.merge_cells(start_row=row, end_row=row, start_column=1, end_column=5)
                sheet.cell(row=row, column=6, value='Итого по программе:').style = footer_style
                sheet.cell(row=row, column=7, value=program_sum).style = footer_style
                sheet.cell(row=row, column=8, value=aggregate_program['january__sum']).style = footer_style
                sheet.cell(row=row, column=9, value=aggregate_program['february__sum']).style = footer_style
                sheet.cell(row=row, column=10, value=aggregate_program['march__sum']).style = footer_style
                sheet.cell(row=row, column=11, value=aggregate_program['april__sum']).style = footer_style
                sheet.cell(row=row, column=12, value=aggregate_program['may__sum']).style = footer_style
                sheet.cell(row=row, column=13, value=aggregate_program['june__sum']).style = footer_style
                sheet.cell(row=row, column=14, value=aggregate_program['july__sum']).style = footer_style
                sheet.cell(row=row, column=15, value=aggregate_program['august__sum']).style = footer_style
                sheet.cell(row=row, column=16, value=aggregate_program['september__sum']).style = footer_style
                sheet.cell(row=row, column=17, value=aggregate_program['october__sum']).style = footer_style
                sheet.cell(row=row, column=18, value=aggregate_program['november__sum']).style = footer_style
                sheet.cell(row=row, column=19, value=aggregate_program['december__sum']).style = footer_style
    row += 1
    sheet.cell(row=row, column=1).style = footer_style
    sheet.merge_cells(start_row=row, end_row=row, start_column=1, end_column=5)
    sheet.cell(row=row, column=6, value='Итого по заявке:').style = footer_style
    sheet.cell(row=row, column=7, value=total_sum['year__sum']).style = footer_style
    sheet.cell(row=row, column=8, value=total_sum['january__sum']).style = footer_style
    sheet.cell(row=row, column=9, value=total_sum['february__sum']).style = footer_style
    sheet.cell(row=row, column=10, value=total_sum['march__sum']).style = footer_style
    sheet.cell(row=row, column=11, value=total_sum['april__sum']).style = footer_style
    sheet.cell(row=row, column=12, value=total_sum['may__sum']).style = footer_style
    sheet.cell(row=row, column=13, value=total_sum['june__sum']).style = footer_style
    sheet.cell(row=row, column=14, value=total_sum['july__sum']).style = footer_style
    sheet.cell(row=row, column=15, value=total_sum['august__sum']).style = footer_style
    sheet.cell(row=row, column=16, value=total_sum['september__sum']).style = footer_style
    sheet.cell(row=row, column=17, value=total_sum['october__sum']).style = footer_style
    sheet.cell(row=row, column=18, value=total_sum['november__sum']).style = footer_style
    sheet.cell(row=row, column=19, value=total_sum['december__sum']).style = footer_style

    workbook.properties.identifier = str(consolidation.org_administrator_id)
    stream = BytesIO()
    workbook.save(stream)
    django_file = DjangoFile(
        file=stream,
        name=f"Заявка на изменение плана финансирования.xlsx"  # TODO придумать имя
    )
    workbook_file = File()
    workbook_file.upload = django_file
    workbook_file.is_confined = True
    workbook_file.save()
    consolidation_file = models.ConsolidationFileModel.objects.create(
        name=f"Заявка на изменение плана финансирования",  # TODO придумать имя
        original_file=workbook_file,
        file_type_id='ipf_proposal',
        sort=100
    )
    upload_for_1C_workbook = get_upload_for_1C(accounting_reports)
    upload_for_1C_workbook.properties.identifier = str(consolidation.org_administrator_id)
    stream = BytesIO()
    upload_for_1C_workbook.save(stream)
    django_file = DjangoFile(
        file=stream,
        name=f"Выгрузка для 1С.xlsx"  # TODO придумать имя
    )
    workbook_file = File()
    workbook_file.upload = django_file
    workbook_file.is_confined = True
    workbook_file.save()
    upload_for_1C_file = models.ConsolidationFileModel.objects.create(
        name=f"Выгрузка для 1С",  # TODO придумать имя
        original_file=workbook_file,
        file_type_id='ipf_proposal',
        sort=200
    )
    return [consolidation_file, upload_for_1C_file]


def get_balance(contractor, year, with_descendants=False, descendants_only=False):
    from users.utils import get_descendants_departments_related_organizations

    if with_descendants and descendants_only:
        drf_exceptions.ValidationError(
            'Невозможно выполнить условия агрегации'
        )

    result = 0
    contractors_seed = set([contractor])

    if descendants_only:
        contractors = get_descendants_departments_related_organizations(
            contractors_seed=contractors_seed,
            include_self=False
        )
    elif with_descendants:
        contractors = get_descendants_departments_related_organizations(
            contractors_seed=contractors_seed,
            include_self=True
        )
    else:
        contractors = contractors_seed

    for contractor in contractors:
        result += get_contractor_balance(contractor, year)

    return result


def get_contractor_balance(contractor, year) -> int:
    try:
        balance = models.ContractorBalanceModel.objects.get(
            is_active=True,
            contractor_id=contractor,
            year=year
        )
    except models.ContractorBalanceModel.DoesNotExist:
        balance = None

    return balance.balance if balance is not None else 0


def f2go_before_approve(report):
    if report.report_files.filter(
        is_active=True,
        original_file__isnull=True
    ).exists():
        raise drf_exceptions.ValidationError(
            'Загружены не все файлы'
        )


def create_calculation_of_changes_consolidation(consolidation):
    from accounting_reports.models import (ChangeCalculationItemModel,
                                           ChangeCalculationReportModel)
    from common.accounting_catalogs.models import BudgetSubprogramModel
    from common.catalogs.models import ContractorModel
    reports = consolidation.source_reports.filter(is_active=True, without_attachments=False)
    if not reports:
        raise drf_exceptions.ValidationError({'message': 'Консолидация не содержит отчетов'})

    change_calculation_reports = ChangeCalculationReportModel.objects.filter(
        is_active=True,
        consolidation_reports__in=reports
    )
    change_calculation_items = ChangeCalculationItemModel.objects.filter(
        is_active=True,
        report__in=change_calculation_reports
    ).order_by('program__code')  #  TODO order_by дописать

    budget_subprograms = BudgetSubprogramModel.objects.filter(
        is_active=True, change_calculation_items__in=change_calculation_items).order_by('code').distinct()

    workbook = openpyxl.Workbook()

    title_font = Font(
        name='Times New Roman',
        size=14,
        bold=True,
        italic=False,
        vertAlign=None,
        underline='none',
        strike=False,
        color='00000000'
    )
    header_font = Font(
        name='Times New Roman',
        size=11,
        bold=False,
        italic=False,
        vertAlign=None,
        underline='none',
        strike=False,
        color='00000000'
    )
    regular_font = Font(
        name='Times New Roman',
        size=8,
        bold=False,
        italic=False,
        vertAlign=None,
        underline='none',
        strike=False,
        color='00000000'
    )
    bold_font = Font(
        name='Times New Roman',
        size=8,
        bold=True,
        italic=False,
        vertAlign=None,
        underline='none',
        strike=False,
        color='00000000'
    )

    center_alignment = Alignment(
        horizontal='center',
        vertical='center',
        wrap_text=True,
        shrink_to_fit=True,
        indent=0
    )

    thin_border = Border(
        left=Side(border_style='thin', color='00000000'),
        right=Side(border_style='thin', color='00000000'),
        top=Side(border_style='thin', color='00000000'),
        bottom=Side(border_style='thin', color='00000000'),
    )

    subprogram_fill = PatternFill(fill_type='solid', start_color='92D050')
    organization_fill = PatternFill(fill_type='solid', start_color='00B0F0')
    total_fill = PatternFill(fill_type='solid', start_color='FFFF00')
    footer_fill = PatternFill(fill_type='solid', start_color='00B050')


    header_style = NamedStyle(name='header_style')
    header_style.font = header_font
    header_style.border = thin_border
    header_style.alignment = center_alignment

    bold_common_style = NamedStyle(name='bold_common_style')
    bold_common_style.font = bold_font
    bold_common_style.border = thin_border
    bold_common_style.alignment = center_alignment

    common_style = NamedStyle(name='common_style')
    common_style.font = regular_font
    common_style.border = thin_border
    common_style.alignment = center_alignment

    title_style = NamedStyle(name='title_style')
    title_style.font = title_font
    title_style.alignment = center_alignment

    total_style = NamedStyle(name='total_style')
    total_style.font = bold_font
    total_style.border = thin_border
    total_style.alignment = center_alignment
    total_style.fill = total_fill

    footer_style = NamedStyle(name='footer_style')
    footer_style.font = bold_font
    footer_style.border = thin_border
    footer_style.alignment = center_alignment
    footer_style.fill = footer_fill

    org_style = NamedStyle(name='org_style')
    org_style.font = regular_font
    org_style.border = thin_border
    org_style.alignment = center_alignment
    org_style.fill = organization_fill

    subprogram_style = NamedStyle(name='subprogram_style')
    subprogram_style.font = regular_font
    subprogram_style.border = thin_border
    subprogram_style.alignment = center_alignment
    subprogram_style.fill = subprogram_fill

    sheet = workbook.worksheets[0]
    sheet.column_dimensions['A'].width = 12
    sheet.column_dimensions['B'].width = 42
    sheet.column_dimensions['C'].width = 18
    sheet.column_dimensions['D'].width = 13
    sheet.column_dimensions['E'].width = 13
    sheet.column_dimensions['F'].width = 13
    sheet.column_dimensions['G'].width = 13
    sheet.column_dimensions['H'].width = 13
    sheet.column_dimensions['I'].width = 13
    sheet.column_dimensions['J'].width = 68

    humanized_start = get_humanized_month(consolidation.start.month)
    humanized_end = get_humanized_month(consolidation.end.month)
    if humanized_start == humanized_end:
        humanized_period = humanized_end
    else:
        humanized_period = f'{humanized_start}-{humanized_end}'
    # Заголовок:
    year_end = consolidation.end.year
    sheet.cell(
        row=3,
        column=1,
        value=f'Сводный расчет на внесение изменений в индивидуальный план финансирования по платежам '
              f'на {humanized_end} {year_end} года'
    ).style = title_style
    sheet.merge_cells(start_row=3, end_row=3, start_column=1, end_column=10)

    # Хэдер таблицы:
    sheet.cell(row=6, column=1, value='КБК расходов').style = header_style
    sheet.merge_cells(start_row=6, end_row=7, start_column=1, end_column=1)
    sheet.cell(row=6, column=2, value='Наимерование программы').style = header_style
    sheet.merge_cells(start_row=6, end_row=7, start_column=2, end_column=2)
    sheet.cell(row=6, column=3, value='Специфика').style = header_style
    sheet.merge_cells(start_row=6, end_row=7, start_column=3, end_column=3)
    sheet.cell(row=6, column=4, value=f'План ({humanized_period})').style = header_style
    sheet.merge_cells(start_row=6, end_row=6, start_column=4, end_column=5)
    sheet.cell(row=6, column=6, value=f'Факт ({humanized_period})').style = header_style
    sheet.merge_cells(start_row=6, end_row=6, start_column=6, end_column=7)
    sheet.cell(row=6, column=8, value='Отклонение (+,-)').style = header_style
    sheet.merge_cells(start_row=6, end_row=6, start_column=8, end_column=9)
    sheet.cell(row=6, column=10, value='Обоснование').style = header_style
    sheet.merge_cells(start_row=6, end_row=7, start_column=10, end_column=10)
    sheet.cell(row=7, column=4, value='к-во').style = header_style
    sheet.cell(row=7, column=5, value='сумма').style = header_style
    sheet.cell(row=7, column=6, value='к-во').style = header_style
    sheet.cell(row=7, column=7, value='сумма').style = header_style
    sheet.cell(row=7, column=8, value='к-во').style = header_style
    sheet.cell(row=7, column=9, value='сумма').style = header_style

    # группируем по подпрограммам:
    row = 7
    for budget_subprogram in budget_subprograms:
        row += 1
        sheet.cell(
            row=row,
            column=1,
            value=f"{budget_subprogram.program.budget_program_administrator.code}"
                  f"{budget_subprogram.program.code}"
                  f"{budget_subprogram.code}"
        ).style = subprogram_style
        sheet.cell(row=row, column=2, value=budget_subprogram.program.name).style = subprogram_style
        sheet.cell(row=row, column=3).style = subprogram_style
        sheet.row_dimensions[row].height = get_height_for_cell(sheet, sheet.cell(row=row, column=2), font_size=8)
        budget_subprogram_items = change_calculation_items.filter(subprogram=budget_subprogram)

        # Сумма по подпрограммам:
        aggregate_subprogram = budget_subprogram_items.aggregate(
            plan_quantity_sum=Sum('plan_quantity'),
            plan_amount_sum=Sum('plan_amount'),
            actual_quantity_sum=Sum('actual_quantity'),
            actual_amount_sum=Sum('actual_amount'),
        )
        sheet.cell(row=row, column=4, value=aggregate_subprogram['plan_quantity_sum']).style = subprogram_style
        sheet.cell(row=row, column=5, value=aggregate_subprogram['plan_amount_sum']).style = subprogram_style
        sheet.cell(row=row, column=6, value=aggregate_subprogram['actual_quantity_sum']).style = subprogram_style
        sheet.cell(row=row, column=7, value=aggregate_subprogram['actual_amount_sum']).style = subprogram_style
        sheet.cell(
            row=row, column=8, value=aggregate_subprogram['actual_quantity_sum'] - aggregate_subprogram['plan_quantity_sum']
        ).style = subprogram_style
        sheet.cell(
            row=row, column=9, value=aggregate_subprogram['actual_amount_sum'] - aggregate_subprogram['plan_amount_sum']
        ).style = subprogram_style
        sheet.cell(row=row, column=10).style = subprogram_style

        organizations = ContractorModel.objects.filter(
            is_active=True,
            accounting_reports__in=budget_subprogram_items.values_list('report', flat=True).distinct()
        ).order_by('name').distinct()

        for organization in organizations:
            row += 1
            sheet.cell(row=row, column=1, value=organization.code_gu).style = common_style
            sheet.cell(row=row, column=2, value=organization.name).style = org_style
            org_height = get_height_for_cell(sheet, sheet.cell(row=row, column=2), font_size=8)
            org_budget_subprogram_items = budget_subprogram_items.filter(report__organization=organization)
            save_row = row
            accumulate_row_dimension = 0
            for org_budget_subprogram_item in org_budget_subprogram_items:
                sheet.cell(row=row, column=3, value=org_budget_subprogram_item.specificity.code).style = common_style
                plan_quantity = org_budget_subprogram_item.plan_quantity
                sheet.cell(row=row, column=4, value=plan_quantity).style = common_style
                plan_amount = org_budget_subprogram_item.plan_amount
                sheet.cell(row=row, column=5, value=plan_amount).style = common_style
                actual_quantity = org_budget_subprogram_item.actual_quantity
                sheet.cell(row=row, column=6, value=actual_quantity).style = common_style
                actual_amount = org_budget_subprogram_item.actual_amount
                sheet.cell(row=row, column=7, value=actual_amount).style = common_style
                sheet.cell(row=row, column=8, value=actual_quantity - plan_quantity).style = common_style
                sheet.cell(row=row, column=9, value=actual_amount - actual_quantity).style = common_style
                sheet.cell(
                    row=row, column=10, value=org_budget_subprogram_item.rationale.rationale
                ).style = common_style
                row_dimension = get_height_for_cell(sheet, sheet.cell(row=row, column=10), font_size=8)
                sheet.row_dimensions[row].height = row_dimension
                accumulate_row_dimension = accumulate_row_dimension + row_dimension
                row += 1
            # Подбираев высоту строки с именем организации:
            if org_height > accumulate_row_dimension:
                sheet.row_dimensions[save_row].height = 2 * org_height - accumulate_row_dimension

            sheet.merge_cells(start_row=save_row, end_row=row-1, start_column=1, end_column=1)
            sheet.merge_cells(start_row=save_row, end_row=row-1, start_column=2, end_column=2)
            # Сумма по организациям:
            aggregate_org = org_budget_subprogram_items.aggregate(
                plan_quantity_sum=Sum('plan_quantity'),
                plan_amount_sum=Sum('plan_amount'),
                actual_quantity_sum=Sum('actual_quantity'),
                actual_amount_sum=Sum('actual_amount'),
            )
            sheet.cell(row=row, column=1).style = total_style
            sheet.cell(row=row, column=2, value='Итого').style = total_style
            sheet.cell(row=row, column=3).style = total_style
            sheet.cell(row=row, column=4, value=aggregate_org['plan_quantity_sum']).style = total_style
            sheet.cell(row=row, column=5, value=aggregate_org['plan_amount_sum']).style = total_style
            sheet.cell(row=row, column=6, value=aggregate_org['actual_quantity_sum']).style = total_style
            sheet.cell(row=row, column=7, value=aggregate_org['actual_amount_sum']).style = total_style
            sheet.cell(
                row=row, column=8, value=aggregate_org['actual_quantity_sum']-aggregate_org['plan_quantity_sum']
            ).style = total_style
            sheet.cell(
                row=row, column=9, value=aggregate_org['actual_amount_sum']-aggregate_org['plan_amount_sum']
            ).style = total_style
            sheet.cell(row=row, column=10).style = total_style
    # Сумма по всей таблице:
    row += 1
    sheet.cell(row=row, column=1).style = footer_style
    sheet.cell(row=row, column=2, value='Итого по программам').style = footer_style
    aggregate_total = change_calculation_items.aggregate(
        plan_quantity_sum=Sum('plan_quantity'),
        plan_amount_sum=Sum('plan_amount'),
        actual_quantity_sum=Sum('actual_quantity'),
        actual_amount_sum=Sum('actual_amount'),
    )
    sheet.cell(row=row, column=3).style = footer_style
    sheet.cell(row=row, column=4, value=aggregate_total['plan_quantity_sum']).style = footer_style
    sheet.cell(row=row, column=5, value=aggregate_total['plan_amount_sum']).style = footer_style
    sheet.cell(row=row, column=6, value=aggregate_total['actual_quantity_sum']).style = footer_style
    sheet.cell(row=row, column=7, value=aggregate_total['actual_amount_sum']).style = footer_style
    sheet.cell(
        row=row, column=8, value=aggregate_total['actual_quantity_sum']-aggregate_total['plan_quantity_sum']
               ).style = footer_style
    sheet.cell(
        row=row, column=9, value=aggregate_total['actual_amount_sum']-aggregate_total['plan_amount_sum']
    ).style = footer_style
    sheet.cell(row=row, column=10).style = footer_style

    sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
    sheet.page_setup.fitToPage = True
    sheet.page_setup.fitToHeight = False
    sheet.page_margins = PageMargins(left=0.75, right=0.25, top=0.25, bottom=0.25)

    sheet.print_options.horizontalCentered = True
    sheet.print_area = f'A1:J{row}'

    workbook.properties.identifier = str(consolidation.org_administrator_id)
    stream = BytesIO()
    workbook.save(stream)
    django_file = DjangoFile(
        file=stream,
        name=f"Расчет на внесение изменений в индивидуальный план финансирования.xlsx"  # TODO придумать имя
    )
    workbook_file = File()
    workbook_file.upload = django_file
    workbook_file.is_confined = True
    workbook_file.save()
    consolidation_file = models.ConsolidationFileModel.objects.create(
        name=f"Расчет на внесение изменений в индивидуальный план финансирования",  # TODO придумать имя
        original_file=workbook_file,
        file_type_id='ipf_proposal',
        sort=100
    )
    return [consolidation_file, ]


def update_saved_reports():
    from common.catalogs.models import ContractorModel

    MONTHS = (
        {
            'name': 'Январь',
            'start': '2024-01-01',
            'end': '2024-01-31'
        },
        {
            'name': 'Февраль',
            'start': '2024-02-01',
            'end': '2024-02-29'
        },
        {
            'name': 'Март',
            'start': '2024-03-01',
            'end': '2024-03-31'
        },
        # {
        #     'name': 'Апрель',
        #     'start': '2024-04-01',
        #     'end': '2024-04-30'
        # }
    )
    result = {
        'summary': dict(),
        'results': dict(),
        'error': dict()

    }
    organizations = ContractorModel.objects.filter(
        is_active=True
    )

    f2go_report_form = models.ReportFormModel.objects.get(
        is_active=True,
        code='f2go'
    )

    for organization in organizations:
        result['results'][f'{str(organization.id)} - {str(organization)}'] = dict()
        for period in MONTHS:
            consolidations = models.ConsolidationModel.objects.filter(
                is_active=True,
                report_form_id=f2go_report_form.id,
                source_reports__contractor_id=organization.id,
                start=period['start'],
                end=period['end']
            )
            if consolidations:
                if len(consolidations) == 1:
                    try:
                        report = consolidations[0].source_reports.get(
                            is_active=True,
                            contractor_id=organization.id
                        )
                    except Exception as e:
                        result['error'][str(consolidation.id)] = {
                            'organization': f'{str(organization.id)} - {str(organization)}',
                            'class': type(e).__name__,
                            'message': str(e)
                        }
                        result['results'][f'{str(organization.id)} - {str(organization)}'][period['name']] = {
                            'error_class': type(e).__name__,
                            'message': str(e),
                        }
                    else:
                        report_file = report.report_files.get(file_type_id='f2go')
                        file = report_file.original_file
                        if file is not None:
                            save_f2go_report(file, report)
                            result['results'][f'{str(organization.id)} - {str(organization)}'][period['name']] = f'К-я одна, файл - {str(file.id)}'
                        else:
                            result['results'][f'{str(organization.id)} - {str(organization)}'][period['name']] = 'К-я одна, файл в отчете не найден'
                else:
                    consolidation = consolidations.filter(
                        Q(org_administrator_id=organization.id) &
                        Q(source_reports__report_files__file_type_id='f2go') &
                        Q(source_reports__report_files__original_file__isnull=False)
                    ).first()
                    if consolidation:
                        try:
                            report = consolidation.source_reports.get(
                                is_active=True,
                                contractor_id=organization.id
                            )
                        except Exception as e:
                            result['error'][str(consolidation.id)] = {
                                'organization': f'{str(organization.id)} - {str(organization)}',
                                'class': type(e).__name__,
                                'message': str(e)
                            }
                            result['results'][f'{str(organization.id)} - {str(organization)}'][period['name']] = {
                                'error_class': type(e).__name__,
                                'message': str(e),
                            }
                        else:
                            report_file = report.report_files.get(file_type_id='f2go')
                            file = report_file.original_file
                            if file is not None:
                                save_f2go_report(file, report)
                                result['results'][f'{str(organization.id)} - {str(organization)}'][period['name']] = f'Орг-админ, файл - {str(file.id)}'
                            else:
                                result['results'][f'{str(organization.id)} - {str(organization)}'][period['name']] = 'Орг-админ, файл в отчете не найден'
                    else:
                        consolidation = consolidations.filter(
                            Q(source_reports__report_files__file_type_id='f2go') & Q (source_reports__report_files__original_file__isnull=False)
                        ).first()
                        if consolidation:
                            try:
                                report = consolidation.source_reports.get(
                                    is_active=True,
                                    contractor_id=organization.id
                                )
                            except Exception as e:
                                result['error'][str(consolidation.id)] = {
                                    'organization': f'{str(organization.id)} - {str(organization)}',
                                    'class': type(e).__name__,
                                    'message': str(e)
                                }
                            else:
                                report_file = report.report_files.get(file_type_id='f2go')
                                file = report_file.original_file
                                if file is not None:
                                    save_f2go_report(file, report)
                                    result['results'][f'{str(organization.id)} - {str(organization)}'][period['name']] = f'Конс-й несколько, файл - {str(file.id)}'
                                else:
                                    result['results'][f'{str(organization.id)} - {str(organization)}'][period['name']] = 'Конс-й несколько, файлы в отчетах не найдены'
                        else:
                            result['results'][f'{str(organization.id)} - {str(organization)}'][period['name']] = 'Конс-й несколько, файлы в отчетах не найдены'
            else:
                result['results'][f'{str(organization.id)} - {str(organization)}'][period['name']] = 'Консолидаций не найдено в этом месяце'

    with open('update_reports_result.txt', 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, indent=4)

    return


def get_risk_matrix_template_workbook_pyexcelerate():
    """
    Возвращает Workbook() с шаблоном карты рисков (pyexcelerate)
    """
    wb = pyexcelerate.Workbook()
    ws = wb.new_sheet('Карта риска')
    ws.set_col_style(1, pyexcelerate.Style(size=5))
    ws.set_col_style(2, pyexcelerate.Style(size=23))
    ws.set_col_style(3, pyexcelerate.Style(size=18))
    ws.set_col_style(4, pyexcelerate.Style(size=12))
    ws.set_col_style(5, pyexcelerate.Style(size=12))
    ws.set_col_style(6, pyexcelerate.Style(size=12))
    ws.set_col_style(7, pyexcelerate.Style(size=15))
    ws.set_col_style(8, pyexcelerate.Style(size=12))
    ws.set_col_style(9, pyexcelerate.Style(size=12))
    ws.set_col_style(10, pyexcelerate.Style(size=14))
    ws.set_col_style(11, pyexcelerate.Style(size=14))
    ws.set_col_style(12, pyexcelerate.Style(size=14))
    ws.set_col_style(13, pyexcelerate.Style(size=14))
    ws.set_col_style(14, pyexcelerate.Style(size=18))
    ws.set_col_style(15, pyexcelerate.Style(size=18))
    ws.set_col_style(16, pyexcelerate.Style(size=18))
    ws.set_row_style(3, pyexcelerate.Style(size=105))

    data = (
        ('№',
         '№ обращения',
         'Вид обращения',
         'Масштаб поднимаемого вопроса',
         'Относится ли заявитель к социально уязвимым слоям населения',
         'Упоминание о ранее направленных обращениях по поднимаемым вопросам',
         'Выражение намерения к публичным протестам, акцентрирование отчаянного положения заявителя, угрозы к крайним мерам',
         'Является ли автор публичной персоной',
         'Срочность решения вопроса: Экстренный, плановый',
         'Упоминание в обращениях о тяжелых последствиях в случае бездействия государственных органов',
         'Сообщает, что ранее был на личном приеме либо о поручении вышестоящих органов решить вопрос',
         'Заявитель сообщает, что не может решить вопрос ввиду уникальности своей проблемы',
         'Упоминание о предвзятости госорганов или коррупционной составляющей в действиях или бездействии госорганов',
         'Итоговое значение',
         'Направлено на рассмотрение Первый руководитель/заместители, руководитель аппарата Примечание: данный критерий не входит в итоговую оценку',
         'Характер вопроса',
         ),
    )
    ws.range('A3', 'P3').value = data
    ws.range('A3', 'P3').style.alignment = pyexcelerate.Alignment(horizontal='left', vertical='top', wrap_text=True,)
    ws.range('A3', 'P3').style.font = pyexcelerate.Font(size=10, family='Times', bold=True,)
    ws.range('A3', 'P3').style.borders.top.style = '_'
    ws.range('A3', 'P3').style.borders.left.style = '_'
    ws.range('A3', 'P3').style.borders.bottom.style = '_'
    ws.range('A3', 'P3').style.borders.right.style = '_'
    return wb


def get_risk_matrix_template_workbook() -> openpyxl.Workbook:
    """
    Возвращает Workbook() с шаблоном карты рисков
    """
    workbook = openpyxl.Workbook()
    # workbook.properties.creator = get_current_authenticated_profile().full_name
    sheet = workbook.active

    sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0

    sheet.row_dimensions[1].height = 60
    sheet.row_dimensions[2].height = 20
    sheet.row_dimensions[3].height = 146
    sheet.column_dimensions['A'].width = 5
    sheet.column_dimensions['B'].width = 23
    sheet.column_dimensions['C'].width = 18
    sheet.column_dimensions['D'].width = 12
    sheet.column_dimensions['E'].width = 12
    sheet.column_dimensions['F'].width = 12
    sheet.column_dimensions['G'].width = 15
    sheet.column_dimensions['H'].width = 12
    sheet.column_dimensions['I'].width = 12
    sheet.column_dimensions['J'].width = 14
    sheet.column_dimensions['K'].width = 14
    sheet.column_dimensions['L'].width = 14
    sheet.column_dimensions['M'].width = 14
    sheet.column_dimensions['N'].width = 18
    sheet.column_dimensions['O'].width = 18
    sheet.column_dimensions['P'].width = 18

    top_table_style = NamedStyle(name="top_table_style")

    times_new_roman_10_bold_font = Font(
        name='Times New Roman',
        size=10,
        bold=True,
        italic=False,
        vertAlign=None,
        underline='none',
        strike=False,
        color='00000000'
    )
    thin_border = Border(
        left=Side(border_style='thin', color='00000000'),
        right=Side(border_style='thin', color='00000000'),
        top=Side(border_style='thin', color='00000000'),
        bottom=Side(border_style='thin', color='00000000'),
    )
    left_alignment = Alignment(
        horizontal='left',
        vertical='top',
        wrap_text=True,
        shrink_to_fit=True,
        indent=0
    )

    top_table_style.alignment = left_alignment
    top_table_style.border = thin_border
    top_table_style.font = times_new_roman_10_bold_font

    sheet.cell(3, 1, '№').style = top_table_style
    sheet.cell(3, 2, '№ обращения').style = top_table_style
    sheet.cell(3, 3, 'Вид обращения').style = top_table_style
    sheet.cell(3, 4, 'Масштаб поднимаемого вопроса').style = top_table_style
    sheet.cell(3, 5, 'Относится ли заявитель к социально уязвимым слоям населения').style = top_table_style
    sheet.cell(3, 6, 'Упоминание о ранее направленных обращениях по поднимаемым вопросам').style = top_table_style
    sheet.cell(3, 7, 'Выражение намерения к публичным протестам, акцентрирование отчаянного положения заявителя, угрозы к крайним мерам').style = top_table_style
    sheet.cell(3, 8, 'Является ли автор публичной персоной').style = top_table_style
    sheet.cell(3, 9, 'Срочность решения вопроса: Экстренный, плановый').style = top_table_style
    sheet.cell(3, 10, 'Упоминание в обращениях о тяжелых последствиях в случае бездействия государственных органов').style = top_table_style
    sheet.cell(3, 11, 'Сообщает, что ранее был на личном приеме либо о поручении вышестоящих органов решить вопрос').style = top_table_style
    sheet.cell(3, 12, 'Заявитель сообщает, что не может решить вопрос ввиду уникальности своей проблемы').style = top_table_style
    sheet.cell(3, 13, 'Упоминание о предвзятости госорганов или коррупционной составляющей в действиях или бездействии госорганов').style = top_table_style
    sheet.cell(3, 14, 'Итоговое значение').style = top_table_style
    sheet.cell(3, 15, 'Направлено на рассмотрение Первый руководитель/заместители, руководитель аппарата Примечание: данный критерий не входит в итоговую оценку').style = top_table_style
    sheet.cell(3, 16, 'Характер вопроса').style = top_table_style

    return workbook


def get_template_personal_reception() -> openpyxl.Workbook:
    """
    Возвращает Workbook() с шаблоном отчета о проведении личного према
    """
    workbook = openpyxl.Workbook()
    workbook.properties.creator = get_current_authenticated_profile().full_name
    sheet = workbook.active

    sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0

    sheet.row_dimensions[2].height = 20
    sheet.row_dimensions[3].height = 20
    sheet.row_dimensions[5].height = 20
    sheet.row_dimensions[6].height = 20
    sheet.row_dimensions[11].height = 45
    sheet.row_dimensions[12].height = 270
    sheet.column_dimensions['A'].width = 53
    sheet.column_dimensions['B'].width = 25
    sheet.column_dimensions['C'].width = 25
    sheet.column_dimensions['D'].width = 25
    sheet.column_dimensions['E'].width = 25
    sheet.column_dimensions['F'].width = 27
    sheet.column_dimensions['G'].width = 25
    sheet.column_dimensions['H'].width = 27
    sheet.column_dimensions['I'].width = 10
    sheet.column_dimensions['J'].width = 10
    sheet.column_dimensions['K'].width = 10
    sheet.column_dimensions['L'].width = 10
    sheet.column_dimensions['M'].width = 10

    rows_2_and_3_style = NamedStyle(name="rows_2_and_3_style")
    row_12_style = NamedStyle(name="row_12_style")
    table_top_style = NamedStyle(name="table_top_style")

    times_new_roman_16_bold_font = Font(
        name='Times New Roman',
        size=16,
        bold=True,
        italic=False,
        vertAlign=None,
        underline='none',
        strike=False,
        color='00000000'
    )
    times_new_roman_14_bold_font = Font(
        name='Times New Roman',
        size=14,
        bold=True,
        italic=False,
        vertAlign=None,
        underline='none',
        strike=False,
        color='00000000'
    )
    thin_border = Border(
        left=Side(border_style='thin', color='00000000'),
        right=Side(border_style='thin', color='00000000'),
        top=Side(border_style='thin', color='00000000'),
        bottom=Side(border_style='thin', color='00000000'),
    )
    center_alignment = Alignment(
        horizontal='center',
        vertical='center',
        wrap_text=True,
        shrink_to_fit=True,
        indent=0
    )
    center_alignment_rotated = Alignment(
        text_rotation=90,
        horizontal='center',
        vertical='bottom',
        wrap_text=True,
        shrink_to_fit=True,
        indent=0
    )
    rows_2_and_3_style.alignment = center_alignment
    rows_2_and_3_style.font = times_new_roman_16_bold_font

    table_top_style.alignment = center_alignment
    table_top_style.font = times_new_roman_14_bold_font
    table_top_style.border = thin_border
    row_12_style.alignment = center_alignment_rotated
    row_12_style.font = times_new_roman_14_bold_font
    row_12_style.border = thin_border

    sheet.cell(2, 1, 'Сведения').style = rows_2_and_3_style
    sheet.merge_cells(start_column=1, end_column=13, start_row=2, end_row=2)
    sheet.cell(3, 1, 'о проведении личного приема в').style = rows_2_and_3_style
    sheet.merge_cells(start_column=1, end_column=13, start_row=3, end_row=3)
    sheet.cell(11, 1, '').style = table_top_style
    sheet.merge_cells(start_column=1, end_column=1, start_row=11, end_row=12)
    sheet.cell(11, 2, 'Количество проведенных приемов руководством за период (руководители, заместители)').style = table_top_style
    sheet.merge_cells(start_column=2, end_column=2, start_row=11, end_row=12)
    sheet.cell(11, 3, 'Всего принято человек').style = table_top_style
    sheet.merge_cells(start_column=3, end_column=3, start_row=11, end_row=12)
    sheet.cell(11, 4, 'В том числе количество принятых граждан на личном приеме первых руководителей').style = table_top_style
    sheet.merge_cells(start_column=4, end_column=4, start_row=11, end_row=12)
    sheet.cell(11, 5, 'В том числе количество принятых граждан на личном приеме заместителей').style = table_top_style
    sheet.merge_cells(start_column=5, end_column=5, start_row=11, end_row=12)
    sheet.cell(11, 6, 'Количество очередников за отчетный период').style = table_top_style
    sheet.merge_cells(start_column=6, end_column=6, start_row=11, end_row=12)
    sheet.cell(11, 7, 'Количество дней в очереди').style = table_top_style
    sheet.merge_cells(start_column=7, end_column=7, start_row=11, end_row=12)
    sheet.cell(11, 8, 'Завершенные карточки приема').style = table_top_style
    sheet.merge_cells(start_column=8, end_column=8, start_row=11, end_row=12)
    sheet.cell(11, 9, 'Социальный статус граждан, принятых на личном приеме за период').style = table_top_style
    sheet.merge_cells(start_column=9, end_column=13, start_row=11, end_row=11)
    sheet.cell(12, 9, ' Пенсионеры').style = row_12_style
    sheet.cell(12, 10, ' Многодетные').style = row_12_style
    sheet.cell(12, 11, ' Инвалиды, самозанятые, безработные').style = row_12_style
    sheet.cell(12, 12, ' Предприниматели').style = row_12_style
    sheet.cell(12, 13, ' Работники бюдж. и внебюдж. сферы').style = row_12_style

    return workbook


def get_template_verification_act() -> openpyxl.Workbook:
    """
    Возвращает Workbook() с шаблоном Акта сверки
    """
    workbook = openpyxl.Workbook()
    workbook.properties.creator = get_current_authenticated_profile().full_name
    sheet = workbook.active

    sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0

    sheet.row_dimensions[2].height = 20
    sheet.row_dimensions[3].height = 20
    sheet.row_dimensions[5].height = 20
    sheet.row_dimensions[6].height = 20
    sheet.row_dimensions[11].height = 70
    sheet.row_dimensions[12].height = 25
    sheet.column_dimensions['A'].width = 53
    sheet.column_dimensions['B'].width = 35
    sheet.column_dimensions['C'].width = 35
    sheet.column_dimensions['D'].width = 35
    sheet.column_dimensions['E'].width = 35
    sheet.column_dimensions['F'].width = 35
    sheet.column_dimensions['G'].width = 35

    rows_2_and_3_style = NamedStyle(name="rows_2_and_3_style")
    cell_A6_style = NamedStyle(name="cell_A6_style")
    cell_G6_style = NamedStyle(name="cell_G6_style")
    row_12_style = NamedStyle(name="row_12_style")
    table_top_style = NamedStyle(name="table_top_style")

    times_new_roman_16_bold_font = Font(
        name='Times New Roman',
        size=16,
        bold=True,
        italic=False,
        vertAlign=None,
        underline='none',
        strike=False,
        color='00000000'
    )
    times_new_roman_14_bold_font = Font(
        name='Times New Roman',
        size=14,
        bold=True,
        italic=False,
        vertAlign=None,
        underline='none',
        strike=False,
        color='00000000'
    )
    thin_border = Border(
        left=Side(border_style='thin', color='00000000'),
        right=Side(border_style='thin', color='00000000'),
        top=Side(border_style='thin', color='00000000'),
        bottom=Side(border_style='thin', color='00000000'),
    )
    left_top_right_thin_border = Border(
        left=Side(border_style='thin', color='00000000'),
        right=Side(border_style='thin', color='00000000'),
        top=Side(border_style='thin', color='00000000')
    )
    left_alignment = Alignment(
        horizontal='left',
        vertical='center',
        wrap_text=True,
        shrink_to_fit=True,
        indent=0
    )
    right_alignment = Alignment(
        horizontal='right',
        vertical='center',
        wrap_text=True,
        shrink_to_fit=True,
        indent=0
    )
    center_alignment = Alignment(
        horizontal='center',
        vertical='center',
        wrap_text=True,
        shrink_to_fit=True,
        indent=0
    )
    rows_2_and_3_style.alignment = center_alignment
    rows_2_and_3_style.font = times_new_roman_16_bold_font
    cell_A6_style.alignment = left_alignment
    cell_A6_style.font = times_new_roman_16_bold_font
    cell_G6_style.alignment = right_alignment
    cell_G6_style.font = times_new_roman_16_bold_font
    table_top_style.alignment = center_alignment
    table_top_style.font = times_new_roman_14_bold_font
    table_top_style.border = thin_border
    row_12_style.alignment = center_alignment
    row_12_style.font = times_new_roman_14_bold_font
    row_12_style.border = left_top_right_thin_border

    sheet.cell(2, 1, 'АКТ СВЕРКИ').style = rows_2_and_3_style
    sheet.merge_cells(start_column=1, end_column=7, start_row=2, end_row=2)
    sheet.cell(3, 1, 'по обращениям физических и юридических лиц, поступивших в').style = rows_2_and_3_style
    sheet.merge_cells(start_column=1, end_column=7, start_row=3, end_row=3)
    sheet.cell(6, 1, 'г.______________________').style = cell_A6_style
    sheet.cell(6, 7, '\"___\"_________20___г.').style = cell_G6_style
    sheet.cell(11, 1, 'Наименование').style = table_top_style
    sheet.cell(11, 2, 'Поступило обращений, сообщений, запросов, откликов, предложений за отчетный период').style = table_top_style
    sheet.cell(11, 3, 'Рассмотрено обращений, сообщений, запросов, откликов, предложений').style = table_top_style
    sheet.cell(11, 4, 'Направлено в другие органы для рассмотрения по компетенции').style = table_top_style
    sheet.cell(11, 5, 'Рассмотрено с нарушением срока').style = table_top_style
    sheet.cell(11, 6, 'Направлено в другие органы по компетенции с нарушением срока').style = table_top_style
    sheet.cell(11, 7, 'Остаток на конец отчетного периода').style = table_top_style
    sheet.cell(12, 1, 'Всего по базе «е-Otinish»').style = row_12_style
    sheet.cell(12, 2, None).style = row_12_style
    sheet.cell(12, 3, None).style = row_12_style
    sheet.cell(12, 4, None).style = row_12_style
    sheet.cell(12, 5, None).style = row_12_style
    sheet.cell(12, 6, None).style = row_12_style
    sheet.cell(12, 7, None).style = row_12_style

    return workbook


def get_template_citizen_inquiries() -> openpyxl.Workbook:
    """
    Возвращает Workbook() с шаблоном отчета Ф2ГО
    """
    workbook = openpyxl.Workbook()
    workbook.properties.creator = get_current_authenticated_profile().full_name
    sheet = workbook.active

    sheet.page_setup.orientation = sheet.ORIENTATION_PORTRAIT
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.fitToWidth = 0
    sheet.page_setup.fitToHeight = 1

    sheet.column_dimensions['A'].width = 42

    cell_A1_style = NamedStyle(name="cell_A1_style")
    calibri_bold_12_style = NamedStyle(name="calibri_bold_12_style")
    calibri_normal_12_style = NamedStyle(name="calibri_normal_12_style")
    calibri_bold_12_red_style = NamedStyle(name="calibri_bold_12_red_style")
    calibri_normal_12_red_style = NamedStyle(name="calibri_normal_12_red_style")

    calibri_10_bold_font = Font(
        name='Calibri',
        size=10,
        bold=True,
        italic=False,
        vertAlign=None,
        underline='none',
        strike=False,
        color='00000000'
    )
    calibri_12_bold_font = Font(
        name='Calibri',
        size=12,
        bold=True,
        italic=False,
        vertAlign=None,
        underline='none',
        strike=False,
        color='00000000'
    )
    calibri_12_bold_red_font = Font(
        name='Calibri',
        size=12,
        bold=True,
        italic=False,
        vertAlign=None,
        underline='none',
        strike=False,
        color='00FF0000'
    )
    calibri_12_normal_font = Font(
        name='Calibri',
        size=12,
        bold=False,
        italic=False,
        vertAlign=None,
        underline='none',
        strike=False,
        color='00000000'
    )
    calibri_12_normal_red_font = Font(
        name='Calibri',
        size=12,
        bold=False,
        italic=False,
        vertAlign=None,
        underline='none',
        strike=False,
        color='00FF0000'
    )
    thin_border = Border(
        left=Side(border_style='thin', color='00000000'),
        right=Side(border_style='thin', color='00000000'),
        top=Side(border_style='thin', color='00000000'),
        bottom=Side(border_style='thin', color='00000000'),
    )
    left_alignment = Alignment(
        horizontal='left',
        vertical='center',
        wrap_text=True,
        shrink_to_fit=True,
        indent=0
    )
    center_alignment = Alignment(
        horizontal='center',
        vertical='center',
        wrap_text=True,
        shrink_to_fit=True,
        indent=0
    )
    cell_A1_style.font = calibri_10_bold_font
    cell_A1_style.alignment = center_alignment
    cell_A1_style.border = thin_border
    calibri_bold_12_style.font = calibri_12_bold_font
    calibri_bold_12_style.alignment = left_alignment
    calibri_bold_12_style.border = thin_border
    calibri_normal_12_style.font = calibri_12_normal_font
    calibri_normal_12_style.alignment = left_alignment
    calibri_normal_12_style.border = thin_border
    calibri_bold_12_red_style.font = calibri_12_bold_red_font
    calibri_bold_12_red_style.alignment = left_alignment
    calibri_bold_12_red_style.border = thin_border
    calibri_normal_12_red_style.font = calibri_12_normal_red_font
    calibri_normal_12_red_style.alignment = left_alignment
    calibri_normal_12_red_style.border = thin_border

    def get_style(row_number):
        if row_number in (1,):
            return cell_A1_style
        if row_number in (2, 6, 11, 17, 26, 35, 54, 63, 64):
            return calibri_bold_12_style
        if row_number in (46, 47):
            return calibri_bold_12_red_style
        if row_number in (53, 56):
            return calibri_normal_12_red_style
        return calibri_normal_12_style

    def get_row_dimension(row_number):
        if row_number == 63:
            return 47
        if row_number in (36, 38, 39, 40, 47, 64):
            return 32
        return 15.6

    cell_values = [
        'Категории',
        'ВСЕГО ЗАРЕГИСТРИРОВАНО ОБРАЩЕНИЙ',
        'в том числе',
        'дубликаты',
        'обжалования',
        'источники поступления:',
        'почта',
        'нарочно',
        'на приеме граждан',
        'ИС "Е-өтініш"',
        'по видам:',
        'заявления',
        'жалобы',
        'предложения',
        'прочие (сообщение, запрос, отклик)',
        'поступивших непосредственно в ГО',
        'перенаправленных из:',
        'Администрации Президента РК',
        'Канцелярии Премьер-Министра РК',
        'Канцелярии Первого Президента РК',
        'ЦГО',
        'МИО',
        'территориальных ведомств',
        'подведомственных организаций',
        'прочих организаций',
        'перенаправленных частично из:',
        'Администрации Президента РК',
        'Канцелярии Премьер-Министра РК',
        'Канцелярии Первого Президента РК',
        'ЦГО',
        'МИО',
        'территориальных ведомств',
        'подведомственных организаций',
        'прочих организаций',
        'РАССМОТРЕНО в ГО',
        'принятое по итогам рассмотрения решение:',
        'принят административный акт',
        'принят благоприятный административный акт',
        'принят обременительный административный акт',
        'в т. ч. по которым заслушивание не проводилось',
        'дан ответ автору',
        'прекращено (ст. 89; 70 АППК)',
        'принято к сведению (ст. 89 АППК)',
        'другое',
        'сроки рассмотрения продлены:',
        'рассмотрены с нарушением срока:',
        'ОБЖАЛОВАНО АДМИНИСТРАТИВНЫХ АКТОВ',
        'в ГО, в том числе',
        'адм. акт отменен и принят новый акт',
        'в суде, в том числе',
        'иск удовлетворен',
        'ОСТАТОК НА РАССМОТРЕНИИ:',
        'в т.ч. с нарушением срока:',
        'ПЕРЕНАПРАВЛЕНО, В ТОМ ЧИСЛЕ В',
        'Администрацию Президента РК',
        'Канцелярию Премьер-Министра РК',
        'Канцелярию Первого Президента РК',
        'ЦГО',
        'МИО',
        'территориальные ведомства',
        'подведомственные организации',
        'прочие организации',
        'ПЕРЕНАПРАВЛЕННЫХ С НАРУШЕНИЕМ УСТАНОВЛЕННОГО АППК 3-Х ДНЕВНОГО СРОКА',
        'ПЕРЕНАПРАВЛЕНО ЧАСТИЧНО, В ТОМ ЧИСЛЕ В',
        'Администрацию Президента РК',
        'Канцелярию Премьер-Министра РК',
        'Канцелярию Первого Президента РК',
        'ЦГО',
        'МИО',
        'территориальные ведомства',
        'подведомственные организации',
        'прочие организации'
    ]

    for row_number, value in enumerate(cell_values, start=1):
        style = get_style(row_number)
        bg_color = get_bg_color(row_number)
        row_dimension = get_row_dimension(row_number)
        sheet.cell(row_number, 1, value).style = style
        if row_number != 1:
            sheet.row_dimensions[row_number].height = row_dimension
            sheet.cell(row_number, 1).fill = PatternFill(
                fill_type='solid',
                start_color=bg_color
            )

    return workbook


def is_february_consolidation(consolidation):
    """
    Проверяет, что консолидация за февраль (от первого до
    последнего дня месяца).
    """
    if not consolidation.start or not consolidation.end:
        return False

    first_day = date(consolidation.start.year, 2, 1)
    last_day = first_day + relativedelta(day=31)

    return consolidation.start == first_day and consolidation.end == last_day


def get_previous_report(consolidation, request):
    contractor = request.data.get('contractor', None)
    if consolidation is None or contractor is None:
        return None
    report_form_list = [consolidation.report_form.code]
    if is_february_consolidation(consolidation):
        report_form_list = ['f2go_with_verification_act', 'f2go']

    reports = models.ReportModel.objects.filter(
        parent__is_active=True,
        contractor_id=request.data.get('contractor'),
        parent__report_form__code__in=report_form_list,
        parent__start__year=consolidation.start.year,
        parent__start__lt=consolidation.start
    ).distinct().order_by("-parent__start", "-created_at")

    return next((r for r in reports if hasattr(r, "disintegration")), None)
