import json
import openpyxl
import uuid
from io import BytesIO
from openpyxl.styles import Font, NamedStyle, PatternFill, Alignment, Border, Side
from math import ceil

from shapely.geometry import Polygon, Point

from django.apps import apps
from django.db import transaction
from django.db.models import F, ExpressionWrapper, DurationField
from django.utils import timezone

from rest_framework.serializers import DateTimeField
from rest_framework import exceptions as drf_exceptions

from bkz3.settings import GOODS_PRICE_BY_CATALOG, SOCKETIO_SYSTEM_CHANNEL, FRONTEND_URL, COMPANY_NAME

from common.current_profile.middleware import get_current_authenticated_profile
from common.redis import socketio_redis

from bpms.event_calendar.models import EventCalendarModel, CalendarModel
from bpms.chat.models import ChatModel, MessageModel


def get_current_price_type():
    """
    Возвращает текущий тип цен текущего пользователя.
    """
    user = get_current_authenticated_profile()
    return get_user_price_type(user)


def get_user_price_type(user):
    """
    Возвращает тип цен пользователя исходя из текущего контрагента пользователя.
    """
    contractor = user.contractors.first()
    PriceTypeModel = apps.get_model('catalogs', 'PriceTypeModel')  # noqa
    if contractor:
        return getattr(
            getattr(
                contractor.contracts.filter(default=True).first(), 'contract', None
            ), 'price_type', PriceTypeModel.objects.filter(default=True).first()
        )
    else:
        return PriceTypeModel.objects.filter(default=True).first()


def get_price_by_catalog_for_serializer(goods):
    """
    Получаем цену по категории при сериализации
    """
    price = goods.get_price()
    if GOODS_PRICE_BY_CATALOG:
        price = goods.price_by_catalog
    return price

def get_contractors_filtered_list(queryset, search_string):
    """
    Фильтруем список объектов queryset и оставляем только те объекты,
    которые содержат search_string в своих полях name, only_digits_phone, email или last_delivery_address
    """

    queryset = [item for item in queryset if (search_string.lower() in item.name.lower() or
                                              str(search_string) in item.only_digits_phone.lower() or
                                              search_string.lower() in item.email.lower() or
                                              search_string.lower() in item.last_delivery_address.lower())]
    return queryset


def import_units_from_csv():
    from common.catalogs.models import MeasureUnitModel
    with open('edizm.csv', 'r', newline='', encoding='utf-8') as file:
        import csv
        reader = csv.DictReader(file, delimiter=';')

        for row in reader:
            id = row['id']
            code = row['code']
            name = row['name']
            mu = MeasureUnitModel.objects.create(id=id, code=code, name=name)
            print(mu)


def create_chat(chat_author, member, is_public):

    chat_name = f'Чат для {member.full_name}'

    chat = ChatModel.objects.create(
        chat_author=chat_author,
        name=chat_name,
        is_public=is_public,
        last_sent=timezone.now(),
    )
    author_member = chat.members.create(
        user=chat_author,
        is_moderator=True
    )
    new_member = chat.members.create(
        user=member
    )
    chat_members = [author_member, new_member]
    data = json.dumps(
        {
            'event': 'chat_create_chat',
            'data': {
                "chat_uid": str(chat.chat_uid),
                "is_public": is_public,
                "members": [{
                    "user": str(each.user.id),
                    "is_moderator": each.is_moderator
                } for each in chat_members],
                "name": chat.name,
                "chat_author": str(chat.chat_author.pk),
                "last_sent": DateTimeField().to_representation(chat.last_sent),
                "new_message_count": 0,
                "is_active": True,
                "member_count": len(chat_members)
            }
        }
    )
    MessageModel.objects.create(
        chat=chat,
        message_author=chat_author,
        text='Приветствуем Вас в нашем бизнес-сообществе. '
             'Рады ответить на Ваши вопросы и оказать всестороннюю поддержку.',
        created=timezone.now()
    )
    socketio_redis.publish(SOCKETIO_SYSTEM_CHANNEL, data)


def get_nearest_event(instance):
    tasks = instance.my_tasks.filter(
        is_active=True
    ).values_list('id', flat=True)
    if tasks:
        calendars = CalendarModel.objects.filter(
            is_active=True,
            related_object_id__in=tasks
        ).values_list('id', flat=True)
    else:
        calendars = None
    if calendars:
        events = EventCalendarModel.objects.filter(
            calendar__is_active=True,
            is_active=True,
            is_finished=False,
            calendar_id__in=calendars
        )
    else:
        events = None
    if events:
        current_time = timezone.now()
        nearest_event = events.annotate(
            time_difference=ExpressionWrapper(
                F('start_at') - current_time,
                output_field=DurationField()
            )
        ).filter(
            time_difference__gte=timezone.timedelta(0)
        ).order_by(
            'time_difference'
        ).first()
        if nearest_event:
            return nearest_event
    return None


def set_uids_for_uploaded_contractors(file):
    workbook = openpyxl.load_workbook(filename=file)
    sheet = workbook.active
    row_count = sheet.max_row + 1
    for each in range(2, row_count):
        some_id = sheet.cell(row=each, column=2).value
        if some_id:
            uid = sheet.cell(row=each, column=4).value
            if not uid:
                sheet.cell(row=each, column=4, value=str(uuid.uuid4()))
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


depth_fills = [
    PatternFill(start_color='4080ff', fill_type='solid',),
    PatternFill(start_color='5898ff', fill_type='solid',),
    PatternFill(start_color='70b0ff', fill_type='solid',),
    PatternFill(start_color='88c8ff', fill_type='solid',),
    PatternFill(start_color='a0e0ff', fill_type='solid'),
    PatternFill(start_color='a8f8ff', fill_type='solid'),
    PatternFill(start_color='c0ffff', fill_type='solid'),
]

common_font = openpyxl.styles.Font(
        name='Arial',
        size=12,
        bold=False,
        italic=False,
        vertAlign=None,
        underline='none',
        strike=False,
        color='00000000'
    )
title_font = Font(
    name='Arial',
    size=12,
    bold=True,
    italic=False,
    vertAlign=None,
    underline='none',
    strike=False,
    color='00000000'
)

left_aligment = Alignment(
    horizontal='left',
    vertical='center',
    wrap_text=True,
    shrink_to_fit=True,
    indent=0,
)
center_aligment = Alignment(
    horizontal='center',
    vertical='center',
    wrap_text=True,
    shrink_to_fit=True,
    indent=0
)
thin_border = Border(
    left=Side(border_style='thin', color='00000000'),
    right=Side(border_style='thin', color='00000000'),
    top=Side(border_style='thin', color='00000000'),
    bottom=Side(border_style='thin', color='00000000'),
)
common_style = NamedStyle(name='common_style')
common_style.font = common_font
common_style.alignment = left_aligment
common_style.border = thin_border

title_style = NamedStyle(name='title_style')
title_style.font = title_font
title_style.alignment = center_aligment
title_style.border = thin_border


def block_contractor_users(root):
    # Корневая организация
    root.profiles.filter(
        is_active=True,
        user__is_active=True,
        user__is_superuser=False,
        temporary_blocked=False,
        is_support=False,
    ).update(temporary_blocked=True)
    # Подчиненные организации:
    children_relations = root.contractor_relations_parent.filter(
        is_active=True,
        relation_type_id='structural_division'
    ).select_related('contractor').order_by('created_at')
    for each in children_relations:
        # Осторожно! Работает рекурсия
        block_contractor_users(each.contractor)


def unblock_contractor_users(root):
    # Корневая организация
    profiles = root.profiles.filter(
        is_active=True,
        user__is_active=True,
        temporary_blocked=True,
        user__is_superuser=False,
        is_support=False,
    )
    profiles.update(temporary_blocked=False)
    # Подчиненные организации:
    children_relations = root.contractor_relations_parent.filter(
        is_active=True,
        relation_type_id='structural_division'
    ).select_related('contractor').order_by('created_at')
    for each in children_relations:
        # Осторожно! Работает рекурсия
        unblock_contractor_users(each.contractor)


def get_contractor_report_file(root):

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.column_dimensions['A'].width = 5
    sheet.column_dimensions['B'].width = 100
    sheet.column_dimensions['C'].width = 45
    sheet.column_dimensions['D'].width = 45
    sheet.column_dimensions['E'].width = 20
    # Хэдер
    row = 1
    sheet.cell(row=row, column=1, value='lvl').style = title_style
    sheet.cell(row=row, column=2, value='Наименование').style = title_style
    sheet.cell(row=row, column=3, value='id').style = title_style
    sheet.cell(row=row, column=4, value='parent').style = title_style
    sheet.cell(row=row, column=5, value='bin').style = title_style
    # Корневая организация
    row = 2
    cell = sheet.cell(row=row, column=1, value=0)
    cell.style = common_style
    cell.fill = depth_fills[0]
    cell.alignment = center_aligment
    cell = sheet.cell(row=row, column=2, value=root.name)
    cell.style = common_style
    cell.fill = depth_fills[0]
    sheet.row_dimensions[row].height = get_height_for_name_cell(sheet, cell)
    cell = sheet.cell(row=row, column=3, value=str(root.pk))
    cell.style = common_style
    cell.fill = depth_fills[0]
    cell = sheet.cell(row=row, column=4, value='')
    cell.style = common_style
    cell.fill = depth_fills[0]
    cell = sheet.cell(row=row, column=5, value=root.member_inn)
    cell.style = common_style
    cell.fill = depth_fills[0]
    # Подчиненные организации:
    set_children_contractors({'row': 2}, 0, root, sheet)
    # Собираем файл:
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


def set_children_contractors(state: dict, depth: int, root, sheet):
    depth += 1
    try:
        fill = depth_fills[depth]
    except IndexError:
        fill = PatternFill(fill_type='solid', start_color='ffffff', )
    children_relations = root.contractor_relations_parent.filter(
        is_active=True,
        relation_type_id='structural_division'
    ).select_related('contractor').order_by('contractor__name',)
    for each in children_relations:
        state['row'] = state['row'] + 1
        contractor = each.contractor
        cell = sheet.cell(row=state['row'], column=1, value=depth)
        cell.style = common_style
        cell.fill = fill
        cell.alignment = center_aligment
        cell = sheet.cell(row=state['row'], column=2, value=contractor.name)
        cell.style = common_style
        cell.fill = fill
        sheet.row_dimensions[state['row']].height = get_height_for_name_cell(sheet, cell)
        cell = sheet.cell(row=state['row'], column=3, value=str(contractor.pk))
        cell.style = common_style
        cell.fill = fill
        cell = sheet.cell(row=state['row'], column=4, value=str(each.contractor_parent_id))
        cell.style = common_style
        cell.fill = fill
        cell = sheet.cell(row=state['row'], column=5, value=contractor.member_inn)
        cell.style = common_style
        cell.fill = fill
        set_children_contractors(state, depth, contractor, sheet)


def get_height_for_name_cell(sheet, cell):
    height = 16
    words_count_at_one_row = sheet.column_dimensions[cell.column_letter].width / 1.2
    cell_value = cell.value
    if isinstance(cell_value, str):
        extra_cells = cell_value.split('\n')
        extra_cells_len = len(extra_cells)
        if extra_cells_len > 1:
            extra_lines = extra_cells_len
            for extra_cell in extra_cells:
                extra_lines = extra_lines + ceil(len(extra_cell) / words_count_at_one_row) - 1
            height = max(height, extra_lines * 16)
            return height
    lines = ceil(len(str(cell_value)) / words_count_at_one_row)
    height = max(height, lines * 16)
    return height


def handle_upload_contractors(file):
    workbook = openpyxl.load_workbook(filename=file)
    sheet = workbook.active
    message = check_uids_for_contractors_file(sheet)
    if message:
        return {"message": message}
    result = dict()
    row_count = sheet.max_row + 1
    for each in range(2, row_count):
        some_id = sheet.cell(row=each, column=2).value
        if some_id:
            some_key = str(some_id).strip()
            if some_key in result:
                return {"message": f'Повторный еуол {some_key}'}
            result[some_key] = {
                "name": sheet.cell(row=each, column=1).value,
                "parent": str(sheet.cell(row=each, column=3).value).strip(),
                "id": str(sheet.cell(row=each, column=4).value).strip(),
                "inn": str(sheet.cell(row=each, column=5).value).strip()
            }
    tree_result = dict()
    keys = list(result.keys())
    for each in keys:
        parent_value = result.pop(each, None)
        if parent_value:
            set_children(each, parent_value, result)
            tree_result[each] = {
                'name': parent_value['name'],
                'id': parent_value['id'],
                'inn': parent_value['inn'],
                'children': parent_value['children'],
            }
    return tree_result


def set_children(parent_key, parent_value, result):
        """
        Рекурсивная функция, которая в словарь parent_value записывает подчиненные организации по ключу "children"
        из словаря result. При этом подчиненные организации удаляются из словаря result.
        """
        children = dict(filter(lambda item: item[1]['parent'] == parent_key, result.items()))
        for each in list(children.keys()):
            result.pop(each, None)
        parent_value['children'] = children
        if children:
            for key, value in children.items():
                set_children(key, value, result)
        else:
            return


def check_uids_for_contractors_file(sheet):
    row_count = sheet.max_row + 1
    for each in range(2, row_count):
        some_id = sheet.cell(row=each, column=2).value
        if some_id:
            uid = sheet.cell(row=each, column=4).value
            if not uid:
                return f'row {each}: UID not found.'
    return None


def create_contractors_from_dict(data, parent=None):
    from . import models
    with transaction.atomic():
        for key, value in data.items():
            contractor, created = models.ContractorModel.objects.update_or_create(
                id=value['id'],
                defaults={'name': value['name']},
            )
            if created:
                contractor_member = models.ContractorMemberModel.objects.create(
                    contractor=contractor,
                    name=value['name'],
                    inn=value.get('inn', ''),
                )
            else:
                contractor_member = contractor.contractor_members.all().order_by('-created_at').first()
                if contractor_member:
                    contractor_member.inn = value.get('inn', '')
                    contractor_member.save(update_fields=('inn',))
            if parent and created:
                models.ContractorRelationModel.objects.create(
                    contractor_parent=parent,
                    contractor=contractor,
                    relation_type_id='structural_division',
                )
            children = value.get('children')
            if children:
                create_contractors_from_dict(children, parent=contractor)


def send_email_about_register_request_approved(user_profile):
    from notifications.utils import send_email
    from notifications.models import EmailNotificationModel, EmailNotificationRecipientModel

    with transaction.atomic():
        notification = EmailNotificationModel.objects.create(
            template='register_request_approved',
            subject='Одобрение заявки на регистрацию',
            context={
                "url": f"{FRONTEND_URL}/user/login",
                "company_name": COMPANY_NAME,
                "user_full_name": user_profile.full_name,
            }
        )
        EmailNotificationRecipientModel.objects.create(
            recipient=user_profile.user.email,
            email_notification=notification
        )
        send_email(notification.pk)


def set_balances(file):
    """
    Процедура добавления остатков обращений за предыдущий год.
    Файл должен быть в формате xlsx. Данные должны начинаться со второй строки,
    id организаций должны быть во второй колонке, значение остатков должны быть
    в третей колонке.
    """
    from consolidation.models import ContractorBalanceModel
    from datetime import datetime

    workbook = openpyxl.load_workbook(filename=file)
    sheet = workbook.active
    year = datetime.now().year - 1
    row_count = sheet.max_row + 1
    result = {
        'success': {
            'created': dict(),
            'updated': dict()
        },
        'error': dict()
    }
    for each in range(2, row_count):
        name = sheet.cell(row=each, column=1).value
        contractor_id = sheet.cell(row=each, column=2).value
        balance = sheet.cell(row=each, column=3).value
        try:
            contractor_balance, created = ContractorBalanceModel.objects.get_or_create(
                is_active=True,
                contractor_id=contractor_id,
                defaults={
                    'year': year,
                    'balance': balance
                }
            )
        except Exception as e:
            result['error'][contractor_id] = {
                'class': str(type(e)),
                'message': str(e)
            }
        else:
            if created:
                result['success']['created'][contractor_id] = {
                    'org_name': name,
                    'set_value': balance,
                    'year': year
                }
            else:
                contractor_balance.balance = balance
                contractor_balance.save(update_fields=('balance',))
                result['success']['updated'][contractor_id] = {
                    'org_name': name,
                    'set_value': balance,
                    'year': year
                }
    return result


def get_wkt_point(location_point):
    return f"POINT({location_point[0]} {location_point[1]})"


def get_admin_area_for_point(location_point, admin_level: int = 6):
    """
    Возвращает область, которой принадлежит точка, или None.
    location_point: пара чисел (lon, lat) - координаты точки
    """
    from .models import LocationAdminAreaModel
    admin_area = LocationAdminAreaModel.objects.filter(
        is_active=True,
        geom__contains=get_wkt_point(location_point),
        admin_level=admin_level,
    ).order_by('-created_at').first()
    return admin_area


def check_point_in_area(location_point, area):
    """
    Проверяет, принадлежит ли точка данной области. True если принадлежит. False если не принадлежит.
    location_point: пара чисел (lon, lat) - координаты точки
    """
    polygon = area.polygon
    include = polygon.get('include', [])
    exclude = polygon.get('exclude', [])
    result = False
    for each in include:
        shapely_polygon = Polygon(each)
        shapely_point = Point(location_point)
        if shapely_polygon.contains(shapely_point):
            result = True
            break
    if result:
        for each in exclude:
            shapely_polygon = Polygon(each)
            shapely_point = Point(location_point)
            if shapely_polygon.contains(shapely_point):
                result = False
                break
    return result
