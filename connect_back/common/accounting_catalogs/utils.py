import openpyxl

from django.db import transaction

from rest_framework import exceptions as drf_exceptions

from common.catalogs.models import ContractorModel

from . import models


def set_budget_administrators_from_xlsx(file):
    workbook = openpyxl.load_workbook(filename=file)
    sheet = workbook.active
    row_count = sheet.max_row + 1
    with transaction.atomic():
        for each in range(2, row_count):
            contractor_id = str(sheet.cell(row=each, column=3).value)
            if not contractor_id:
                raise drf_exceptions.ValidationError(f'Id does not exist in row {each}.')
            try:
                contractor = ContractorModel.objects.get(
                    is_active=True,
                    pk=contractor_id,
                )
            except ContractorModel.DoesNotExist:
                raise drf_exceptions.ValidationError(
                    f'Row {each}: contractor {contractor_id} not found.'
                )
            f_group_code = str(sheet.cell(row=each, column=6).value)
            if not f_group_code:
                raise drf_exceptions.ValidationError(f'Functional group does not exist in row {each}.')
            f_subgroup_code = str(sheet.cell(row=each, column=7).value)
            if not f_subgroup_code:
                raise drf_exceptions.ValidationError(f'Functional subgroup does not exist in row {each}.')
            budget_administrator_code = str(sheet.cell(row=each, column=8).value)
            if not budget_administrator_code:
                raise drf_exceptions.ValidationError(f'Budget program administrator does not exist in row {each}.')
            code_gu = str(sheet.cell(row=each, column=9).value)
            if not code_gu:
                raise drf_exceptions.ValidationError(f'Row {each}: code_gu is required.')
            try:
                budget_administrator = models.BudgetProgramAdministratorModel.objects.get(
                    is_active=True,
                    code=budget_administrator_code,
                    functional_subgroup__code=f_subgroup_code,
                    functional_subgroup__functional_group__code=f_group_code,
                )
            except models.BudgetProgramAdministratorModel.DoesNotExist:
                raise drf_exceptions.ValidationError(
                    f'Budget administrator {f_group_code} {f_subgroup_code} {budget_administrator_code} not found.'
                )
            except models.BudgetProgramAdministratorModel.MultipleObjectsReturned:
                raise drf_exceptions.ValidationError(
                    f'Budget administrator {f_group_code} {f_subgroup_code} {budget_administrator_code} is multiple.'
                )
            contractor.budget_program_administrator = budget_administrator
            contractor.code_gu = code_gu
            contractor.save(update_fields=('budget_program_administrator', 'code_gu'),)


def set_kato_codes_from_xlsx(file):
    """Загрузка локаций из справочника КАТО. Выполняется один раз."""
    workbook = openpyxl.load_workbook(filename=file)
    sheet = workbook.active
    rows = tuple(sheet.iter_rows(min_row=2, max_col=9))
    with transaction.atomic():
        old_entries = models.KATOCodesModel.objects.filter(is_active=True)
        old_entries.delete()
        for row in rows:
            models.KATOCodesModel(
                code=row[0].value if row[0].value is not None else '',
                ab=row[1].value if row[1].value is not None else '',
                cd=row[2].value if row[2].value is not None else '',
                ef=row[3].value if row[3].value is not None else '',
                hij=row[4].value if row[4].value is not None else '',
                k=row[5].value if row[5].value is not None else '',
                name_ru=row[6].value if row[6].value is not None else '',
                name_kk=row[7].value if row[7].value is not None else '',
                nn=row[8].value if row[8].value is not None else ''
            ).save()
