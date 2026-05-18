import json
import os

import openpyxl
from rest_framework import exceptions as drf_exceptions

from . import models

APP_NAME = 'proposals'


def cost_classification_parsing():
    xlsx_file_name = os.path.join(APP_NAME, 'cost_classification.xlsx')
    result_file_name = os.path.join(APP_NAME, 'cost_classification.py')

    try:
        workbook = openpyxl.load_workbook(xlsx_file_name)
    except Exception:
        raise drf_exceptions.ValidationError(
            'Не удалось прочитать данные из файла'
        )
    sheet = workbook.worksheets[0]

    cost_classification = dict()

    for row in sheet.iter_rows(min_row=7, max_col=6, values_only=True):
        if row[0]:
            functional_group = row[0]
            cost_classification[functional_group] = dict()
            cost_classification[functional_group]['label'] = row[5]
            cost_classification[functional_group]['items'] = dict()
            continue
        if row[1]:
            functional_subgroup = row[1]
            cost_classification[functional_group]['items'][functional_subgroup] = dict()
            cost_classification[functional_group]['items'][functional_subgroup]['label'] = row[5]
            cost_classification[functional_group]['items'][functional_subgroup]['items'] = dict()
            continue
        if row[2]:
            budget_programme_administrator = row[2]
            cost_classification[functional_group]['items'][functional_subgroup]['items'][budget_programme_administrator] = dict()
            cost_classification[functional_group]['items'][functional_subgroup]['items'][budget_programme_administrator]['label'] = row[5]
            cost_classification[functional_group]['items'][functional_subgroup]['items'][budget_programme_administrator]['items'] = dict()
            continue
        if row[3]:
            programme = row[3]
            cost_classification[functional_group]['items'][functional_subgroup]['items'][budget_programme_administrator]['items'][programme] = dict()
            cost_classification[functional_group]['items'][functional_subgroup]['items'][budget_programme_administrator]['items'][programme]['label'] = row[5]
            cost_classification[functional_group]['items'][functional_subgroup]['items'][budget_programme_administrator]['items'][programme]['items'] = dict()
            continue
        if row[4]:
            subprogramme = row[4]
            cost_classification[functional_group]['items'][functional_subgroup]['items'][budget_programme_administrator]['items'][programme]['items'][subprogramme] = row[5]
            continue

    with open(result_file_name, 'w', encoding='utf-8') as outfile:
        json.dump(cost_classification, outfile, ensure_ascii=False, indent=4)
    return


def specificity_structure_parsing():
    xlsx_file_name = os.path.join(APP_NAME, 'specificity_structure.xlsx')
    result_file_name = os.path.join(APP_NAME, 'specificity_structure.py')

    try:
        workbook = openpyxl.load_workbook(xlsx_file_name)
    except Exception:
        raise drf_exceptions.ValidationError(
            f'Не удалось прочитать данные из файла {xlsx_file_name}'
        )
    sheet = workbook.worksheets[0]
    specificity_structure = dict()

    for row in sheet.iter_rows(min_row=1, min_col=1, max_col=3, values_only=True):
        if row[0]:
            continue
        if row[1] and row[2]:
            specificity_structure[row[1]] = row[2]

    with open(result_file_name, 'w', encoding='utf-8') as outfile:
        json.dump(specificity_structure, outfile, ensure_ascii=False, indent=4)
    return


def cost_classification_to_bd():
    try:
        from accounting_reports.cost_classification import COST_CLASSIFICATION
    except ImportError:
        return

    if not isinstance(COST_CLASSIFICATION, dict):
        return

    for group in COST_CLASSIFICATION.items():
        functional_group, created = models.FunctionalGroupModel.objects.get_or_create(
            code=group[0],
            defaults={
                'name': group[1]['label']
            }
        )
        for f_subgroup in group[1]['items'].items():
            functional_subgroup, created = models.FunctionalSubgroupModel.objects.get_or_create(
                code=f_subgroup[0],
                functional_group=functional_group,
                defaults={
                    'name': f_subgroup[1]['label']
                }
            )
            for bp_administrator in f_subgroup[1]['items'].items():
                budget_program_administrator, created = models.BudgetProgramAdministratorModel.objects.get_or_create(
                    code=bp_administrator[0],
                    functional_subgroup=functional_subgroup,
                    defaults={
                        'name': bp_administrator[1]['label']
                    }
                )
                for prog in bp_administrator[1]['items'].items():
                    program, created = models.ProgramModel.objects.get_or_create(
                        code=prog[0],
                        budget_program_administrator=budget_program_administrator,
                        defaults={
                            'name': prog[1]['label']
                        }
                    )
                    for s_prog in prog[1]['items'].items():
                        subprogram, created = models.SubprogramModel.objects.get_or_create(
                            code=s_prog[0],
                            program=program,
                            defaults={
                                'name': s_prog[1]
                            }
                        )
    return


def specificity_structure_to_bd():
    try:
        from accounting_reports.specificity_structure import SPECIFICITY_STRUCTURE
    except ImportError:
        pass

    if not isinstance(SPECIFICITY_STRUCTURE, dict):
        return

    for item in SPECIFICITY_STRUCTURE.items():
        if item[0] and item[1]:
            obj, created = models.SpecificityStructureModel.objects.get_or_create(
                code=item[0],
                defaults={
                    'name': item[1]
                }
            )
    return
