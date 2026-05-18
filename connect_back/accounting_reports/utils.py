import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, NamedStyle, Side
from rest_framework import exceptions as drf_exceptions

from bkz3.settings import PDF_PARSER_URL
from common.accounting_catalogs.models import (BudgetProgramModel,
                                               BudgetSubprogramModel)
from common.accounting_catalogs.serializers import (
    BudgetProgramModelSerializer, BudgetSubprogramModelSerializer)

from . import models, serializers


def parse_expense_report_from_pdf(file, budget_program_admin):
    resp = requests.post(url=PDF_PARSER_URL, files={'file': file})
    resp_data = resp.json()

    return get_budget_expense_report(resp_data, budget_program_admin)


def get_budget_expense_report(resp_data, budget_program_admin):
    is_initial = False
    budget_program = None
    budget_program_dict = dict()
    budget_subprogram_dict = dict()
    data = []

    for table_key, table_value in resp_data.items():
        subjects = table_value['0']
        for subj_key, subj_value in subjects.items():
            if not is_initial and table_key == '0':
                budget_program_admin_code = subj_value[:3]
                if budget_program_admin_code == budget_program_admin.code:
                    is_initial = True
                continue
            else:
                specific_code = subj_value[:3]
                specific_name = subj_value[3:].strip().replace(' \n', ' ').replace('\n', '')[:8]
                try:
                    specific = models.SpecificityStructureModel.objects.get(code=specific_code, is_active=True)
                except models.SpecificityStructureModel.DoesNotExist:
                    specific = None

                if not specific or not specific.name[:8] == specific_name:
                    try:
                        budget_subprogram = BudgetSubprogramModel.objects.get(
                            program=budget_program,
                            code=specific_code,
                        )
                    except BudgetSubprogramModel.DoesNotExist:
                        budget_subprogram = None
                    if not budget_subprogram or not budget_subprogram.name[:8] == specific_name:
                        try:
                            budget_program = BudgetProgramModel.objects.get(
                                budget_program_administrator=budget_program_admin,
                                code=specific_code
                            )
                        except BudgetProgramModel.DoesNotExist:
                            continue
                        budget_program_dict = BudgetProgramModelSerializer(budget_program).data
                        budget_program_dict['subprograms'] = []
                        data.append(budget_program_dict)
                        continue

                    budget_subprogram_dict = BudgetSubprogramModelSerializer(budget_subprogram).data
                    budget_subprogram_dict['specifics'] = []
                    budget_program_dict['subprograms'].append(budget_subprogram_dict)
                    continue

                specific_dict = serializers.SpecificityStructureModelSerializer(specific).data
                specific_dict['plan_amount'] = resp_data[table_key]['2'][subj_key].replace(',', '')
                specific_dict['actual_amount_accumulated'] = resp_data[table_key]['6'][subj_key].replace(',', '')
                specific_dict['actual_amount'] = resp_data[table_key]['7'][subj_key].replace(',', '')
                budget_subprogram_dict['specifics'].append(specific_dict)
                continue

    return data


def get_upload_for_1C(reports: list) -> Workbook:
    """
    Утилита формирует документ для экспорта данных в 1С.
    """
    TITLE_ROW = [
        'Месторасположение',
        'АдмПрПодпр',
        'Специфика',
        'Госучреждение',
        'ВСЕГО:',
        'ЯНВ',
        'ФЕВ',
        'МАР',
        'АПР',
        'МАЙ',
        'ИЮН',
        'ИЮЛ',
        'АВГ',
        'СЕН',
        'ОКТ',
        'НОЯ',
        'ДЕК'
    ]

    offset = 0
    workbook = Workbook()
    sheet_payments = workbook.active
    sheet_payments.title = "Платежи"
    title_cell_style = NamedStyle(name="title_cell_style")
    value_cell_style = NamedStyle(name="value_cell_style")
    border = Border(
        left=Side(border_style='thin', color='00000000'),
        right=Side(border_style='thin', color='00000000'),
        top=Side(border_style='thin', color='00000000'),
        bottom=Side(border_style='thin', color='00000000'),
    )
    center = Alignment(horizontal='center')
    bold = Font(bold=True)
    title_cell_style.alignment = center
    title_cell_style.border = border
    title_cell_style.font = bold
    value_cell_style.alignment = center
    value_cell_style.border = border

    for report in reports:
        proposals = models.ProposalItemModel.objects.filter(
            is_active=True,
            report_id=report.id
        ).select_related(
            'program',
            'subprogram',
            'specificity',
        )
        if not proposals:
            raise drf_exceptions.ValidationError(
                'Не удалось получить заявки'
            )
        organization = report.organization
        bpa_code = organization.budget_program_administrator.code
        government_agency = f'{bpa_code}{organization.code_gu}'

        for column, value in enumerate(TITLE_ROW, 1):
            sheet_payments.cell(row=1, column=column, value=value).style = title_cell_style

        for row, proposal in enumerate(proposals, 2+offset):
            program_code = proposal.program.code if proposal.program else '000'
            subprogram_code = proposal.subprogram.code if proposal.subprogram else '000'
            specificity_code = proposal.specificity.code if proposal.specificity else '000'
            sheet_payments.cell(row=row, column=1, value=None).style = value_cell_style
            sheet_payments.cell(row=row, column=2, value=f'{bpa_code}{program_code}{subprogram_code}').style = value_cell_style
            sheet_payments.cell(row=row, column=3, value=f'000{specificity_code}').style = value_cell_style
            sheet_payments.cell(row=row, column=4, value=government_agency).style = value_cell_style
            sheet_payments.cell(row=row, column=5, value=proposal.get_sum()).style = value_cell_style
            sheet_payments.cell(row=row, column=6, value=proposal.january).style = value_cell_style
            sheet_payments.cell(row=row, column=7, value=proposal.february).style = value_cell_style
            sheet_payments.cell(row=row, column=8, value=proposal.march).style = value_cell_style
            sheet_payments.cell(row=row, column=9, value=proposal.april).style = value_cell_style
            sheet_payments.cell(row=row, column=10, value=proposal.may).style = value_cell_style
            sheet_payments.cell(row=row, column=11, value=proposal.june).style = value_cell_style
            sheet_payments.cell(row=row, column=12, value=proposal.july).style = value_cell_style
            sheet_payments.cell(row=row, column=13, value=proposal.august).style = value_cell_style
            sheet_payments.cell(row=row, column=14, value=proposal.september).style = value_cell_style
            sheet_payments.cell(row=row, column=15, value=proposal.october).style = value_cell_style
            sheet_payments.cell(row=row, column=16, value=proposal.november).style = value_cell_style
            sheet_payments.cell(row=row, column=17, value=proposal.december).style = value_cell_style
        offset += len(proposals)

    ws2 = workbook.copy_worksheet(sheet_payments)
    ws2.title = "Обязательства"

    return workbook
